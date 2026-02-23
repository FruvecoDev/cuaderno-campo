"""
POC Test Script for Agricultural Field Management Core Functionality
Tests: AI Report Generation + PDF Export + Excel Export

This script tests the three most critical features:
1. AI-powered agricultural report generation from campaign data
2. PDF generation for field notebooks (Cuaderno de Campo)
3. Excel export for treatment/irrigation/cost data
"""

import asyncio
import json
import os
from datetime import datetime, timedelta
from io import BytesIO
from dotenv import load_dotenv

# Load environment variables
load_dotenv('/app/backend/.env')

# Test imports for AI
from emergentintegrations.llm.chat import LlmChat, UserMessage

# Test imports for PDF
from weasyprint import HTML

# Test imports for Excel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ============================================================================
# 1. AI REPORT GENERATION TEST
# ============================================================================

async def test_ai_report_generation():
    """Test AI-powered agricultural report generation"""
    print("\n" + "="*80)
    print("TEST 1: AI-Powered Agricultural Report Generation")
    print("="*80)
    
    try:
        # Get API key
        api_key = os.getenv('EMERGENT_LLM_KEY')
        if not api_key:
            print("❌ ERROR: EMERGENT_LLM_KEY not found in environment")
            return False
            
        print(f"✓ API Key loaded: {api_key[:20]}...")
        
        # Sample campaign data (simulating a real agricultural campaign)
        campaign_data = {
            "contrato": {
                "id": "MP-2026-001",
                "proveedor": "Agrícola López S.L.",
                "cultivo": "Tomate Cherry",
                "variedad": "Kumato",
                "campana": "2025/26",
                "superficie": 2.5,  # hectáreas
                "fecha_inicio": "2025-09-15",
                "fecha_fin": "2026-06-30"
            },
            "parcela": {
                "codigo": "PAR-001",
                "finca": "Finca El Olivo",
                "municipio": "Almería",
                "superficie": 2.5,
                "num_plantas": 15000,
                "sistema_riego": "Goteo"
            },
            "visitas": [
                {
                    "fecha": "2025-10-15",
                    "tipo": "Control Rutinario",
                    "observaciones": "Desarrollo vegetativo normal, sin incidencias"
                },
                {
                    "fecha": "2025-11-20",
                    "tipo": "Plagas y Enfermedades",
                    "observaciones": "Detectada presencia leve de mosca blanca"
                },
                {
                    "fecha": "2026-01-10",
                    "tipo": "Evaluación",
                    "observaciones": "Floración óptima, buen cuajado"
                }
            ],
            "tratamientos": [
                {
                    "fecha": "2025-11-22",
                    "tipo": "Insecticida",
                    "producto": "Confidor 20 LS",
                    "dosis": "0.5 L/ha",
                    "coste": 45.50,
                    "plazo_seguridad": 3
                },
                {
                    "fecha": "2025-12-10",
                    "tipo": "Fungicida",
                    "producto": "Amistar 25 SC",
                    "dosis": "0.8 L/ha",
                    "coste": 62.30,
                    "plazo_seguridad": 7
                }
            ],
            "riegos": [
                {
                    "fecha": "2025-10-01",
                    "volumen": 350,  # m³
                    "duracion": 4,  # horas
                    "coste": 28.50
                },
                {
                    "fecha": "2025-10-15",
                    "volumen": 380,
                    "duracion": 4.5,
                    "coste": 31.00
                },
                {
                    "fecha": "2025-11-01",
                    "volumen": 420,
                    "duracion": 5,
                    "coste": 34.20
                }
            ],
            "cosechas": [
                {
                    "fecha": "2026-03-15",
                    "cantidad": 18500,  # kg
                    "precio": 1.85,  # €/kg
                    "ingreso": 34225.00
                },
                {
                    "fecha": "2026-05-20",
                    "cantidad": 22000,
                    "precio": 1.65,
                    "ingreso": 36300.00
                }
            ],
            "costes_totales": {
                "tratamientos": 214.60,
                "riegos": 287.30,
                "mano_obra": 4500.00,
                "materiales": 1250.00,
                "total": 6251.90
            },
            "ingresos_totales": 70525.00,
            "beneficio": 64273.10
        }
        
        # Create prompt for agricultural analysis
        prompt = f"""Actúa como un agrónomo experto especializado en análisis de campañas agrícolas. 

Analiza los siguientes datos de una campaña agrícola de {campaign_data['contrato']['cultivo']} y genera un INFORME EJECUTIVO completo:

**DATOS DE LA CAMPAÑA:**
```json
{json.dumps(campaign_data, indent=2, ensure_ascii=False)}
```

**GENERA UN INFORME QUE INCLUYA:**

1. **RESUMEN EJECUTIVO** (3-4 líneas clave)
2. **ANÁLISIS DE PRODUCCIÓN**
   - Rendimiento por hectárea (kg/ha)
   - Comparativa con media del sector
   - Calidad y evolución de cosechas
3. **ANÁLISIS DE COSTES**
   - Desglose por categoría
   - Coste/kg producido
   - Identificación de partidas significativas
4. **ANÁLISIS DE TRATAMIENTOS**
   - Cumplimiento de plazos de seguridad
   - Eficacia contra plagas detectadas
   - Recomendaciones de mejora
5. **ANÁLISIS DE RIEGOS**
   - Consumo de agua (m³/ha)
   - Eficiencia hídrica
   - Recomendaciones de optimización
6. **RENTABILIDAD**
   - Margen bruto (€/ha)
   - ROI de la campaña
   - Comparativa con objetivos
7. **ALERTAS Y ANOMALÍAS** (si las hay)
8. **RECOMENDACIONES PARA PRÓXIMA CAMPAÑA**

**FORMATO:** Usa markdown con secciones claras. Sé preciso con los números. Incluye cálculos específicos."""

        print("\n📤 Enviando datos a IA para análisis...")
        print(f"   Datos: {len(json.dumps(campaign_data))} caracteres")
        print(f"   Cultivo: {campaign_data['contrato']['cultivo']}")
        print(f"   Superficie: {campaign_data['parcela']['superficie']} ha")
        
        # Initialize AI chat
        chat = LlmChat(
            api_key=api_key,
            session_id="poc-agricultural-report",
            system_message="Eres un agrónomo experto en análisis de datos agrícolas. Generas informes precisos, estructurados y accionables."
        )
        chat.with_model("openai", "gpt-5.2")
        
        # Send message
        user_message = UserMessage(text=prompt)
        response = await chat.send_message(user_message)
        
        print("\n✅ Informe IA generado exitosamente!")
        print("\n" + "-"*80)
        print("INFORME GENERADO:")
        print("-"*80)
        print(response[:1000] + "..." if len(response) > 1000 else response)
        print("-"*80)
        
        # Save report
        report_path = "/app/tests/poc_ai_report.md"
        with open(report_path, 'w', encoding='utf-8') as f:
            f.write(f"# Informe de Campaña - {campaign_data['contrato']['id']}\n\n")
            f.write(f"**Generado:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
            f.write(response)
        
        print(f"\n💾 Informe guardado en: {report_path}")
        print("\n✅ TEST 1 PASSED: AI report generation working correctly!")
        return True
        
    except Exception as e:
        print(f"\n❌ TEST 1 FAILED: {str(e)}")
        import traceback
        traceback.print_exc()
        return False


# ============================================================================
# 2. PDF GENERATION TEST
# ============================================================================

def test_pdf_generation():
    """Test PDF generation from HTML template"""
    print("\n" + "="*80)
    print("TEST 2: PDF Generation for Field Notebook (Cuaderno de Campo)")
    print("="*80)
    
    try:
        # Sample field notebook data
        notebook_data = {
            "contrato": "MP-2026-001",
            "proveedor": "Agrícola López S.L.",
            "cultivo": "Tomate Cherry",
            "variedad": "Kumato",
            "parcela": "PAR-001",
            "finca": "Finca El Olivo",
            "superficie": "2.5 ha",
            "fecha_generacion": datetime.now().strftime("%d/%m/%Y %H:%M"),
            "tratamientos": [
                {
                    "fecha": "22/11/2025",
                    "tipo": "Insecticida",
                    "producto": "Confidor 20 LS",
                    "dosis": "0.5 L/ha",
                    "justificacion": "Control de mosca blanca",
                    "plazo_seguridad": "3 días"
                },
                {
                    "fecha": "10/12/2025",
                    "tipo": "Fungicida",
                    "producto": "Amistar 25 SC",
                    "dosis": "0.8 L/ha",
                    "justificacion": "Prevención mildiu",
                    "plazo_seguridad": "7 días"
                }
            ],
            "riegos": [
                {
                    "fecha": "01/10/2025",
                    "volumen": "350 m³",
                    "duracion": "4 horas"
                },
                {
                    "fecha": "15/10/2025",
                    "volumen": "380 m³",
                    "duracion": "4.5 horas"
                }
            ],
            "cosechas": [
                {
                    "fecha": "15/03/2026",
                    "cantidad": "18,500 kg",
                    "calidad": "Extra"
                },
                {
                    "fecha": "20/05/2026",
                    "cantidad": "22,000 kg",
                    "calidad": "Primera"
                }
            ]
        }
        
        # Create HTML template
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <style>
                @page {{
                    size: A4;
                    margin: 2cm;
                }}
                body {{
                    font-family: Arial, sans-serif;
                    font-size: 11pt;
                    color: #333;
                }}
                h1 {{
                    color: #2c5f2d;
                    border-bottom: 3px solid #97bf0d;
                    padding-bottom: 10px;
                    font-size: 24pt;
                }}
                h2 {{
                    color: #2c5f2d;
                    margin-top: 20px;
                    font-size: 16pt;
                    border-bottom: 1px solid #ddd;
                    padding-bottom: 5px;
                }}
                .header-info {{
                    background-color: #f0f8f0;
                    padding: 15px;
                    border-radius: 5px;
                    margin-bottom: 20px;
                }}
                .info-row {{
                    display: flex;
                    margin-bottom: 5px;
                }}
                .info-label {{
                    font-weight: bold;
                    width: 150px;
                }}
                table {{
                    width: 100%;
                    border-collapse: collapse;
                    margin-top: 10px;
                    margin-bottom: 20px;
                }}
                th {{
                    background-color: #2c5f2d;
                    color: white;
                    padding: 10px;
                    text-align: left;
                    font-size: 10pt;
                }}
                td {{
                    border: 1px solid #ddd;
                    padding: 8px;
                    font-size: 10pt;
                }}
                tr:nth-child(even) {{
                    background-color: #f9f9f9;
                }}
                .footer {{
                    margin-top: 30px;
                    padding-top: 10px;
                    border-top: 1px solid #ddd;
                    font-size: 9pt;
                    color: #666;
                    text-align: center;
                }}
            </style>
        </head>
        <body>
            <h1>📋 CUADERNO DE CAMPO</h1>
            
            <div class="header-info">
                <div class="info-row">
                    <div class="info-label">Contrato:</div>
                    <div>{notebook_data['contrato']}</div>
                </div>
                <div class="info-row">
                    <div class="info-label">Proveedor:</div>
                    <div>{notebook_data['proveedor']}</div>
                </div>
                <div class="info-row">
                    <div class="info-label">Cultivo:</div>
                    <div>{notebook_data['cultivo']} - {notebook_data['variedad']}</div>
                </div>
                <div class="info-row">
                    <div class="info-label">Parcela:</div>
                    <div>{notebook_data['parcela']} ({notebook_data['finca']})</div>
                </div>
                <div class="info-row">
                    <div class="info-label">Superficie:</div>
                    <div>{notebook_data['superficie']}</div>
                </div>
            </div>
            
            <h2>🌿 Tratamientos Fitosanitarios</h2>
            <table>
                <thead>
                    <tr>
                        <th>Fecha</th>
                        <th>Tipo</th>
                        <th>Producto</th>
                        <th>Dosis</th>
                        <th>Justificación</th>
                        <th>P.S.</th>
                    </tr>
                </thead>
                <tbody>
                    {''.join(f'''
                    <tr>
                        <td>{t['fecha']}</td>
                        <td>{t['tipo']}</td>
                        <td>{t['producto']}</td>
                        <td>{t['dosis']}</td>
                        <td>{t['justificacion']}</td>
                        <td>{t['plazo_seguridad']}</td>
                    </tr>
                    ''' for t in notebook_data['tratamientos'])}
                </tbody>
            </table>
            
            <h2>💧 Irrigaciones</h2>
            <table>
                <thead>
                    <tr>
                        <th>Fecha</th>
                        <th>Volumen</th>
                        <th>Duración</th>
                    </tr>
                </thead>
                <tbody>
                    {''.join(f'''
                    <tr>
                        <td>{r['fecha']}</td>
                        <td>{r['volumen']}</td>
                        <td>{r['duracion']}</td>
                    </tr>
                    ''' for r in notebook_data['riegos'])}
                </tbody>
            </table>
            
            <h2>🌾 Cosechas</h2>
            <table>
                <thead>
                    <tr>
                        <th>Fecha</th>
                        <th>Cantidad</th>
                        <th>Calidad</th>
                    </tr>
                </thead>
                <tbody>
                    {''.join(f'''
                    <tr>
                        <td>{c['fecha']}</td>
                        <td>{c['cantidad']}</td>
                        <td>{c['calidad']}</td>
                    </tr>
                    ''' for c in notebook_data['cosechas'])}
                </tbody>
            </table>
            
            <div class="footer">
                Documento generado el {notebook_data['fecha_generacion']}<br>
                Sistema de Gestión Agrícola - Cuaderno de Campo Digital
            </div>
        </body>
        </html>
        """
        
        print("\n📄 Generando PDF desde plantilla HTML...")
        
        # Generate PDF
        html = HTML(string=html_content)
        pdf_bytes = html.write_pdf()
        
        # Save PDF
        pdf_path = "/app/tests/poc_cuaderno_campo.pdf"
        with open(pdf_path, 'wb') as f:
            f.write(pdf_bytes)
        
        pdf_size = len(pdf_bytes) / 1024  # KB
        print(f"✅ PDF generado: {pdf_size:.1f} KB")
        print(f"💾 PDF guardado en: {pdf_path}")
        print("\n✅ TEST 2 PASSED: PDF generation working correctly!")
        return True
        
    except Exception as e:
        print(f"\n❌ TEST 2 FAILED: {str(e)}")
        import traceback
        traceback.print_exc()
        return False


# ============================================================================
# 3. EXCEL GENERATION TEST
# ============================================================================

def test_excel_generation():
    """Test Excel generation with multiple sheets"""
    print("\n" + "="*80)
    print("TEST 3: Excel Export for Agricultural Data")
    print("="*80)
    
    try:
        # Create workbook
        wb = Workbook()
        
        # -------------------- SHEET 1: TREATMENTS --------------------
        ws_treatments = wb.active
        ws_treatments.title = "Tratamientos"
        
        # Headers
        headers = ["Fecha", "Tipo", "Producto", "Dosis", "Superficie (ha)", "Coste (€)", "P.S. (días)"]
        ws_treatments.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="2C5F2D", end_color="2C5F2D", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for col in range(1, len(headers) + 1):
            cell = ws_treatments.cell(1, col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Sample data
        treatments = [
            ["22/11/2025", "Insecticida", "Confidor 20 LS", "0.5 L/ha", 2.5, 45.50, 3],
            ["10/12/2025", "Fungicida", "Amistar 25 SC", "0.8 L/ha", 2.5, 62.30, 7],
            ["15/01/2026", "Fertilizante", "NPK 15-15-15", "300 kg/ha", 2.5, 187.50, 0],
            ["20/02/2026", "Herbicida", "Glifosato 36%", "3 L/ha", 2.5, 28.80, 14]
        ]
        
        for row in treatments:
            ws_treatments.append(row)
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            ws_treatments.column_dimensions[get_column_letter(col)].width = 15
        
        print("✓ Hoja 'Tratamientos' creada con 4 registros")
        
        # -------------------- SHEET 2: IRRIGATIONS --------------------
        ws_irrigations = wb.create_sheet("Irrigaciones")
        
        headers = ["Fecha", "Sistema", "Duración (h)", "Volumen (m³)", "Coste (€)", "Observaciones"]
        ws_irrigations.append(headers)
        
        # Style headers
        for col in range(1, len(headers) + 1):
            cell = ws_irrigations.cell(1, col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        irrigations = [
            ["01/10/2025", "Goteo", 4.0, 350, 28.50, "Riego estándar"],
            ["15/10/2025", "Goteo", 4.5, 380, 31.00, "Riego estándar"],
            ["01/11/2025", "Goteo", 5.0, 420, 34.20, "Incremento por calor"],
            ["15/11/2025", "Goteo", 4.0, 350, 28.50, "Riego estándar"]
        ]
        
        for row in irrigations:
            ws_irrigations.append(row)
        
        for col in range(1, len(headers) + 1):
            ws_irrigations.column_dimensions[get_column_letter(col)].width = 15
        
        print("✓ Hoja 'Irrigaciones' creada con 4 registros")
        
        # -------------------- SHEET 3: COSTS --------------------
        ws_costs = wb.create_sheet("Costes")
        
        headers = ["Categoría", "Concepto", "Cantidad", "Unidad", "Precio Unit. (€)", "Total (€)"]
        ws_costs.append(headers)
        
        # Style headers
        for col in range(1, len(headers) + 1):
            cell = ws_costs.cell(1, col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        costs = [
            ["Tratamientos", "Fitosanitarios", 4, "aplicaciones", 53.65, 214.60],
            ["Riegos", "Agua y electricidad", 12, "riegos", 23.94, 287.30],
            ["Mano de obra", "Jornales", 120, "horas", 37.50, 4500.00],
            ["Materiales", "Varios", 1, "lote", 1250.00, 1250.00],
            ["", "", "", "", "TOTAL:", 6251.90]
        ]
        
        for row in costs:
            ws_costs.append(row)
        
        # Bold and highlight total
        total_row = len(costs) + 1
        ws_costs.cell(total_row, 5).font = Font(bold=True)
        ws_costs.cell(total_row, 6).font = Font(bold=True)
        ws_costs.cell(total_row, 6).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        for col in range(1, len(headers) + 1):
            ws_costs.column_dimensions[get_column_letter(col)].width = 18
        
        print("✓ Hoja 'Costes' creada con resumen financiero")
        
        # -------------------- SHEET 4: HARVESTS --------------------
        ws_harvests = wb.create_sheet("Cosechas")
        
        headers = ["Fecha", "Cantidad (kg)", "Calidad", "Precio (€/kg)", "Ingreso (€)", "Lote"]
        ws_harvests.append(headers)
        
        # Style headers
        for col in range(1, len(headers) + 1):
            cell = ws_harvests.cell(1, col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        harvests = [
            ["15/03/2026", 18500, "Extra", 1.85, 34225.00, "LOT-001"],
            ["20/05/2026", 22000, "Primera", 1.65, 36300.00, "LOT-002"],
            ["", "", "", "", 70525.00, "TOTAL"]
        ]
        
        for row in harvests:
            ws_harvests.append(row)
        
        # Highlight total
        total_row = len(harvests) + 1
        ws_harvests.cell(total_row, 5).font = Font(bold=True)
        ws_harvests.cell(total_row, 6).font = Font(bold=True)
        ws_harvests.cell(total_row, 5).fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        
        for col in range(1, len(headers) + 1):
            ws_harvests.column_dimensions[get_column_letter(col)].width = 15
        
        print("✓ Hoja 'Cosechas' creada con ingresos")
        
        # Save Excel file
        excel_path = "/app/tests/poc_datos_agricolas.xlsx"
        wb.save(excel_path)
        
        print(f"\n💾 Excel guardado en: {excel_path}")
        print("✅ TEST 3 PASSED: Excel generation working correctly!")
        return True
        
    except Exception as e:
        print(f"\n❌ TEST 3 FAILED: {str(e)}")
        import traceback
        traceback.print_exc()
        return False


# ============================================================================
# MAIN POC TEST RUNNER
# ============================================================================

async def run_all_tests():
    """Run all POC tests"""
    print("\n" + "="*80)
    print(" 🚀 AGRICULTURAL FIELD MANAGEMENT - CORE POC TESTS")
    print("="*80)
    print("\nTesting 3 critical features:")
    print("  1. AI-powered report generation")
    print("  2. PDF field notebook generation")
    print("  3. Excel data export")
    print("\n" + "="*80)
    
    results = []
    
    # Test 1: AI Report Generation
    result1 = await test_ai_report_generation()
    results.append(("AI Report Generation", result1))
    
    # Test 2: PDF Generation
    result2 = test_pdf_generation()
    results.append(("PDF Generation", result2))
    
    # Test 3: Excel Generation
    result3 = test_excel_generation()
    results.append(("Excel Export", result3))
    
    # Summary
    print("\n" + "="*80)
    print(" 📊 POC TEST SUMMARY")
    print("="*80)
    
    for test_name, passed in results:
        status = "✅ PASSED" if passed else "❌ FAILED"
        print(f"  {status} - {test_name}")
    
    all_passed = all(result for _, result in results)
    
    print("\n" + "="*80)
    if all_passed:
        print(" 🎉 ALL POC TESTS PASSED!")
        print(" Core functionality verified. Ready to build full application.")
    else:
        print(" ⚠️  SOME TESTS FAILED")
        print(" Please review errors above before proceeding.")
    print("="*80 + "\n")
    
    return all_passed


if __name__ == "__main__":
    success = asyncio.run(run_all_tests())
    exit(0 if success else 1)
