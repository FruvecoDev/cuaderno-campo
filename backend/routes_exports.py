from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import List, Optional
from datetime import datetime
import io

from database import (
    db, serialize_docs,
    contratos_collection, parcelas_collection, fincas_collection,
    visitas_collection, tareas_collection, cosechas_collection,
    tratamientos_collection, irrigaciones_collection, recetas_collection,
    albaranes_collection, maquinaria_collection, evaluaciones_collection
)
from rbac_guards import get_current_user

router = APIRouter(prefix="/api", tags=["exports"])

tecnicos_aplicadores_collection = db['tecnicos_aplicadores']

MODULES_CONFIG = {
    "contratos": {
        "label": "Contratos",
        "collection": "contratos",
        "sort_field": "created_at",
        "columns": ["numero_contrato", "tipo", "proveedor", "cultivo", "variedad", "campana", "estado", "superficie_total", "precio"],
        "headers": ["N Contrato", "Tipo", "Proveedor", "Cultivo", "Variedad", "Campana", "Estado", "Superficie", "Precio"],
    },
    "parcelas": {
        "label": "Parcelas",
        "collection": "parcelas",
        "sort_field": "created_at",
        "columns": ["codigo_plantacion", "proveedor", "finca", "cultivo", "variedad", "campana", "superficie_total", "estado"],
        "headers": ["Codigo", "Proveedor", "Finca", "Cultivo", "Variedad", "Campana", "Superficie", "Estado"],
    },
    "fincas": {
        "label": "Fincas",
        "collection": "fincas",
        "sort_field": "nombre",
        "columns": ["nombre", "provincia", "municipio", "superficie_total", "tipo_suelo", "sistema_riego"],
        "headers": ["Nombre", "Provincia", "Municipio", "Superficie", "Tipo Suelo", "Sistema Riego"],
    },
    "visitas": {
        "label": "Visitas",
        "collection": "visitas",
        "sort_field": "fecha_visita",
        "columns": ["fecha_visita", "objetivo", "proveedor", "cultivo", "finca", "campana", "realizado"],
        "headers": ["Fecha", "Objetivo", "Proveedor", "Cultivo", "Finca", "Campana", "Realizado"],
    },
    "tareas": {
        "label": "Tareas",
        "collection": "tareas",
        "sort_field": "created_at",
        "columns": ["titulo", "tipo", "prioridad", "estado", "fecha_limite", "asignado_a", "parcela_nombre"],
        "headers": ["Titulo", "Tipo", "Prioridad", "Estado", "Fecha Limite", "Asignado", "Parcela"],
    },
    "cosechas": {
        "label": "Cosechas",
        "collection": "cosechas",
        "sort_field": "created_at",
        "columns": ["nombre", "cultivo", "variedad", "campana", "cosecha_total", "estado"],
        "headers": ["Nombre", "Cultivo", "Variedad", "Campana", "Total Kg", "Estado"],
    },
    "tratamientos": {
        "label": "Tratamientos",
        "collection": "tratamientos",
        "sort_field": "fecha_tratamiento",
        "columns": ["tipo", "fecha_tratamiento", "fecha_aplicacion", "aplicador_nombre", "campana", "coste_total", "realizado"],
        "headers": ["Tipo", "Fecha Trat.", "Fecha Aplic.", "Aplicador", "Campana", "Coste", "Realizado"],
    },
    "irrigaciones": {
        "label": "Irrigaciones",
        "collection": "irrigaciones",
        "sort_field": "fecha",
        "columns": ["fecha", "sistema", "volumen", "duracion", "parcela_nombre", "campana"],
        "headers": ["Fecha", "Sistema", "Volumen m3", "Duracion h", "Parcela", "Campana"],
    },
    "recetas": {
        "label": "Recetas",
        "collection": "recetas",
        "sort_field": "created_at",
        "columns": ["numero_receta", "tipo", "cultivo", "estado", "fecha_prescripcion", "tecnico"],
        "headers": ["N Receta", "Tipo", "Cultivo", "Estado", "Fecha", "Tecnico"],
    },
    "albaranes": {
        "label": "Albaranes",
        "collection": "albaranes",
        "sort_field": "created_at",
        "columns": ["numero_albaran", "tipo", "proveedor", "fecha", "estado", "total"],
        "headers": ["N Albaran", "Tipo", "Proveedor", "Fecha", "Estado", "Total"],
    },
    "evaluaciones": {
        "label": "Evaluaciones",
        "collection": "evaluaciones",
        "sort_field": "created_at",
        "columns": ["codigo_plantacion", "cultivo", "proveedor", "campana", "estado", "tecnico"],
        "headers": ["Codigo", "Cultivo", "Proveedor", "Campana", "Estado", "Tecnico"],
    },
    "tecnicos_aplicadores": {
        "label": "Tecnicos Aplicadores",
        "collection": "tecnicos_aplicadores",
        "sort_field": "nombre",
        "columns": ["nombre", "apellidos", "dni", "nivel_capacitacion", "num_carnet", "activo"],
        "headers": ["Nombre", "Apellidos", "DNI", "Nivel", "N Carnet", "Activo"],
    },
    "maquinaria": {
        "label": "Maquinaria",
        "collection": "maquinaria",
        "sort_field": "nombre",
        "columns": ["nombre", "tipo", "marca", "modelo", "matricula", "estado"],
        "headers": ["Nombre", "Tipo", "Marca", "Modelo", "Matricula", "Estado"],
    },
}


class CombinedExportRequest(BaseModel):
    modules: List[str]
    format: str = "excel"


@router.get("/exports/modules")
async def get_export_modules(current_user: dict = Depends(get_current_user)):
    """Get available modules for export with record counts"""
    result = []
    for key, config in MODULES_CONFIG.items():
        coll = db[config["collection"]]
        count = await coll.count_documents({})
        result.append({
            "key": key,
            "label": config["label"],
            "count": count
        })
    return {"modules": result}


@router.post("/exports/combined")
async def export_combined(
    request: CombinedExportRequest,
    current_user: dict = Depends(get_current_user)
):
    """Generate a combined export with multiple modules"""
    valid_modules = [m for m in request.modules if m in MODULES_CONFIG]
    if not valid_modules:
        from fastapi import HTTPException
        raise HTTPException(status_code=400, detail="No valid modules selected")

    if request.format == "excel":
        return await _generate_combined_excel(valid_modules)
    else:
        return await _generate_combined_pdf(valid_modules)


async def _generate_combined_excel(modules: List[str]):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    wb.remove(wb.active)

    colors_map = {
        "contratos": "1B5E20", "parcelas": "2E7D32", "fincas": "388E3C",
        "visitas": "1A5276", "tareas": "0D47A1", "cosechas": "1E8449",
        "tratamientos": "E65100", "irrigaciones": "2874A6", "recetas": "4A148C",
        "albaranes": "795548", "evaluaciones": "6A1B9A", "tecnicos_aplicadores": "BF360C",
        "maquinaria": "37474F",
    }

    for mod_key in modules:
        config = MODULES_CONFIG[mod_key]
        coll = db[config["collection"]]
        docs = await coll.find({}).sort(config["sort_field"], -1).to_list(5000)
        data = serialize_docs(docs)

        ws = wb.create_sheet(title=config["label"][:31])
        color = colors_map.get(mod_key, "333333")
        header_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for col_idx, h in enumerate(config["headers"], 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        for row_idx, doc in enumerate(data, 2):
            for col_idx, col_key in enumerate(config["columns"], 1):
                val = doc.get(col_key, "")
                if isinstance(val, bool):
                    val = "Si" if val else "No"
                elif val is None:
                    val = ""
                cell = ws.cell(row=row_idx, column=col_idx, value=str(val)[:100])
                cell.border = thin_border

        for col_idx in range(1, len(config["headers"]) + 1):
            ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else 'A'].width = 18

    # Add summary sheet at the beginning
    summary = wb.create_sheet(title="Resumen", index=0)
    summary_fill = PatternFill(start_color="2D5A27", end_color="2D5A27", fill_type="solid")
    summary_font = Font(bold=True, color="FFFFFF", size=11)

    summary.cell(row=1, column=1, value="FRUVECO - Informe Combinado").font = Font(bold=True, size=16, color="2D5A27")
    summary.cell(row=2, column=1, value=f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}").font = Font(size=10, color="666666")
    summary.cell(row=4, column=1, value="Modulo").font = summary_font
    summary.cell(row=4, column=1).fill = summary_fill
    summary.cell(row=4, column=2, value="Registros").font = summary_font
    summary.cell(row=4, column=2).fill = summary_fill

    total_records = 0
    for idx, mod_key in enumerate(modules, 5):
        config = MODULES_CONFIG[mod_key]
        count = await db[config["collection"]].count_documents({})
        summary.cell(row=idx, column=1, value=config["label"])
        summary.cell(row=idx, column=2, value=count)
        total_records += count

    summary.cell(row=5 + len(modules) + 1, column=1, value="TOTAL").font = Font(bold=True, size=11)
    summary.cell(row=5 + len(modules) + 1, column=2, value=total_records).font = Font(bold=True, size=11)
    summary.column_dimensions['A'].width = 25
    summary.column_dimensions['B'].width = 15

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"fruveco_informe_combinado_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


async def _generate_combined_pdf(modules: List[str]):
    from reportlab.lib import colors as rl_colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak

    output = io.BytesIO()
    pdf = SimpleDocTemplate(output, pagesize=landscape(A4), topMargin=12*mm, bottomMargin=12*mm)
    elements = []
    styles = getSampleStyleSheet()

    # Cover page
    title_style = ParagraphStyle('CoverTitle', parent=styles['Heading1'], fontSize=24, textColor=rl_colors.HexColor('#2D5A27'), alignment=1, spaceAfter=5*mm)
    subtitle_style = ParagraphStyle('CoverSub', parent=styles['Normal'], fontSize=12, textColor=rl_colors.gray, alignment=1, spaceAfter=3*mm)
    section_style = ParagraphStyle('Section', parent=styles['Heading2'], fontSize=16, textColor=rl_colors.HexColor('#2D5A27'), spaceAfter=4*mm, spaceBefore=2*mm)

    elements.append(Spacer(1, 30*mm))
    elements.append(Paragraph("FRUVECO", title_style))
    elements.append(Paragraph("Informe Combinado - Cuaderno de Campo", subtitle_style))
    elements.append(Paragraph(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", subtitle_style))
    elements.append(Spacer(1, 15*mm))

    # Summary table
    summary_data = [["Modulo", "Registros"]]
    total = 0
    for mod_key in modules:
        config = MODULES_CONFIG[mod_key]
        count = await db[config["collection"]].count_documents({})
        summary_data.append([config["label"], str(count)])
        total += count
    summary_data.append(["TOTAL", str(total)])

    summary_table = Table(summary_data, colWidths=[80*mm, 40*mm])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), rl_colors.HexColor('#2D5A27')),
        ('TEXTCOLOR', (0, 0), (-1, 0), rl_colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, rl_colors.gray),
        ('BACKGROUND', (0, -1), (-1, -1), rl_colors.HexColor('#E8F5E9')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [rl_colors.white, rl_colors.HexColor('#F5F5F5')]),
        ('ALIGN', (1, 0), (1, -1), 'CENTER'),
    ]))
    elements.append(summary_table)

    colors_map = {
        "contratos": '#1B5E20', "parcelas": '#2E7D32', "fincas": '#388E3C',
        "visitas": '#1A5276', "tareas": '#0D47A1', "cosechas": '#1E8449',
        "tratamientos": '#E65100', "irrigaciones": '#2874A6', "recetas": '#4A148C',
        "albaranes": '#795548', "evaluaciones": '#6A1B9A', "tecnicos_aplicadores": '#BF360C',
        "maquinaria": '#37474F',
    }

    for mod_key in modules:
        config = MODULES_CONFIG[mod_key]
        coll = db[config["collection"]]
        docs = await coll.find({}).sort(config["sort_field"], -1).to_list(5000)
        data = serialize_docs(docs)

        elements.append(PageBreak())
        color = rl_colors.HexColor(colors_map.get(mod_key, '#333333'))
        mod_title = ParagraphStyle('ModTitle', parent=styles['Heading2'], fontSize=16, textColor=color, spaceAfter=4*mm)
        elements.append(Paragraph(config["label"].upper(), mod_title))
        elements.append(Paragraph(f"{len(data)} registros", ParagraphStyle('ModSub', parent=styles['Normal'], textColor=rl_colors.gray, spaceAfter=3*mm)))

        table_data = [config["headers"]]
        for doc in data[:200]:
            row = []
            for col_key in config["columns"]:
                val = doc.get(col_key, "")
                if isinstance(val, bool):
                    val = "Si" if val else "No"
                elif val is None:
                    val = ""
                row.append(str(val)[:25])
            table_data.append(row)

        n_cols = len(config["headers"])
        available_width = 267 * mm
        col_w = available_width / n_cols
        col_widths = [col_w] * n_cols

        doc_table = Table(table_data, colWidths=col_widths, repeatRows=1)
        doc_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), color),
            ('TEXTCOLOR', (0, 0), (-1, 0), rl_colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('GRID', (0, 0), (-1, -1), 0.5, rl_colors.HexColor('#CCCCCC')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [rl_colors.white, rl_colors.HexColor('#F5F5F5')]),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(doc_table)

    pdf.build(elements)
    output.seek(0)
    filename = f"fruveco_informe_combinado_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
