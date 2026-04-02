"""
Routes for Visitas (Visits) - CRUD operations
Extracted from routes_main.py for better code organization
Uses simplified model: only parcela_id required, rest inherited
"""

from fastapi import APIRouter, HTTPException, Depends
from typing import Optional
from bson import ObjectId
from datetime import datetime, timedelta

from models import VisitaCreate
from database import (
    visitas_collection, parcelas_collection, contratos_collection,
    serialize_doc, serialize_docs
)
from rbac_guards import (
    RequireCreate, RequireEdit, RequireDelete,
    RequireVisitasAccess, get_current_user
)

router = APIRouter(prefix="/api", tags=["visitas"])


@router.post("/visitas", response_model=dict)
async def create_visita(
    visita: VisitaCreate,
    current_user: dict = Depends(RequireCreate),
    _access: dict = Depends(RequireVisitasAccess)
):
    # MODELO SIMPLIFICADO: parcela_id es obligatorio, el resto se hereda
    if not visita.parcela_id:
        raise HTTPException(status_code=400, detail="parcela_id es obligatorio")
    
    if not ObjectId.is_valid(visita.parcela_id):
        raise HTTPException(status_code=400, detail="parcela_id inválido")
    
    # Obtener la parcela para heredar datos
    parcela = await parcelas_collection.find_one({"_id": ObjectId(visita.parcela_id)})
    if not parcela:
        raise HTTPException(status_code=400, detail="Parcela no encontrada")
    
    # Heredar datos desde la parcela
    contrato_id = parcela.get("contrato_id", "")
    proveedor = parcela.get("proveedor", "")
    cultivo = parcela.get("cultivo", "")
    campana = parcela.get("campana", "")
    variedad = parcela.get("variedad", "")
    codigo_plantacion = parcela.get("codigo_plantacion", "")
    finca = parcela.get("finca", "")
    
    # Si la parcela tiene contrato_id, buscar cultivo_id desde el contrato
    cultivo_id = ""
    if contrato_id:
        contrato = await contratos_collection.find_one({"_id": ObjectId(contrato_id)})
        if contrato:
            cultivo_id = contrato.get("cultivo_id", "")
            # Actualizar datos heredados si el contrato tiene más información
            if not proveedor:
                proveedor = contrato.get("proveedor", "")
            if not cultivo:
                cultivo = contrato.get("cultivo", "")
            if not campana:
                campana = contrato.get("campana", "")
    
    visita_dict = {
        "objetivo": visita.objetivo,
        "parcela_id": visita.parcela_id,
        "contrato_id": contrato_id,
        "cultivo_id": cultivo_id,
        "proveedor": proveedor,
        "cultivo": cultivo,
        "campana": campana,
        "variedad": variedad,
        "codigo_plantacion": codigo_plantacion,
        "finca": finca,
        "fecha_visita": visita.fecha_visita or "",
        "fecha_planificada": visita.fecha_planificada or "",
        "observaciones": visita.observaciones or "",
        "realizado": False,
        "planificado": bool(visita.fecha_planificada),
        "documentos": [],
        "formularios": [],
        "cuestionario_plagas": visita.cuestionario_plagas if visita.objetivo == "Plagas y Enfermedades" else None,
        "created_at": datetime.now(),
        "updated_at": datetime.now()
    }
    
    result = await visitas_collection.insert_one(visita_dict)
    created = await visitas_collection.find_one({"_id": result.inserted_id})
    
    return {"success": True, "data": serialize_doc(created)}


@router.get("/visitas")
async def get_visitas(
    skip: int = 0,
    limit: int = 100,
    parcela_id: Optional[str] = None,
    campana: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
    _access: dict = Depends(RequireVisitasAccess)
):
    query = {}
    if parcela_id:
        query["parcela_id"] = parcela_id
    if campana:
        query["campana"] = campana
    
    visitas = await visitas_collection.find(query).skip(skip).limit(limit).to_list(limit)
    return {"visitas": serialize_docs(visitas), "total": await visitas_collection.count_documents(query)}


@router.get("/visitas/planificadas")
async def get_visitas_planificadas(
    current_user: dict = Depends(get_current_user),
    _access: dict = Depends(RequireVisitasAccess)
):
    """Obtener visitas planificadas (con fecha_visita o fecha_planificada futura o reciente)"""
    # Buscar visitas con fecha en los próximos 30 días o los últimos 7 días
    hoy = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    hace_7_dias = hoy - timedelta(days=7)
    en_30_dias = hoy + timedelta(days=30)
    
    hace_7_dias_str = hace_7_dias.strftime("%Y-%m-%d")
    en_30_dias_str = en_30_dias.strftime("%Y-%m-%d")
    
    query = {
        "$or": [
            # Visitas con fecha_planificada en el rango
            {"fecha_planificada": {"$gte": hace_7_dias_str, "$lte": en_30_dias_str}},
            # Visitas con fecha_visita en el rango (próximas visitas)
            {"fecha_visita": {"$gte": hace_7_dias_str, "$lte": en_30_dias_str}},
            # Visitas marcadas como planificadas no realizadas
            {"planificado": True, "realizado": False}
        ]
    }
    
    visitas = await visitas_collection.find(query).sort("fecha_visita", 1).to_list(50)
    
    # Procesar visitas y usar fecha_visita si no hay fecha_planificada
    visitas_filtradas = []
    for v in visitas:
        # Usar fecha_visita como fecha principal si no hay fecha_planificada
        fecha = v.get("fecha_planificada") or v.get("fecha_visita")
        if fecha:
            v["fecha_planificada"] = fecha  # Para compatibilidad con el frontend
            v["parcela"] = v.get("codigo_plantacion", "")
            visitas_filtradas.append(v)
    
    # Ordenar por fecha
    visitas_filtradas.sort(key=lambda x: x.get("fecha_planificada", ""))
    
    return {"visitas": serialize_docs(visitas_filtradas)}


@router.get("/visitas/{visita_id}")
async def get_visita(
    visita_id: str,
    current_user: dict = Depends(get_current_user),
    _access: dict = Depends(RequireVisitasAccess)
):
    if not ObjectId.is_valid(visita_id):
        raise HTTPException(status_code=400, detail="Invalid ID")
    
    visita = await visitas_collection.find_one({"_id": ObjectId(visita_id)})
    if not visita:
        raise HTTPException(status_code=404, detail="Visita not found")
    
    return serialize_doc(visita)


@router.put("/visitas/{visita_id}")
async def update_visita(
    visita_id: str,
    visita: VisitaCreate,
    current_user: dict = Depends(RequireEdit),
    _access: dict = Depends(RequireVisitasAccess)
):
    if not ObjectId.is_valid(visita_id):
        raise HTTPException(status_code=400, detail="Invalid ID")
    
    # MODELO SIMPLIFICADO: si cambia la parcela, re-heredar datos
    update_data = {"updated_at": datetime.now()}
    
    if visita.parcela_id:
        if not ObjectId.is_valid(visita.parcela_id):
            raise HTTPException(status_code=400, detail="parcela_id inválido")
        
        parcela = await parcelas_collection.find_one({"_id": ObjectId(visita.parcela_id)})
        if not parcela:
            raise HTTPException(status_code=400, detail="Parcela no encontrada")
        
        # Heredar datos desde la parcela
        contrato_id = parcela.get("contrato_id", "")
        cultivo_id = ""
        if contrato_id:
            contrato = await contratos_collection.find_one({"_id": ObjectId(contrato_id)})
            if contrato:
                cultivo_id = contrato.get("cultivo_id", "")
        
        update_data.update({
            "parcela_id": visita.parcela_id,
            "contrato_id": contrato_id,
            "cultivo_id": cultivo_id,
            "proveedor": parcela.get("proveedor", ""),
            "cultivo": parcela.get("cultivo", ""),
            "campana": parcela.get("campana", ""),
            "variedad": parcela.get("variedad", ""),
            "codigo_plantacion": parcela.get("codigo_plantacion", ""),
            "finca": parcela.get("finca", "")
        })
    
    # Actualizar campos editables por el usuario
    update_data["objetivo"] = visita.objetivo
    if visita.fecha_visita:
        update_data["fecha_visita"] = visita.fecha_visita
    if visita.observaciones is not None:
        update_data["observaciones"] = visita.observaciones
    
    # Actualizar cuestionario de plagas si el objetivo es "Plagas y Enfermedades"
    if visita.objetivo == "Plagas y Enfermedades" and visita.cuestionario_plagas:
        update_data["cuestionario_plagas"] = visita.cuestionario_plagas
    elif visita.objetivo != "Plagas y Enfermedades":
        # Si el objetivo cambió, eliminar el cuestionario
        update_data["cuestionario_plagas"] = None
    
    result = await visitas_collection.update_one(
        {"_id": ObjectId(visita_id)},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Visita not found")
    
    updated = await visitas_collection.find_one({"_id": ObjectId(visita_id)})
    return {"success": True, "data": serialize_doc(updated)}


@router.delete("/visitas/{visita_id}")
async def delete_visita(
    visita_id: str,
    current_user: dict = Depends(RequireDelete),
    _access: dict = Depends(RequireVisitasAccess)
):
    if not ObjectId.is_valid(visita_id):
        raise HTTPException(status_code=400, detail="Invalid ID")
    
    result = await visitas_collection.delete_one({"_id": ObjectId(visita_id)})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Visita not found")
    
    return {"success": True, "message": "Visita deleted"}



@router.get("/visitas/export/excel")
async def export_visitas_excel(
    campana: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Export visitas to Excel"""
    from fastapi.responses import StreamingResponse
    import io as _io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    query = {}
    if campana:
        query["campana"] = campana

    visitas_list = await visitas_collection.find(query).sort("fecha", -1).to_list(5000)
    visitas_raw = serialize_docs(visitas_list)

    wb = Workbook()
    ws = wb.active
    ws.title = "Visitas"

    headers = ["Fecha", "Parcela", "Proveedor", "Cultivo", "Campaña", "Objetivo", "Observaciones", "Estado Cultivo"]
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for idx, v in enumerate(visitas_raw, 2):
        row = [
            v.get("fecha", "")[:10] if v.get("fecha") else "",
            v.get("parcela_codigo", v.get("codigo_plantacion", "")),
            v.get("proveedor", ""), v.get("cultivo", ""), v.get("campana", ""),
            v.get("objetivo", ""), v.get("observaciones", "")[:100], v.get("estado_cultivo", "")
        ]
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=idx, column=col, value=val)
            cell.border = thin_border

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 18

    output = _io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"visitas_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename={filename}"})


@router.get("/visitas/export/pdf")
async def export_visitas_pdf(
    campana: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Export visitas to PDF"""
    from fastapi.responses import StreamingResponse
    import io as _io
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    query = {}
    if campana:
        query["campana"] = campana

    visitas_list = await visitas_collection.find(query).sort("fecha", -1).to_list(5000)
    visitas_raw = serialize_docs(visitas_list)

    output = _io.BytesIO()
    pdf = SimpleDocTemplate(output, pagesize=landscape(A4), topMargin=15*mm, bottomMargin=15*mm)
    elements = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=16, textColor=colors.HexColor('#2E7D32'), alignment=1, spaceAfter=8*mm)
    elements.append(Paragraph("Informe de Visitas - FRUVECO", title_style))
    elements.append(Paragraph(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", ParagraphStyle('Sub', parent=styles['Normal'], alignment=1, textColor=colors.gray)))
    elements.append(Spacer(1, 8*mm))

    table_data = [["Fecha", "Parcela", "Proveedor", "Cultivo", "Objetivo", "Estado"]]
    for v in visitas_raw:
        table_data.append([
            v.get("fecha", "")[:10] if v.get("fecha") else "",
            v.get("parcela_codigo", "")[:20],
            v.get("proveedor", "")[:20], v.get("cultivo", "")[:15],
            v.get("objetivo", "")[:30], v.get("estado_cultivo", "")[:15]
        ])

    col_widths = [25*mm, 35*mm, 35*mm, 30*mm, 55*mm, 30*mm]
    doc_table = Table(table_data, colWidths=col_widths)
    doc_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E7D32')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.gray),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
    ]))
    elements.append(doc_table)

    pdf.build(elements)
    output.seek(0)
    filename = f"visitas_{datetime.now().strftime('%Y%m%d')}.pdf"
    return StreamingResponse(output, media_type="application/pdf",
                             headers={"Content-Disposition": f"attachment; filename={filename}"})
