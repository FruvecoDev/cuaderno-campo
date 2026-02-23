from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import List, Optional
from datetime import datetime
from bson import ObjectId
import os
from dotenv import load_dotenv
import json
from io import BytesIO

# Load environment
load_dotenv()

# AI imports
from emergentintegrations.llm.chat import LlmChat, UserMessage

# PDF imports
from weasyprint import HTML

# Excel imports
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Import routes
from routes_main import router as main_router
from routes_extended import router as extended_router
from routes_auth import router as auth_router
from routes_catalogos import router as catalogos_router

# Import database
from database import (
    parcelas_collection, contratos_collection, tratamientos_collection,
    irrigaciones_collection, cosechas_collection, visitas_collection,
    tareas_collection, fincas_collection, serialize_doc, serialize_docs
)

app = FastAPI(title="Agricultural Management System V1")

# CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Include routers
app.include_router(auth_router)
app.include_router(main_router)
app.include_router(extended_router)

# ============================================================================
# ROOT
# ============================================================================

@app.get("/")
async def root():
    return {
        "message": "Agricultural Management System V1 API",
        "version": "1.0.0",
        "status": "running"
    }

# ============================================================================
# DASHBOARD - KPIs
# ============================================================================

@app.get("/api/dashboard/kpis")
async def get_dashboard_kpis():
    # Count documents
    total_contratos = await contratos_collection.count_documents({})
    total_parcelas = await parcelas_collection.count_documents({})
    parcelas_activas = await parcelas_collection.count_documents({"activo": True})
    total_fincas = await fincas_collection.count_documents({})
    total_tratamientos = await tratamientos_collection.count_documents({})
    total_riegos = await irrigaciones_collection.count_documents({})
    total_visitas = await visitas_collection.count_documents({})
    total_cosechas = await cosechas_collection.count_documents({})
    
    # Calculate production
    cosechas = await cosechas_collection.find().to_list(1000)
    total_produccion = sum(c.get("cosecha_total", 0) for c in cosechas)
    total_ingresos = sum(c.get("ingreso_total", 0) for c in cosechas)
    
    # Calculate costs
    tratamientos = await tratamientos_collection.find().to_list(1000)
    total_coste_tratamientos = sum(t.get("coste_total", 0) for t in tratamientos)
    
    riegos = await irrigaciones_collection.find().to_list(1000)
    total_coste_riegos = sum(r.get("coste", 0) for r in riegos)
    
    tareas = await tareas_collection.find().to_list(1000)
    total_coste_tareas = sum(t.get("coste_total", 0) for t in tareas)
    
    total_costes = total_coste_tratamientos + total_coste_riegos + total_coste_tareas
    
    # Calculate surface
    parcelas = await parcelas_collection.find().to_list(1000)
    total_superficie = sum(p.get("superficie_total", 0) for p in parcelas)
    
    # Production by crop
    produccion_por_cultivo = {}
    for parcela in parcelas:
        cultivo = parcela.get("cultivo", "Unknown")
        if cultivo not in produccion_por_cultivo:
            produccion_por_cultivo[cultivo] = {
                "superficie": 0,
                "parcelas": 0,
                "produccion": 0
            }
        produccion_por_cultivo[cultivo]["superficie"] += parcela.get("superficie_total", 0)
        produccion_por_cultivo[cultivo]["parcelas"] += 1
    
    # Recent activity
    recent_visitas = await visitas_collection.find().sort("created_at", -1).limit(5).to_list(5)
    recent_tratamientos = await tratamientos_collection.find().sort("created_at", -1).limit(5).to_list(5)
    
    return {
        "totales": {
            "contratos": total_contratos,
            "parcelas": total_parcelas,
            "parcelas_activas": parcelas_activas,
            "fincas": total_fincas,
            "tratamientos": total_tratamientos,
            "riegos": total_riegos,
            "visitas": total_visitas,
            "cosechas": total_cosechas
        },
        "produccion": {
            "total_kg": total_produccion,
            "total_ingresos": total_ingresos,
            "por_cultivo": produccion_por_cultivo
        },
        "costes": {
            "tratamientos": total_coste_tratamientos,
            "riegos": total_coste_riegos,
            "tareas": total_coste_tareas,
            "total": total_costes
        },
        "superficie": {
            "total_ha": total_superficie,
            "promedio_ha_parcela": total_superficie / total_parcelas if total_parcelas > 0 else 0
        },
        "rentabilidad": {
            "margen_bruto": total_ingresos - total_costes,
            "margen_por_ha": (total_ingresos - total_costes) / total_superficie if total_superficie > 0 else 0
        },
        "actividad_reciente": {
            "visitas": serialize_docs(recent_visitas),
            "tratamientos": serialize_docs(recent_tratamientos)
        }
    }

# ============================================================================
# AI REPORT GENERATION
# ============================================================================

class AIReportRequest(BaseModel):
    parcela_id: Optional[str] = None
    contrato_id: Optional[str] = None
    tipo: str = "parcela"  # parcela, contrato, finca, general

@app.post("/api/generate-ai-report")
async def generate_ai_report(request: AIReportRequest):
    try:
        api_key = os.getenv('EMERGENT_LLM_KEY')
        if not api_key:
            raise HTTPException(status_code=500, detail="AI API key not configured")
        
        # Gather data based on type
        if request.tipo == "parcela" and request.parcela_id:
            if not ObjectId.is_valid(request.parcela_id):
                raise HTTPException(status_code=400, detail="Invalid parcela ID")
            
            parcela = await parcelas_collection.find_one({"_id": ObjectId(request.parcela_id)})
            if not parcela:
                raise HTTPException(status_code=404, detail="Parcela not found")
            
            # Get related data
            tratamientos = await tratamientos_collection.find({"parcelas_ids": request.parcela_id}).to_list(100)
            riegos = await irrigaciones_collection.find({"parcela_id": request.parcela_id}).to_list(100)
            visitas = await visitas_collection.find({"parcela_id": request.parcela_id}).to_list(100)
            
            campaign_data = {
                "parcela": serialize_doc(parcela),
                "tratamientos": serialize_docs(tratamientos),
                "riegos": serialize_docs(riegos),
                "visitas": serialize_docs(visitas),
                "total_tratamientos": len(tratamientos),
                "total_riegos": len(riegos)
            }
            
            prompt = f"""Actúa como un agrónomo experto. Analiza estos datos de la parcela {parcela.get('codigo_plantacion')} y genera un INFORME EJECUTIVO:

**DATOS:**
```json
{json.dumps(campaign_data, indent=2, ensure_ascii=False)}
```

**GENERA UN INFORME CON:**
1. Resumen ejecutivo (3-4 líneas)
2. Análisis de la parcela (superficie, cultivo, estado)
3. Análisis de tratamientos aplicados (eficacia, cumplimiento)
4. Análisis de riegos (eficiencia hídrica, optimización)
5. Actividades y visitas realizadas
6. Recomendaciones específicas para esta parcela

Usa markdown con secciones claras. Sé preciso con los números."""
        
        elif request.tipo == "general":
            # General report of all operations
            total_parcelas = await parcelas_collection.count_documents({})
            total_tratamientos = await tratamientos_collection.count_documents({})
            total_riegos = await irrigaciones_collection.count_documents({})
            total_cosechas = await cosechas_collection.count_documents({})
            
            cosechas = await cosechas_collection.find().to_list(100)
            total_produccion = sum(c.get("cosecha_total", 0) for c in cosechas)
            total_ingresos = sum(c.get("ingreso_total", 0) for c in cosechas)
            
            campaign_data = {
                "totales": {
                    "parcelas": total_parcelas,
                    "tratamientos": total_tratamientos,
                    "riegos": total_riegos,
                    "cosechas": total_cosechas,
                    "produccion_kg": total_produccion,
                    "ingresos_euros": total_ingresos
                }
            }
            
            prompt = f"""Actúa como un gerente agrícola experto. Analiza estos datos generales de la explotación y genera un INFORME EJECUTIVO:

**DATOS:**
```json
{json.dumps(campaign_data, indent=2, ensure_ascii=False)}
```

**GENERA UN INFORME CON:**
1. Resumen ejecutivo general
2. Análisis de producción global
3. Análisis de gestión de cultivos
4. Eficiencia operacional
5. Indicadores clave de rendimiento (KPIs)
6. Recomendaciones estratégicas

Usa markdown con secciones claras."""
        
        else:
            raise HTTPException(status_code=400, detail="Invalid report type or missing IDs")
        
        # Call AI
        chat = LlmChat(
            api_key=api_key,
            session_id=f"report-{request.tipo}-{datetime.now().timestamp()}",
            system_message="Eres un agrónomo experto en análisis de datos agrícolas."
        )
        chat.with_model("openai", "gpt-5.2")
        
        user_message = UserMessage(text=prompt)
        response = await chat.send_message(user_message)
        
        return {"success": True, "report": response}
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ============================================================================
# PDF GENERATION
# ============================================================================

class PDFRequest(BaseModel):
    parcela_id: str
    incluir_tratamientos: bool = True
    incluir_riegos: bool = True
    incluir_visitas: bool = True
    incluir_cosechas: bool = True

@app.post("/api/generate-pdf")
async def generate_pdf(request: PDFRequest):
    try:
        if not ObjectId.is_valid(request.parcela_id):
            raise HTTPException(status_code=400, detail="Invalid parcela ID")
        
        parcela = await parcelas_collection.find_one({"_id": ObjectId(request.parcela_id)})
        if not parcela:
            raise HTTPException(status_code=404, detail="Parcela not found")
        
        # Get related data
        tratamientos = []
        riegos = []
        visitas = []
        cosechas = []
        
        if request.incluir_tratamientos:
            tratamientos = await tratamientos_collection.find({"parcelas_ids": request.parcela_id}).to_list(100)
        if request.incluir_riegos:
            riegos = await irrigaciones_collection.find({"parcela_id": request.parcela_id}).to_list(100)
        if request.incluir_visitas:
            visitas = await visitas_collection.find({"parcela_id": request.parcela_id}).to_list(100)
        if request.incluir_cosechas:
            # Note: cosechas link to parcelas via parcelas_ids array
            cosechas = await cosechas_collection.find({"parcelas_ids": request.parcela_id}).to_list(100)
        
        # Create HTML
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <style>
                @page {{ size: A4; margin: 2cm; }}
                body {{ font-family: Arial, sans-serif; font-size: 10pt; color: #333; }}
                h1 {{ color: #2c5f2d; border-bottom: 3px solid #97bf0d; padding-bottom: 10px; font-size: 20pt; }}
                h2 {{ color: #2c5f2d; margin-top: 20px; font-size: 14pt; border-bottom: 1px solid #ddd; padding-bottom: 5px; }}
                .header-info {{ background-color: #f0f8f0; padding: 15px; border-radius: 5px; margin-bottom: 20px; }}
                .info-row {{ margin-bottom: 5px; font-size: 9pt; }}
                .info-label {{ font-weight: bold; display: inline-block; width: 150px; }}
                table {{ width: 100%; border-collapse: collapse; margin: 10px 0 20px 0; font-size: 8pt; }}
                th {{ background-color: #2c5f2d; color: white; padding: 8px; text-align: left; }}
                td {{ border: 1px solid #ddd; padding: 6px; }}
                tr:nth-child(even) {{ background-color: #f9f9f9; }}
                .footer {{ margin-top: 30px; padding-top: 10px; border-top: 1px solid #ddd; font-size: 8pt; color: #666; text-align: center; }}
            </style>
        </head>
        <body>
            <h1>📋 CUADERNO DE CAMPO - Sistema de Gestión Agrícola</h1>
            <div class="header-info">
                <div class="info-row"><span class="info-label">Código Plantación:</span> {parcela.get('codigo_plantacion', 'N/A')}</div>
                <div class="info-row"><span class="info-label">Proveedor:</span> {parcela.get('proveedor', 'N/A')}</div>
                <div class="info-row"><span class="info-label">Finca:</span> {parcela.get('finca', 'N/A')}</div>
                <div class="info-row"><span class="info-label">Cultivo:</span> {parcela.get('cultivo', 'N/A')} - {parcela.get('variedad', 'N/A')}</div>
                <div class="info-row"><span class="info-label">Campaña:</span> {parcela.get('campana', 'N/A')}</div>
                <div class="info-row"><span class="info-label">Superficie:</span> {parcela.get('superficie_total', 0)} {parcela.get('unidad_medida', 'ha')}</div>
                <div class="info-row"><span class="info-label">Nº Plantas:</span> {parcela.get('num_plantas', 0)}</div>
                <div class="info-row"><span class="info-label">Estado:</span> {'Activa' if parcela.get('activo') else 'Inactiva'}</div>
            </div>
        """
        
        if tratamientos:
            html_content += """
            <h2>🌿 Tratamientos Fitosanitarios</h2>
            <table>
                <thead><tr><th>Fecha</th><th>Tipo</th><th>Método</th><th>Superficie</th><th>Coste Total</th></tr></thead>
                <tbody>
            """
            for t in tratamientos:
                html_content += f"""
                    <tr>
                        <td>{t.get('fecha_inicio', 'N/A')}</td>
                        <td>{t.get('tipo_tratamiento', 'N/A')} - {t.get('subtipo', '')}</td>
                        <td>{t.get('metodo_aplicacion', 'N/A')}</td>
                        <td>{t.get('superficie_aplicacion', 0)} ha</td>
                        <td>{t.get('coste_total', 0):.2f} €</td>
                    </tr>
                """
            html_content += "</tbody></table>"
        
        if riegos:
            html_content += """
            <h2>💧 Irrigaciones</h2>
            <table>
                <thead><tr><th>Fecha</th><th>Sistema</th><th>Duración (h)</th><th>Volumen (m³)</th><th>Coste (€)</th></tr></thead>
                <tbody>
            """
            for r in riegos:
                html_content += f"""
                    <tr>
                        <td>{r.get('fecha', 'N/A')}</td>
                        <td>{r.get('sistema', 'N/A')}</td>
                        <td>{r.get('duracion', 0)}</td>
                        <td>{r.get('volumen', 0)}</td>
                        <td>{r.get('coste', 0):.2f}</td>
                    </tr>
                """
            html_content += "</tbody></table>"
        
        if visitas:
            html_content += """
            <h2>👨‍🌾 Visitas</h2>
            <table>
                <thead><tr><th>Fecha</th><th>Objetivo</th><th>Realizado</th><th>Observaciones</th></tr></thead>
                <tbody>
            """
            for v in visitas:
                html_content += f"""
                    <tr>
                        <td>{v.get('fecha_visita', 'N/A')}</td>
                        <td>{v.get('objetivo', 'N/A')}</td>
                        <td>{'Sí' if v.get('realizado') else 'No'}</td>
                        <td>{v.get('observaciones', '')[:50]}...</td>
                    </tr>
                """
            html_content += "</tbody></table>"
        
        if cosechas:
            html_content += """
            <h2>🌾 Cosechas</h2>
            <table>
                <thead><tr><th>Nombre</th><th>Superficie</th><th>Total Cosecha</th><th>Ingreso Total</th></tr></thead>
                <tbody>
            """
            for c in cosechas:
                html_content += f"""
                    <tr>
                        <td>{c.get('nombre', 'N/A')}</td>
                        <td>{c.get('superficie_total', 0)} {c.get('unidad_medida', 'ha')}</td>
                        <td>{c.get('cosecha_total', 0):.0f} kg</td>
                        <td>{c.get('ingreso_total', 0):.2f} €</td>
                    </tr>
                """
            html_content += "</tbody></table>"
        
        html_content += f"""
            <div class="footer">
                Documento generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}<br>
                Sistema de Gestión Agrícola V1 - Cuaderno de Campo Digital
            </div>
        </body>
        </html>
        """
        
        # Generate PDF
        html = HTML(string=html_content)
        pdf_bytes = html.write_pdf()
        
        return StreamingResponse(
            BytesIO(pdf_bytes),
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=cuaderno_{parcela.get('codigo_plantacion', 'parcela')}.pdf"}
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ============================================================================\n# EXCEL GENERATION
# ============================================================================

class ExcelRequest(BaseModel):
    parcela_id: Optional[str] = None
    tipo: str = "parcela"  # parcela, general

@app.post("/api/generate-excel")
async def generate_excel(request: ExcelRequest):
    try:
        wb = Workbook()
        
        # Header style
        header_fill = PatternFill(start_color="2C5F2D", end_color="2C5F2D", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        if request.tipo == "parcela" and request.parcela_id:
            if not ObjectId.is_valid(request.parcela_id):
                raise HTTPException(status_code=400, detail="Invalid parcela ID")
            
            # Get data
            tratamientos = await tratamientos_collection.find({"parcelas_ids": request.parcela_id}).to_list(100)
            riegos = await irrigaciones_collection.find({"parcela_id": request.parcela_id}).to_list(100)
            
            # Tratamientos sheet
            ws = wb.active
            ws.title = "Tratamientos"
            headers = ["Fecha", "Tipo", "Subtipo", "Método", "Superficie (ha)", "Coste (€)"]
            ws.append(headers)
            for col in range(1, len(headers) + 1):
                cell = ws.cell(1, col)
                cell.fill = header_fill
                cell.font = header_font
                ws.column_dimensions[get_column_letter(col)].width = 15
            
            for t in tratamientos:
                ws.append([
                    t.get("fecha_inicio", ""),
                    t.get("tipo_tratamiento", ""),
                    t.get("subtipo", ""),
                    t.get("metodo_aplicacion", ""),
                    t.get("superficie_aplicacion", 0),
                    t.get("coste_total", 0)
                ])
            
            # Riegos sheet
            ws2 = wb.create_sheet("Irrigaciones")
            headers2 = ["Fecha", "Sistema", "Duración (h)", "Volumen (m³)", "Coste (€)"]
            ws2.append(headers2)
            for col in range(1, len(headers2) + 1):
                cell = ws2.cell(1, col)
                cell.fill = header_fill
                cell.font = header_font
                ws2.column_dimensions[get_column_letter(col)].width = 15
            
            for r in riegos:
                ws2.append([
                    r.get("fecha", ""),
                    r.get("sistema", ""),
                    r.get("duracion", 0),
                    r.get("volumen", 0),
                    r.get("coste", 0)
                ])
        
        else:
            # General export - all parcelas
            parcelas = await parcelas_collection.find().to_list(1000)
            
            ws = wb.active
            ws.title = "Parcelas"
            headers = ["Código", "Proveedor", "Cultivo", "Variedad", "Campaña", "Superficie", "Activo"]
            ws.append(headers)
            for col in range(1, len(headers) + 1):
                cell = ws.cell(1, col)
                cell.fill = header_fill
                cell.font = header_font
                ws.column_dimensions[get_column_letter(col)].width = 18
            
            for p in parcelas:
                ws.append([
                    p.get("codigo_plantacion", ""),
                    p.get("proveedor", ""),
                    p.get("cultivo", ""),
                    p.get("variedad", ""),
                    p.get("campana", ""),
                    p.get("superficie_total", 0),
                    "Sí" if p.get("activo") else "No"
                ])
        
        # Save to bytes
        excel_bytes = BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)
        
        return StreamingResponse(
            excel_bytes,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=datos_agricolas.xlsx"}
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8001)
