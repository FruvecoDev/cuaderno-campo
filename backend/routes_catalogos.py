"""
Rutas para gestión de catálogos maestros: Proveedores y Cultivos
"""
from fastapi import APIRouter, HTTPException, Depends
from typing import List, Optional
from bson import ObjectId
from datetime import datetime

from models_catalogos import (
    ProveedorCreate, ProveedorInDB,
    CultivoCreate, CultivoInDB
)
from database import db, serialize_doc, serialize_docs
from rbac_guards import (
    RequireCreate, RequireEdit, RequireDelete,
    get_current_user
)

router = APIRouter(prefix="/api", tags=["catalogos"])

# Collections
proveedores_collection = db['proveedores']
cultivos_collection = db['cultivos']
tipos_proveedor_collection = db['tipos_proveedor']
tipos_operacion_collection = db['tipos_operacion_proveedor']
documentos_proveedor_collection = db['documentos_proveedor']
changelog_proveedor_collection = db['proveedor_changelog']
formas_pago_collection = db['formas_pago']
tipos_iva_collection = db['tipos_iva']
tipos_cultivo_collection = db['tipos_cultivo']
changelog_cultivo_collection = db['cultivo_changelog']
categorias_articulo_collection = db['categorias_articulo']


async def log_proveedor_change(proveedor_id, action, user, changes=None):
    entry = {
        "proveedor_id": proveedor_id,
        "action": action,
        "user_email": user.get("email", ""),
        "user_name": user.get("full_name", user.get("email", "")),
        "changes": changes or [],
        "timestamp": datetime.now(),
    }
    await changelog_proveedor_collection.insert_one(entry)


def diff_dicts(old, new, prefix=""):
    changes = []
    skip = {"_id", "created_at", "updated_at"}
    all_keys = set(list(old.keys()) + list(new.keys())) - skip
    for key in sorted(all_keys):
        old_val = old.get(key)
        new_val = new.get(key)
        field = f"{prefix}{key}" if not prefix else f"{prefix}.{key}"
        if isinstance(old_val, dict) and isinstance(new_val, dict):
            changes.extend(diff_dicts(old_val, new_val, field))
        elif isinstance(old_val, list) and isinstance(new_val, list):
            if old_val != new_val:
                changes.append({"field": field, "old": str(old_val)[:200], "new": str(new_val)[:200]})
        elif str(old_val) != str(new_val):
            changes.append({"field": field, "old": str(old_val or ""), "new": str(new_val or "")})
    return changes


# ============================================================================
# TIPOS DE PROVEEDOR
# ============================================================================

@router.get("/tipos-proveedor")
async def get_tipos_proveedor(current_user: dict = Depends(get_current_user)):
    tipos = await tipos_proveedor_collection.find().sort("nombre", 1).to_list(200)
    return {"tipos": serialize_docs(tipos)}

@router.post("/tipos-proveedor")
async def create_tipo_proveedor(data: dict, current_user: dict = Depends(RequireCreate)):
    nombre = data.get("nombre", "").strip()
    if not nombre:
        raise HTTPException(status_code=400, detail="El nombre es obligatorio")
    existing = await tipos_proveedor_collection.find_one({"nombre": {"$regex": f"^{nombre}$", "$options": "i"}})
    if existing:
        raise HTTPException(status_code=400, detail="Ya existe un tipo con ese nombre")
    result = await tipos_proveedor_collection.insert_one({"nombre": nombre, "created_at": datetime.now()})
    created = await tipos_proveedor_collection.find_one({"_id": result.inserted_id})
    return {"success": True, "tipo": serialize_doc(created)}

@router.delete("/tipos-proveedor/{tipo_id}")
async def delete_tipo_proveedor(tipo_id: str, current_user: dict = Depends(RequireDelete)):
    if not ObjectId.is_valid(tipo_id):
        raise HTTPException(status_code=400, detail="ID no valido")
    result = await tipos_proveedor_collection.delete_one({"_id": ObjectId(tipo_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Tipo no encontrado")
    return {"success": True}


# ============================================================================
# TIPOS DE OPERACION PROVEEDOR
# ============================================================================

@router.get("/tipos-operacion-proveedor")
async def get_tipos_operacion(current_user: dict = Depends(get_current_user)):
    tipos = await tipos_operacion_collection.find().sort("nombre", 1).to_list(200)
    return {"tipos": serialize_docs(tipos)}

@router.post("/tipos-operacion-proveedor")
async def create_tipo_operacion(data: dict, current_user: dict = Depends(RequireCreate)):
    nombre = data.get("nombre", "").strip()
    if not nombre:
        raise HTTPException(status_code=400, detail="El nombre es obligatorio")
    existing = await tipos_operacion_collection.find_one({"nombre": {"$regex": f"^{nombre}$", "$options": "i"}})
    if existing:
        raise HTTPException(status_code=400, detail="Ya existe un tipo con ese nombre")
    result = await tipos_operacion_collection.insert_one({"nombre": nombre, "created_at": datetime.now()})
    created = await tipos_operacion_collection.find_one({"_id": result.inserted_id})
    return {"success": True, "tipo": serialize_doc(created)}

@router.delete("/tipos-operacion-proveedor/{tipo_id}")
async def delete_tipo_operacion(tipo_id: str, current_user: dict = Depends(RequireDelete)):
    if not ObjectId.is_valid(tipo_id):
        raise HTTPException(status_code=400, detail="ID no valido")
    result = await tipos_operacion_collection.delete_one({"_id": ObjectId(tipo_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Tipo no encontrado")
    return {"success": True}



# ============================================================================
# FORMAS DE PAGO
# ============================================================================

@router.get("/formas-pago")
async def get_formas_pago(current_user: dict = Depends(get_current_user)):
    items = await formas_pago_collection.find().sort("nombre", 1).to_list(200)
    return {"items": serialize_docs(items)}

@router.post("/formas-pago")
async def create_forma_pago(data: dict, current_user: dict = Depends(RequireCreate)):
    nombre = data.get("nombre", "").strip()
    if not nombre:
        raise HTTPException(status_code=400, detail="El nombre es obligatorio")
    result = await formas_pago_collection.insert_one({"nombre": nombre, "created_at": datetime.now()})
    created = await formas_pago_collection.find_one({"_id": result.inserted_id})
    return {"success": True, "item": serialize_doc(created)}

@router.delete("/formas-pago/{item_id}")
async def delete_forma_pago(item_id: str, current_user: dict = Depends(RequireDelete)):
    if not ObjectId.is_valid(item_id):
        raise HTTPException(status_code=400, detail="ID no valido")
    await formas_pago_collection.delete_one({"_id": ObjectId(item_id)})
    return {"success": True}


# ============================================================================
# TIPOS DE IVA
# ============================================================================

@router.get("/tipos-iva")
async def get_tipos_iva(current_user: dict = Depends(get_current_user)):
    items = await tipos_iva_collection.find().sort("valor", 1).to_list(200)
    return {"items": serialize_docs(items)}

@router.post("/tipos-iva")
async def create_tipo_iva(data: dict, current_user: dict = Depends(RequireCreate)):
    nombre = data.get("nombre", "").strip()
    valor = data.get("valor", "")
    if not nombre:
        raise HTTPException(status_code=400, detail="El nombre es obligatorio")
    result = await tipos_iva_collection.insert_one({"nombre": nombre, "valor": valor, "created_at": datetime.now()})
    created = await tipos_iva_collection.find_one({"_id": result.inserted_id})
    return {"success": True, "item": serialize_doc(created)}

@router.delete("/tipos-iva/{item_id}")
async def delete_tipo_iva(item_id: str, current_user: dict = Depends(RequireDelete)):
    if not ObjectId.is_valid(item_id):
        raise HTTPException(status_code=400, detail="ID no valido")
    await tipos_iva_collection.delete_one({"_id": ObjectId(item_id)})
    return {"success": True}


# ============================================================================
# DOCUMENTOS PROVEEDOR
# ============================================================================

@router.get("/proveedores/{proveedor_id}/documentos")
async def get_documentos_proveedor(proveedor_id: str, current_user: dict = Depends(get_current_user)):
    docs = await documentos_proveedor_collection.find({"proveedor_id": proveedor_id}).sort("created_at", -1).to_list(200)
    return {"documentos": serialize_docs(docs)}

@router.post("/proveedores/{proveedor_id}/documentos")
async def upload_documento_proveedor(proveedor_id: str, current_user: dict = Depends(RequireCreate)):
    from fastapi import UploadFile, File, Form
    # This will be handled via multipart form
    return {"success": True}

@router.delete("/proveedores/documentos/{doc_id}")
async def delete_documento_proveedor(doc_id: str, current_user: dict = Depends(RequireDelete)):
    if not ObjectId.is_valid(doc_id):
        raise HTTPException(status_code=400, detail="ID no valido")
    result = await documentos_proveedor_collection.delete_one({"_id": ObjectId(doc_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Documento no encontrado")
    return {"success": True}


# ============================================================================
# PROVEEDORES
# ============================================================================

@router.post("/proveedores", response_model=dict)
async def create_proveedor(
    proveedor: ProveedorCreate,
    current_user: dict = Depends(RequireCreate)
):
    proveedor_dict = proveedor.dict()
    
    # Auto-generate codigo_proveedor
    last = await proveedores_collection.find_one(
        {"codigo_proveedor": {"$exists": True, "$ne": None}},
        sort=[("codigo_proveedor", -1)]
    )
    if last and last.get("codigo_proveedor"):
        try:
            next_num = int(last["codigo_proveedor"]) + 1
        except ValueError:
            next_num = 1
    else:
        count = await proveedores_collection.count_documents({})
        next_num = count + 1
    proveedor_dict['codigo_proveedor'] = str(next_num).zfill(6)
    
    proveedor_dict['created_at'] = datetime.now()
    proveedor_dict['updated_at'] = datetime.now()
    
    result = await proveedores_collection.insert_one(proveedor_dict)
    created = await proveedores_collection.find_one({"_id": result.inserted_id})
    
    await log_proveedor_change(str(result.inserted_id), "creacion", current_user, [{"field": "nombre", "old": "", "new": proveedor_dict.get("nombre", "")}])
    
    return {"success": True, "proveedor": serialize_doc(created)}


@router.get("/proveedores")
async def get_proveedores(
    skip: int = 0,
    limit: int = 10000,
    activo: Optional[bool] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if activo is not None:
        query['activo'] = activo
    
    proveedores = await proveedores_collection.find(query).skip(skip).limit(limit).to_list(limit)
    total = await proveedores_collection.count_documents(query)
    
    return {
        "proveedores": serialize_docs(proveedores),
        "total": total
    }


# ============================================================================
# DEDUPLICACION DE PROVEEDORES (declarado antes de /proveedores/{id} para
# que FastAPI no capture "duplicados" como path parameter)
# ============================================================================

def _normalize_nombre_proveedor(nombre: str) -> str:
    """Normaliza un nombre para deteccion de duplicados.
    Aplana espacios, elimina puntuacion final, unifica sufijos societarios
    (S.L., SL, S.A., SA, S.COOP, S.L.U, etc.) y quita tildes.
    """
    if not nombre:
        return ""
    import re
    import unicodedata
    s = nombre.strip().lower()
    s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    suffix_pattern = re.compile(
        r"[\s,\.\-]*("
        r"s\.?l\.?u\.?|"
        r"s\.?l\.?l\.?|"
        r"s\.?l\.?|"
        r"s\.?a\.?u\.?|"
        r"s\.?a\.?|"
        r"s\.?coop\.?|"
        r"s\.?c\.?a\.?|"
        r"s\.?a\.?t\.?"
        r")\s*$"
    )
    prev = None
    while prev != s:
        prev = s
        s = suffix_pattern.sub("", s)
    s = re.sub(r"[\.,;:\-_]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


@router.get("/proveedores/duplicados")
async def detectar_duplicados_proveedores(current_user: dict = Depends(get_current_user)):
    """Escanea la coleccion de proveedores y agrupa aquellos cuyo nombre
    normalizado coincida. Devuelve grupos con >=2 registros.
    """
    provs = await proveedores_collection.find({}).to_list(20000)
    grupos: dict = {}
    for p in provs:
        key = _normalize_nombre_proveedor(p.get("nombre", ""))
        if not key:
            continue
        grupos.setdefault(key, []).append(p)

    resultado = []
    for key, items in grupos.items():
        if len(items) < 2:
            continue
        items_sorted = sorted(
            items,
            key=lambda x: (
                int(x.get("codigo_proveedor", "999999") or 999999)
                if str(x.get("codigo_proveedor", "")).isdigit()
                else 999999
            ),
        )
        keep = items_sorted[0]

        enriched = []
        for it in items_sorted:
            prov_id = str(it["_id"])
            prov_name = it.get("nombre", "")
            refs = {
                "contratos": await db["contratos"].count_documents({
                    "$or": [{"proveedor_id": prov_id}, {"proveedor": prov_name}]
                }),
                "parcelas": await db["parcelas"].count_documents({
                    "$or": [{"proveedor_id": prov_id}, {"proveedor": prov_name}]
                }),
                "albaranes": await db["albaranes"].count_documents({
                    "$or": [{"proveedor_id": prov_id}, {"proveedor": prov_name}]
                }),
            }
            enriched.append({
                "_id": prov_id,
                "codigo_proveedor": it.get("codigo_proveedor", ""),
                "nombre": prov_name,
                "cif_nif": it.get("cif_nif", ""),
                "activo": it.get("activo", True),
                "referencias": refs,
                "total_referencias": sum(refs.values()),
            })

        resultado.append({
            "clave_normalizada": key,
            "keep_id_sugerido": str(keep["_id"]),
            "proveedores": enriched,
        })

    resultado.sort(key=lambda g: -len(g["proveedores"]))
    return {"success": True, "total_grupos": len(resultado), "grupos": resultado}


@router.post("/proveedores/merge")
async def fusionar_proveedores(
    payload: dict,
    current_user: dict = Depends(RequireDelete),
):
    """Fusiona proveedores duplicados en uno canonico.
    Body: { keep_id: str, merge_ids: [str, ...] }
    """
    keep_id = payload.get("keep_id")
    merge_ids = payload.get("merge_ids") or []

    if not keep_id or not ObjectId.is_valid(keep_id):
        raise HTTPException(status_code=400, detail="keep_id invalido")
    if not merge_ids or not isinstance(merge_ids, list):
        raise HTTPException(status_code=400, detail="merge_ids requerido")
    if keep_id in merge_ids:
        raise HTTPException(status_code=400, detail="keep_id no puede estar en merge_ids")

    for mid in merge_ids:
        if not ObjectId.is_valid(mid):
            raise HTTPException(status_code=400, detail=f"merge_id invalido: {mid}")

    keep = await proveedores_collection.find_one({"_id": ObjectId(keep_id)})
    if not keep:
        raise HTTPException(status_code=404, detail="Proveedor canonico no encontrado")

    keep_nombre = keep.get("nombre", "")
    resumen = {
        "keep_id": keep_id,
        "keep_nombre": keep_nombre,
        "merged": [],
        "referencias_actualizadas": {"contratos": 0, "parcelas": 0, "albaranes": 0},
    }

    for mid in merge_ids:
        merged = await proveedores_collection.find_one({"_id": ObjectId(mid)})
        if not merged:
            continue
        merged_nombre = merged.get("nombre", "")

        for col_name in ("contratos", "parcelas", "albaranes"):
            col = db[col_name]
            r1 = await col.update_many(
                {"proveedor_id": mid},
                {"$set": {"proveedor_id": keep_id, "proveedor": keep_nombre}},
            )
            r2 = await col.update_many(
                {"proveedor": merged_nombre, "proveedor_id": {"$ne": keep_id}},
                {"$set": {"proveedor": keep_nombre, "proveedor_id": keep_id}},
            )
            resumen["referencias_actualizadas"][col_name] += (r1.modified_count + r2.modified_count)

        await documentos_proveedor_collection.update_many(
            {"proveedor_id": mid}, {"$set": {"proveedor_id": keep_id}}
        )
        await changelog_proveedor_collection.update_many(
            {"proveedor_id": mid}, {"$set": {"proveedor_id": keep_id}}
        )

        await proveedores_collection.delete_one({"_id": ObjectId(mid)})

        resumen["merged"].append({
            "_id": mid,
            "nombre": merged_nombre,
            "codigo_proveedor": merged.get("codigo_proveedor", ""),
        })

    if resumen["merged"]:
        await log_proveedor_change(
            keep_id,
            "fusion",
            current_user,
            [{
                "field": "fusion_duplicados",
                "old": "",
                "new": f"Fusionados {len(resumen['merged'])}: " + ", ".join(
                    f"{m['codigo_proveedor']} {m['nombre']}" for m in resumen["merged"]
                ),
            }],
        )

    return {"success": True, "resumen": resumen}


@router.get("/proveedores/{proveedor_id}")
async def get_proveedor(
    proveedor_id: str,
    current_user: dict = Depends(get_current_user)
):
    if not ObjectId.is_valid(proveedor_id):
        raise HTTPException(status_code=400, detail="Invalid ID")
    
    proveedor = await proveedores_collection.find_one({"_id": ObjectId(proveedor_id)})
    if not proveedor:
        raise HTTPException(status_code=404, detail="Proveedor not found")
    
    return {"proveedor": serialize_doc(proveedor)}


@router.put("/proveedores/{proveedor_id}")
async def update_proveedor(
    proveedor_id: str,
    proveedor: ProveedorCreate,
    current_user: dict = Depends(RequireEdit)
):
    if not ObjectId.is_valid(proveedor_id):
        raise HTTPException(status_code=400, detail="Invalid ID")
    
    proveedor_dict = proveedor.dict()
    proveedor_dict['updated_at'] = datetime.now()
    
    # Get old data for diff
    old_doc = await proveedores_collection.find_one({"_id": ObjectId(proveedor_id)})
    if not old_doc:
        raise HTTPException(status_code=404, detail="Proveedor not found")
    old_clean = {k: v for k, v in old_doc.items() if k != "_id"}
    changes = diff_dicts(old_clean, proveedor_dict)
    
    result = await proveedores_collection.update_one(
        {"_id": ObjectId(proveedor_id)},
        {"$set": proveedor_dict}
    )
    
    if changes:
        await log_proveedor_change(proveedor_id, "modificacion", current_user, changes)
    
    updated = await proveedores_collection.find_one({"_id": ObjectId(proveedor_id)})
    return {"success": True, "proveedor": serialize_doc(updated)}


@router.delete("/proveedores/{proveedor_id}")
async def delete_proveedor(
    proveedor_id: str,
    current_user: dict = Depends(RequireDelete)
):
    if not ObjectId.is_valid(proveedor_id):
        raise HTTPException(status_code=400, detail="Invalid ID")
    
    result = await proveedores_collection.delete_one({"_id": ObjectId(proveedor_id)})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Proveedor not found")
    
    await log_proveedor_change(proveedor_id, "eliminacion", current_user)
    
    return {"success": True, "message": "Proveedor deleted"}


@router.get("/proveedores/{proveedor_id}/changelog")
async def get_proveedor_changelog(proveedor_id: str, current_user: dict = Depends(get_current_user)):
    logs = await changelog_proveedor_collection.find({"proveedor_id": proveedor_id}).sort("timestamp", -1).to_list(200)
    return {"changelog": serialize_docs(logs)}


@router.get("/proveedores/stats/resumen")
async def get_proveedores_stats(current_user: dict = Depends(get_current_user)):
    """Obtener estadísticas de proveedores"""
    total = await proveedores_collection.count_documents({})
    activos = await proveedores_collection.count_documents({"activo": True})
    
    # Proveedores por provincia
    pipeline = [
        {"$match": {"activo": True}},
        {"$group": {"_id": "$provincia", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}},
        {"$limit": 10}
    ]
    por_provincia = []
    async for doc in proveedores_collection.aggregate(pipeline):
        por_provincia.append({"provincia": doc["_id"] or "Sin especificar", "count": doc["count"]})
    
    # Últimos proveedores añadidos
    ultimos = await proveedores_collection.find({"activo": True}).sort("created_at", -1).limit(5).to_list(5)
    
    # Volumen de compras por proveedor (si existe la colección de gastos)
    gastos_collection = db.get_collection('gastos')
    volumen_compras = []
    try:
        pipeline_gastos = [
            {"$match": {"proveedor_id": {"$exists": True, "$ne": None}}},
            {"$group": {"_id": "$proveedor_id", "total": {"$sum": "$importe"}, "count": {"$sum": 1}}},
            {"$sort": {"total": -1}},
            {"$limit": 10}
        ]
        async for doc in gastos_collection.aggregate(pipeline_gastos):
            prov = await proveedores_collection.find_one({"_id": ObjectId(doc["_id"])}) if ObjectId.is_valid(doc["_id"]) else None
            volumen_compras.append({
                "proveedor_id": doc["_id"],
                "proveedor_nombre": prov.get("nombre") if prov else "Desconocido",
                "total_compras": round(doc["total"], 2),
                "num_operaciones": doc["count"]
            })
    except Exception:
        pass
    
    return {
        "success": True,
        "stats": {
            "total": total,
            "activos": activos,
            "inactivos": total - activos,
            "por_provincia": por_provincia,
            "ultimos_agregados": serialize_docs(ultimos),
            "top_proveedores_compras": volumen_compras
        }
    }


@router.get("/proveedores/{proveedor_id}/historial")
async def get_proveedor_historial(
    proveedor_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Obtener historial de compras de un proveedor"""
    if not ObjectId.is_valid(proveedor_id):
        raise HTTPException(status_code=400, detail="Invalid ID")
    
    proveedor = await proveedores_collection.find_one({"_id": ObjectId(proveedor_id)})
    if not proveedor:
        raise HTTPException(status_code=404, detail="Proveedor not found")
    
    # Buscar gastos/compras asociadas
    gastos_collection = db.get_collection('gastos')
    gastos = []
    try:
        cursor = gastos_collection.find({"proveedor_id": proveedor_id}).sort("fecha", -1).limit(50)
        async for g in cursor:
            gastos.append(serialize_doc(g))
    except Exception:
        pass
    
    # Buscar albaranes de compra
    albaranes_collection = db.get_collection('albaranes')
    albaranes = []
    try:
        cursor = albaranes_collection.find({
            "$or": [
                {"proveedor_id": proveedor_id},
                {"proveedor": proveedor.get("nombre")}
            ],
            "tipo": "compra"
        }).sort("fecha", -1).limit(50)
        async for a in cursor:
            albaranes.append(serialize_doc(a))
    except Exception:
        pass
    
    # Calcular totales
    total_compras = sum(g.get("importe", 0) for g in gastos)
    
    return {
        "success": True,
        "proveedor": serialize_doc(proveedor),
        "historial": {
            "gastos": gastos,
            "albaranes": albaranes,
            "total_compras": round(total_compras, 2),
            "num_operaciones": len(gastos) + len(albaranes)
        }
    }


@router.get("/proveedores/export/excel")
async def export_proveedores_excel(
    activo: Optional[bool] = None,
    current_user: dict = Depends(get_current_user)
):
    """Exportar proveedores a Excel"""
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io
    
    query = {}
    if activo is not None:
        query['activo'] = activo
    
    proveedores = await proveedores_collection.find(query).sort("nombre", 1).to_list(1000)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Proveedores"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Encabezados
    headers = ["Nombre", "CIF/NIF", "Dirección", "Población", "Provincia", "C.P.", "Teléfono", "Email", "Contacto", "Estado"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    
    # Datos
    for row_idx, prov in enumerate(proveedores, 2):
        data = [
            prov.get("nombre", ""),
            prov.get("cif_nif", ""),
            prov.get("direccion", ""),
            prov.get("poblacion", ""),
            prov.get("provincia", ""),
            prov.get("codigo_postal", ""),
            prov.get("telefono", ""),
            prov.get("email", ""),
            prov.get("persona_contacto", ""),
            "Activo" if prov.get("activo", True) else "Inactivo"
        ]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
    
    # Ajustar anchos
    column_widths = [25, 15, 30, 20, 15, 10, 15, 25, 20, 10]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=proveedores_{datetime.now().strftime('%Y%m%d')}.xlsx"}
    )


# ============================================================================
# TIPOS DE CULTIVO
# ============================================================================

@router.get("/tipos-cultivo")
async def get_tipos_cultivo(current_user: dict = Depends(get_current_user)):
    tipos = await tipos_cultivo_collection.find().sort("nombre", 1).to_list(200)
    return {"tipos": serialize_docs(tipos)}

@router.post("/tipos-cultivo")
async def create_tipo_cultivo(data: dict, current_user: dict = Depends(RequireCreate)):
    nombre = data.get("nombre", "").strip()
    if not nombre:
        raise HTTPException(status_code=400, detail="El nombre es obligatorio")
    existing = await tipos_cultivo_collection.find_one({"nombre": {"$regex": f"^{nombre}$", "$options": "i"}})
    if existing:
        raise HTTPException(status_code=400, detail="Ya existe un tipo con ese nombre")
    result = await tipos_cultivo_collection.insert_one({"nombre": nombre, "created_at": datetime.now()})
    created = await tipos_cultivo_collection.find_one({"_id": result.inserted_id})
    return {"success": True, "tipo": serialize_doc(created)}

@router.delete("/tipos-cultivo/{tipo_id}")
async def delete_tipo_cultivo(tipo_id: str, current_user: dict = Depends(RequireDelete)):
    if not ObjectId.is_valid(tipo_id):
        raise HTTPException(status_code=400, detail="ID no valido")
    result = await tipos_cultivo_collection.delete_one({"_id": ObjectId(tipo_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Tipo no encontrado")
    return {"success": True}


async def log_cultivo_change(cultivo_id, action, user, changes=None):
    entry = {
        "cultivo_id": cultivo_id,
        "action": action,
        "user_email": user.get("email", ""),
        "user_name": user.get("full_name", user.get("email", "")),
        "changes": changes or [],
        "timestamp": datetime.now(),
    }
    await changelog_cultivo_collection.insert_one(entry)


# ============================================================================
# CATEGORIAS DE ARTICULO
# ============================================================================

@router.get("/categorias-articulo")
async def get_categorias_articulo(current_user: dict = Depends(get_current_user)):
    cats = await categorias_articulo_collection.find().sort("nombre", 1).to_list(200)
    return {"categorias": serialize_docs(cats)}

@router.post("/categorias-articulo")
async def create_categoria_articulo(data: dict, current_user: dict = Depends(RequireCreate)):
    nombre = data.get("nombre", "").strip()
    if not nombre:
        raise HTTPException(status_code=400, detail="El nombre es obligatorio")
    existing = await categorias_articulo_collection.find_one({"nombre": {"$regex": f"^{nombre}$", "$options": "i"}})
    if existing:
        raise HTTPException(status_code=400, detail="Ya existe una categoria con ese nombre")
    result = await categorias_articulo_collection.insert_one({"nombre": nombre, "created_at": datetime.now()})
    created = await categorias_articulo_collection.find_one({"_id": result.inserted_id})
    return {"success": True, "categoria": serialize_doc(created)}

@router.delete("/categorias-articulo/{cat_id}")
async def delete_categoria_articulo(cat_id: str, current_user: dict = Depends(RequireDelete)):
    if not ObjectId.is_valid(cat_id):
        raise HTTPException(status_code=400, detail="ID no valido")
    result = await categorias_articulo_collection.delete_one({"_id": ObjectId(cat_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Categoria no encontrada")
    return {"success": True}


# ============================================================================
# CULTIVOS
# ============================================================================

@router.post("/cultivos", response_model=dict)
async def create_cultivo(
    cultivo: CultivoCreate,
    current_user: dict = Depends(RequireCreate)
):
    cultivo_dict = cultivo.dict()
    
    # Auto-generate codigo_cultivo
    last = await cultivos_collection.find_one(
        {"codigo_cultivo": {"$exists": True, "$ne": None}},
        sort=[("codigo_cultivo", -1)]
    )
    if last and last.get("codigo_cultivo"):
        try:
            next_num = int(last["codigo_cultivo"]) + 1
        except ValueError:
            next_num = 1
    else:
        count = await cultivos_collection.count_documents({})
        next_num = count + 1
    cultivo_dict['codigo_cultivo'] = str(next_num).zfill(6)
    
    cultivo_dict['created_at'] = datetime.now()
    cultivo_dict['updated_at'] = datetime.now()
    
    result = await cultivos_collection.insert_one(cultivo_dict)
    created = await cultivos_collection.find_one({"_id": result.inserted_id})
    
    await log_cultivo_change(str(result.inserted_id), "creacion", current_user, [{"field": "nombre", "old": "", "new": cultivo_dict.get("nombre", "")}])
    
    return {"success": True, "cultivo": serialize_doc(created)}


@router.get("/cultivos")
async def get_cultivos(
    skip: int = 0,
    limit: int = 10000,
    activo: Optional[bool] = None,
    tipo: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if activo is not None:
        query['activo'] = activo
    if tipo:
        query['tipo'] = tipo
    
    cultivos = await cultivos_collection.find(query).skip(skip).limit(limit).to_list(limit)
    total = await cultivos_collection.count_documents(query)
    
    return {
        "cultivos": serialize_docs(cultivos),
        "total": total
    }


@router.get("/cultivos/{cultivo_id}")
async def get_cultivo(
    cultivo_id: str,
    current_user: dict = Depends(get_current_user)
):
    if not ObjectId.is_valid(cultivo_id):
        raise HTTPException(status_code=400, detail="Invalid ID")
    
    cultivo = await cultivos_collection.find_one({"_id": ObjectId(cultivo_id)})
    if not cultivo:
        raise HTTPException(status_code=404, detail="Cultivo not found")
    
    return {"cultivo": serialize_doc(cultivo)}


@router.put("/cultivos/{cultivo_id}")
async def update_cultivo(
    cultivo_id: str,
    cultivo: CultivoCreate,
    current_user: dict = Depends(RequireEdit)
):
    if not ObjectId.is_valid(cultivo_id):
        raise HTTPException(status_code=400, detail="Invalid ID")
    
    cultivo_dict = cultivo.dict()
    cultivo_dict['updated_at'] = datetime.now()
    
    # Get old data for diff
    old_doc = await cultivos_collection.find_one({"_id": ObjectId(cultivo_id)})
    if not old_doc:
        raise HTTPException(status_code=404, detail="Cultivo not found")
    old_clean = {k: v for k, v in old_doc.items() if k != "_id"}
    changes = diff_dicts(old_clean, cultivo_dict)
    
    result = await cultivos_collection.update_one(
        {"_id": ObjectId(cultivo_id)},
        {"$set": cultivo_dict}
    )
    
    if changes:
        await log_cultivo_change(cultivo_id, "modificacion", current_user, changes)
    
    updated = await cultivos_collection.find_one({"_id": ObjectId(cultivo_id)})
    return {"success": True, "cultivo": serialize_doc(updated)}


@router.delete("/cultivos/{cultivo_id}")
async def delete_cultivo(
    cultivo_id: str,
    current_user: dict = Depends(RequireDelete)
):
    if not ObjectId.is_valid(cultivo_id):
        raise HTTPException(status_code=400, detail="Invalid ID")
    
    result = await cultivos_collection.delete_one({"_id": ObjectId(cultivo_id)})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Cultivo not found")


@router.post("/cultivos/{cultivo_id}/variedades")
async def add_variedad_to_cultivo(
    cultivo_id: str,
    payload: dict,
    current_user: dict = Depends(RequireEdit)
):
    """Add a new variedad to the cultivo's `variedades` list (idempotent).

    Body: {"variedad": "Nombre de la variedad"}
    """
    if not ObjectId.is_valid(cultivo_id):
        raise HTTPException(status_code=400, detail="Invalid ID")

    raw = (payload or {}).get("variedad", "")
    variedad = str(raw).strip()
    if not variedad:
        raise HTTPException(status_code=400, detail="`variedad` no puede estar vacía")

    cultivo = await cultivos_collection.find_one({"_id": ObjectId(cultivo_id)})
    if not cultivo:
        raise HTTPException(status_code=404, detail="Cultivo not found")

    # Idempotente: si ya existe (case-insensitive) no añade duplicado.
    existing = cultivo.get("variedades") or []
    if any(v.lower() == variedad.lower() for v in existing):
        return {"success": True, "variedades": existing, "message": "Variedad ya existía"}

    await cultivos_collection.update_one(
        {"_id": ObjectId(cultivo_id)},
        {
            "$addToSet": {"variedades": variedad},
            "$set": {"updated_at": datetime.now()},
        },
    )

    updated = await cultivos_collection.find_one({"_id": ObjectId(cultivo_id)})
    await log_cultivo_change(
        cultivo_id, "modificacion", current_user,
        [{"campo": "variedades", "antes": existing, "despues": updated.get("variedades", [])}]
    )
    return {"success": True, "variedades": updated.get("variedades", [])}
    
    await log_cultivo_change(cultivo_id, "eliminacion", current_user)
    
    return {"success": True, "message": "Cultivo deleted"}


@router.get("/cultivos/{cultivo_id}/changelog")
async def get_cultivo_changelog(cultivo_id: str, current_user: dict = Depends(get_current_user)):
    logs = await changelog_cultivo_collection.find({"cultivo_id": cultivo_id}).sort("timestamp", -1).to_list(200)
    return {"changelog": serialize_docs(logs)}
