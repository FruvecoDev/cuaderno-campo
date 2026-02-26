"""
Routes for Irrigaciones module - Enhanced with planning, calculations, history, alerts
"""

from fastapi import APIRouter, HTTPException, Depends, Query
from fastapi.responses import StreamingResponse
from typing import Optional, List
from bson import ObjectId
from datetime import datetime, timedelta
import io

from models_tratamientos import IrrigacionCreate
from database import (
    irrigaciones_collection, parcelas_collection, 
    serialize_doc, serialize_docs
)
from rbac_guards import (
    RequireCreate, RequireEdit, RequireDelete,
    RequireIrrigacionesAccess, get_current_user
)

router = APIRouter(prefix="/api", tags=["irrigaciones"])


# ============================================================================
# IRRIGACIONES CRUD - ENHANCED
# ============================================================================

@router.post("/irrigaciones", response_model=dict)
async def create_irrigacion(
    irrigacion: IrrigacionCreate,
    current_user: dict = Depends(RequireCreate),
    _access: dict = Depends(RequireIrrigacionesAccess)
):
    irrigacion_dict = irrigacion.dict()
    
    # Auto-inherit from parcela if available
    if irrigacion_dict.get("parcela_id"):
        parcela = await parcelas_collection.find_one({"_id": ObjectId(irrigacion_dict["parcela_id"])})
        if parcela:
            irrigacion_dict["parcela_codigo"] = parcela.get("codigo_plantacion")
            irrigacion_dict["cultivo"] = parcela.get("cultivo")
            irrigacion_dict["campana"] = parcela.get("campana")
            # Calculate consumption per hectare
            superficie = parcela.get("superficie_total", 0)
            if superficie > 0 and irrigacion_dict.get("volumen"):
                irrigacion_dict["consumo_por_ha"] = round(irrigacion_dict["volumen"] / superficie, 2)
            irrigacion_dict["superficie_regada"] = superficie
    
    # Set estado based on fecha
    if irrigacion_dict.get("es_planificada") or irrigacion_dict.get("fecha_planificada"):
        irrigacion_dict["estado"] = "planificado"
    else:
        irrigacion_dict["estado"] = "completado"
    
    irrigacion_dict.update({
        "created_at": datetime.now(),
        "updated_at": datetime.now(),
        "created_by": str(current_user.get("_id", ""))
    })
    
    result = await irrigaciones_collection.insert_one(irrigacion_dict)
    created = await irrigaciones_collection.find_one({"_id": result.inserted_id})
    
    return {"success": True, "data": serialize_doc(created)}


@router.get("/irrigaciones")
async def get_irrigaciones(
    skip: int = 0,
    limit: int = 100,
    parcela_id: Optional[str] = None,
    sistema: Optional[str] = None,
    estado: Optional[str] = None,
    fecha_desde: Optional[str] = None,
    fecha_hasta: Optional[str] = None,
    cultivo: Optional[str] = None,
    campana: Optional[str] = None,
    es_planificada: Optional[bool] = None,
    current_user: dict = Depends(get_current_user),
    _access: dict = Depends(RequireIrrigacionesAccess)
):
    query = {}
    
    if parcela_id:
        query["parcela_id"] = parcela_id
    
    if sistema:
        query["sistema"] = sistema
    
    if estado:
        query["estado"] = estado
    
    if fecha_desde:
        query["fecha"] = {"$gte": fecha_desde}
    
    if fecha_hasta:
        if "fecha" in query:
            query["fecha"]["$lte"] = fecha_hasta
        else:
            query["fecha"] = {"$lte": fecha_hasta}
    
    if cultivo:
        query["cultivo"] = cultivo
    
    if campana:
        query["campana"] = campana
    
    if es_planificada is not None:
        query["es_planificada"] = es_planificada
    
    irrigaciones = await irrigaciones_collection.find(query).sort("fecha", -1).skip(skip).limit(limit).to_list(limit)
    total = await irrigaciones_collection.count_documents(query)
    
    return {"irrigaciones": serialize_docs(irrigaciones), "total": total}


@router.get("/irrigaciones/planificadas")
async def get_irrigaciones_planificadas(
    dias: int = Query(default=14, ge=1, le=90),
    current_user: dict = Depends(get_current_user),
    _access: dict = Depends(RequireIrrigacionesAccess)
):
    """Get planned irrigations for the next N days"""
    hoy = datetime.now().strftime("%Y-%m-%d")
    fecha_limite = (datetime.now() + timedelta(days=dias)).strftime("%Y-%m-%d")
    
    query = {
        "$or": [
            {"es_planificada": True, "estado": "planificado"},
            {"fecha_planificada": {"$gte": hoy, "$lte": fecha_limite}}
        ]
    }
    
    irrigaciones = await irrigaciones_collection.find(query).sort("fecha_planificada", 1).to_list(100)
    
    return {"planificadas": serialize_docs(irrigaciones), "dias": dias}


@router.get("/irrigaciones/historial/{parcela_id}")
async def get_historial_parcela(
    parcela_id: str,
    limit: int = Query(default=50, ge=1, le=200),
    current_user: dict = Depends(get_current_user),
    _access: dict = Depends(RequireIrrigacionesAccess)
):
    """Get irrigation history for a specific parcel"""
    if not ObjectId.is_valid(parcela_id):
        raise HTTPException(status_code=400, detail="ID de parcela inválido")
    
    # Get parcela info
    parcela = await parcelas_collection.find_one({"_id": ObjectId(parcela_id)})
    if not parcela:
        raise HTTPException(status_code=404, detail="Parcela no encontrada")
    
    # Get irrigations
    irrigaciones = await irrigaciones_collection.find(
        {"parcela_id": parcela_id}
    ).sort("fecha", -1).limit(limit).to_list(limit)
    
    # Calculate totals
    total_volumen = sum(i.get("volumen", 0) for i in irrigaciones)
    total_horas = sum(i.get("duracion", 0) for i in irrigaciones)
    total_coste = sum(i.get("coste", 0) for i in irrigaciones)
    
    # By system
    por_sistema = {}
    for i in irrigaciones:
        sistema = i.get("sistema", "Otro")
        if sistema not in por_sistema:
            por_sistema[sistema] = {"count": 0, "volumen": 0, "horas": 0}
        por_sistema[sistema]["count"] += 1
        por_sistema[sistema]["volumen"] += i.get("volumen", 0)
        por_sistema[sistema]["horas"] += i.get("duracion", 0)
    
    return {
        "parcela": {
            "id": str(parcela["_id"]),
            "codigo": parcela.get("codigo_plantacion"),
            "cultivo": parcela.get("cultivo"),
            "superficie": parcela.get("superficie_total", 0)
        },
        "historial": serialize_docs(irrigaciones),
        "totales": {
            "riegos": len(irrigaciones),
            "volumen_total": round(total_volumen, 2),
            "horas_total": round(total_horas, 2),
            "coste_total": round(total_coste, 2),
            "volumen_por_ha": round(total_volumen / parcela.get("superficie_total", 1), 2) if parcela.get("superficie_total") else 0
        },
        "por_sistema": por_sistema
    }


@router.get("/irrigaciones/stats")
async def get_irrigaciones_stats(
    campana: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
    _access: dict = Depends(RequireIrrigacionesAccess)
):
    """Get irrigation statistics"""
    query = {}
    if campana:
        query["campana"] = campana
    
    total = await irrigaciones_collection.count_documents(query)
    
    # By status
    completados = await irrigaciones_collection.count_documents({**query, "estado": "completado"})
    planificados = await irrigaciones_collection.count_documents({**query, "estado": "planificado"})
    en_curso = await irrigaciones_collection.count_documents({**query, "estado": "en_curso"})
    
    # Aggregations
    pipeline = [
        {"$match": query},
        {"$group": {
            "_id": None,
            "volumen_total": {"$sum": {"$ifNull": ["$volumen", 0]}},
            "horas_total": {"$sum": {"$ifNull": ["$duracion", 0]}},
            "coste_total": {"$sum": {"$ifNull": ["$coste", 0]}},
            "superficie_total": {"$sum": {"$ifNull": ["$superficie_regada", 0]}}
        }}
    ]
    totales = await irrigaciones_collection.aggregate(pipeline).to_list(1)
    totales_data = totales[0] if totales else {
        "volumen_total": 0, "horas_total": 0, "coste_total": 0, "superficie_total": 0
    }
    
    # By system
    pipeline_sistema = [
        {"$match": query},
        {"$group": {
            "_id": "$sistema",
            "count": {"$sum": 1},
            "volumen": {"$sum": {"$ifNull": ["$volumen", 0]}}
        }},
        {"$sort": {"count": -1}}
    ]
    por_sistema = await irrigaciones_collection.aggregate(pipeline_sistema).to_list(10)
    
    # By month (last 12 months)
    hace_12_meses = (datetime.now() - timedelta(days=365)).strftime("%Y-%m-%d")
    pipeline_mes = [
        {"$match": {**query, "fecha": {"$gte": hace_12_meses}}},
        {"$addFields": {
            "mes": {"$substr": ["$fecha", 0, 7]}
        }},
        {"$group": {
            "_id": "$mes",
            "count": {"$sum": 1},
            "volumen": {"$sum": {"$ifNull": ["$volumen", 0]}}
        }},
        {"$sort": {"_id": 1}}
    ]
    por_mes = await irrigaciones_collection.aggregate(pipeline_mes).to_list(12)
    
    # Upcoming (next 7 days)
    hoy = datetime.now().strftime("%Y-%m-%d")
    proxima_semana = (datetime.now() + timedelta(days=7)).strftime("%Y-%m-%d")
    proximos = await irrigaciones_collection.count_documents({
        "estado": "planificado",
        "$or": [
            {"fecha_planificada": {"$gte": hoy, "$lte": proxima_semana}},
            {"fecha": {"$gte": hoy, "$lte": proxima_semana}, "es_planificada": True}
        ]
    })
    
    return {
        "total": total,
        "completados": completados,
        "planificados": planificados,
        "en_curso": en_curso,
        "proximos_7_dias": proximos,
        "totales": {
            "volumen": round(totales_data.get("volumen_total", 0), 2),
            "horas": round(totales_data.get("horas_total", 0), 2),
            "coste": round(totales_data.get("coste_total", 0), 2),
            "superficie": round(totales_data.get("superficie_total", 0), 2)
        },
        "por_sistema": [{"sistema": s["_id"] or "Otro", "count": s["count"], "volumen": round(s["volumen"], 2)} for s in por_sistema],
        "por_mes": [{"mes": m["_id"], "count": m["count"], "volumen": round(m["volumen"], 2)} for m in por_mes]
    }


@router.get("/irrigaciones/calcular-consumo")
async def calcular_consumo(
    parcela_id: str = Query(...),
    volumen: float = Query(..., ge=0),
    current_user: dict = Depends(get_current_user)
):
    """Calculate consumption per hectare for a parcel"""
    if not ObjectId.is_valid(parcela_id):
        raise HTTPException(status_code=400, detail="ID de parcela inválido")
    
    parcela = await parcelas_collection.find_one({"_id": ObjectId(parcela_id)})
    if not parcela:
        raise HTTPException(status_code=404, detail="Parcela no encontrada")
    
    superficie = parcela.get("superficie_total", 0)
    consumo_por_ha = round(volumen / superficie, 2) if superficie > 0 else 0
    
    return {
        "parcela_codigo": parcela.get("codigo_plantacion"),
        "superficie_ha": superficie,
        "volumen_m3": volumen,
        "consumo_por_ha": consumo_por_ha,
        "cultivo": parcela.get("cultivo")
    }


@router.get("/irrigaciones/sistemas")
async def get_sistemas():
    """Get irrigation systems"""
    return {
        "sistemas": [
            {"id": "goteo", "nombre": "Goteo", "descripcion": "Riego localizado por goteo"},
            {"id": "aspersion", "nombre": "Aspersión", "descripcion": "Riego por aspersores"},
            {"id": "microaspersion", "nombre": "Microaspersión", "descripcion": "Riego por microaspersores"},
            {"id": "inundacion", "nombre": "Inundación", "descripcion": "Riego por inundación"},
            {"id": "pivot", "nombre": "Pivot", "descripcion": "Riego por pivot central"},
            {"id": "manguera", "nombre": "Manguera", "descripcion": "Riego manual con manguera"},
            {"id": "surcos", "nombre": "Surcos", "descripcion": "Riego por surcos"}
        ],
        "fuentes_agua": [
            {"id": "pozo", "nombre": "Pozo"},
            {"id": "embalse", "nombre": "Embalse"},
            {"id": "red", "nombre": "Red municipal"},
            {"id": "canal", "nombre": "Canal"},
            {"id": "balsa", "nombre": "Balsa"},
            {"id": "rio", "nombre": "Río"}
        ],
        "estados": [
            {"id": "planificado", "nombre": "Planificado"},
            {"id": "en_curso", "nombre": "En Curso"},
            {"id": "completado", "nombre": "Completado"},
            {"id": "cancelado", "nombre": "Cancelado"}
        ]
    }


@router.get("/irrigaciones/{irrigacion_id}")
async def get_irrigacion(
    irrigacion_id: str,
    current_user: dict = Depends(get_current_user),
    _access: dict = Depends(RequireIrrigacionesAccess)
):
    if not ObjectId.is_valid(irrigacion_id):
        raise HTTPException(status_code=400, detail="ID de irrigación inválido")
    
    irrigacion = await irrigaciones_collection.find_one({"_id": ObjectId(irrigacion_id)})
    if not irrigacion:
        raise HTTPException(status_code=404, detail="Irrigación no encontrada")
    
    return {"success": True, "data": serialize_doc(irrigacion)}


@router.put("/irrigaciones/{irrigacion_id}")
async def update_irrigacion(
    irrigacion_id: str,
    irrigacion: IrrigacionCreate,
    current_user: dict = Depends(RequireEdit),
    _access: dict = Depends(RequireIrrigacionesAccess)
):
    if not ObjectId.is_valid(irrigacion_id):
        raise HTTPException(status_code=400, detail="ID de irrigación inválido")
    
    irrigacion_dict = irrigacion.dict()
    
    # Recalculate consumption if parcela and volume changed
    if irrigacion_dict.get("parcela_id") and irrigacion_dict.get("volumen"):
        parcela = await parcelas_collection.find_one({"_id": ObjectId(irrigacion_dict["parcela_id"])})
        if parcela:
            superficie = parcela.get("superficie_total", 0)
            if superficie > 0:
                irrigacion_dict["consumo_por_ha"] = round(irrigacion_dict["volumen"] / superficie, 2)
            irrigacion_dict["superficie_regada"] = superficie
            irrigacion_dict["parcela_codigo"] = parcela.get("codigo_plantacion")
            irrigacion_dict["cultivo"] = parcela.get("cultivo")
            irrigacion_dict["campana"] = parcela.get("campana")
    
    irrigacion_dict["updated_at"] = datetime.now()
    
    result = await irrigaciones_collection.update_one(
        {"_id": ObjectId(irrigacion_id)},
        {"$set": irrigacion_dict}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Irrigación no encontrada")
    
    updated = await irrigaciones_collection.find_one({"_id": ObjectId(irrigacion_id)})
    return {"success": True, "data": serialize_doc(updated)}


@router.patch("/irrigaciones/{irrigacion_id}/estado")
async def update_irrigacion_estado(
    irrigacion_id: str,
    estado: str = Query(...),
    current_user: dict = Depends(RequireEdit),
    _access: dict = Depends(RequireIrrigacionesAccess)
):
    """Quick status update"""
    if not ObjectId.is_valid(irrigacion_id):
        raise HTTPException(status_code=400, detail="ID de irrigación inválido")
    
    update_data = {
        "estado": estado,
        "updated_at": datetime.now()
    }
    
    if estado == "completado":
        update_data["fecha_completado"] = datetime.now().strftime("%Y-%m-%d")
    elif estado == "en_curso":
        update_data["hora_inicio"] = datetime.now().strftime("%H:%M")
    
    result = await irrigaciones_collection.update_one(
        {"_id": ObjectId(irrigacion_id)},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Irrigación no encontrada")
    
    return {"success": True, "message": f"Estado actualizado a {estado}"}


@router.delete("/irrigaciones/{irrigacion_id}")
async def delete_irrigacion(
    irrigacion_id: str,
    current_user: dict = Depends(RequireDelete),
    _access: dict = Depends(RequireIrrigacionesAccess)
):
    if not ObjectId.is_valid(irrigacion_id):
        raise HTTPException(status_code=400, detail="ID de irrigación inválido")
    
    result = await irrigaciones_collection.delete_one({"_id": ObjectId(irrigacion_id)})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Irrigación no encontrada")
    
    return {"success": True, "message": "Irrigación eliminada correctamente"}


# ============================================================================
# EXPORT
# ============================================================================

@router.get("/irrigaciones/export/excel")
async def export_irrigaciones_excel(
    parcela_id: Optional[str] = None,
    fecha_desde: Optional[str] = None,
    fecha_hasta: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
    _access: dict = Depends(RequireIrrigacionesAccess)
):
    """Export irrigations to Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    
    query = {}
    if parcela_id:
        query["parcela_id"] = parcela_id
    if fecha_desde:
        query["fecha"] = {"$gte": fecha_desde}
    if fecha_hasta:
        if "fecha" in query:
            query["fecha"]["$lte"] = fecha_hasta
        else:
            query["fecha"] = {"$lte": fecha_hasta}
    
    irrigaciones = await irrigaciones_collection.find(query).sort("fecha", -1).to_list(1000)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Irrigaciones"
    
    # Headers
    headers = ["Fecha", "Parcela", "Cultivo", "Sistema", "Duración (h)", "Volumen (m³)", "m³/ha", "Coste (€)", "Estado", "Observaciones"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="1976d2", end_color="1976d2", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    
    # Data
    for row, irrig in enumerate(irrigaciones, 2):
        ws.cell(row=row, column=1, value=irrig.get("fecha", ""))
        ws.cell(row=row, column=2, value=irrig.get("parcela_codigo", ""))
        ws.cell(row=row, column=3, value=irrig.get("cultivo", ""))
        ws.cell(row=row, column=4, value=irrig.get("sistema", ""))
        ws.cell(row=row, column=5, value=irrig.get("duracion", 0))
        ws.cell(row=row, column=6, value=irrig.get("volumen", 0))
        ws.cell(row=row, column=7, value=irrig.get("consumo_por_ha", 0))
        ws.cell(row=row, column=8, value=irrig.get("coste", 0))
        ws.cell(row=row, column=9, value=irrig.get("estado", "completado"))
        ws.cell(row=row, column=10, value=irrig.get("observaciones", ""))
    
    # Auto-width
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 40)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"irrigaciones_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
