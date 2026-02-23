from fastapi import APIRouter, HTTPException, Depends, Query
from typing import Optional, List
from datetime import datetime
from bson import ObjectId

from database import (
    albaranes_collection, contratos_collection, serialize_doc, serialize_docs, db
)
from rbac_guards import RequireAlbaranesAccess, get_current_user

router = APIRouter(prefix="/api/gastos", tags=["gastos"])

def parse_date(date_str: str) -> datetime:
    """Parse date string to datetime"""
    try:
        return datetime.strptime(date_str, "%Y-%m-%d")
    except ValueError:
        return None

# ============================================================================
# INFORMES DE GASTOS
# ============================================================================

@router.get("/resumen")
async def get_resumen_gastos(
    fecha_desde: Optional[str] = None,
    fecha_hasta: Optional[str] = None,
    campana: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
    _access: dict = Depends(RequireAlbaranesAccess)
):
    """
    Obtiene un resumen general de gastos con totales por proveedor, contrato, cultivo y parcela.
    """
    # Build match query
    match_query = {}
    
    if fecha_desde:
        match_query["fecha"] = {"$gte": fecha_desde}
    if fecha_hasta:
        if "fecha" in match_query:
            match_query["fecha"]["$lte"] = fecha_hasta
        else:
            match_query["fecha"] = {"$lte": fecha_hasta}
    if campana:
        match_query["campana"] = campana
    
    # Get all albaranes matching criteria
    albaranes = await albaranes_collection.find(match_query).to_list(1000)
    
    # Aggregate by different dimensions
    por_proveedor = {}
    por_contrato = {}
    por_cultivo = {}
    por_parcela = {}
    total_general = 0
    total_albaranes = len(albaranes)
    
    for albaran in albaranes:
        total = albaran.get("total_albaran", 0) or 0
        total_general += total
        
        # Por proveedor
        proveedor = albaran.get("proveedor") or "Sin proveedor"
        if proveedor not in por_proveedor:
            por_proveedor[proveedor] = {"total": 0, "count": 0, "albaranes": []}
        por_proveedor[proveedor]["total"] += total
        por_proveedor[proveedor]["count"] += 1
        
        # Por contrato
        contrato_id = albaran.get("contrato_id")
        if contrato_id:
            if contrato_id not in por_contrato:
                por_contrato[contrato_id] = {"total": 0, "count": 0, "proveedor": proveedor}
            por_contrato[contrato_id]["total"] += total
            por_contrato[contrato_id]["count"] += 1
        
        # Por cultivo
        cultivo = albaran.get("cultivo") or "Sin cultivo"
        if cultivo not in por_cultivo:
            por_cultivo[cultivo] = {"total": 0, "count": 0}
        por_cultivo[cultivo]["total"] += total
        por_cultivo[cultivo]["count"] += 1
        
        # Por parcela
        parcela = albaran.get("parcela_codigo") or "Sin parcela"
        if parcela not in por_parcela:
            por_parcela[parcela] = {"total": 0, "count": 0, "cultivo": cultivo}
        por_parcela[parcela]["total"] += total
        por_parcela[parcela]["count"] += 1
    
    # Enrich contract data with contract info
    contratos_enriched = []
    for contrato_id, data in por_contrato.items():
        try:
            contrato = await contratos_collection.find_one({"_id": ObjectId(contrato_id)})
            contratos_enriched.append({
                "contrato_id": contrato_id,
                "numero_contrato": contrato.get("numero_contrato") if contrato else None,
                "proveedor": data["proveedor"],
                "cultivo": contrato.get("cultivo") if contrato else None,
                "campana": contrato.get("campana") if contrato else None,
                "total": data["total"],
                "count": data["count"]
            })
        except Exception:
            contratos_enriched.append({
                "contrato_id": contrato_id,
                "total": data["total"],
                "count": data["count"],
                "proveedor": data["proveedor"]
            })
    
    return {
        "total_general": total_general,
        "total_albaranes": total_albaranes,
        "por_proveedor": [{"proveedor": k, **v} for k, v in sorted(por_proveedor.items(), key=lambda x: -x[1]["total"])],
        "por_contrato": sorted(contratos_enriched, key=lambda x: -x["total"]),
        "por_cultivo": [{"cultivo": k, **v} for k, v in sorted(por_cultivo.items(), key=lambda x: -x[1]["total"])],
        "por_parcela": [{"parcela": k, **v} for k, v in sorted(por_parcela.items(), key=lambda x: -x[1]["total"])]
    }


@router.get("/por-proveedor")
async def get_gastos_por_proveedor(
    proveedor: Optional[str] = None,
    fecha_desde: Optional[str] = None,
    fecha_hasta: Optional[str] = None,
    campana: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
    _access: dict = Depends(RequireAlbaranesAccess)
):
    """
    Obtiene detalle de gastos agrupados por proveedor.
    Si se especifica proveedor, devuelve el detalle de ese proveedor.
    """
    match_query = {}
    if proveedor:
        match_query["proveedor"] = proveedor
    if fecha_desde:
        match_query["fecha"] = {"$gte": fecha_desde}
    if fecha_hasta:
        if "fecha" in match_query:
            match_query["fecha"]["$lte"] = fecha_hasta
        else:
            match_query["fecha"] = {"$lte": fecha_hasta}
    if campana:
        match_query["campana"] = campana
    
    pipeline = [
        {"$match": match_query},
        {"$group": {
            "_id": "$proveedor",
            "total": {"$sum": "$total_albaran"},
            "count": {"$sum": 1},
            "cultivos": {"$addToSet": "$cultivo"},
            "contratos": {"$addToSet": "$contrato_id"},
            "primer_albaran": {"$min": "$fecha"},
            "ultimo_albaran": {"$max": "$fecha"}
        }},
        {"$sort": {"total": -1}}
    ]
    
    results = await albaranes_collection.aggregate(pipeline).to_list(100)
    
    return {
        "gastos_por_proveedor": [
            {
                "proveedor": r["_id"] or "Sin proveedor",
                "total": r["total"],
                "count": r["count"],
                "cultivos": [c for c in r["cultivos"] if c],
                "num_contratos": len([c for c in r["contratos"] if c]),
                "primer_albaran": r["primer_albaran"],
                "ultimo_albaran": r["ultimo_albaran"]
            }
            for r in results
        ]
    }


@router.get("/por-contrato")
async def get_gastos_por_contrato(
    contrato_id: Optional[str] = None,
    fecha_desde: Optional[str] = None,
    fecha_hasta: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
    _access: dict = Depends(RequireAlbaranesAccess)
):
    """
    Obtiene detalle de gastos agrupados por contrato.
    Si se especifica contrato_id, devuelve el detalle de ese contrato.
    """
    match_query = {}
    if contrato_id:
        match_query["contrato_id"] = contrato_id
    if fecha_desde:
        match_query["fecha"] = {"$gte": fecha_desde}
    if fecha_hasta:
        if "fecha" in match_query:
            match_query["fecha"]["$lte"] = fecha_hasta
        else:
            match_query["fecha"] = {"$lte": fecha_hasta}
    
    pipeline = [
        {"$match": match_query},
        {"$group": {
            "_id": "$contrato_id",
            "total": {"$sum": "$total_albaran"},
            "count": {"$sum": 1},
            "proveedor": {"$first": "$proveedor"},
            "cultivo": {"$first": "$cultivo"},
            "campana": {"$first": "$campana"},
            "parcela": {"$first": "$parcela_codigo"},
            "primer_albaran": {"$min": "$fecha"},
            "ultimo_albaran": {"$max": "$fecha"}
        }},
        {"$sort": {"total": -1}}
    ]
    
    results = await albaranes_collection.aggregate(pipeline).to_list(100)
    
    # Enrich with contract details
    enriched_results = []
    for r in results:
        contrato_info = {}
        if r["_id"]:
            try:
                contrato = await contratos_collection.find_one({"_id": ObjectId(r["_id"])})
                if contrato:
                    contrato_info = {
                        "numero_contrato": contrato.get("numero_contrato"),
                        "precio": contrato.get("precio"),
                        "superficie": contrato.get("superficie")
                    }
            except Exception:
                pass
        
        enriched_results.append({
            "contrato_id": r["_id"],
            **contrato_info,
            "proveedor": r["proveedor"],
            "cultivo": r["cultivo"],
            "campana": r["campana"],
            "parcela": r["parcela"],
            "total": r["total"],
            "count": r["count"],
            "primer_albaran": r["primer_albaran"],
            "ultimo_albaran": r["ultimo_albaran"]
        })
    
    return {"gastos_por_contrato": enriched_results}


@router.get("/por-cultivo")
async def get_gastos_por_cultivo(
    cultivo: Optional[str] = None,
    fecha_desde: Optional[str] = None,
    fecha_hasta: Optional[str] = None,
    campana: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
    _access: dict = Depends(RequireAlbaranesAccess)
):
    """
    Obtiene detalle de gastos agrupados por cultivo.
    """
    match_query = {}
    if cultivo:
        match_query["cultivo"] = cultivo
    if fecha_desde:
        match_query["fecha"] = {"$gte": fecha_desde}
    if fecha_hasta:
        if "fecha" in match_query:
            match_query["fecha"]["$lte"] = fecha_hasta
        else:
            match_query["fecha"] = {"$lte": fecha_hasta}
    if campana:
        match_query["campana"] = campana
    
    pipeline = [
        {"$match": match_query},
        {"$group": {
            "_id": "$cultivo",
            "total": {"$sum": "$total_albaran"},
            "count": {"$sum": 1},
            "proveedores": {"$addToSet": "$proveedor"},
            "parcelas": {"$addToSet": "$parcela_codigo"},
            "primer_albaran": {"$min": "$fecha"},
            "ultimo_albaran": {"$max": "$fecha"}
        }},
        {"$sort": {"total": -1}}
    ]
    
    results = await albaranes_collection.aggregate(pipeline).to_list(100)
    
    return {
        "gastos_por_cultivo": [
            {
                "cultivo": r["_id"] or "Sin cultivo",
                "total": r["total"],
                "count": r["count"],
                "num_proveedores": len([p for p in r["proveedores"] if p]),
                "num_parcelas": len([p for p in r["parcelas"] if p]),
                "primer_albaran": r["primer_albaran"],
                "ultimo_albaran": r["ultimo_albaran"]
            }
            for r in results
        ]
    }


@router.get("/por-parcela")
async def get_gastos_por_parcela(
    parcela_codigo: Optional[str] = None,
    fecha_desde: Optional[str] = None,
    fecha_hasta: Optional[str] = None,
    campana: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
    _access: dict = Depends(RequireAlbaranesAccess)
):
    """
    Obtiene detalle de gastos agrupados por parcela.
    """
    match_query = {}
    if parcela_codigo:
        match_query["parcela_codigo"] = parcela_codigo
    if fecha_desde:
        match_query["fecha"] = {"$gte": fecha_desde}
    if fecha_hasta:
        if "fecha" in match_query:
            match_query["fecha"]["$lte"] = fecha_hasta
        else:
            match_query["fecha"] = {"$lte": fecha_hasta}
    if campana:
        match_query["campana"] = campana
    
    # Get parcelas info for enrichment
    parcelas_collection = db['parcelas']
    
    pipeline = [
        {"$match": match_query},
        {"$group": {
            "_id": "$parcela_codigo",
            "total": {"$sum": "$total_albaran"},
            "count": {"$sum": 1},
            "cultivo": {"$first": "$cultivo"},
            "proveedor": {"$first": "$proveedor"},
            "campana": {"$first": "$campana"},
            "parcela_id": {"$first": "$parcela_id"},
            "primer_albaran": {"$min": "$fecha"},
            "ultimo_albaran": {"$max": "$fecha"}
        }},
        {"$sort": {"total": -1}}
    ]
    
    results = await albaranes_collection.aggregate(pipeline).to_list(100)
    
    # Enrich with parcela details (superficie)
    enriched_results = []
    for r in results:
        parcela_info = {}
        if r.get("parcela_id"):
            try:
                parcela = await parcelas_collection.find_one({"_id": ObjectId(r["parcela_id"])})
                if parcela:
                    parcela_info = {
                        "superficie": parcela.get("superficie"),
                        "finca": parcela.get("finca")
                    }
            except Exception:
                pass
        
        coste_por_ha = None
        if parcela_info.get("superficie") and parcela_info["superficie"] > 0:
            coste_por_ha = round(r["total"] / parcela_info["superficie"], 2)
        
        enriched_results.append({
            "parcela_codigo": r["_id"] or "Sin parcela",
            "cultivo": r["cultivo"],
            "proveedor": r["proveedor"],
            "campana": r["campana"],
            **parcela_info,
            "total": r["total"],
            "count": r["count"],
            "coste_por_ha": coste_por_ha,
            "primer_albaran": r["primer_albaran"],
            "ultimo_albaran": r["ultimo_albaran"]
        })
    
    return {"gastos_por_parcela": enriched_results}


@router.get("/detalle-albaranes")
async def get_detalle_albaranes(
    proveedor: Optional[str] = None,
    contrato_id: Optional[str] = None,
    cultivo: Optional[str] = None,
    parcela_codigo: Optional[str] = None,
    fecha_desde: Optional[str] = None,
    fecha_hasta: Optional[str] = None,
    campana: Optional[str] = None,
    skip: int = 0,
    limit: int = 50,
    current_user: dict = Depends(get_current_user),
    _access: dict = Depends(RequireAlbaranesAccess)
):
    """
    Obtiene el detalle de albaranes filtrados por cualquier combinación de criterios.
    Útil para ver los albaranes específicos de un proveedor, contrato, cultivo o parcela.
    """
    match_query = {}
    if proveedor:
        match_query["proveedor"] = proveedor
    if contrato_id:
        match_query["contrato_id"] = contrato_id
    if cultivo:
        match_query["cultivo"] = cultivo
    if parcela_codigo:
        match_query["parcela_codigo"] = parcela_codigo
    if fecha_desde:
        match_query["fecha"] = {"$gte": fecha_desde}
    if fecha_hasta:
        if "fecha" in match_query:
            match_query["fecha"]["$lte"] = fecha_hasta
        else:
            match_query["fecha"] = {"$lte": fecha_hasta}
    if campana:
        match_query["campana"] = campana
    
    albaranes = await albaranes_collection.find(match_query).sort("fecha", -1).skip(skip).limit(limit).to_list(limit)
    total = await albaranes_collection.count_documents(match_query)
    
    # Calculate total sum
    total_sum_pipeline = [
        {"$match": match_query},
        {"$group": {"_id": None, "total": {"$sum": "$total_albaran"}}}
    ]
    total_sum_result = await albaranes_collection.aggregate(total_sum_pipeline).to_list(1)
    total_sum = total_sum_result[0]["total"] if total_sum_result else 0
    
    return {
        "albaranes": serialize_docs(albaranes),
        "total_count": total,
        "total_sum": total_sum,
        "skip": skip,
        "limit": limit
    }


@router.get("/campanas")
async def get_campanas_disponibles(
    current_user: dict = Depends(get_current_user),
    _access: dict = Depends(RequireAlbaranesAccess)
):
    """
    Obtiene lista de campañas disponibles para filtrar.
    """
    pipeline = [
        {"$match": {"$and": [{"campana": {"$ne": None}}, {"campana": {"$ne": ""}}]}},
        {"$group": {"_id": "$campana"}},
        {"$sort": {"_id": -1}}
    ]
    results = await albaranes_collection.aggregate(pipeline).to_list(100)
    return {"campanas": [r["_id"] for r in results if r["_id"]]}


# ============================================================================
# EXPORTACIÓN DE INFORMES
# ============================================================================

@router.get("/export/excel")
async def export_gastos_excel(
    fecha_desde: Optional[str] = None,
    fecha_hasta: Optional[str] = None,
    campana: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
    _access: dict = Depends(RequireAlbaranesAccess)
):
    """
    Exporta el informe de gastos a Excel.
    """
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io
    
    # Get resumen data
    match_query = {}
    if fecha_desde:
        match_query["fecha"] = {"$gte": fecha_desde}
    if fecha_hasta:
        if "fecha" in match_query:
            match_query["fecha"]["$lte"] = fecha_hasta
        else:
            match_query["fecha"] = {"$lte": fecha_hasta}
    if campana:
        match_query["campana"] = campana
    
    albaranes = await albaranes_collection.find(match_query).to_list(1000)
    
    # Aggregate data
    por_proveedor = {}
    por_cultivo = {}
    por_parcela = {}
    total_general = 0
    
    for albaran in albaranes:
        total = albaran.get("total_albaran", 0) or 0
        total_general += total
        
        proveedor = albaran.get("proveedor") or "Sin proveedor"
        if proveedor not in por_proveedor:
            por_proveedor[proveedor] = {"total": 0, "count": 0}
        por_proveedor[proveedor]["total"] += total
        por_proveedor[proveedor]["count"] += 1
        
        cultivo = albaran.get("cultivo") or "Sin cultivo"
        if cultivo not in por_cultivo:
            por_cultivo[cultivo] = {"total": 0, "count": 0}
        por_cultivo[cultivo]["total"] += total
        por_cultivo[cultivo]["count"] += 1
        
        parcela = albaran.get("parcela_codigo") or "Sin parcela"
        if parcela not in por_parcela:
            por_parcela[parcela] = {"total": 0, "count": 0, "cultivo": cultivo}
        por_parcela[parcela]["total"] += total
        por_parcela[parcela]["count"] += 1
    
    # Create workbook
    wb = Workbook()
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="166534", end_color="166534", fill_type="solid")
    currency_format = '#,##0.00 €'
    percent_format = '0.0%'
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Sheet 1: Resumen
    ws = wb.active
    ws.title = "Resumen"
    
    ws['A1'] = "INFORME DE GASTOS - FRUVECO"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:D1')
    
    ws['A3'] = "Período:"
    ws['B3'] = f"{fecha_desde or 'Inicio'} - {fecha_hasta or 'Fin'}"
    ws['A4'] = "Campaña:"
    ws['B4'] = campana or "Todas"
    ws['A5'] = "Total Gastos:"
    ws['B5'] = total_general
    ws['B5'].number_format = currency_format
    ws['B5'].font = Font(bold=True, size=14, color="166534")
    ws['A6'] = "Total Albaranes:"
    ws['B6'] = len(albaranes)
    
    # Sheet 2: Por Proveedor
    ws_prov = wb.create_sheet("Por Proveedor")
    headers = ["Proveedor", "Albaranes", "Total", "%"]
    for col, header in enumerate(headers, 1):
        cell = ws_prov.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    row = 2
    for prov, data in sorted(por_proveedor.items(), key=lambda x: -x[1]["total"]):
        ws_prov.cell(row=row, column=1, value=prov).border = thin_border
        ws_prov.cell(row=row, column=2, value=data["count"]).border = thin_border
        cell_total = ws_prov.cell(row=row, column=3, value=data["total"])
        cell_total.number_format = currency_format
        cell_total.border = thin_border
        cell_pct = ws_prov.cell(row=row, column=4, value=data["total"]/total_general if total_general > 0 else 0)
        cell_pct.number_format = percent_format
        cell_pct.border = thin_border
        row += 1
    
    for col in range(1, 5):
        ws_prov.column_dimensions[get_column_letter(col)].width = 20
    
    # Sheet 3: Por Cultivo
    ws_cult = wb.create_sheet("Por Cultivo")
    headers = ["Cultivo", "Albaranes", "Total", "%"]
    for col, header in enumerate(headers, 1):
        cell = ws_cult.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    row = 2
    for cult, data in sorted(por_cultivo.items(), key=lambda x: -x[1]["total"]):
        ws_cult.cell(row=row, column=1, value=cult).border = thin_border
        ws_cult.cell(row=row, column=2, value=data["count"]).border = thin_border
        cell_total = ws_cult.cell(row=row, column=3, value=data["total"])
        cell_total.number_format = currency_format
        cell_total.border = thin_border
        cell_pct = ws_cult.cell(row=row, column=4, value=data["total"]/total_general if total_general > 0 else 0)
        cell_pct.number_format = percent_format
        cell_pct.border = thin_border
        row += 1
    
    for col in range(1, 5):
        ws_cult.column_dimensions[get_column_letter(col)].width = 20
    
    # Sheet 4: Por Parcela
    ws_parc = wb.create_sheet("Por Parcela")
    headers = ["Parcela", "Cultivo", "Albaranes", "Total"]
    for col, header in enumerate(headers, 1):
        cell = ws_parc.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    row = 2
    for parc, data in sorted(por_parcela.items(), key=lambda x: -x[1]["total"]):
        ws_parc.cell(row=row, column=1, value=parc).border = thin_border
        ws_parc.cell(row=row, column=2, value=data["cultivo"]).border = thin_border
        ws_parc.cell(row=row, column=3, value=data["count"]).border = thin_border
        cell_total = ws_parc.cell(row=row, column=4, value=data["total"])
        cell_total.number_format = currency_format
        cell_total.border = thin_border
        row += 1
    
    for col in range(1, 5):
        ws_parc.column_dimensions[get_column_letter(col)].width = 20
    
    # Sheet 5: Detalle Albaranes
    ws_det = wb.create_sheet("Detalle Albaranes")
    headers = ["Fecha", "Tipo", "Proveedor", "Cultivo", "Parcela", "Total"]
    for col, header in enumerate(headers, 1):
        cell = ws_det.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    row = 2
    for alb in sorted(albaranes, key=lambda x: x.get("fecha", ""), reverse=True):
        ws_det.cell(row=row, column=1, value=alb.get("fecha", "")).border = thin_border
        ws_det.cell(row=row, column=2, value=alb.get("tipo", "")).border = thin_border
        ws_det.cell(row=row, column=3, value=alb.get("proveedor", "")).border = thin_border
        ws_det.cell(row=row, column=4, value=alb.get("cultivo", "")).border = thin_border
        ws_det.cell(row=row, column=5, value=alb.get("parcela_codigo", "")).border = thin_border
        cell_total = ws_det.cell(row=row, column=6, value=alb.get("total_albaran", 0))
        cell_total.number_format = currency_format
        cell_total.border = thin_border
        row += 1
    
    for col in range(1, 7):
        ws_det.column_dimensions[get_column_letter(col)].width = 18
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"informe_gastos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/export/pdf")
async def export_gastos_pdf(
    fecha_desde: Optional[str] = None,
    fecha_hasta: Optional[str] = None,
    campana: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
    _access: dict = Depends(RequireAlbaranesAccess)
):
    """
    Exporta el informe de gastos a PDF.
    """
    from fastapi.responses import StreamingResponse
    from weasyprint import HTML
    import io
    
    # Get data
    match_query = {}
    if fecha_desde:
        match_query["fecha"] = {"$gte": fecha_desde}
    if fecha_hasta:
        if "fecha" in match_query:
            match_query["fecha"]["$lte"] = fecha_hasta
        else:
            match_query["fecha"] = {"$lte": fecha_hasta}
    if campana:
        match_query["campana"] = campana
    
    albaranes = await albaranes_collection.find(match_query).to_list(1000)
    
    # Aggregate
    por_proveedor = {}
    por_cultivo = {}
    por_parcela = {}
    total_general = 0
    
    for albaran in albaranes:
        total = albaran.get("total_albaran", 0) or 0
        total_general += total
        
        proveedor = albaran.get("proveedor") or "Sin proveedor"
        if proveedor not in por_proveedor:
            por_proveedor[proveedor] = {"total": 0, "count": 0}
        por_proveedor[proveedor]["total"] += total
        por_proveedor[proveedor]["count"] += 1
        
        cultivo = albaran.get("cultivo") or "Sin cultivo"
        if cultivo not in por_cultivo:
            por_cultivo[cultivo] = {"total": 0, "count": 0}
        por_cultivo[cultivo]["total"] += total
        por_cultivo[cultivo]["count"] += 1
        
        parcela = albaran.get("parcela_codigo") or "Sin parcela"
        if parcela not in por_parcela:
            por_parcela[parcela] = {"total": 0, "count": 0, "cultivo": cultivo}
        por_parcela[parcela]["total"] += total
        por_parcela[parcela]["count"] += 1
    
    # Generate HTML
    def format_currency(val):
        return f"{val:,.2f} €".replace(",", "X").replace(".", ",").replace("X", ".")
    
    def format_percent(val, total):
        if total == 0:
            return "0,0%"
        return f"{(val/total)*100:.1f}%".replace(".", ",")
    
    # Proveedor rows
    prov_rows = ""
    for prov, data in sorted(por_proveedor.items(), key=lambda x: -x[1]["total"]):
        prov_rows += f"""
        <tr>
            <td>{prov}</td>
            <td style="text-align:center">{data['count']}</td>
            <td style="text-align:right; font-weight:600; color:#166534">{format_currency(data['total'])}</td>
            <td style="text-align:right">{format_percent(data['total'], total_general)}</td>
        </tr>"""
    
    # Cultivo rows
    cult_rows = ""
    for cult, data in sorted(por_cultivo.items(), key=lambda x: -x[1]["total"]):
        cult_rows += f"""
        <tr>
            <td>{cult}</td>
            <td style="text-align:center">{data['count']}</td>
            <td style="text-align:right; font-weight:600; color:#166534">{format_currency(data['total'])}</td>
            <td style="text-align:right">{format_percent(data['total'], total_general)}</td>
        </tr>"""
    
    # Parcela rows
    parc_rows = ""
    for parc, data in sorted(por_parcela.items(), key=lambda x: -x[1]["total"]):
        parc_rows += f"""
        <tr>
            <td>{parc}</td>
            <td>{data['cultivo']}</td>
            <td style="text-align:center">{data['count']}</td>
            <td style="text-align:right; font-weight:600; color:#166534">{format_currency(data['total'])}</td>
        </tr>"""
    
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <style>
            @page {{ size: A4; margin: 1.5cm; }}
            body {{ font-family: Arial, sans-serif; font-size: 11px; color: #333; }}
            .header {{ text-align: center; margin-bottom: 20px; border-bottom: 3px solid #166534; padding-bottom: 15px; }}
            .header h1 {{ color: #166534; margin: 0; font-size: 24px; }}
            .header p {{ margin: 5px 0 0 0; color: #666; }}
            .kpis {{ display: flex; justify-content: space-around; margin: 20px 0; }}
            .kpi {{ text-align: center; padding: 15px; background: #f0fdf4; border-radius: 8px; min-width: 120px; }}
            .kpi-value {{ font-size: 24px; font-weight: bold; color: #166534; }}
            .kpi-label {{ font-size: 10px; color: #666; margin-top: 5px; }}
            .section {{ margin: 25px 0; }}
            .section h2 {{ color: #166534; font-size: 14px; border-bottom: 2px solid #166534; padding-bottom: 5px; margin-bottom: 10px; }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 10px; }}
            th {{ background: #166534; color: white; padding: 8px; text-align: left; font-size: 10px; }}
            td {{ padding: 6px 8px; border-bottom: 1px solid #e5e7eb; font-size: 10px; }}
            tr:nth-child(even) {{ background: #f9fafb; }}
            .footer {{ text-align: center; margin-top: 30px; padding-top: 15px; border-top: 1px solid #e5e7eb; font-size: 9px; color: #666; }}
        </style>
    </head>
    <body>
        <div class="header">
            <h1>FRUVECO - Informe de Gastos</h1>
            <p>Período: {fecha_desde or 'Inicio'} - {fecha_hasta or 'Fin'} | Campaña: {campana or 'Todas'}</p>
        </div>
        
        <div class="kpis">
            <div class="kpi">
                <div class="kpi-value">{format_currency(total_general)}</div>
                <div class="kpi-label">TOTAL GASTOS</div>
            </div>
            <div class="kpi">
                <div class="kpi-value">{len(albaranes)}</div>
                <div class="kpi-label">ALBARANES</div>
            </div>
            <div class="kpi">
                <div class="kpi-value">{len(por_proveedor)}</div>
                <div class="kpi-label">PROVEEDORES</div>
            </div>
            <div class="kpi">
                <div class="kpi-value">{len(por_cultivo)}</div>
                <div class="kpi-label">CULTIVOS</div>
            </div>
        </div>
        
        <div class="section">
            <h2>Gastos por Proveedor</h2>
            <table>
                <thead>
                    <tr>
                        <th>Proveedor</th>
                        <th style="text-align:center">Albaranes</th>
                        <th style="text-align:right">Total</th>
                        <th style="text-align:right">%</th>
                    </tr>
                </thead>
                <tbody>
                    {prov_rows}
                </tbody>
            </table>
        </div>
        
        <div class="section">
            <h2>Gastos por Cultivo</h2>
            <table>
                <thead>
                    <tr>
                        <th>Cultivo</th>
                        <th style="text-align:center">Albaranes</th>
                        <th style="text-align:right">Total</th>
                        <th style="text-align:right">%</th>
                    </tr>
                </thead>
                <tbody>
                    {cult_rows}
                </tbody>
            </table>
        </div>
        
        <div class="section">
            <h2>Gastos por Parcela</h2>
            <table>
                <thead>
                    <tr>
                        <th>Parcela</th>
                        <th>Cultivo</th>
                        <th style="text-align:center">Albaranes</th>
                        <th style="text-align:right">Total</th>
                    </tr>
                </thead>
                <tbody>
                    {parc_rows}
                </tbody>
            </table>
        </div>
        
        <div class="footer">
            <p>Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')} | FRUVECO - Cuaderno de Campo</p>
        </div>
    </body>
    </html>
    """
    
    # Generate PDF
    pdf_buffer = io.BytesIO()
    HTML(string=html_content).write_pdf(pdf_buffer)
    pdf_buffer.seek(0)
    
    filename = f"informe_gastos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    
    return StreamingResponse(
        pdf_buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
