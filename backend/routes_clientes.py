"""
Routes for Clientes (Customers/Clients) - CRUD operations
For sales contracts and delivery notes
"""

from fastapi import APIRouter, HTTPException, Depends, UploadFile, File
from typing import Optional, List
from bson import ObjectId
from datetime import datetime
import os
import shutil

from database import db, serialize_doc, serialize_docs
from rbac_guards import (
    RequireCreate, RequireEdit, RequireDelete,
    get_current_user
)

router = APIRouter(prefix="/api", tags=["clientes"])

clientes_collection = db['clientes']
contratos_collection = db['contratos']
albaranes_collection = db['albaranes']

# Directorio para fotos de clientes
UPLOAD_DIR = "/app/uploads/clientes"
os.makedirs(UPLOAD_DIR, exist_ok=True)


@router.post("/clientes", response_model=dict)
async def create_cliente(
    cliente: dict,
    current_user: dict = Depends(RequireCreate)
):
    """Crear nuevo cliente"""
    # Generar código automático si no se proporciona
    if not cliente.get("codigo"):
        last_cliente = await clientes_collection.find_one(sort=[("codigo_num", -1)])
        next_num = (last_cliente.get("codigo_num", 0) if last_cliente else 0) + 1
        cliente["codigo"] = str(next_num).zfill(4)
        cliente["codigo_num"] = next_num
    else:
        # Extraer número del código para ordenamiento
        try:
            cliente["codigo_num"] = int(cliente["codigo"])
        except:
            cliente["codigo_num"] = 0
    
    # Verificar código único
    existing = await clientes_collection.find_one({"codigo": cliente["codigo"]})
    if existing:
        raise HTTPException(status_code=400, detail="Ya existe un cliente con este código")
    
    cliente["activo"] = cliente.get("activo", True)
    cliente["created_at"] = datetime.now()
    cliente["updated_at"] = datetime.now()
    cliente["created_by"] = current_user.get("email")
    
    result = await clientes_collection.insert_one(cliente)
    created = await clientes_collection.find_one({"_id": result.inserted_id})
    
    return {"success": True, "data": serialize_doc(created)}


@router.get("/clientes")
async def get_clientes(
    skip: int = 0,
    limit: int = 100,
    search: Optional[str] = None,
    activo: Optional[bool] = None,
    tipo: Optional[str] = None,
    provincia: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Listar clientes con filtros"""
    query = {}
    
    if search:
        query["$or"] = [
            {"nombre": {"$regex": search, "$options": "i"}},
            {"razon": {"$regex": search, "$options": "i"}},
            {"codigo": {"$regex": search, "$options": "i"}},
            {"nif": {"$regex": search, "$options": "i"}},
            {"email": {"$regex": search, "$options": "i"}}
        ]
    
    if activo is not None:
        query["activo"] = activo
    
    if tipo:
        query["tipo"] = tipo
    
    if provincia:
        query["provincia"] = provincia
    
    clientes = await clientes_collection.find(query).sort("codigo_num", 1).skip(skip).limit(limit).to_list(limit)
    total = await clientes_collection.count_documents(query)
    
    return {"clientes": serialize_docs(clientes), "total": total}


@router.get("/clientes/activos")
async def get_clientes_activos(
    current_user: dict = Depends(get_current_user)
):
    """Listar solo clientes activos para selectores"""
    clientes = await clientes_collection.find({"activo": True}).sort("nombre", 1).to_list(500)
    return {"clientes": serialize_docs(clientes)}


@router.get("/clientes/tipos")
async def get_tipos_cliente(
    current_user: dict = Depends(get_current_user)
):
    """Obtener lista de tipos de cliente"""
    return {
        "tipos": [
            "Mayorista",
            "Minorista",
            "Distribuidor",
            "Exportador",
            "Importador",
            "Industria",
            "Cooperativa",
            "Particular",
            "Otros"
        ]
    }


@router.get("/clientes/provincias")
async def get_provincias_clientes(
    current_user: dict = Depends(get_current_user)
):
    """Obtener lista de provincias con clientes"""
    provincias = await clientes_collection.distinct("provincia", {"provincia": {"$ne": None, "$ne": ""}})
    return {"provincias": sorted([p for p in provincias if p])}


@router.get("/clientes/{cliente_id}")
async def get_cliente(
    cliente_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Obtener cliente por ID"""
    if not ObjectId.is_valid(cliente_id):
        raise HTTPException(status_code=400, detail="ID inválido")
    
    cliente = await clientes_collection.find_one({"_id": ObjectId(cliente_id)})
    if not cliente:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    
    return serialize_doc(cliente)


@router.put("/clientes/{cliente_id}")
async def update_cliente(
    cliente_id: str,
    cliente: dict,
    current_user: dict = Depends(RequireEdit)
):
    """Actualizar cliente"""
    if not ObjectId.is_valid(cliente_id):
        raise HTTPException(status_code=400, detail="ID inválido")
    
    # Si se cambia el código, verificar unicidad
    if cliente.get("codigo"):
        existing = await clientes_collection.find_one({
            "codigo": cliente["codigo"],
            "_id": {"$ne": ObjectId(cliente_id)}
        })
        if existing:
            raise HTTPException(status_code=400, detail="Ya existe un cliente con este código")
    
    cliente["updated_at"] = datetime.now()
    
    result = await clientes_collection.update_one(
        {"_id": ObjectId(cliente_id)},
        {"$set": cliente}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    
    updated = await clientes_collection.find_one({"_id": ObjectId(cliente_id)})
    return {"success": True, "data": serialize_doc(updated)}


@router.patch("/clientes/{cliente_id}/toggle-activo")
async def toggle_cliente_activo(
    cliente_id: str,
    current_user: dict = Depends(RequireEdit)
):
    """Activar/desactivar cliente"""
    if not ObjectId.is_valid(cliente_id):
        raise HTTPException(status_code=400, detail="ID inválido")
    
    cliente = await clientes_collection.find_one({"_id": ObjectId(cliente_id)})
    if not cliente:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    
    new_status = not cliente.get("activo", True)
    
    await clientes_collection.update_one(
        {"_id": ObjectId(cliente_id)},
        {"$set": {"activo": new_status, "updated_at": datetime.now()}}
    )
    
    return {"success": True, "activo": new_status}


@router.post("/clientes/{cliente_id}/foto")
async def upload_foto_cliente(
    cliente_id: str,
    file: UploadFile = File(...),
    current_user: dict = Depends(RequireEdit)
):
    """Subir foto de cliente"""
    if not ObjectId.is_valid(cliente_id):
        raise HTTPException(status_code=400, detail="ID inválido")
    
    cliente = await clientes_collection.find_one({"_id": ObjectId(cliente_id)})
    if not cliente:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    
    # Validar tipo de archivo
    allowed_types = ["image/jpeg", "image/png", "image/webp"]
    if file.content_type not in allowed_types:
        raise HTTPException(status_code=400, detail="Tipo de archivo no permitido. Use JPG, PNG o WebP")
    
    # Guardar archivo
    file_extension = file.filename.split(".")[-1] if "." in file.filename else "jpg"
    filename = f"cliente_{cliente_id}.{file_extension}"
    file_path = os.path.join(UPLOAD_DIR, filename)
    
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    
    # Actualizar URL en BD
    foto_url = f"/api/uploads/clientes/{filename}"
    await clientes_collection.update_one(
        {"_id": ObjectId(cliente_id)},
        {"$set": {"foto_url": foto_url, "updated_at": datetime.now()}}
    )
    
    return {"success": True, "foto_url": foto_url}


@router.get("/clientes/{cliente_id}/resumen-ventas")
async def get_resumen_ventas_cliente(
    cliente_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Obtener resumen de ventas de un cliente"""
    if not ObjectId.is_valid(cliente_id):
        raise HTTPException(status_code=400, detail="ID inválido")
    
    cliente = await clientes_collection.find_one({"_id": ObjectId(cliente_id)})
    if not cliente:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    
    # Buscar contratos de venta asociados al cliente
    contratos = await contratos_collection.find({
        "$or": [
            {"cliente_id": cliente_id},
            {"cliente": cliente.get("nombre", "")}
        ],
        "tipo": "Venta"
    }).to_list(500)
    
    # Agrupar por campaña
    ventas_por_campana = {}
    total_cantidad = 0
    total_importe = 0
    
    for contrato in contratos:
        campana = contrato.get("campana", "Sin campaña")
        cantidad = contrato.get("cantidad", 0) or 0
        precio = contrato.get("precio", 0) or 0
        importe = cantidad * precio
        
        if campana not in ventas_por_campana:
            ventas_por_campana[campana] = {
                "campana": campana,
                "num_contratos": 0,
                "cantidad_total": 0,
                "importe_total": 0,
                "cultivos": set()
            }
        
        ventas_por_campana[campana]["num_contratos"] += 1
        ventas_por_campana[campana]["cantidad_total"] += cantidad
        ventas_por_campana[campana]["importe_total"] += importe
        ventas_por_campana[campana]["cultivos"].add(contrato.get("cultivo", ""))
        
        total_cantidad += cantidad
        total_importe += importe
    
    # Convertir sets a listas para serialización
    for campana in ventas_por_campana:
        ventas_por_campana[campana]["cultivos"] = list(ventas_por_campana[campana]["cultivos"])
    
    # Buscar albaranes de venta asociados
    albaranes = await albaranes_collection.find({
        "$or": [
            {"cliente": cliente.get("nombre", "")},
            {"cliente_contrato": cliente.get("nombre", "")}
        ],
        "tipo": "Albarán de venta"
    }).to_list(500)
    
    total_albaranes = len(albaranes)
    importe_albaranes = sum(
        sum(item.get("total", 0) or 0 for item in alb.get("items", []))
        for alb in albaranes
    )
    
    return {
        "cliente_id": cliente_id,
        "cliente_nombre": cliente.get("nombre", ""),
        "resumen": {
            "total_contratos": len(contratos),
            "total_cantidad_kg": total_cantidad,
            "total_importe": total_importe,
            "total_albaranes": total_albaranes,
            "importe_albaranes": importe_albaranes
        },
        "ventas_por_campana": list(ventas_por_campana.values()),
        "contratos": serialize_docs(contratos)
    }


@router.delete("/clientes/{cliente_id}")
async def delete_cliente(
    cliente_id: str,
    current_user: dict = Depends(RequireDelete)
):
    """Eliminar cliente"""
    if not ObjectId.is_valid(cliente_id):
        raise HTTPException(status_code=400, detail="ID inválido")
    
    # Verificar si tiene contratos asociados
    contratos_count = await db['contratos'].count_documents({"cliente_id": cliente_id})
    if contratos_count > 0:
        raise HTTPException(
            status_code=400, 
            detail=f"No se puede eliminar: el cliente tiene {contratos_count} contrato(s) asociado(s)"
        )
    
    result = await clientes_collection.delete_one({"_id": ObjectId(cliente_id)})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    
    return {"success": True, "message": "Cliente eliminado"}


@router.get("/clientes/stats/resumen")
async def get_clientes_stats(current_user: dict = Depends(get_current_user)):
    """Obtener estadísticas de clientes"""
    total = await clientes_collection.count_documents({})
    activos = await clientes_collection.count_documents({"activo": {"$ne": False}})
    
    # Clientes por provincia
    pipeline = [
        {"$match": {"activo": {"$ne": False}}},
        {"$group": {"_id": "$provincia", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}},
        {"$limit": 10}
    ]
    por_provincia = []
    async for doc in clientes_collection.aggregate(pipeline):
        por_provincia.append({"provincia": doc["_id"] or "Sin especificar", "count": doc["count"]})
    
    # Clientes con contratos
    clientes_con_contratos = await contratos_collection.distinct("cliente_id")
    
    # Volumen de ventas por cliente
    pipeline_ventas = [
        {"$match": {"tipo": "venta"}},
        {"$group": {"_id": "$cliente_id", "total": {"$sum": "$total"}, "count": {"$sum": 1}}},
        {"$sort": {"total": -1}},
        {"$limit": 10}
    ]
    top_clientes = []
    async for doc in albaranes_collection.aggregate(pipeline_ventas):
        cliente = await clientes_collection.find_one({"_id": ObjectId(doc["_id"])}) if doc["_id"] and ObjectId.is_valid(doc["_id"]) else None
        top_clientes.append({
            "cliente_id": doc["_id"],
            "cliente_nombre": cliente.get("nombre") if cliente else "Desconocido",
            "total_ventas": round(doc["total"], 2),
            "num_operaciones": doc["count"]
        })
    
    return {
        "success": True,
        "stats": {
            "total": total,
            "activos": activos,
            "inactivos": total - activos,
            "con_contratos": len(clientes_con_contratos),
            "por_provincia": por_provincia,
            "top_clientes_ventas": top_clientes
        }
    }


@router.get("/clientes/{cliente_id}/historial")
async def get_cliente_historial(
    cliente_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Obtener historial de ventas y contratos de un cliente"""
    if not ObjectId.is_valid(cliente_id):
        raise HTTPException(status_code=400, detail="ID inválido")
    
    cliente = await clientes_collection.find_one({"_id": ObjectId(cliente_id)})
    if not cliente:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    
    # Contratos del cliente
    contratos = []
    cursor = contratos_collection.find({"cliente_id": cliente_id}).sort("fecha_inicio", -1).limit(20)
    async for c in cursor:
        contratos.append(serialize_doc(c))
    
    # Albaranes de venta del cliente
    albaranes = []
    cursor = albaranes_collection.find({
        "$or": [
            {"cliente_id": cliente_id},
            {"cliente": cliente.get("nombre")}
        ],
        "tipo": "venta"
    }).sort("fecha", -1).limit(50)
    async for a in cursor:
        albaranes.append(serialize_doc(a))
    
    # Calcular totales
    total_ventas = sum(a.get("total", 0) for a in albaranes)
    
    return {
        "success": True,
        "cliente": serialize_doc(cliente),
        "historial": {
            "contratos": contratos,
            "albaranes": albaranes,
            "total_ventas": round(total_ventas, 2),
            "num_contratos": len(contratos),
            "num_albaranes": len(albaranes)
        }
    }


@router.get("/clientes/export/excel")
async def export_clientes_excel(
    activo: Optional[bool] = None,
    current_user: dict = Depends(get_current_user)
):
    """Exportar clientes a Excel"""
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io
    
    query = {}
    if activo is not None:
        query['activo'] = activo
    
    clientes = await clientes_collection.find(query).sort("nombre", 1).to_list(1000)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Clientes"
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    headers = ["Código", "Nombre", "CIF/NIF", "Dirección", "Población", "Provincia", "C.P.", "Teléfono", "Email", "Contacto", "Estado"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    
    for row_idx, cli in enumerate(clientes, 2):
        data = [
            cli.get("codigo", ""),
            cli.get("nombre", ""),
            cli.get("cif_nif", ""),
            cli.get("direccion", ""),
            cli.get("poblacion", ""),
            cli.get("provincia", ""),
            cli.get("codigo_postal", ""),
            cli.get("telefono", ""),
            cli.get("email", ""),
            cli.get("persona_contacto", ""),
            "Activo" if cli.get("activo", True) else "Inactivo"
        ]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
    
    column_widths = [10, 25, 15, 30, 20, 15, 10, 15, 25, 20, 10]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=clientes_{datetime.now().strftime('%Y%m%d')}.xlsx"}
    )
