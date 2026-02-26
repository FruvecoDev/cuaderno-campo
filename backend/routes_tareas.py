"""
Routes for Tareas module - Enhanced with priorities, assignments, subtasks, calendar
"""

from fastapi import APIRouter, HTTPException, Depends, Query
from fastapi.responses import StreamingResponse
from typing import Optional, List
from bson import ObjectId
from datetime import datetime, timedelta
import uuid
import io

from models import TareaCreate, SubTarea
from database import (
    tareas_collection, parcelas_collection, users_collection,
    serialize_doc, serialize_docs
)
from rbac_guards import (
    RequireCreate, RequireEdit, RequireDelete,
    RequireTareasAccess, get_current_user
)

router = APIRouter(prefix="/api", tags=["tareas"])


# ============================================================================
# TAREAS CRUD - ENHANCED
# ============================================================================

@router.post("/tareas", response_model=dict)
async def create_tarea(
    tarea: TareaCreate,
    current_user: dict = Depends(RequireCreate),
    _access: dict = Depends(RequireTareasAccess)
):
    tarea_dict = tarea.dict()
    
    # Auto-inherit from first parcela if available
    if tarea_dict.get("parcelas_ids") and len(tarea_dict["parcelas_ids"]) > 0:
        parcela = await parcelas_collection.find_one({"_id": ObjectId(tarea_dict["parcelas_ids"][0])})
        if parcela:
            tarea_dict["cultivo"] = parcela.get("cultivo")
            tarea_dict["campana"] = parcela.get("campana")
            tarea_dict["proveedor"] = parcela.get("proveedor")
    
    # Get assigned user name
    if tarea_dict.get("asignado_a"):
        user = await users_collection.find_one({"_id": ObjectId(tarea_dict["asignado_a"])})
        if user:
            tarea_dict["asignado_nombre"] = user.get("name", user.get("email"))
    
    # Generate IDs for subtasks
    for subtarea in tarea_dict.get("subtareas", []):
        if not subtarea.get("id"):
            subtarea["id"] = str(uuid.uuid4())[:8]
    
    tarea_dict.update({
        "realizada": False,
        "created_at": datetime.now(),
        "updated_at": datetime.now(),
        "created_by": str(current_user.get("_id", ""))
    })
    
    result = await tareas_collection.insert_one(tarea_dict)
    created = await tareas_collection.find_one({"_id": result.inserted_id})
    
    return {"success": True, "data": serialize_doc(created)}


@router.get("/tareas")
async def get_tareas(
    skip: int = 0,
    limit: int = 100,
    estado: Optional[str] = None,
    prioridad: Optional[str] = None,
    asignado_a: Optional[str] = None,
    tipo_tarea: Optional[str] = None,
    fecha_desde: Optional[str] = None,
    fecha_hasta: Optional[str] = None,
    parcela_id: Optional[str] = None,
    cultivo: Optional[str] = None,
    campana: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
    _access: dict = Depends(RequireTareasAccess)
):
    query = {}
    
    if estado:
        if estado == "pendiente":
            query["$or"] = [{"estado": "pendiente"}, {"estado": {"$exists": False}, "realizada": {"$ne": True}}]
        elif estado == "completada":
            query["$or"] = [{"estado": "completada"}, {"realizada": True}]
        else:
            query["estado"] = estado
    
    if prioridad:
        query["prioridad"] = prioridad
    
    if asignado_a:
        query["asignado_a"] = asignado_a
    
    if tipo_tarea:
        query["tipo_tarea"] = tipo_tarea
    
    if fecha_desde:
        query["fecha_inicio"] = {"$gte": fecha_desde}
    
    if fecha_hasta:
        if "fecha_inicio" in query:
            query["fecha_inicio"]["$lte"] = fecha_hasta
        else:
            query["fecha_inicio"] = {"$lte": fecha_hasta}
    
    if parcela_id:
        query["parcelas_ids"] = parcela_id
    
    if cultivo:
        query["cultivo"] = cultivo
    
    if campana:
        query["campana"] = campana
    
    tareas = await tareas_collection.find(query).sort("fecha_inicio", -1).skip(skip).limit(limit).to_list(limit)
    total = await tareas_collection.count_documents(query)
    
    return {"tareas": serialize_docs(tareas), "total": total}


@router.get("/tareas/calendario")
async def get_tareas_calendario(
    mes: int = Query(..., ge=1, le=12),
    ano: int = Query(..., ge=2020, le=2100),
    current_user: dict = Depends(get_current_user),
    _access: dict = Depends(RequireTareasAccess)
):
    """Get tasks for calendar view - returns tasks grouped by day"""
    # Calculate date range for the month
    fecha_inicio = f"{ano}-{mes:02d}-01"
    if mes == 12:
        fecha_fin = f"{ano + 1}-01-01"
    else:
        fecha_fin = f"{ano}-{mes + 1:02d}-01"
    
    query = {
        "$or": [
            {"fecha_inicio": {"$gte": fecha_inicio, "$lt": fecha_fin}},
            {"fecha_fin": {"$gte": fecha_inicio, "$lt": fecha_fin}},
            {"fecha_vencimiento": {"$gte": fecha_inicio, "$lt": fecha_fin}}
        ]
    }
    
    tareas = await tareas_collection.find(query).to_list(500)
    
    # Group by day
    tareas_por_dia = {}
    for tarea in tareas:
        fecha = tarea.get("fecha_inicio") or tarea.get("fecha_vencimiento")
        if fecha:
            dia = fecha[:10]  # YYYY-MM-DD
            if dia not in tareas_por_dia:
                tareas_por_dia[dia] = []
            tareas_por_dia[dia].append(serialize_doc(tarea))
    
    return {"calendario": tareas_por_dia, "mes": mes, "ano": ano}


@router.get("/tareas/stats")
async def get_tareas_stats(
    current_user: dict = Depends(get_current_user),
    _access: dict = Depends(RequireTareasAccess)
):
    """Get task statistics"""
    total = await tareas_collection.count_documents({})
    
    # By status
    pendientes = await tareas_collection.count_documents({
        "$or": [{"estado": "pendiente"}, {"estado": {"$exists": False}, "realizada": {"$ne": True}}]
    })
    en_progreso = await tareas_collection.count_documents({"estado": "en_progreso"})
    completadas = await tareas_collection.count_documents({
        "$or": [{"estado": "completada"}, {"realizada": True}]
    })
    canceladas = await tareas_collection.count_documents({"estado": "cancelada"})
    
    # By priority
    alta = await tareas_collection.count_documents({"prioridad": "alta", "estado": {"$ne": "completada"}, "realizada": {"$ne": True}})
    media = await tareas_collection.count_documents({"prioridad": "media", "estado": {"$ne": "completada"}, "realizada": {"$ne": True}})
    baja = await tareas_collection.count_documents({"prioridad": "baja", "estado": {"$ne": "completada"}, "realizada": {"$ne": True}})
    
    # Overdue tasks
    hoy = datetime.now().strftime("%Y-%m-%d")
    vencidas = await tareas_collection.count_documents({
        "fecha_vencimiento": {"$lt": hoy},
        "$or": [{"estado": "pendiente"}, {"estado": "en_progreso"}, {"estado": {"$exists": False}, "realizada": {"$ne": True}}]
    })
    
    # This week
    semana_inicio = (datetime.now() - timedelta(days=datetime.now().weekday())).strftime("%Y-%m-%d")
    semana_fin = (datetime.now() + timedelta(days=6 - datetime.now().weekday())).strftime("%Y-%m-%d")
    esta_semana = await tareas_collection.count_documents({
        "$or": [
            {"fecha_inicio": {"$gte": semana_inicio, "$lte": semana_fin}},
            {"fecha_vencimiento": {"$gte": semana_inicio, "$lte": semana_fin}}
        ]
    })
    
    # Costs
    pipeline = [
        {"$group": {
            "_id": None,
            "coste_estimado_total": {"$sum": {"$ifNull": ["$coste_estimado", 0]}},
            "coste_real_total": {"$sum": {"$ifNull": ["$coste_real", 0]}}
        }}
    ]
    costes = await tareas_collection.aggregate(pipeline).to_list(1)
    costes_data = costes[0] if costes else {"coste_estimado_total": 0, "coste_real_total": 0}
    
    return {
        "total": total,
        "pendientes": pendientes,
        "en_progreso": en_progreso,
        "completadas": completadas,
        "canceladas": canceladas,
        "prioridad": {"alta": alta, "media": media, "baja": baja},
        "vencidas": vencidas,
        "esta_semana": esta_semana,
        "costes": {
            "estimado": costes_data.get("coste_estimado_total", 0),
            "real": costes_data.get("coste_real_total", 0)
        }
    }


@router.get("/tareas/usuarios-asignables")
async def get_usuarios_asignables(
    current_user: dict = Depends(get_current_user)
):
    """Get users that can be assigned to tasks (technicians, managers, admins)"""
    users = await users_collection.find({
        "role": {"$in": ["admin", "manager", "technician"]},
        "activo": {"$ne": False}
    }, {"_id": 1, "name": 1, "email": 1, "role": 1}).to_list(100)
    
    return {"usuarios": serialize_docs(users)}


@router.get("/tareas/tipos")
async def get_tipos_tarea():
    """Get task types"""
    return {
        "tipos": [
            {"id": "general", "nombre": "General"},
            {"id": "tratamiento", "nombre": "Tratamiento"},
            {"id": "riego", "nombre": "Riego"},
            {"id": "cosecha", "nombre": "Cosecha"},
            {"id": "mantenimiento", "nombre": "Mantenimiento"},
            {"id": "siembra", "nombre": "Siembra"},
            {"id": "poda", "nombre": "Poda"},
            {"id": "fertilizacion", "nombre": "Fertilización"},
            {"id": "otro", "nombre": "Otro"}
        ],
        "prioridades": [
            {"id": "alta", "nombre": "Alta", "color": "#ef4444"},
            {"id": "media", "nombre": "Media", "color": "#f59e0b"},
            {"id": "baja", "nombre": "Baja", "color": "#22c55e"}
        ],
        "estados": [
            {"id": "pendiente", "nombre": "Pendiente"},
            {"id": "en_progreso", "nombre": "En Progreso"},
            {"id": "completada", "nombre": "Completada"},
            {"id": "cancelada", "nombre": "Cancelada"}
        ]
    }


@router.get("/tareas/{tarea_id}")
async def get_tarea(
    tarea_id: str,
    current_user: dict = Depends(get_current_user),
    _access: dict = Depends(RequireTareasAccess)
):
    if not ObjectId.is_valid(tarea_id):
        raise HTTPException(status_code=400, detail="ID de tarea inválido")
    
    tarea = await tareas_collection.find_one({"_id": ObjectId(tarea_id)})
    if not tarea:
        raise HTTPException(status_code=404, detail="Tarea no encontrada")
    
    return {"success": True, "data": serialize_doc(tarea)}


@router.put("/tareas/{tarea_id}")
async def update_tarea(
    tarea_id: str,
    tarea: TareaCreate,
    current_user: dict = Depends(RequireEdit),
    _access: dict = Depends(RequireTareasAccess)
):
    if not ObjectId.is_valid(tarea_id):
        raise HTTPException(status_code=400, detail="ID de tarea inválido")
    
    tarea_dict = tarea.dict()
    
    # Get assigned user name
    if tarea_dict.get("asignado_a"):
        user = await users_collection.find_one({"_id": ObjectId(tarea_dict["asignado_a"])})
        if user:
            tarea_dict["asignado_nombre"] = user.get("name", user.get("email"))
    
    # Generate IDs for new subtasks
    for subtarea in tarea_dict.get("subtareas", []):
        if not subtarea.get("id"):
            subtarea["id"] = str(uuid.uuid4())[:8]
    
    # Set realizada based on estado
    if tarea_dict.get("estado") == "completada":
        tarea_dict["realizada"] = True
    
    tarea_dict["updated_at"] = datetime.now()
    
    result = await tareas_collection.update_one(
        {"_id": ObjectId(tarea_id)},
        {"$set": tarea_dict}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Tarea no encontrada")
    
    updated = await tareas_collection.find_one({"_id": ObjectId(tarea_id)})
    return {"success": True, "data": serialize_doc(updated)}


@router.patch("/tareas/{tarea_id}/estado")
async def update_tarea_estado(
    tarea_id: str,
    estado: str = Query(...),
    current_user: dict = Depends(RequireEdit),
    _access: dict = Depends(RequireTareasAccess)
):
    """Quick status update"""
    if not ObjectId.is_valid(tarea_id):
        raise HTTPException(status_code=400, detail="ID de tarea inválido")
    
    update_data = {
        "estado": estado,
        "updated_at": datetime.now()
    }
    
    if estado == "completada":
        update_data["realizada"] = True
        update_data["fecha_completada"] = datetime.now().strftime("%Y-%m-%d")
    
    result = await tareas_collection.update_one(
        {"_id": ObjectId(tarea_id)},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Tarea no encontrada")
    
    return {"success": True, "message": f"Estado actualizado a {estado}"}


@router.patch("/tareas/{tarea_id}/subtarea/{subtarea_id}")
async def toggle_subtarea(
    tarea_id: str,
    subtarea_id: str,
    completada: bool = Query(...),
    current_user: dict = Depends(RequireEdit),
    _access: dict = Depends(RequireTareasAccess)
):
    """Toggle subtask completion"""
    if not ObjectId.is_valid(tarea_id):
        raise HTTPException(status_code=400, detail="ID de tarea inválido")
    
    tarea = await tareas_collection.find_one({"_id": ObjectId(tarea_id)})
    if not tarea:
        raise HTTPException(status_code=404, detail="Tarea no encontrada")
    
    subtareas = tarea.get("subtareas", [])
    for st in subtareas:
        if st.get("id") == subtarea_id:
            st["completada"] = completada
            if completada:
                st["completada_por"] = current_user.get("name", current_user.get("email"))
                st["completada_fecha"] = datetime.now().strftime("%Y-%m-%d %H:%M")
            else:
                st["completada_por"] = None
                st["completada_fecha"] = None
            break
    
    await tareas_collection.update_one(
        {"_id": ObjectId(tarea_id)},
        {"$set": {"subtareas": subtareas, "updated_at": datetime.now()}}
    )
    
    return {"success": True, "subtareas": subtareas}


@router.delete("/tareas/{tarea_id}")
async def delete_tarea(
    tarea_id: str,
    current_user: dict = Depends(RequireDelete),
    _access: dict = Depends(RequireTareasAccess)
):
    if not ObjectId.is_valid(tarea_id):
        raise HTTPException(status_code=400, detail="ID de tarea inválido")
    
    result = await tareas_collection.delete_one({"_id": ObjectId(tarea_id)})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Tarea no encontrada")
    
    return {"success": True, "message": "Tarea eliminada correctamente"}


# ============================================================================
# EXPORT
# ============================================================================

@router.get("/tareas/export/excel")
async def export_tareas_excel(
    estado: Optional[str] = None,
    prioridad: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
    _access: dict = Depends(RequireTareasAccess)
):
    """Export tasks to Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    
    query = {}
    if estado:
        query["estado"] = estado
    if prioridad:
        query["prioridad"] = prioridad
    
    tareas = await tareas_collection.find(query).to_list(1000)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tareas"
    
    # Headers
    headers = ["Nombre", "Tipo", "Prioridad", "Estado", "Asignado a", "Fecha Inicio", "Fecha Vencimiento", "Cultivo", "Campaña", "Coste Est.", "Coste Real", "Observaciones"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="2d5a27", end_color="2d5a27", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    
    # Data
    for row, tarea in enumerate(tareas, 2):
        ws.cell(row=row, column=1, value=tarea.get("nombre", ""))
        ws.cell(row=row, column=2, value=tarea.get("tipo_tarea", "general"))
        ws.cell(row=row, column=3, value=tarea.get("prioridad", "media"))
        ws.cell(row=row, column=4, value=tarea.get("estado", "pendiente"))
        ws.cell(row=row, column=5, value=tarea.get("asignado_nombre", ""))
        ws.cell(row=row, column=6, value=tarea.get("fecha_inicio", ""))
        ws.cell(row=row, column=7, value=tarea.get("fecha_vencimiento", ""))
        ws.cell(row=row, column=8, value=tarea.get("cultivo", ""))
        ws.cell(row=row, column=9, value=tarea.get("campana", ""))
        ws.cell(row=row, column=10, value=tarea.get("coste_estimado", 0))
        ws.cell(row=row, column=11, value=tarea.get("coste_real", 0))
        ws.cell(row=row, column=12, value=tarea.get("observaciones", ""))
    
    # Auto-width columns
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"tareas_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
