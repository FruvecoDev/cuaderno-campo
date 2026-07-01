"""
Routes for Contratos (Contracts) - CRUD operations
Extracted from routes_main.py for better code organization
"""

from fastapi import APIRouter, HTTPException, Depends, UploadFile, File
from typing import Optional
from bson import ObjectId
from datetime import datetime
import re
import io

from models import ContratoCreate
from database import contratos_collection, serialize_doc, serialize_docs, db
from rbac_guards import (
    RequireCreate, RequireEdit, RequireDelete,
    RequireContratosAccess, get_current_user, ensure_tipo_operacion
)
from services.audit_service import create_audit_log, calculate_changes

router = APIRouter(prefix="/api", tags=["contratos"])

# Collections for lookups
proveedores_collection = db['proveedores']
clientes_collection = db['clientes']
cultivos_collection = db['cultivos']


@router.post("/contratos", response_model=dict)
async def create_contrato(
    contrato: ContratoCreate,
    current_user: dict = Depends(RequireCreate),
    _access: dict = Depends(RequireContratosAccess)
):
    # Validar tipo de operación del usuario (Compra/Venta/Ambos)
    ensure_tipo_operacion(current_user, contrato.tipo)

    # Generación atómica del número de contrato dentro del año actual:
    # MP-{año}-{numero:06d}. Se busca el mayor "numero" del año en curso
    # para evitar colisiones con contratos de otros años.
    current_year = datetime.now().year
    last_contrato_year = await contratos_collection.find_one(
        {"año": current_year},
        sort=[("numero", -1)],
    )
    next_numero = (last_contrato_year.get("numero", 0) if last_contrato_year else 0) + 1
    numero_contrato = f"MP-{current_year}-{str(next_numero).zfill(6)}"

    # Lookup proveedor name (para contratos de Compra)
    proveedor_name = contrato.proveedor or ""
    if contrato.proveedor_id:
        prov = await proveedores_collection.find_one({"_id": ObjectId(contrato.proveedor_id)})
        if prov:
            proveedor_name = prov.get("nombre", "")
    
    # Lookup cliente name (para contratos de Venta)
    cliente_name = ""
    cliente_id_str = getattr(contrato, 'cliente_id', None) or ""
    if cliente_id_str:
        cli = await clientes_collection.find_one({"_id": ObjectId(cliente_id_str)})
        if cli:
            cliente_name = cli.get("nombre", "")
    
    # Lookup cultivo name
    cultivo_name = contrato.cultivo or ""
    cultivo_id_str = contrato.cultivo_id or ""
    if contrato.cultivo_id:
        cult = await cultivos_collection.find_one({"_id": ObjectId(contrato.cultivo_id)})
        if cult:
            cultivo_name = cult.get("nombre", "")
    
    contrato_dict = contrato.dict()
    contrato_dict.update({
        "serie": "MP",
        "año": current_year,
        "numero": next_numero,
        "numero_contrato": numero_contrato,
        "proveedor": proveedor_name,
        "cliente": cliente_name,
        "cliente_id": cliente_id_str,
        "cultivo": cultivo_name,
        "cultivo_id": cultivo_id_str,
        "created_at": datetime.now(),
        "updated_at": datetime.now()
    })
    
    result = await contratos_collection.insert_one(contrato_dict)
    created = await contratos_collection.find_one({"_id": result.inserted_id})
    
    # Registrar en auditoría
    await create_audit_log(
        collection_name="contratos",
        document_id=str(result.inserted_id),
        action="create",
        user_email=current_user.get("email", "unknown"),
        user_name=current_user.get("full_name", current_user.get("username", "Usuario")),
        new_data=serialize_doc(created.copy())
    )
    
    return {"success": True, "data": serialize_doc(created)}


@router.post("/contratos/import-excel")
async def import_contratos_excel(
    file: UploadFile = File(...),
    current_user: dict = Depends(RequireCreate),
    _access: dict = Depends(RequireContratosAccess),
):
    """Importa contratos desde un fichero Excel (.xlsx) con las columnas:
    Numero Contrato · Tipo Contrato · Campaña · Procedencia · Fecha ·
    Nombre Proveedor · Cultivo · Cantidad (Kg).
    
    Comportamiento:
    - Los proveedores/cultivos se buscan por `nombre` (case-insensitive con trim).
      Si no existen, se crean automáticamente con los campos mínimos.
    - Contratos con `Numero Contrato` ya presente se omiten (dedup).
    - Fila con "Total" o todas las columnas vacías salvo Cantidad → se ignora.
    - Devuelve `{imported, skipped_duplicates, created_proveedores, created_cultivos, errors}`.
    """
    from openpyxl import load_workbook

    if not file.filename or not file.filename.lower().endswith(('.xlsx', '.xlsm')):
        raise HTTPException(status_code=400, detail="Solo se aceptan ficheros .xlsx")

    contenido = await file.read()
    try:
        wb = load_workbook(io.BytesIO(contenido), data_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error leyendo el Excel: {e}")

    # Mapeo de headers a keys internas (tolerante a mayúsculas/acentos).
    header_row = [(c.value or '') for c in ws[1]]
    def _norm(s: str) -> str:
        return re.sub(r'[^a-z0-9]', '', str(s).lower())
    hidx = {}
    aliases = {
        'numero_contrato': ['numerocontrato', 'contrato', 'numero'],
        'tipo': ['tipocontrato', 'tipo'],
        'campana': ['campana', 'campaña', 'campana2', 'campaa'],
        'procedencia': ['procedencia'],
        'fecha': ['fecha', 'fechacontrato'],
        'proveedor': ['nombreproveedor', 'proveedor'],
        'cultivo': ['cultivo', 'articulo'],
        'cantidad': ['cantidadkg', 'cantidad'],
    }
    for i, h in enumerate(header_row):
        hn = _norm(h)
        for key, alist in aliases.items():
            if hn in alist and key not in hidx:
                hidx[key] = i
                break
    faltantes = [k for k in ('numero_contrato', 'campana', 'fecha', 'proveedor', 'cultivo', 'cantidad') if k not in hidx]
    if faltantes:
        raise HTTPException(status_code=400, detail=f"Faltan columnas obligatorias en el Excel: {faltantes}")

    imported = 0
    skipped_dupes = 0
    created_provs = 0
    created_cults = 0
    errors: list[dict] = []

    # Cache de lookups por nombre normalizado
    prov_cache: dict[str, str] = {}
    cult_cache: dict[str, str] = {}

    async def resolve_proveedor(nombre: str) -> str:
        nonlocal created_provs
        key = _norm(nombre)
        if key in prov_cache:
            return prov_cache[key]
        doc = await proveedores_collection.find_one({"nombre": {"$regex": f"^{re.escape(nombre.strip())}$", "$options": "i"}})
        if not doc:
            new_doc = {
                "nombre": nombre.strip(),
                "tipo_proveedor": "Materia Prima",
                "created_at": datetime.now(),
                "updated_at": datetime.now(),
            }
            r = await proveedores_collection.insert_one(new_doc)
            created_provs += 1
            _id = str(r.inserted_id)
        else:
            _id = str(doc["_id"])
        prov_cache[key] = _id
        return _id

    async def resolve_cultivo(nombre: str) -> str:
        nonlocal created_cults
        key = _norm(nombre)
        if key in cult_cache:
            return cult_cache[key]
        doc = await cultivos_collection.find_one({"nombre": {"$regex": f"^{re.escape(nombre.strip())}$", "$options": "i"}})
        if not doc:
            new_doc = {
                "nombre": nombre.strip(),
                "tipo": "Horticola",
                "created_at": datetime.now(),
                "updated_at": datetime.now(),
            }
            r = await cultivos_collection.insert_one(new_doc)
            created_cults += 1
            _id = str(r.inserted_id)
        else:
            _id = str(doc["_id"])
        cult_cache[key] = _id
        return _id

    def _parse_numero_contrato(s: str):
        # Formato esperado: "MP-2025-000001" → (serie, año, numero)
        m = re.match(r'^\s*([A-Za-z]+)[- ]?(\d{4})[- ]?(\d+)\s*$', str(s or ''))
        if not m:
            return None
        return (m.group(1).upper(), int(m.group(2)), int(m.group(3)))

    def _parse_fecha(v):
        if v is None:
            return ''
        if isinstance(v, datetime):
            return v.strftime('%Y-%m-%d')
        s = str(v).strip()
        # dd/mm/yyyy
        m = re.match(r'^(\d{1,2})[/-](\d{1,2})[/-](\d{4})$', s)
        if m:
            return f"{m.group(3)}-{m.group(2).zfill(2)}-{m.group(1).zfill(2)}"
        return s

    def _get(row, key):
        idx = hidx.get(key)
        return row[idx] if idx is not None and idx < len(row) else None

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            numero_txt = _get(row, 'numero_contrato')
            cantidad = _get(row, 'cantidad')
            # Filas totalizadoras o vacías
            if not numero_txt:
                continue
            parsed = _parse_numero_contrato(numero_txt)
            if not parsed:
                errors.append({"row": row_num, "error": f"Número contrato inválido: {numero_txt!r}"})
                continue
            serie, año, numero = parsed
            numero_contrato_str = f"{serie}-{año}-{str(numero).zfill(6)}"
            # Dedup por número contrato exacto
            existing = await contratos_collection.find_one({"$or": [
                {"numero_contrato": numero_contrato_str},
                {"serie": serie, "año": año, "numero": numero},
                {"serie": serie, "ano": año, "numero": numero},
            ]})
            if existing:
                skipped_dupes += 1
                continue

            nombre_prov = _get(row, 'proveedor')
            nombre_cult = _get(row, 'cultivo')
            if not nombre_prov or not nombre_cult:
                errors.append({"row": row_num, "error": "Proveedor o cultivo vacíos"})
                continue

            proveedor_id = await resolve_proveedor(str(nombre_prov))
            cultivo_id = await resolve_cultivo(str(nombre_cult))

            tipo_str = str(_get(row, 'tipo') or 'Compra').strip().capitalize()
            if tipo_str not in ('Compra', 'Venta'):
                tipo_str = 'Compra'
            campana = str(_get(row, 'campana') or '').strip()
            procedencia = str(_get(row, 'procedencia') or 'Campo').strip()
            fecha_str = _parse_fecha(_get(row, 'fecha'))
            try:
                cantidad_val = float(cantidad or 0)
            except Exception:
                cantidad_val = 0.0

            doc = {
                "numero_contrato": numero_contrato_str,
                "serie": serie,
                "año": año,
                "ano": año,  # alias sin tilde (frontend lee `c.ano`)
                "numero": numero,
                "tipo": tipo_str,
                "tipo_contrato": "Por Kilos",
                "campana": campana,
                "procedencia": procedencia,
                "fecha_contrato": fecha_str,
                "proveedor_id": proveedor_id if tipo_str == 'Compra' else None,
                "cliente_id": None,
                "cultivo_id": cultivo_id,
                "proveedor": str(nombre_prov).strip(),
                "cultivo": str(nombre_cult).strip(),
                "cantidad": cantidad_val,
                "precio": 0.0,
                "moneda": "EUR",
                "cambio": 1.0,
                "periodo_desde": fecha_str,
                "periodo_hasta": fecha_str,
                "observaciones": "Importado desde Excel",
                "created_at": datetime.now(),
                "updated_at": datetime.now(),
            }
            await contratos_collection.insert_one(doc)
            imported += 1
        except Exception as e:
            errors.append({"row": row_num, "error": str(e)})

    return {
        "success": True,
        "imported": imported,
        "skipped_duplicates": skipped_dupes,
        "created_proveedores": created_provs,
        "created_cultivos": created_cults,
        "errors": errors[:50],  # solo las 50 primeras para no reventar respuesta
        "total_errors": len(errors),
    }


@router.get("/contratos")
async def get_contratos(
    skip: int = 0,
    limit: int = 100,
    campana: Optional[str] = None,
    proveedor: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
    _access: dict = Depends(RequireContratosAccess)
):
    query = {}
    if campana:
        query["campana"] = campana
    if proveedor:
        query["proveedor"] = {"$regex": proveedor, "$options": "i"}
    
    contratos = await contratos_collection.find(query).skip(skip).limit(limit).to_list(limit)
    return {"contratos": serialize_docs(contratos), "total": await contratos_collection.count_documents(query)}


@router.get("/contratos/next-numero")
async def get_next_numero_contrato(
    current_user: dict = Depends(get_current_user),
    _access: dict = Depends(RequireContratosAccess)
):
    """Previsualiza el próximo número de contrato del año actual (MP-{año}-{numero:06d}).

    El frontend lo usa para mostrar al usuario qué número tendrá el contrato
    ANTES de guardarlo. El valor real se reasigna atómicamente en POST /contratos
    para evitar colisiones.

    IMPORTANTE: este endpoint debe declararse ANTES de /contratos/{contrato_id}
    en este archivo; de lo contrario FastAPI interpreta "next-numero" como un
    ObjectId y devuelve "Invalid ID".
    """
    current_year = datetime.now().year
    last_contrato = await contratos_collection.find_one(
        {"año": current_year},
        sort=[("numero", -1)],
    )
    next_numero = (last_contrato.get("numero", 0) if last_contrato else 0) + 1
    return {
        "year": current_year,
        "numero": next_numero,
        "numero_contrato": f"MP-{current_year}-{str(next_numero).zfill(6)}",
    }


@router.get("/contratos/{contrato_id}")
async def get_contrato(
    contrato_id: str,
    current_user: dict = Depends(get_current_user),
    _access: dict = Depends(RequireContratosAccess)
):
    if not ObjectId.is_valid(contrato_id):
        raise HTTPException(status_code=400, detail="Invalid ID")
    
    contrato = await contratos_collection.find_one({"_id": ObjectId(contrato_id)})
    if not contrato:
        raise HTTPException(status_code=404, detail="Contrato not found")
    
    return serialize_doc(contrato)


@router.put("/contratos/{contrato_id}")
async def update_contrato(
    contrato_id: str,
    contrato: ContratoCreate,
    current_user: dict = Depends(RequireEdit),
    _access: dict = Depends(RequireContratosAccess)
):
    if not ObjectId.is_valid(contrato_id):
        raise HTTPException(status_code=400, detail="Invalid ID")
    
    # Obtener documento anterior para auditoría
    old_doc = await contratos_collection.find_one({"_id": ObjectId(contrato_id)})
    if not old_doc:
        raise HTTPException(status_code=404, detail="Contrato not found")

    # Validar tipo de operación tanto del documento existente como del nuevo
    ensure_tipo_operacion(current_user, old_doc.get("tipo"))
    ensure_tipo_operacion(current_user, contrato.tipo)
    
    # Lookup proveedor name (para contratos de Compra)
    proveedor_name = contrato.proveedor or ""
    if contrato.proveedor_id:
        prov = await proveedores_collection.find_one({"_id": ObjectId(contrato.proveedor_id)})
        if prov:
            proveedor_name = prov.get("nombre", "")
    
    # Lookup cliente name (para contratos de Venta)
    cliente_name = ""
    cliente_id_str = getattr(contrato, 'cliente_id', None) or ""
    if cliente_id_str:
        cli = await clientes_collection.find_one({"_id": ObjectId(cliente_id_str)})
        if cli:
            cliente_name = cli.get("nombre", "")
    
    # Lookup cultivo name
    cultivo_name = contrato.cultivo or ""
    if contrato.cultivo_id:
        cult = await cultivos_collection.find_one({"_id": ObjectId(contrato.cultivo_id)})
        if cult:
            cultivo_name = cult.get("nombre", "")
    
    update_data = contrato.dict()
    update_data.update({
        "proveedor": proveedor_name,
        "cliente": cliente_name,
        "cliente_id": cliente_id_str,
        "cultivo": cultivo_name,
        "updated_at": datetime.now()
    })
    
    result = await contratos_collection.update_one(
        {"_id": ObjectId(contrato_id)},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Contrato not found")
    
    updated = await contratos_collection.find_one({"_id": ObjectId(contrato_id)})
    
    # Calcular cambios y registrar en auditoría
    changes = calculate_changes(serialize_doc(old_doc.copy()), serialize_doc(updated.copy()))
    if changes:  # Solo registrar si hay cambios reales
        await create_audit_log(
            collection_name="contratos",
            document_id=contrato_id,
            action="update",
            user_email=current_user.get("email", "unknown"),
            user_name=current_user.get("full_name", current_user.get("username", "Usuario")),
            changes=changes
        )
    
    return {"success": True, "data": serialize_doc(updated)}


@router.delete("/contratos/{contrato_id}")
async def delete_contrato(
    contrato_id: str,
    current_user: dict = Depends(RequireDelete),
    _access: dict = Depends(RequireContratosAccess)
):
    if not ObjectId.is_valid(contrato_id):
        raise HTTPException(status_code=400, detail="Invalid ID")
    
    # Obtener documento antes de eliminar para auditoría
    old_doc = await contratos_collection.find_one({"_id": ObjectId(contrato_id)})
    if not old_doc:
        raise HTTPException(status_code=404, detail="Contrato not found")
    
    result = await contratos_collection.delete_one({"_id": ObjectId(contrato_id)})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Contrato not found")
    
    # Registrar eliminación en auditoría
    await create_audit_log(
        collection_name="contratos",
        document_id=contrato_id,
        action="delete",
        user_email=current_user.get("email", "unknown"),
        user_name=current_user.get("full_name", current_user.get("username", "Usuario")),
        previous_data=serialize_doc(old_doc.copy())
    )
    
    return {"success": True, "message": "Contrato deleted"}



@router.post("/contratos/regenerar-numeros")
async def regenerar_numeros_contratos(
    current_user: dict = Depends(RequireEdit)
):
    """
    Regenera el campo numero_contrato para todos los contratos existentes
    basándose en serie, año y numero.
    Formato: MP-{año}-{numero_6_digitos}
    """
    # Solo admin puede ejecutar esta operación
    if current_user.get('role') != 'Admin':
        raise HTTPException(status_code=403, detail="Solo administradores pueden regenerar números")
    
    # Obtener todos los contratos
    contratos = await contratos_collection.find({}).to_list(10000)
    
    actualizados = 0
    errores = []
    
    for contrato in contratos:
        try:
            serie = contrato.get("serie", "MP")
            año = contrato.get("año", datetime.now().year)
            numero = contrato.get("numero", 0)
            
            # Generar numero_contrato formateado
            numero_contrato = f"{serie}-{año}-{str(numero).zfill(6)}"
            
            # Actualizar el documento
            await contratos_collection.update_one(
                {"_id": contrato["_id"]},
                {"$set": {"numero_contrato": numero_contrato}}
            )
            actualizados += 1
        except Exception as e:
            errores.append({
                "id": str(contrato.get("_id")),
                "error": str(e)
            })
    
    return {
        "success": True,
        "message": f"Regenerados {actualizados} números de contrato",
        "actualizados": actualizados,
        "errores": errores,
        "total_contratos": len(contratos)
    }



# ============================================================================
# EXPORT FUNCTIONS - PDF & EXCEL
# ============================================================================

from utils.formatters import format_number_es  # noqa: E402


@router.get("/contratos/export/pdf")
async def export_contratos_pdf(
    proveedor: Optional[str] = None,
    cultivo: Optional[str] = None,
    campana: Optional[str] = None,
    tipo: Optional[str] = None,
    fecha_desde: Optional[str] = None,
    fecha_hasta: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Exporta el listado de contratos filtrado a PDF"""
    from weasyprint import HTML
    from fastapi.responses import Response
    
    # Build query
    query = {}
    if proveedor:
        query["proveedor"] = proveedor
    if cultivo:
        query["cultivo"] = cultivo
    if campana:
        query["campana"] = campana
    if tipo:
        query["tipo"] = tipo
    if fecha_desde or fecha_hasta:
        date_filter = {}
        if fecha_desde:
            date_filter["$gte"] = fecha_desde
        if fecha_hasta:
            date_filter["$lte"] = fecha_hasta
        query["fecha_contrato"] = date_filter
    
    # Get contracts
    contratos = await contratos_collection.find(query).sort("fecha_contrato", -1).to_list(1000)
    
    # Calculate totals - usando los campos correctos
    total_cantidad = sum(c.get("cantidad", 0) or 0 for c in contratos)
    total_importe = sum((c.get("cantidad", 0) or 0) * (c.get("precio", 0) or 0) for c in contratos)
    
    # Build filter description
    filtros_texto = []
    if proveedor:
        filtros_texto.append(f"Proveedor: {proveedor}")
    if cultivo:
        filtros_texto.append(f"Cultivo: {cultivo}")
    if campana:
        filtros_texto.append(f"Campaña: {campana}")
    if tipo:
        filtros_texto.append(f"Tipo: {tipo}")
    if fecha_desde:
        filtros_texto.append(f"Desde: {fecha_desde}")
    if fecha_hasta:
        filtros_texto.append(f"Hasta: {fecha_hasta}")
    
    # Generate HTML
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="utf-8">
        <style>
            body {{ font-family: Arial, sans-serif; font-size: 10pt; margin: 20px; }}
            h1 {{ color: #2563eb; font-size: 18pt; margin-bottom: 5px; }}
            .header {{ margin-bottom: 20px; border-bottom: 2px solid #2563eb; padding-bottom: 10px; }}
            .filters {{ background: #f3f4f6; padding: 10px; border-radius: 5px; margin-bottom: 15px; font-size: 9pt; }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 10px; }}
            th {{ background: #2563eb; color: white; padding: 8px 5px; text-align: left; font-size: 9pt; }}
            td {{ padding: 6px 5px; border-bottom: 1px solid #e5e7eb; font-size: 9pt; }}
            tr:nth-child(even) {{ background: #f9fafb; }}
            .tipo-compra {{ background: #dcfce7; color: #166534; padding: 2px 6px; border-radius: 3px; font-size: 8pt; }}
            .tipo-venta {{ background: #fef3c7; color: #92400e; padding: 2px 6px; border-radius: 3px; font-size: 8pt; }}
            .totals {{ margin-top: 20px; background: #eff6ff; padding: 15px; border-radius: 5px; }}
            .totals-grid {{ display: flex; justify-content: space-between; }}
            .total-item {{ text-align: center; }}
            .total-value {{ font-size: 14pt; font-weight: bold; color: #2563eb; }}
            .total-label {{ font-size: 9pt; color: #6b7280; }}
            .right {{ text-align: right; }}
            .date {{ font-size: 9pt; color: #6b7280; }}
        </style>
    </head>
    <body>
        <div class="header">
            <h1>LISTADO DE CONTRATOS</h1>
            <div class="date">Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}</div>
        </div>
        
        {f'<div class="filters"><strong>Filtros aplicados:</strong> {" | ".join(filtros_texto)}</div>' if filtros_texto else ''}
        
        <table>
            <thead>
                <tr>
                    <th>Nº Contrato</th>
                    <th>Tipo</th>
                    <th>Campaña</th>
                    <th>Proveedor/Cliente</th>
                    <th>Cultivo</th>
                    <th class="right">Cantidad (kg)</th>
                    <th class="right">Precio (€/kg)</th>
                    <th class="right">Total (€)</th>
                    <th>Fecha</th>
                </tr>
            </thead>
            <tbody>
    """
    
    for c in contratos:
        cantidad = c.get("cantidad", 0) or 0
        precio = c.get("precio", 0) or 0
        total = cantidad * precio
        tipo_class = "tipo-compra" if c.get("tipo") == "Compra" else "tipo-venta"
        proveedor_cliente = c.get("proveedor") or c.get("cliente") or "-"
        
        html_content += f"""
            <tr>
                <td><strong>{c.get('numero_contrato', '-')}</strong></td>
                <td><span class="{tipo_class}">{c.get('tipo', '-')}</span></td>
                <td>{c.get('campana', '-')}</td>
                <td>{proveedor_cliente}</td>
                <td>{c.get('cultivo', '-')}</td>
                <td class="right">{format_number_es(cantidad, 0)}</td>
                <td class="right">{format_number_es(precio, 2)}</td>
                <td class="right"><strong>{format_number_es(total, 2)}</strong></td>
                <td>{c.get('fecha_contrato', '-')}</td>
            </tr>
        """
    
    html_content += f"""
            </tbody>
        </table>
        
        <div class="totals">
            <div class="totals-grid">
                <div class="total-item">
                    <div class="total-value">{len(contratos)}</div>
                    <div class="total-label">Contratos</div>
                </div>
                <div class="total-item">
                    <div class="total-value">{format_number_es(total_cantidad, 0)} kg</div>
                    <div class="total-label">Cantidad Total</div>
                </div>
                <div class="total-item">
                    <div class="total-value">{format_number_es(total_importe, 2)} €</div>
                    <div class="total-label">Importe Total</div>
                </div>
            </div>
        </div>
    </body>
    </html>
    """
    
    # Generate PDF
    pdf = HTML(string=html_content).write_pdf()
    
    return Response(
        content=pdf,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=contratos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"}
    )


@router.get("/contratos/export/excel")
async def export_contratos_excel(
    proveedor: Optional[str] = None,
    cultivo: Optional[str] = None,
    campana: Optional[str] = None,
    tipo: Optional[str] = None,
    fecha_desde: Optional[str] = None,
    fecha_hasta: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Exporta el listado de contratos filtrado a Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    from fastapi.responses import Response
    
    # Build query
    query = {}
    if proveedor:
        query["proveedor"] = proveedor
    if cultivo:
        query["cultivo"] = cultivo
    if campana:
        query["campana"] = campana
    if tipo:
        query["tipo"] = tipo
    if fecha_desde or fecha_hasta:
        date_filter = {}
        if fecha_desde:
            date_filter["$gte"] = fecha_desde
        if fecha_hasta:
            date_filter["$lte"] = fecha_hasta
        query["fecha_contrato"] = date_filter
    
    # Get contracts
    contratos = await contratos_collection.find(query).sort("fecha_contrato", -1).to_list(1000)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Contratos"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ["Nº Contrato", "Tipo", "Campaña", "Proveedor/Cliente", "Cultivo", "Cantidad (kg)", "Precio (€/kg)", "Total (€)", "Fecha"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Data rows
    for row, c in enumerate(contratos, 2):
        cantidad = c.get("cantidad", 0) or 0
        precio = c.get("precio", 0) or 0
        total = cantidad * precio
        proveedor_cliente = c.get("proveedor") or c.get("cliente") or ""
        
        ws.cell(row=row, column=1, value=c.get('numero_contrato', '')).border = border
        ws.cell(row=row, column=2, value=c.get('tipo', '')).border = border
        ws.cell(row=row, column=3, value=c.get('campana', '')).border = border
        ws.cell(row=row, column=4, value=proveedor_cliente).border = border
        ws.cell(row=row, column=5, value=c.get('cultivo', '')).border = border
        ws.cell(row=row, column=6, value=cantidad).border = border
        ws.cell(row=row, column=7, value=precio).border = border
        ws.cell(row=row, column=8, value=total).border = border
        ws.cell(row=row, column=9, value=c.get('fecha_contrato', '')).border = border
    
    # Totals row
    total_row = len(contratos) + 2
    total_cantidad = sum(c.get("cantidad", 0) or 0 for c in contratos)
    total_importe = sum((c.get("cantidad", 0) or 0) * (c.get("precio", 0) or 0) for c in contratos)
    
    ws.cell(row=total_row, column=5, value="TOTALES:").font = Font(bold=True)
    ws.cell(row=total_row, column=6, value=total_cantidad).font = Font(bold=True)
    ws.cell(row=total_row, column=8, value=total_importe).font = Font(bold=True)
    
    # Adjust column widths
    column_widths = [18, 10, 12, 25, 15, 15, 15, 15, 12]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=contratos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"}
    )