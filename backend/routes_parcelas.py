"""
Routes for Parcelas (Parcels) - CRUD operations
Extracted from routes_main.py for better code organization
"""

from fastapi import APIRouter, HTTPException, Depends
from typing import Optional
from bson import ObjectId
from datetime import datetime

from models import ParcelaCreate, ParcelaUpdate
from database import parcelas_collection, serialize_doc, serialize_docs
from rbac_guards import (
    RequireCreate, RequireEdit, RequireDelete,
    RequireParcelasAccess, get_current_user
)

router = APIRouter(prefix="/api", tags=["parcelas"])


@router.post("/parcelas", response_model=dict)
async def create_parcela(
    parcela: ParcelaCreate,
    current_user: dict = Depends(RequireCreate),
    _access: dict = Depends(RequireParcelasAccess)
):
    parcela_dict = parcela.dict()
    parcela_dict.update({
        "activo": True,
        "unidad_medida": "ha",
        "created_at": datetime.now(),
        "updated_at": datetime.now()
    })
    
    result = await parcelas_collection.insert_one(parcela_dict)
    created = await parcelas_collection.find_one({"_id": result.inserted_id})
    
    return {"success": True, "data": serialize_doc(created)}


@router.get("/parcelas")
async def get_parcelas(
    skip: int = 0,
    limit: int = 100,
    campana: Optional[str] = None,
    proveedor: Optional[str] = None,
    contrato_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
    _access: dict = Depends(RequireParcelasAccess)
):
    query = {}
    if campana:
        query["campana"] = campana
    if proveedor:
        query["proveedor"] = {"$regex": proveedor, "$options": "i"}
    if contrato_id:
        query["contrato_id"] = contrato_id
    
    parcelas = await parcelas_collection.find(query).skip(skip).limit(limit).to_list(limit)
    return {"parcelas": serialize_docs(parcelas), "total": await parcelas_collection.count_documents(query)}


@router.get("/parcelas/{parcela_id}")
async def get_parcela(
    parcela_id: str,
    current_user: dict = Depends(get_current_user),
    _access: dict = Depends(RequireParcelasAccess)
):
    if not ObjectId.is_valid(parcela_id):
        raise HTTPException(status_code=400, detail="Invalid ID")
    
    parcela = await parcelas_collection.find_one({"_id": ObjectId(parcela_id)})
    if not parcela:
        raise HTTPException(status_code=404, detail="Parcela not found")
    
    return serialize_doc(parcela)


@router.put("/parcelas/{parcela_id}")
async def update_parcela(
    parcela_id: str,
    parcela: ParcelaUpdate,
    current_user: dict = Depends(RequireEdit),
    _access: dict = Depends(RequireParcelasAccess)
):
    if not ObjectId.is_valid(parcela_id):
        raise HTTPException(status_code=400, detail="Invalid ID")
    
    # Only include fields that were actually provided
    update_data = {k: v for k, v in parcela.dict().items() if v is not None}
    update_data["updated_at"] = datetime.now()
    
    result = await parcelas_collection.update_one(
        {"_id": ObjectId(parcela_id)},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Parcela not found")
    
    updated = await parcelas_collection.find_one({"_id": ObjectId(parcela_id)})
    return {"success": True, "data": serialize_doc(updated)}


@router.delete("/parcelas/{parcela_id}")
async def delete_parcela(
    parcela_id: str,
    current_user: dict = Depends(RequireDelete),
    _access: dict = Depends(RequireParcelasAccess)
):
    if not ObjectId.is_valid(parcela_id):
        raise HTTPException(status_code=400, detail="Invalid ID")
    
    result = await parcelas_collection.delete_one({"_id": ObjectId(parcela_id)})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Parcela not found")
    
    return {"success": True, "message": "Parcela deleted"}



@router.get("/parcelas/export/excel")
async def export_parcelas_excel(
    campana: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Export parcelas to Excel"""
    from fastapi.responses import StreamingResponse
    import io as _io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    query = {}
    if campana:
        query["campana"] = campana

    parcelas_list = await parcelas_collection.find(query).sort("created_at", -1).to_list(5000)
    parcelas_raw = serialize_docs(parcelas_list)

    wb = Workbook()
    ws = wb.active
    ws.title = "Parcelas"

    headers = ["Codigo", "Proveedor", "Finca", "Cultivo", "Variedad", "Superficie (ha)", "N Plantas", "Campana", "Zonas"]
    header_fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for idx, p in enumerate(parcelas_raw, 2):
        row = [
            p.get("codigo_plantacion", ""), p.get("proveedor", ""), p.get("finca", ""),
            p.get("cultivo", ""), p.get("variedad", ""),
            p.get("superficie_total", 0), p.get("num_plantas", 0),
            p.get("campana", ""), len(p.get("recintos", []))
        ]
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=idx, column=col, value=val)
            cell.border = thin_border

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 18

    output = _io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"parcelas_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename={filename}"})


@router.get("/parcelas/export/pdf")
async def export_parcelas_pdf(
    campana: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Export parcelas to PDF"""
    from fastapi.responses import StreamingResponse
    import io as _io
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    query = {}
    if campana:
        query["campana"] = campana

    parcelas_list = await parcelas_collection.find(query).sort("created_at", -1).to_list(5000)
    parcelas_raw = serialize_docs(parcelas_list)

    output = _io.BytesIO()
    pdf = SimpleDocTemplate(output, pagesize=landscape(A4), topMargin=15*mm, bottomMargin=15*mm)
    elements = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=16, textColor=colors.HexColor('#1B5E20'), alignment=1, spaceAfter=8*mm)
    elements.append(Paragraph("Informe de Parcelas - FRUVECO", title_style))
    elements.append(Paragraph(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", ParagraphStyle('Sub', parent=styles['Normal'], alignment=1, textColor=colors.gray)))
    elements.append(Spacer(1, 8*mm))

    table_data = [["Codigo", "Proveedor", "Finca", "Cultivo", "Variedad", "Sup. (ha)", "Plantas", "Campana"]]
    total_sup = total_plantas = 0
    for p in parcelas_raw:
        sup = p.get("superficie_total", 0)
        plantas = p.get("num_plantas", 0)
        total_sup += sup or 0
        total_plantas += plantas or 0
        table_data.append([
            p.get("codigo_plantacion", "")[:15], p.get("proveedor", "")[:20],
            p.get("finca", "")[:15], p.get("cultivo", "")[:15], p.get("variedad", "")[:15],
            f"{sup:.2f}" if sup else "0", f"{plantas:,}" if plantas else "0", p.get("campana", "")
        ])
    table_data.append(["TOTAL", "", "", "", "", f"{total_sup:.2f}", f"{total_plantas:,}", ""])

    col_widths = [30*mm, 35*mm, 25*mm, 25*mm, 25*mm, 20*mm, 20*mm, 25*mm]
    doc_table = Table(table_data, colWidths=col_widths)
    doc_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1B5E20')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.gray),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F5F5F5')]),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E8F5E9')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
    ]))
    elements.append(doc_table)

    pdf.build(elements)
    output.seek(0)
    filename = f"parcelas_{datetime.now().strftime('%Y%m%d')}.pdf"
    return StreamingResponse(output, media_type="application/pdf",
                             headers={"Content-Disposition": f"attachment; filename={filename}"})
