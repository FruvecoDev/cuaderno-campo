from fastapi import APIRouter, HTTPException, Depends, Query
from fastapi.responses import StreamingResponse
from typing import Optional, List
from datetime import datetime
from bson import ObjectId
import io

from database import (
    albaranes_collection, contratos_collection, clientes_collection,
    serialize_doc, serialize_docs, db
)
from rbac_guards import RequireAlbaranesAccess, get_current_user

router = APIRouter(prefix="/api/ingresos", tags=["ingresos"])

def parse_date(date_str: str) -> datetime:
    """Parse date string to datetime"""
    try:
        return datetime.strptime(date_str, "%Y-%m-%d")
    except ValueError:
        return None

# ============================================================================
# INFORMES DE INGRESOS (Albaranes de Venta)
# ============================================================================

@router.get("/resumen")
async def get_resumen_ingresos(
    fecha_desde: Optional[str] = None,
    fecha_hasta: Optional[str] = None,
    campana: Optional[str] = None,
    contrato_id: Optional[str] = None,
    cultivo: Optional[str] = None,
    cliente: Optional[str] = None,
    parcela_codigo: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
    _access: dict = Depends(RequireAlbaranesAccess)
):
    """
    Obtiene un resumen general de ingresos (albaranes de venta) con totales por cliente, contrato, cultivo y parcela.
    """
    # Build match query - Solo albaranes de venta
    match_query = {"tipo": "Albarán de venta"}
    
    if fecha_desde:
        match_query["fecha"] = {"$gte": fecha_desde}
    if fecha_hasta:
        if "fecha" in match_query:
            match_query["fecha"]["$lte"] = fecha_hasta
        else:
            match_query["fecha"] = {"$lte": fecha_hasta}
    if campana:
        match_query["campana"] = campana
    if contrato_id:
        match_query["contrato_id"] = contrato_id
    if cultivo:
        match_query["cultivo"] = cultivo
    if cliente:
        match_query["cliente"] = cliente
    if parcela_codigo:
        match_query["parcela_codigo"] = parcela_codigo
    
    # Get all albaranes de venta matching criteria
    albaranes = await albaranes_collection.find(match_query).to_list(1000)
    
    # Aggregate by different dimensions
    por_cliente = {}
    por_contrato = {}
    por_cultivo = {}
    por_parcela = {}
    total_general = 0
    total_albaranes = len(albaranes)
    
    for albaran in albaranes:
        total = albaran.get("total_albaran", 0) or 0
        total_general += total
        
        # Por cliente
        cliente_nombre = albaran.get("cliente") or "Sin cliente"
        if cliente_nombre not in por_cliente:
            por_cliente[cliente_nombre] = {"total": 0, "count": 0, "albaranes": []}
        por_cliente[cliente_nombre]["total"] += total
        por_cliente[cliente_nombre]["count"] += 1
        
        # Por contrato
        contrato_id = albaran.get("contrato_id")
        if contrato_id:
            if contrato_id not in por_contrato:
                por_contrato[contrato_id] = {"total": 0, "count": 0, "cliente": cliente_nombre}
            por_contrato[contrato_id]["total"] += total
            por_contrato[contrato_id]["count"] += 1
        
        # Por cultivo
        cultivo = albaran.get("cultivo") or "Sin cultivo"
        if cultivo not in por_cultivo:
            por_cultivo[cultivo] = {"total": 0, "count": 0}
        por_cultivo[cultivo]["total"] += total
        por_cultivo[cultivo]["count"] += 1
        
        # Por parcela
        parcela = albaran.get("parcela_codigo") or "Sin parcela"
        if parcela not in por_parcela:
            por_parcela[parcela] = {"total": 0, "count": 0, "cultivo": cultivo}
        por_parcela[parcela]["total"] += total
        por_parcela[parcela]["count"] += 1
    
    # Enrich contract data with contract info
    contratos_enriched = []
    for contrato_id, data in por_contrato.items():
        try:
            contrato = await contratos_collection.find_one({"_id": ObjectId(contrato_id)})
            contratos_enriched.append({
                "contrato_id": contrato_id,
                "numero_contrato": contrato.get("numero_contrato") if contrato else None,
                "cliente": data["cliente"],
                "cultivo": contrato.get("cultivo") if contrato else None,
                "campana": contrato.get("campana") if contrato else None,
                "total": data["total"],
                "count": data["count"]
            })
        except Exception:
            contratos_enriched.append({
                "contrato_id": contrato_id,
                "total": data["total"],
                "count": data["count"],
                "cliente": data["cliente"]
            })
    
    return {
        "total_general": total_general,
        "total_albaranes": total_albaranes,
        "por_cliente": [{"cliente": k, **v} for k, v in sorted(por_cliente.items(), key=lambda x: -x[1]["total"])],
        "por_contrato": sorted(contratos_enriched, key=lambda x: -x["total"]),
        "por_cultivo": [{"cultivo": k, **v} for k, v in sorted(por_cultivo.items(), key=lambda x: -x[1]["total"])],
        "por_parcela": [{"parcela": k, **v} for k, v in sorted(por_parcela.items(), key=lambda x: -x[1]["total"])]
    }


@router.get("/por-cliente")
async def get_ingresos_por_cliente(
    cliente: Optional[str] = None,
    fecha_desde: Optional[str] = None,
    fecha_hasta: Optional[str] = None,
    campana: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
    _access: dict = Depends(RequireAlbaranesAccess)
):
    """
    Obtiene detalle de ingresos agrupados por cliente.
    Si se especifica cliente, devuelve el detalle de ese cliente.
    """
    match_query = {"tipo": "Albarán de venta"}
    if cliente:
        match_query["cliente"] = cliente
    if fecha_desde:
        match_query["fecha"] = {"$gte": fecha_desde}
    if fecha_hasta:
        if "fecha" in match_query:
            match_query["fecha"]["$lte"] = fecha_hasta
        else:
            match_query["fecha"] = {"$lte": fecha_hasta}
    if campana:
        match_query["campana"] = campana
    
    pipeline = [
        {"$match": match_query},
        {"$group": {
            "_id": "$cliente",
            "total": {"$sum": "$total_albaran"},
            "count": {"$sum": 1},
            "cultivos": {"$addToSet": "$cultivo"},
            "contratos": {"$addToSet": "$contrato_id"},
            "primer_albaran": {"$min": "$fecha"},
            "ultimo_albaran": {"$max": "$fecha"}
        }},
        {"$sort": {"total": -1}}
    ]
    
    results = await albaranes_collection.aggregate(pipeline).to_list(100)
    
    return {
        "ingresos_por_cliente": [
            {
                "cliente": r["_id"] or "Sin cliente",
                "total": r["total"],
                "count": r["count"],
                "cultivos": [c for c in r["cultivos"] if c],
                "num_contratos": len([c for c in r["contratos"] if c]),
                "primer_albaran": r["primer_albaran"],
                "ultimo_albaran": r["ultimo_albaran"]
            }
            for r in results
        ]
    }


@router.get("/por-contrato")
async def get_ingresos_por_contrato(
    contrato_id: Optional[str] = None,
    fecha_desde: Optional[str] = None,
    fecha_hasta: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
    _access: dict = Depends(RequireAlbaranesAccess)
):
    """
    Obtiene detalle de ingresos agrupados por contrato de venta.
    """
    match_query = {"tipo": "Albarán de venta"}
    if contrato_id:
        match_query["contrato_id"] = contrato_id
    if fecha_desde:
        match_query["fecha"] = {"$gte": fecha_desde}
    if fecha_hasta:
        if "fecha" in match_query:
            match_query["fecha"]["$lte"] = fecha_hasta
        else:
            match_query["fecha"] = {"$lte": fecha_hasta}
    
    pipeline = [
        {"$match": match_query},
        {"$group": {
            "_id": "$contrato_id",
            "total": {"$sum": "$total_albaran"},
            "count": {"$sum": 1},
            "cliente": {"$first": "$cliente"},
            "cultivo": {"$first": "$cultivo"},
            "campana": {"$first": "$campana"},
            "parcela": {"$first": "$parcela_codigo"},
            "primer_albaran": {"$min": "$fecha"},
            "ultimo_albaran": {"$max": "$fecha"}
        }},
        {"$sort": {"total": -1}}
    ]
    
    results = await albaranes_collection.aggregate(pipeline).to_list(100)
    
    # Enrich with contract details
    enriched_results = []
    for r in results:
        contrato_info = {}
        if r["_id"]:
            try:
                contrato = await contratos_collection.find_one({"_id": ObjectId(r["_id"])})
                if contrato:
                    contrato_info = {
                        "numero_contrato": contrato.get("numero_contrato"),
                        "precio": contrato.get("precio"),
                        "cantidad": contrato.get("cantidad")
                    }
            except Exception:
                pass
        
        enriched_results.append({
            "contrato_id": r["_id"] or "Sin contrato",
            "cliente": r["cliente"] or "Sin cliente",
            "cultivo": r["cultivo"],
            "campana": r["campana"],
            "parcela": r["parcela"],
            "total": r["total"],
            "count": r["count"],
            "primer_albaran": r["primer_albaran"],
            "ultimo_albaran": r["ultimo_albaran"],
            **contrato_info
        })
    
    return {"ingresos_por_contrato": enriched_results}


@router.get("/detalle-albaranes")
async def get_detalle_albaranes_venta(
    cliente: Optional[str] = None,
    contrato_id: Optional[str] = None,
    cultivo: Optional[str] = None,
    parcela_codigo: Optional[str] = None,
    fecha_desde: Optional[str] = None,
    fecha_hasta: Optional[str] = None,
    campana: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
    _access: dict = Depends(RequireAlbaranesAccess)
):
    """
    Obtiene el detalle de albaranes de venta según los filtros especificados.
    """
    match_query = {"tipo": "Albarán de venta"}
    
    if cliente:
        match_query["cliente"] = cliente
    if contrato_id:
        match_query["contrato_id"] = contrato_id
    if cultivo:
        match_query["cultivo"] = cultivo
    if parcela_codigo:
        match_query["parcela_codigo"] = parcela_codigo
    if fecha_desde:
        match_query["fecha"] = {"$gte": fecha_desde}
    if fecha_hasta:
        if "fecha" in match_query:
            match_query["fecha"]["$lte"] = fecha_hasta
        else:
            match_query["fecha"] = {"$lte": fecha_hasta}
    if campana:
        match_query["campana"] = campana
    
    albaranes = await albaranes_collection.find(match_query).sort("fecha", -1).to_list(500)
    
    return {
        "albaranes": serialize_docs(albaranes),
        "total": sum(a.get("total_albaran", 0) or 0 for a in albaranes),
        "count": len(albaranes)
    }


@router.get("/campanas")
async def get_campanas_ingresos(
    current_user: dict = Depends(get_current_user)
):
    """Obtiene las campañas disponibles en albaranes de venta"""
    pipeline = [
        {"$match": {"tipo": "Albarán de venta", "campana": {"$ne": None}}},
        {"$group": {"_id": "$campana"}},
        {"$sort": {"_id": -1}}
    ]
    results = await albaranes_collection.aggregate(pipeline).to_list(50)
    return {"campanas": [r["_id"] for r in results if r["_id"]]}


@router.get("/filtros-opciones")
async def get_filtros_opciones_ingresos(
    current_user: dict = Depends(get_current_user)
):
    """Obtiene las opciones disponibles para los filtros de ingresos"""
    # Contratos de venta
    contratos = await contratos_collection.find(
        {"tipo": "Venta"},
        {"_id": 1, "numero_contrato": 1, "cliente": 1, "cultivo": 1}
    ).to_list(100)
    
    # Cultivos únicos de albaranes de venta
    cultivos_pipeline = [
        {"$match": {"tipo": "Albarán de venta", "cultivo": {"$ne": None}}},
        {"$group": {"_id": "$cultivo"}},
        {"$sort": {"_id": 1}}
    ]
    cultivos = await albaranes_collection.aggregate(cultivos_pipeline).to_list(50)
    
    # Clientes únicos
    clientes_pipeline = [
        {"$match": {"tipo": "Albarán de venta", "cliente": {"$ne": None}}},
        {"$group": {"_id": "$cliente"}},
        {"$sort": {"_id": 1}}
    ]
    clientes = await albaranes_collection.aggregate(clientes_pipeline).to_list(100)
    
    # Parcelas únicas
    parcelas_pipeline = [
        {"$match": {"tipo": "Albarán de venta", "parcela_codigo": {"$ne": None}}},
        {"$group": {"_id": "$parcela_codigo"}},
        {"$sort": {"_id": 1}}
    ]
    parcelas = await albaranes_collection.aggregate(parcelas_pipeline).to_list(100)
    
    return {
        "contratos": [{"_id": str(c["_id"]), "numero_contrato": c.get("numero_contrato"), "cliente": c.get("cliente"), "cultivo": c.get("cultivo")} for c in contratos],
        "cultivos": [c["_id"] for c in cultivos if c["_id"]],
        "clientes": [c["_id"] for c in clientes if c["_id"]],
        "parcelas": [p["_id"] for p in parcelas if p["_id"]]
    }


@router.get("/export/excel")
async def export_ingresos_excel(
    fecha_desde: Optional[str] = None,
    fecha_hasta: Optional[str] = None,
    campana: Optional[str] = None,
    cliente: Optional[str] = None,
    cultivo: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
    _access: dict = Depends(RequireAlbaranesAccess)
):
    """Exporta los ingresos a Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # Build match query
        match_query = {"tipo": "Albarán de venta"}
        if fecha_desde:
            match_query["fecha"] = {"$gte": fecha_desde}
        if fecha_hasta:
            if "fecha" in match_query:
                match_query["fecha"]["$lte"] = fecha_hasta
            else:
                match_query["fecha"] = {"$lte": fecha_hasta}
        if campana:
            match_query["campana"] = campana
        if cliente:
            match_query["cliente"] = cliente
        if cultivo:
            match_query["cultivo"] = cultivo
        
        albaranes = await albaranes_collection.find(match_query).sort("fecha", -1).to_list(1000)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Ingresos"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="16a34a", end_color="16a34a", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = ["Fecha", "Nº Albarán", "Cliente", "Cultivo", "Parcela", "Campaña", "Total (€)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Data
        total_general = 0
        for row, albaran in enumerate(albaranes, 2):
            total = albaran.get("total_albaran", 0) or 0
            total_general += total
            
            ws.cell(row=row, column=1, value=albaran.get("fecha", "")).border = border
            ws.cell(row=row, column=2, value=albaran.get("numero_albaran", "")).border = border
            ws.cell(row=row, column=3, value=albaran.get("cliente", "")).border = border
            ws.cell(row=row, column=4, value=albaran.get("cultivo", "")).border = border
            ws.cell(row=row, column=5, value=albaran.get("parcela_codigo", "")).border = border
            ws.cell(row=row, column=6, value=albaran.get("campana", "")).border = border
            cell = ws.cell(row=row, column=7, value=total)
            cell.number_format = '#,##0.00 €'
            cell.border = border
        
        # Total row
        total_row = len(albaranes) + 2
        ws.cell(row=total_row, column=6, value="TOTAL:").font = Font(bold=True)
        total_cell = ws.cell(row=total_row, column=7, value=total_general)
        total_cell.font = Font(bold=True)
        total_cell.number_format = '#,##0.00 €'
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 15
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f"ingresos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generando Excel: {str(e)}")


@router.get("/export/pdf")
async def export_ingresos_pdf(
    fecha_desde: Optional[str] = None,
    fecha_hasta: Optional[str] = None,
    campana: Optional[str] = None,
    cliente: Optional[str] = None,
    cultivo: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
    _access: dict = Depends(RequireAlbaranesAccess)
):
    """Exporta los ingresos a PDF"""
    try:
        from weasyprint import HTML, CSS
        
        # Build match query
        match_query = {"tipo": "Albarán de venta"}
        if fecha_desde:
            match_query["fecha"] = {"$gte": fecha_desde}
        if fecha_hasta:
            if "fecha" in match_query:
                match_query["fecha"]["$lte"] = fecha_hasta
            else:
                match_query["fecha"] = {"$lte": fecha_hasta}
        if campana:
            match_query["campana"] = campana
        if cliente:
            match_query["cliente"] = cliente
        if cultivo:
            match_query["cultivo"] = cultivo
        
        albaranes = await albaranes_collection.find(match_query).sort("fecha", -1).to_list(1000)
        
        # Calculate totals
        total_general = sum(a.get("total_albaran", 0) or 0 for a in albaranes)
        
        # Group by cliente
        por_cliente = {}
        for a in albaranes:
            cli = a.get("cliente") or "Sin cliente"
            if cli not in por_cliente:
                por_cliente[cli] = {"total": 0, "count": 0}
            por_cliente[cli]["total"] += a.get("total_albaran", 0) or 0
            por_cliente[cli]["count"] += 1
        
        # Generate HTML
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <style>
                body {{ font-family: Arial, sans-serif; font-size: 10px; margin: 20px; }}
                h1 {{ color: #16a34a; font-size: 18px; margin-bottom: 5px; }}
                h2 {{ color: #333; font-size: 14px; margin-top: 20px; border-bottom: 2px solid #16a34a; padding-bottom: 5px; }}
                .header {{ margin-bottom: 20px; }}
                .periodo {{ color: #666; font-size: 11px; }}
                .kpis {{ display: flex; gap: 20px; margin: 15px 0; }}
                .kpi {{ background: #f0fdf4; padding: 10px 15px; border-radius: 8px; text-align: center; }}
                .kpi-value {{ font-size: 20px; font-weight: bold; color: #16a34a; }}
                .kpi-label {{ font-size: 9px; color: #666; }}
                table {{ width: 100%; border-collapse: collapse; margin-top: 10px; }}
                th {{ background: #16a34a; color: white; padding: 8px; text-align: left; font-size: 9px; }}
                td {{ padding: 6px 8px; border-bottom: 1px solid #ddd; font-size: 9px; }}
                tr:nth-child(even) {{ background: #f9f9f9; }}
                .total-row {{ background: #f0fdf4 !important; font-weight: bold; }}
                .amount {{ text-align: right; }}
                .footer {{ margin-top: 30px; text-align: center; color: #999; font-size: 8px; }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>INFORME DE INGRESOS</h1>
                <div class="periodo">
                    Período: {fecha_desde or 'Inicio'} - {fecha_hasta or 'Actual'}
                    {f' | Campaña: {campana}' if campana else ''}
                    {f' | Cliente: {cliente}' if cliente else ''}
                </div>
            </div>
            
            <div class="kpis">
                <div class="kpi">
                    <div class="kpi-value">€{total_general:,.2f}</div>
                    <div class="kpi-label">TOTAL INGRESOS</div>
                </div>
                <div class="kpi">
                    <div class="kpi-value">{len(albaranes)}</div>
                    <div class="kpi-label">ALBARANES</div>
                </div>
                <div class="kpi">
                    <div class="kpi-value">{len(por_cliente)}</div>
                    <div class="kpi-label">CLIENTES</div>
                </div>
            </div>
            
            <h2>Resumen por Cliente</h2>
            <table>
                <thead>
                    <tr>
                        <th>Cliente</th>
                        <th>Nº Albaranes</th>
                        <th class="amount">Total (€)</th>
                        <th class="amount">% del Total</th>
                    </tr>
                </thead>
                <tbody>
        """
        
        for cli, data in sorted(por_cliente.items(), key=lambda x: -x[1]["total"]):
            porcentaje = (data["total"] / total_general * 100) if total_general > 0 else 0
            html_content += f"""
                    <tr>
                        <td>{cli}</td>
                        <td>{data["count"]}</td>
                        <td class="amount">€{data["total"]:,.2f}</td>
                        <td class="amount">{porcentaje:.1f}%</td>
                    </tr>
            """
        
        html_content += f"""
                    <tr class="total-row">
                        <td>TOTAL</td>
                        <td>{len(albaranes)}</td>
                        <td class="amount">€{total_general:,.2f}</td>
                        <td class="amount">100%</td>
                    </tr>
                </tbody>
            </table>
            
            <h2>Detalle de Albaranes</h2>
            <table>
                <thead>
                    <tr>
                        <th>Fecha</th>
                        <th>Nº Albarán</th>
                        <th>Cliente</th>
                        <th>Cultivo</th>
                        <th>Parcela</th>
                        <th class="amount">Total (€)</th>
                    </tr>
                </thead>
                <tbody>
        """
        
        for a in albaranes[:100]:  # Limit to 100 for PDF
            html_content += f"""
                    <tr>
                        <td>{a.get("fecha", "")}</td>
                        <td>{a.get("numero_albaran", "")}</td>
                        <td>{a.get("cliente", "")}</td>
                        <td>{a.get("cultivo", "")}</td>
                        <td>{a.get("parcela_codigo", "")}</td>
                        <td class="amount">€{(a.get("total_albaran", 0) or 0):,.2f}</td>
                    </tr>
            """
        
        html_content += f"""
                </tbody>
            </table>
            
            <div class="footer">
                Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')} | FRUVECO - Sistema de Gestión Agrícola
            </div>
        </body>
        </html>
        """
        
        # Generate PDF
        pdf = HTML(string=html_content).write_pdf()
        
        filename = f"ingresos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
        return StreamingResponse(
            io.BytesIO(pdf),
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generando PDF: {str(e)}")
