"""
Routes for Técnicos Aplicadores - Applicator Technicians Management
CRUD operations for managing certified applicators used in treatments
"""

from fastapi import APIRouter, HTTPException, UploadFile, File, Form, Depends
from typing import Optional
from bson import ObjectId
from datetime import datetime, timedelta
from pydantic import BaseModel
import os
import uuid

from database import db, serialize_doc, serialize_docs
from rbac_guards import (
    RequireCreate, RequireEdit, RequireDelete,
    get_current_user
)

router = APIRouter(prefix="/api", tags=["tecnicos_aplicadores"])

# Collection
tecnicos_aplicadores_collection = db['tecnicos_aplicadores']

# Niveles de capacitación según normativa española
NIVELES_CAPACITACION = [
    "Básico",
    "Cualificado", 
    "Fumigador",
    "Piloto Aplicador"
]


class TecnicoAplicadorCreate(BaseModel):
    nombre: str
    apellidos: str
    dni: str
    nivel_capacitacion: str
    num_carnet: str
    fecha_certificacion: str  # YYYY-MM-DD
    observaciones: Optional[str] = ""


@router.post("/tecnicos-aplicadores", response_model=dict)
async def create_tecnico_aplicador(
    tecnico: TecnicoAplicadorCreate,
    current_user: dict = Depends(RequireCreate)
):
    """Crear un nuevo técnico aplicador"""
    # Validar nivel de capacitación
    if tecnico.nivel_capacitacion not in NIVELES_CAPACITACION:
        raise HTTPException(
            status_code=400, 
            detail=f"Nivel de capacitación inválido. Opciones: {NIVELES_CAPACITACION}"
        )
    
    # Verificar si ya existe un técnico con el mismo DNI
    existing = await tecnicos_aplicadores_collection.find_one({"dni": tecnico.dni})
    if existing:
        raise HTTPException(status_code=400, detail="Ya existe un técnico con este DNI")
    
    # Calcular fecha de validez (10 años después de certificación)
    try:
        fecha_cert = datetime.strptime(tecnico.fecha_certificacion, "%Y-%m-%d")
        fecha_validez = fecha_cert + timedelta(days=365*10)  # 10 años
    except ValueError:
        raise HTTPException(status_code=400, detail="Formato de fecha inválido. Use YYYY-MM-DD")
    
    tecnico_dict = tecnico.dict()
    tecnico_dict.update({
        "fecha_validez": fecha_validez.strftime("%Y-%m-%d"),
        "imagen_certificado_url": None,
        "activo": True,
        "created_at": datetime.now(),
        "updated_at": datetime.now()
    })
    
    result = await tecnicos_aplicadores_collection.insert_one(tecnico_dict)
    created = await tecnicos_aplicadores_collection.find_one({"_id": result.inserted_id})
    
    return {"success": True, "data": serialize_doc(created)}


@router.get("/tecnicos-aplicadores")
async def get_tecnicos_aplicadores(
    skip: int = 0,
    limit: int = 100,
    activo: Optional[bool] = None,
    nivel: Optional[str] = None,
    search: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Listar técnicos aplicadores con filtros opcionales"""
    query = {}
    
    if activo is not None:
        query["activo"] = activo
    
    if nivel:
        query["nivel_capacitacion"] = nivel
    
    if search:
        query["$or"] = [
            {"nombre": {"$regex": search, "$options": "i"}},
            {"apellidos": {"$regex": search, "$options": "i"}},
            {"dni": {"$regex": search, "$options": "i"}},
            {"num_carnet": {"$regex": search, "$options": "i"}}
        ]
    
    tecnicos = await tecnicos_aplicadores_collection.find(query).sort("apellidos", 1).skip(skip).limit(limit).to_list(limit)
    total = await tecnicos_aplicadores_collection.count_documents(query)
    
    return {"tecnicos": serialize_docs(tecnicos), "total": total}


@router.get("/tecnicos-aplicadores/activos")
async def get_tecnicos_activos(
    current_user: dict = Depends(get_current_user)
):
    """Obtener lista de técnicos activos para selectores (dropdown en tratamientos)"""
    hoy = datetime.now().strftime("%Y-%m-%d")
    
    # Solo técnicos activos con certificación válida
    query = {
        "activo": True,
        "fecha_validez": {"$gte": hoy}
    }
    
    tecnicos = await tecnicos_aplicadores_collection.find(query).sort("apellidos", 1).to_list(500)
    
    result = []
    for t in tecnicos:
        result.append({
            "_id": str(t["_id"]),
            "nombre_completo": f"{t.get('nombre', '')} {t.get('apellidos', '')}".strip(),
            "dni": t.get("dni", ""),
            "nivel_capacitacion": t.get("nivel_capacitacion", ""),
            "num_carnet": t.get("num_carnet", ""),
            "fecha_validez": t.get("fecha_validez", "")
        })
    
    return {"tecnicos": result}


@router.get("/tecnicos-aplicadores/niveles")
async def get_niveles_capacitacion():
    """Obtener lista de niveles de capacitación disponibles"""
    return {"niveles": NIVELES_CAPACITACION}


@router.get("/tecnicos-aplicadores/{tecnico_id}")
async def get_tecnico_aplicador(
    tecnico_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Obtener un técnico aplicador por ID"""
    if not ObjectId.is_valid(tecnico_id):
        raise HTTPException(status_code=400, detail="ID inválido")
    
    tecnico = await tecnicos_aplicadores_collection.find_one({"_id": ObjectId(tecnico_id)})
    if not tecnico:
        raise HTTPException(status_code=404, detail="Técnico aplicador no encontrado")
    
    return serialize_doc(tecnico)


@router.put("/tecnicos-aplicadores/{tecnico_id}")
async def update_tecnico_aplicador(
    tecnico_id: str,
    tecnico: TecnicoAplicadorCreate,
    current_user: dict = Depends(RequireEdit)
):
    """Actualizar un técnico aplicador"""
    if not ObjectId.is_valid(tecnico_id):
        raise HTTPException(status_code=400, detail="ID inválido")
    
    # Validar nivel de capacitación
    if tecnico.nivel_capacitacion not in NIVELES_CAPACITACION:
        raise HTTPException(
            status_code=400, 
            detail=f"Nivel de capacitación inválido. Opciones: {NIVELES_CAPACITACION}"
        )
    
    # Verificar si existe otro técnico con el mismo DNI
    existing = await tecnicos_aplicadores_collection.find_one({
        "dni": tecnico.dni,
        "_id": {"$ne": ObjectId(tecnico_id)}
    })
    if existing:
        raise HTTPException(status_code=400, detail="Ya existe otro técnico con este DNI")
    
    # Recalcular fecha de validez
    try:
        fecha_cert = datetime.strptime(tecnico.fecha_certificacion, "%Y-%m-%d")
        fecha_validez = fecha_cert + timedelta(days=365*10)
    except ValueError:
        raise HTTPException(status_code=400, detail="Formato de fecha inválido. Use YYYY-MM-DD")
    
    update_data = tecnico.dict()
    update_data.update({
        "fecha_validez": fecha_validez.strftime("%Y-%m-%d"),
        "updated_at": datetime.now()
    })
    
    result = await tecnicos_aplicadores_collection.update_one(
        {"_id": ObjectId(tecnico_id)},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Técnico aplicador no encontrado")
    
    updated = await tecnicos_aplicadores_collection.find_one({"_id": ObjectId(tecnico_id)})
    return {"success": True, "data": serialize_doc(updated)}


@router.post("/tecnicos-aplicadores/{tecnico_id}/certificado")
async def upload_certificado(
    tecnico_id: str,
    file: UploadFile = File(...),
    current_user: dict = Depends(RequireEdit)
):
    """Subir imagen del certificado"""
    if not ObjectId.is_valid(tecnico_id):
        raise HTTPException(status_code=400, detail="ID inválido")
    
    tecnico = await tecnicos_aplicadores_collection.find_one({"_id": ObjectId(tecnico_id)})
    if not tecnico:
        raise HTTPException(status_code=404, detail="Técnico aplicador no encontrado")
    
    # Validar tipo de archivo
    allowed_types = ["image/jpeg", "image/png", "image/webp", "application/pdf"]
    if file.content_type not in allowed_types:
        raise HTTPException(
            status_code=400, 
            detail=f"Tipo de archivo no permitido. Permitidos: {allowed_types}"
        )
    
    # Crear directorio si no existe - usando directorio persistente
    upload_dir = "/app/uploads/certificados"
    os.makedirs(upload_dir, exist_ok=True)
    
    # Guardar archivo
    file_ext = file.filename.split(".")[-1] if "." in file.filename else "jpg"
    file_id = str(uuid.uuid4())
    filename = f"{file_id}.{file_ext}"
    file_path = f"{upload_dir}/{filename}"
    
    with open(file_path, "wb") as f:
        content = await file.read()
        f.write(content)
    
    # Guardar URL relativa para acceso web
    web_url = f"/api/uploads/certificados/{filename}"
    
    # Actualizar URL en la base de datos
    await tecnicos_aplicadores_collection.update_one(
        {"_id": ObjectId(tecnico_id)},
        {"$set": {
            "imagen_certificado_url": web_url,
            "imagen_certificado_path": file_path,
            "imagen_certificado_nombre": file.filename,
            "updated_at": datetime.now()
        }}
    )
    
    updated = await tecnicos_aplicadores_collection.find_one({"_id": ObjectId(tecnico_id)})
    return {"success": True, "data": serialize_doc(updated)}


@router.delete("/tecnicos-aplicadores/{tecnico_id}/certificado")
async def delete_certificado(
    tecnico_id: str,
    current_user: dict = Depends(RequireEdit)
):
    """Eliminar imagen del certificado"""
    if not ObjectId.is_valid(tecnico_id):
        raise HTTPException(status_code=400, detail="ID inválido")
    
    tecnico = await tecnicos_aplicadores_collection.find_one({"_id": ObjectId(tecnico_id)})
    if not tecnico:
        raise HTTPException(status_code=404, detail="Técnico aplicador no encontrado")
    
    # Eliminar archivo si existe
    file_path = tecnico.get("imagen_certificado_url")
    if file_path and os.path.exists(file_path):
        os.remove(file_path)
    
    # Actualizar base de datos
    await tecnicos_aplicadores_collection.update_one(
        {"_id": ObjectId(tecnico_id)},
        {"$set": {
            "imagen_certificado_url": None,
            "imagen_certificado_nombre": None,
            "updated_at": datetime.now()
        }}
    )
    
    return {"success": True, "message": "Certificado eliminado"}


@router.put("/tecnicos-aplicadores/{tecnico_id}/toggle-activo")
async def toggle_activo(
    tecnico_id: str,
    current_user: dict = Depends(RequireEdit)
):
    """Activar/desactivar un técnico aplicador"""
    if not ObjectId.is_valid(tecnico_id):
        raise HTTPException(status_code=400, detail="ID inválido")
    
    tecnico = await tecnicos_aplicadores_collection.find_one({"_id": ObjectId(tecnico_id)})
    if not tecnico:
        raise HTTPException(status_code=404, detail="Técnico aplicador no encontrado")
    
    new_status = not tecnico.get("activo", True)
    
    await tecnicos_aplicadores_collection.update_one(
        {"_id": ObjectId(tecnico_id)},
        {"$set": {"activo": new_status, "updated_at": datetime.now()}}
    )
    
    updated = await tecnicos_aplicadores_collection.find_one({"_id": ObjectId(tecnico_id)})
    return {"success": True, "data": serialize_doc(updated)}


@router.delete("/tecnicos-aplicadores/{tecnico_id}")
async def delete_tecnico_aplicador(
    tecnico_id: str,
    current_user: dict = Depends(RequireDelete)
):
    """Eliminar un técnico aplicador"""
    if not ObjectId.is_valid(tecnico_id):
        raise HTTPException(status_code=400, detail="ID inválido")
    
    # Eliminar archivo de certificado si existe
    tecnico = await tecnicos_aplicadores_collection.find_one({"_id": ObjectId(tecnico_id)})
    if tecnico:
        file_path = tecnico.get("imagen_certificado_url")
        if file_path and os.path.exists(file_path):
            os.remove(file_path)
    
    result = await tecnicos_aplicadores_collection.delete_one({"_id": ObjectId(tecnico_id)})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Técnico aplicador no encontrado")
    
    return {"success": True, "message": "Técnico aplicador eliminado"}



@router.get("/tecnicos-aplicadores/export/excel")
async def export_tecnicos_excel(
    current_user: dict = Depends(get_current_user)
):
    """Export tecnicos aplicadores to Excel"""
    from fastapi.responses import StreamingResponse
    import io as _io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    tecnicos_list = await tecnicos_aplicadores_collection.find({}).sort("nombre", 1).to_list(5000)
    tecnicos_raw = serialize_docs(tecnicos_list)

    wb = Workbook()
    ws = wb.active
    ws.title = "Tecnicos Aplicadores"

    headers = ["Nombre", "Apellidos", "DNI/NIE", "Nivel", "N Carnet", "Caducidad Carnet", "Telefono", "Email", "Activo"]
    header_fill = PatternFill(start_color="E65100", end_color="E65100", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for idx, t in enumerate(tecnicos_raw, 2):
        row = [
            t.get("nombre", ""), t.get("apellidos", ""), t.get("dni_nie", ""),
            t.get("nivel_capacitacion", ""), t.get("numero_carnet", ""),
            t.get("fecha_caducidad_carnet", "")[:10] if t.get("fecha_caducidad_carnet") else "",
            t.get("telefono", ""), t.get("email", ""),
            "Si" if t.get("activo") else "No"
        ]
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=idx, column=col, value=val)
            cell.border = thin_border

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 20

    output = _io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"tecnicos_aplicadores_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename={filename}"})


@router.get("/tecnicos-aplicadores/export/pdf")
async def export_tecnicos_pdf(
    current_user: dict = Depends(get_current_user)
):
    """Export tecnicos aplicadores to PDF"""
    from fastapi.responses import StreamingResponse
    import io as _io
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    tecnicos_list = await tecnicos_aplicadores_collection.find({}).sort("nombre", 1).to_list(5000)
    tecnicos_raw = serialize_docs(tecnicos_list)

    output = _io.BytesIO()
    pdf = SimpleDocTemplate(output, pagesize=landscape(A4), topMargin=15*mm, bottomMargin=15*mm)
    elements = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=16, textColor=colors.HexColor('#E65100'), alignment=1, spaceAfter=8*mm)
    elements.append(Paragraph("Tecnicos Aplicadores - FRUVECO", title_style))
    elements.append(Paragraph(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", ParagraphStyle('Sub', parent=styles['Normal'], alignment=1, textColor=colors.gray)))
    elements.append(Spacer(1, 8*mm))

    table_data = [["Nombre", "DNI/NIE", "Nivel", "N Carnet", "Caducidad", "Telefono", "Activo"]]
    for t in tecnicos_raw:
        table_data.append([
            f"{t.get('nombre', '')} {t.get('apellidos', '')}".strip()[:30],
            t.get("dni_nie", "")[:12], t.get("nivel_capacitacion", "")[:15],
            t.get("numero_carnet", "")[:15],
            t.get("fecha_caducidad_carnet", "")[:10] if t.get("fecha_caducidad_carnet") else "",
            t.get("telefono", "")[:12],
            "Si" if t.get("activo") else "No"
        ])

    col_widths = [50*mm, 25*mm, 30*mm, 30*mm, 25*mm, 25*mm, 15*mm]
    doc_table = Table(table_data, colWidths=col_widths)
    doc_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E65100')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.gray),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#FFF3E0')]),
    ]))
    elements.append(doc_table)

    pdf.build(elements)
    output.seek(0)
    filename = f"tecnicos_aplicadores_{datetime.now().strftime('%Y%m%d')}.pdf"
    return StreamingResponse(output, media_type="application/pdf",
                             headers={"Content-Disposition": f"attachment; filename={filename}"})
