from fastapi import APIRouter, HTTPException, Depends, UploadFile, File
from typing import Optional
from bson import ObjectId
from datetime import datetime
import os
import uuid

from models_tratamientos import MaquinariaCreate, MaquinariaInDB
from database import maquinaria_collection, serialize_doc, serialize_docs
from rbac_guards import RequireCreate, RequireEdit, RequireDelete, get_current_user

router = APIRouter(prefix="/api", tags=["maquinaria"])

# ============================================================================
# MAQUINARIA CRUD
# ============================================================================

@router.post("/maquinaria", response_model=dict)
async def create_maquinaria(
    maquinaria: MaquinariaCreate,
    current_user: dict = Depends(RequireCreate)
):
    """Crear nueva maquinaria en el catálogo"""
    maquinaria_dict = maquinaria.dict()
    maquinaria_dict["created_at"] = datetime.now()
    maquinaria_dict["updated_at"] = datetime.now()
    
    result = await maquinaria_collection.insert_one(maquinaria_dict)
    created = await maquinaria_collection.find_one({"_id": result.inserted_id})
    
    return {"success": True, "data": serialize_doc(created)}


@router.get("/maquinaria")
async def get_maquinaria_list(
    skip: int = 0,
    limit: int = 100,
    tipo: Optional[str] = None,
    estado: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Listar toda la maquinaria con filtros opcionales"""
    query = {}
    if tipo:
        query["tipo"] = tipo
    if estado:
        query["estado"] = estado
    
    maquinaria = await maquinaria_collection.find(query).skip(skip).limit(limit).to_list(limit)
    total = await maquinaria_collection.count_documents(query)
    
    return {"maquinaria": serialize_docs(maquinaria), "total": total}


@router.get("/maquinaria/{maquinaria_id}")
async def get_maquinaria(
    maquinaria_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Obtener una maquinaria por ID"""
    if not ObjectId.is_valid(maquinaria_id):
        raise HTTPException(status_code=400, detail="ID inválido")
    
    maquinaria = await maquinaria_collection.find_one({"_id": ObjectId(maquinaria_id)})
    if not maquinaria:
        raise HTTPException(status_code=404, detail="Maquinaria no encontrada")
    
    return serialize_doc(maquinaria)


@router.put("/maquinaria/{maquinaria_id}")
async def update_maquinaria(
    maquinaria_id: str,
    maquinaria: MaquinariaCreate,
    current_user: dict = Depends(RequireEdit)
):
    """Actualizar una maquinaria existente"""
    if not ObjectId.is_valid(maquinaria_id):
        raise HTTPException(status_code=400, detail="ID inválido")
    
    update_data = maquinaria.dict()
    update_data["updated_at"] = datetime.now()
    
    result = await maquinaria_collection.update_one(
        {"_id": ObjectId(maquinaria_id)},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Maquinaria no encontrada")
    
    updated = await maquinaria_collection.find_one({"_id": ObjectId(maquinaria_id)})
    return {"success": True, "data": serialize_doc(updated)}


@router.delete("/maquinaria/{maquinaria_id}")
async def delete_maquinaria(
    maquinaria_id: str,
    current_user: dict = Depends(RequireDelete)
):
    """Eliminar una maquinaria del catálogo"""
    if not ObjectId.is_valid(maquinaria_id):
        raise HTTPException(status_code=400, detail="ID inválido")
    
    # Eliminar imagen si existe
    maquinaria = await maquinaria_collection.find_one({"_id": ObjectId(maquinaria_id)})
    if maquinaria:
        file_path = maquinaria.get("imagen_placa_ce_url")
        if file_path and os.path.exists(file_path):
            os.remove(file_path)
    
    result = await maquinaria_collection.delete_one({"_id": ObjectId(maquinaria_id)})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Maquinaria no encontrada")
    
    return {"success": True, "message": "Maquinaria eliminada correctamente"}


# ============================================================================
# IMAGEN PLACA CE
# ============================================================================

@router.post("/maquinaria/{maquinaria_id}/imagen-placa-ce")
async def upload_imagen_placa_ce(
    maquinaria_id: str,
    file: UploadFile = File(...),
    current_user: dict = Depends(RequireEdit)
):
    """Subir imagen de la placa CE de la maquinaria"""
    if not ObjectId.is_valid(maquinaria_id):
        raise HTTPException(status_code=400, detail="ID inválido")
    
    maquinaria = await maquinaria_collection.find_one({"_id": ObjectId(maquinaria_id)})
    if not maquinaria:
        raise HTTPException(status_code=404, detail="Maquinaria no encontrada")
    
    # Validar tipo de archivo
    allowed_types = ["image/jpeg", "image/png", "image/webp"]
    if file.content_type not in allowed_types:
        raise HTTPException(
            status_code=400, 
            detail="Tipo de archivo no permitido. Permitidos: JPEG, PNG, WEBP"
        )
    
    # Validar tamaño (max 10MB)
    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="El archivo excede el tamaño máximo de 10MB")
    
    # Crear directorio si no existe - usando directorio persistente
    upload_dir = "/app/uploads/maquinaria_placas"
    os.makedirs(upload_dir, exist_ok=True)
    
    # Eliminar imagen anterior si existe
    old_file_path = maquinaria.get("imagen_placa_ce_path")
    if old_file_path and os.path.exists(old_file_path):
        os.remove(old_file_path)
    
    # Guardar archivo
    file_ext = file.filename.split(".")[-1] if "." in file.filename else "jpg"
    file_id = str(uuid.uuid4())
    filename = f"{file_id}.{file_ext}"
    file_path = f"{upload_dir}/{filename}"
    
    with open(file_path, "wb") as f:
        f.write(content)
    
    # Guardar URL relativa para acceso web
    web_url = f"/api/uploads/maquinaria_placas/{filename}"
    
    # Actualizar URL en la base de datos
    await maquinaria_collection.update_one(
        {"_id": ObjectId(maquinaria_id)},
        {"$set": {
            "imagen_placa_ce_url": web_url,
            "imagen_placa_ce_path": file_path,
            "imagen_placa_ce_nombre": file.filename,
            "updated_at": datetime.now()
        }}
    )
    
    updated = await maquinaria_collection.find_one({"_id": ObjectId(maquinaria_id)})
    return {"success": True, "data": serialize_doc(updated)}


@router.delete("/maquinaria/{maquinaria_id}/imagen-placa-ce")
async def delete_imagen_placa_ce(
    maquinaria_id: str,
    current_user: dict = Depends(RequireEdit)
):
    """Eliminar imagen de la placa CE"""
    if not ObjectId.is_valid(maquinaria_id):
        raise HTTPException(status_code=400, detail="ID inválido")
    
    maquinaria = await maquinaria_collection.find_one({"_id": ObjectId(maquinaria_id)})
    if not maquinaria:
        raise HTTPException(status_code=404, detail="Maquinaria no encontrada")
    
    # Eliminar archivo si existe
    file_path = maquinaria.get("imagen_placa_ce_path")
    if file_path and os.path.exists(file_path):
        os.remove(file_path)
    
    # Actualizar base de datos
    await maquinaria_collection.update_one(
        {"_id": ObjectId(maquinaria_id)},
        {"$set": {
            "imagen_placa_ce_url": None,
            "imagen_placa_ce_path": None,
            "imagen_placa_ce_nombre": None,
            "updated_at": datetime.now()
        }}
    )
    
    return {"success": True, "message": "Imagen de placa CE eliminada"}


@router.get("/maquinaria/{maquinaria_id}/imagen-placa-ce")
async def get_imagen_placa_ce(
    maquinaria_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Obtener la imagen de la placa CE"""
    from fastapi.responses import FileResponse
    
    if not ObjectId.is_valid(maquinaria_id):
        raise HTTPException(status_code=400, detail="ID inválido")
    
    maquinaria = await maquinaria_collection.find_one({"_id": ObjectId(maquinaria_id)})
    if not maquinaria:
        raise HTTPException(status_code=404, detail="Maquinaria no encontrada")
    
    # Intentar obtener el path directamente
    file_path = maquinaria.get("imagen_placa_ce_path")
    
    # Si no existe o el archivo no existe, intentar derivar desde la URL
    if not file_path or not os.path.exists(file_path):
        url = maquinaria.get("imagen_placa_ce_url")
        if url:
            # Si la URL es relativa a /api/uploads/, derivar el path
            if url.startswith("/api/uploads/"):
                file_path = "/app/uploads/" + url.replace("/api/uploads/", "")
            # Si la URL es una ruta absoluta del sistema, usarla directamente
            elif url.startswith("/app/uploads/"):
                file_path = url
    
    if not file_path or not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="No hay imagen de placa CE")
    
    return FileResponse(file_path)


# ============================================================================
# ESTADÍSTICAS Y EXPORTACIÓN
# ============================================================================

@router.get("/maquinaria/stats/resumen")
async def get_maquinaria_stats(current_user: dict = Depends(get_current_user)):
    """Obtener estadísticas de maquinaria"""
    from database import db
    
    total = await maquinaria_collection.count_documents({})
    activa = await maquinaria_collection.count_documents({"estado": "activa"})
    en_mantenimiento = await maquinaria_collection.count_documents({"estado": "en_mantenimiento"})
    baja = await maquinaria_collection.count_documents({"estado": "baja"})
    
    # Por tipo
    pipeline = [
        {"$group": {"_id": "$tipo", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}}
    ]
    por_tipo = []
    async for doc in maquinaria_collection.aggregate(pipeline):
        por_tipo.append({"tipo": doc["_id"] or "Sin especificar", "count": doc["count"]})
    
    # Próximas revisiones (ITV, mantenimiento)
    hoy = datetime.now().strftime("%Y-%m-%d")
    proximas_itv = await maquinaria_collection.find({
        "fecha_proxima_itv": {"$exists": True, "$ne": None, "$lte": hoy}
    }).to_list(10)
    
    # Uso en tratamientos
    tratamientos_collection = db.get_collection('tratamientos')
    uso_maquinaria = []
    try:
        pipeline_uso = [
            {"$match": {"maquinaria_id": {"$exists": True, "$ne": None}}},
            {"$group": {"_id": "$maquinaria_id", "count": {"$sum": 1}}},
            {"$sort": {"count": -1}},
            {"$limit": 10}
        ]
        async for doc in tratamientos_collection.aggregate(pipeline_uso):
            maq = await maquinaria_collection.find_one({"_id": ObjectId(doc["_id"])}) if ObjectId.is_valid(doc["_id"]) else None
            uso_maquinaria.append({
                "maquinaria_id": doc["_id"],
                "nombre": maq.get("nombre") if maq else "Desconocida",
                "usos": doc["count"]
            })
    except Exception:
        pass
    
    return {
        "success": True,
        "stats": {
            "total": total,
            "activa": activa,
            "en_mantenimiento": en_mantenimiento,
            "baja": baja,
            "por_tipo": por_tipo,
            "con_itv_vencida": len(proximas_itv),
            "top_uso": uso_maquinaria
        }
    }


@router.get("/maquinaria/{maquinaria_id}/historial")
async def get_maquinaria_historial(
    maquinaria_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Obtener historial de uso de una maquinaria"""
    from database import db
    
    if not ObjectId.is_valid(maquinaria_id):
        raise HTTPException(status_code=400, detail="ID inválido")
    
    maquinaria = await maquinaria_collection.find_one({"_id": ObjectId(maquinaria_id)})
    if not maquinaria:
        raise HTTPException(status_code=404, detail="Maquinaria no encontrada")
    
    # Tratamientos donde se usó
    tratamientos_collection = db.get_collection('tratamientos')
    tratamientos = []
    try:
        cursor = tratamientos_collection.find({"maquinaria_id": maquinaria_id}).sort("fecha", -1).limit(50)
        async for t in cursor:
            tratamientos.append(serialize_doc(t))
    except Exception:
        pass
    
    # Mantenimientos registrados (si existe colección)
    mantenimientos = []
    try:
        mantenimientos_collection = db.get_collection('mantenimientos_maquinaria')
        cursor = mantenimientos_collection.find({"maquinaria_id": maquinaria_id}).sort("fecha", -1).limit(20)
        async for m in cursor:
            mantenimientos.append(serialize_doc(m))
    except Exception:
        pass
    
    return {
        "success": True,
        "maquinaria": serialize_doc(maquinaria),
        "historial": {
            "tratamientos": tratamientos,
            "mantenimientos": mantenimientos,
            "total_usos": len(tratamientos)
        }
    }


@router.get("/maquinaria/export/excel")
async def export_maquinaria_excel(
    estado: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Exportar maquinaria a Excel"""
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io
    
    query = {}
    if estado:
        query['estado'] = estado
    
    maquinaria = await maquinaria_collection.find(query).sort("nombre", 1).to_list(1000)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Maquinaria"
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF5722", end_color="FF5722", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    headers = ["Nombre", "Tipo", "Marca", "Modelo", "Matrícula", "Nº Serie", "Año", "Potencia", "Capacidad", "Estado", "ITV", "Seguro"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    
    for row_idx, maq in enumerate(maquinaria, 2):
        data = [
            maq.get("nombre", ""),
            maq.get("tipo", ""),
            maq.get("marca", ""),
            maq.get("modelo", ""),
            maq.get("matricula", ""),
            maq.get("numero_serie", ""),
            maq.get("ano_fabricacion", ""),
            maq.get("potencia_cv", ""),
            maq.get("capacidad", ""),
            maq.get("estado", "activa").replace("_", " ").title(),
            maq.get("fecha_proxima_itv", "-"),
            maq.get("fecha_vencimiento_seguro", "-")
        ]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
    
    column_widths = [25, 15, 15, 15, 12, 18, 8, 10, 12, 15, 12, 12]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=maquinaria_{datetime.now().strftime('%Y%m%d')}.xlsx"}
    )
