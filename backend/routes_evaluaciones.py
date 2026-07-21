from fastapi import APIRouter, HTTPException, Depends, UploadFile, File
from typing import Optional, List, Dict, Any
from bson import ObjectId
from datetime import datetime
from pydantic import BaseModel, Field
import os
import uuid
import shutil

from database import db, serialize_doc, serialize_docs, parcelas_collection
from rbac_guards import RequireCreate, RequireEdit, RequireDelete, get_current_user
from models_evaluaciones import (
    SeccionRespuesta, EvaluacionCreate, PreguntaConfig, PREGUNTAS_DEFAULT,
)

router = APIRouter(prefix="/api", tags=["evaluaciones"])

# Collection
evaluaciones_collection = db['evaluaciones']
evaluaciones_config_collection = db['evaluaciones_config']

# ---------------------------------------------------------------------------
# LOGO — se embebe como data-URI base64 en el PDF para que aparezca en la
# cabecera superior-izquierda de todas las páginas (WeasyPrint + position:fixed).
# Cacheado a nivel de módulo para no releer el disco en cada request.
# ---------------------------------------------------------------------------
_FRUVECO_LOGO_DATA_URI: Optional[str] = None

def _get_fruveco_logo_data_uri() -> str:
    global _FRUVECO_LOGO_DATA_URI
    if _FRUVECO_LOGO_DATA_URI is None:
        import base64
        logo_path = os.path.join(os.path.dirname(__file__), 'static', 'fruveco_logo.png')
        try:
            with open(logo_path, 'rb') as f:
                encoded = base64.b64encode(f.read()).decode('ascii')
            _FRUVECO_LOGO_DATA_URI = f"data:image/png;base64,{encoded}"
        except FileNotFoundError:
            _FRUVECO_LOGO_DATA_URI = ""  # PDF sigue generándose sin logo
    return _FRUVECO_LOGO_DATA_URI

# ============================================================================
# MODELS Y PREGUNTAS POR DEFECTO -> extraidos a models_evaluaciones.py
# Los endpoints de configuracion -> extraidos a routes_evaluaciones_config.py
# ============================================================================

# ============================================================================
# CRUD EVALUACIONES
# ============================================================================

# ----------------------------------------------------------------------------
# Anexos (Impresos) — File uploads attached to an evaluación's impresos.
# Se guardan en la colección MongoDB `evaluaciones_anexos` (bytes inline)
# porque `/app/uploads/` es EFÍMERO en el pod de producción y se borra en
# cada redeploy. MongoDB es persistente entre deploys.
# El endpoint GET streamea los bytes con el content-type correcto.
# Backward-compat: si el URL antiguo `/api/uploads/evaluaciones/anexos/...`
# se llama y el archivo todavía existe en disco, se sirve vía StaticFiles.
# ----------------------------------------------------------------------------

anexos_collection = db['evaluaciones_anexos']

ANEXOS_DIR = "/app/uploads/evaluaciones/anexos"  # kept only para backward-compat
os.makedirs(ANEXOS_DIR, exist_ok=True)

ALLOWED_ANEXO_TYPES = {
    "application/pdf",
    "image/jpeg", "image/jpg", "image/png", "image/webp",
    "application/msword",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    "application/vnd.ms-excel",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
}
MAX_ANEXO_SIZE_BYTES = 15 * 1024 * 1024  # 15 MB


@router.post("/evaluaciones/anexos/upload")
async def upload_evaluacion_anexo(
    file: UploadFile = File(...),
    current_user: dict = Depends(RequireCreate),
):
    """Subir un anexo. Guarda los bytes en MongoDB (colección
    `evaluaciones_anexos`) para que persistan entre redeploys. Devuelve
    metadatos con `url` que apunta al endpoint GET /api/evaluaciones/anexos/{id}.
    """
    if file.content_type not in ALLOWED_ANEXO_TYPES:
        raise HTTPException(
            status_code=400,
            detail=f"Tipo de archivo no permitido ({file.content_type}). Use PDF, imagen o documento Office.",
        )

    contents = await file.read(MAX_ANEXO_SIZE_BYTES + 1)
    size = len(contents)
    if size > MAX_ANEXO_SIZE_BYTES:
        raise HTTPException(status_code=413, detail="El archivo supera el tamaño máximo permitido (15 MB).")

    original = file.filename or "anexo"
    doc = {
        "filename": original,
        "content_type": file.content_type,
        "size": size,
        "data": contents,   # bytes → BSON Binary
        "uploaded_at": datetime.now(),
        "uploaded_by": current_user.get("username") or current_user.get("email") or "",
    }
    result = await anexos_collection.insert_one(doc)
    anexo_id = str(result.inserted_id)

    return {
        "success": True,
        "data": {
            "filename": original,
            "stored_name": anexo_id,   # ahora es el ObjectId
            "url": f"/api/evaluaciones/anexos/{anexo_id}",
            "size": size,
            "content_type": file.content_type,
            "uploaded_at": datetime.now().isoformat(),
            "uploaded_by": current_user.get("username") or current_user.get("email") or "",
        },
    }


@router.get("/evaluaciones/anexos/{anexo_id}")
async def download_evaluacion_anexo(
    anexo_id: str,
    # No requerir login: los PDFs se abren desde <iframe>/<a> y el navegador
    # no envía Authorization header. La URL contiene un ObjectId aleatorio
    # (24 hex chars) que actúa como token de acceso implícito.
):
    """Descargar/visualizar un anexo. Sirve los bytes con Content-Disposition
    inline para que los PDFs y las imágenes se rendericen en el navegador
    en lugar de forzar descarga."""
    from fastapi.responses import Response
    if not ObjectId.is_valid(anexo_id):
        raise HTTPException(status_code=404, detail="Anexo no encontrado")
    doc = await anexos_collection.find_one({"_id": ObjectId(anexo_id)})
    if not doc:
        raise HTTPException(status_code=404, detail="Anexo no encontrado")
    return Response(
        content=doc.get("data", b""),
        media_type=doc.get("content_type") or "application/octet-stream",
        headers={
            "Content-Disposition": f'inline; filename="{doc.get("filename", "anexo")}"',
            "Cache-Control": "private, max-age=3600",
        },
    )


@router.delete("/evaluaciones/anexos/{stored_name}")
async def delete_evaluacion_anexo(
    stored_name: str,
    current_user: dict = Depends(RequireDelete),
):
    """Eliminar un anexo subido previamente.

    Acepta:
    - Un ObjectId de MongoDB (anexos nuevos guardados en `evaluaciones_anexos`).
    - Un nombre-de-fichero legacy con formato `<uuid>__<filename>` (anexos
      antiguos guardados en disco antes de migrar a Mongo).
    """
    # Defence: prevent path traversal en el legacy
    if "/" in stored_name or ".." in stored_name:
        raise HTTPException(status_code=400, detail="Nombre de archivo inválido")

    # Nuevo formato: ObjectId de Mongo
    if ObjectId.is_valid(stored_name):
        result = await anexos_collection.delete_one({"_id": ObjectId(stored_name)})
        if result.deleted_count > 0:
            return {"success": True}
        # Si no está en Mongo, seguimos probando el disco (podría ser legacy).

    # Legacy: fichero en disco
    path = os.path.join(ANEXOS_DIR, stored_name)
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail="Anexo no encontrado")
    try:
        os.remove(path)
    except OSError as e:
        raise HTTPException(status_code=500, detail=f"No se pudo eliminar el anexo: {e}")
    return {"success": True}




@router.post("/evaluaciones", response_model=dict)
async def create_evaluacion(
    evaluacion: EvaluacionCreate,
    current_user: dict = Depends(RequireCreate)
):
    """Crear nueva hoja de evaluación"""
    if not evaluacion.parcela_id:
        raise HTTPException(status_code=400, detail="parcela_id es obligatorio")
    
    if not ObjectId.is_valid(evaluacion.parcela_id):
        raise HTTPException(status_code=400, detail="parcela_id inválido")
    
    # Obtener datos de la parcela
    parcela = await parcelas_collection.find_one({"_id": ObjectId(evaluacion.parcela_id)})
    if not parcela:
        raise HTTPException(status_code=400, detail="Parcela no encontrada")
    
    # Construir documento con datos heredados de parcela
    evaluacion_dict = {
        "parcela_id": evaluacion.parcela_id,
        "fecha_inicio": evaluacion.fecha_inicio or datetime.now().strftime("%Y-%m-%d"),
        "fecha_fin": evaluacion.fecha_fin,
        "tecnico": evaluacion.tecnico or current_user.get("full_name", current_user.get("username", "")),
        
        # Datos heredados de parcela
        "proveedor": parcela.get("proveedor", ""),
        "codigo_plantacion": parcela.get("codigo_plantacion", ""),
        "finca": parcela.get("finca", ""),
        "cultivo": parcela.get("cultivo", ""),
        "variedad": parcela.get("variedad", ""),
        "superficie": parcela.get("superficie_total", 0),
        "campana": parcela.get("campana", ""),
        "contrato_id": parcela.get("contrato_id", ""),
        
        # Secciones de cuestionarios
        "toma_datos": evaluacion.toma_datos or [],
        "impresos": evaluacion.impresos or {
            "fecha_inicio": evaluacion.fecha_inicio,
            "fecha_fin": evaluacion.fecha_fin,
            "tecnico": evaluacion.tecnico
        },
        "analisis_suelo": evaluacion.analisis_suelo or [],
        "pasos_precampana": evaluacion.pasos_precampana or [],
        "calidad_cepellones": evaluacion.calidad_cepellones or [],
        "inspeccion_maquinaria": evaluacion.inspeccion_maquinaria or [],
        "observaciones": evaluacion.observaciones or [],
        "calibracion_mantenimiento": evaluacion.calibracion_mantenimiento or [],
        # Mapa completo de respuestas — espejo de los campos top-level.
        # Persistirlo es lo que permite a handleEdit del frontend releer las
        # respuestas con un único Object.values(secciones).
        "secciones": evaluacion.secciones if evaluacion.secciones else {
            "toma_datos": evaluacion.toma_datos or [],
            "analisis_suelo": evaluacion.analisis_suelo or [],
            "pasos_precampana": evaluacion.pasos_precampana or [],
            "calidad_cepellones": evaluacion.calidad_cepellones or [],
            "inspeccion_maquinaria": evaluacion.inspeccion_maquinaria or [],
            "observaciones": evaluacion.observaciones or [],
            "calibracion_mantenimiento": evaluacion.calibracion_mantenimiento or [],
        },

        # Metadatos
        "estado": "borrador",  # borrador, completada, archivada
        "created_by": str(current_user.get("_id", "")),
        "created_at": datetime.now(),
        "updated_at": datetime.now()
    }
    
    result = await evaluaciones_collection.insert_one(evaluacion_dict)
    created = await evaluaciones_collection.find_one({"_id": result.inserted_id})
    
    return {"success": True, "data": serialize_doc(created)}


@router.get("/evaluaciones")
async def get_evaluaciones(
    skip: int = 0,
    limit: int = 100,
    parcela_id: Optional[str] = None,
    campana: Optional[str] = None,
    estado: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Listar hojas de evaluación"""
    query = {}
    if parcela_id:
        query["parcela_id"] = parcela_id
    if campana:
        query["campana"] = campana
    if estado:
        query["estado"] = estado
    
    evaluaciones = await evaluaciones_collection.find(query).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    total = await evaluaciones_collection.count_documents(query)
    
    return {"evaluaciones": serialize_docs(evaluaciones), "total": total}


@router.get("/evaluaciones/{evaluacion_id}")
async def get_evaluacion(
    evaluacion_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Obtener una hoja de evaluación por ID"""
    if not ObjectId.is_valid(evaluacion_id):
        raise HTTPException(status_code=400, detail="ID inválido")
    
    evaluacion = await evaluaciones_collection.find_one({"_id": ObjectId(evaluacion_id)})
    if not evaluacion:
        raise HTTPException(status_code=404, detail="Evaluación no encontrada")
    
    return serialize_doc(evaluacion)


@router.put("/evaluaciones/{evaluacion_id}")
async def update_evaluacion(
    evaluacion_id: str,
    evaluacion: EvaluacionCreate,
    current_user: dict = Depends(RequireEdit)
):
    """Actualizar hoja de evaluación"""
    if not ObjectId.is_valid(evaluacion_id):
        raise HTTPException(status_code=400, detail="ID inválido")
    
    update_data = {
        "fecha_inicio": evaluacion.fecha_inicio,
        "fecha_fin": evaluacion.fecha_fin,
        "tecnico": evaluacion.tecnico,
        "toma_datos": evaluacion.toma_datos or [],
        "impresos": evaluacion.impresos or {},
        "analisis_suelo": evaluacion.analisis_suelo or [],
        "pasos_precampana": evaluacion.pasos_precampana or [],
        "calidad_cepellones": evaluacion.calidad_cepellones or [],
        "inspeccion_maquinaria": evaluacion.inspeccion_maquinaria or [],
        "observaciones": evaluacion.observaciones or [],
        "calibracion_mantenimiento": evaluacion.calibracion_mantenimiento or [],
        # Mapa completo {seccion: [respuestas]} — usado por el frontend en handleEdit
        # para releer todas las respuestas. Si el cliente envía un objeto vacío,
        # reconstruimos el mapa desde los campos top-level para no perder datos.
        "secciones": evaluacion.secciones if evaluacion.secciones else {
            "toma_datos": evaluacion.toma_datos or [],
            "analisis_suelo": evaluacion.analisis_suelo or [],
            "pasos_precampana": evaluacion.pasos_precampana or [],
            "calidad_cepellones": evaluacion.calidad_cepellones or [],
            "inspeccion_maquinaria": evaluacion.inspeccion_maquinaria or [],
            "observaciones": evaluacion.observaciones or [],
            "calibracion_mantenimiento": evaluacion.calibracion_mantenimiento or [],
        },
        "updated_at": datetime.now()
    }
    
    result = await evaluaciones_collection.update_one(
        {"_id": ObjectId(evaluacion_id)},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Evaluación no encontrada")
    
    updated = await evaluaciones_collection.find_one({"_id": ObjectId(evaluacion_id)})
    return {"success": True, "data": serialize_doc(updated)}


@router.patch("/evaluaciones/{evaluacion_id}/estado")
async def update_evaluacion_estado(
    evaluacion_id: str,
    estado: str,
    current_user: dict = Depends(RequireEdit)
):
    """Cambiar estado de la evaluación"""
    if not ObjectId.is_valid(evaluacion_id):
        raise HTTPException(status_code=400, detail="ID inválido")
    
    if estado not in ["borrador", "completada", "archivada"]:
        raise HTTPException(status_code=400, detail="Estado inválido")
    
    result = await evaluaciones_collection.update_one(
        {"_id": ObjectId(evaluacion_id)},
        {"$set": {"estado": estado, "updated_at": datetime.now()}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Evaluación no encontrada")
    
    return {"success": True, "message": f"Estado actualizado a {estado}"}


@router.delete("/evaluaciones/{evaluacion_id}")
async def delete_evaluacion(
    evaluacion_id: str,
    current_user: dict = Depends(RequireDelete)
):
    """Eliminar hoja de evaluación"""
    if not ObjectId.is_valid(evaluacion_id):
        raise HTTPException(status_code=400, detail="ID inválido")
    
    result = await evaluaciones_collection.delete_one({"_id": ObjectId(evaluacion_id)})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Evaluación no encontrada")
    
    return {"success": True, "message": "Evaluación eliminada"}



@router.get("/evaluaciones/export/pdf")
async def export_evaluaciones_pdf(
    current_user: dict = Depends(get_current_user)
):
    """Export evaluaciones to PDF"""
    from fastapi.responses import StreamingResponse
    import io as _io
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    evaluaciones_list = await evaluaciones_collection.find({}).sort("created_at", -1).to_list(5000)
    evaluaciones_raw = serialize_docs(evaluaciones_list)

    output = _io.BytesIO()
    pdf = SimpleDocTemplate(output, pagesize=landscape(A4), topMargin=15*mm, bottomMargin=15*mm)
    elements = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=16, textColor=colors.HexColor('#4A148C'), alignment=1, spaceAfter=8*mm)
    elements.append(Paragraph("Hojas de Evaluacion - FRUVECO", title_style))
    elements.append(Paragraph(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", ParagraphStyle('Sub', parent=styles['Normal'], alignment=1, textColor=colors.gray)))
    elements.append(Spacer(1, 8*mm))

    table_data = [["Parcela", "Cultivo", "Proveedor", "Campana", "Estado", "Tecnico", "Fecha"]]
    for e in evaluaciones_raw:
        table_data.append([
            e.get("codigo_plantacion", "")[:20],
            e.get("cultivo", "")[:15],
            e.get("proveedor", "")[:20],
            e.get("campana", "")[:10],
            e.get("estado", "")[:12],
            e.get("tecnico", "")[:20],
            e.get("created_at", "")[:10] if e.get("created_at") else "",
        ])

    col_widths = [40*mm, 30*mm, 40*mm, 25*mm, 25*mm, 40*mm, 25*mm]
    doc_table = Table(table_data, colWidths=col_widths)
    doc_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4A148C')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.gray),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F3E5F5')]),
    ]))
    elements.append(doc_table)

    pdf.build(elements)
    output.seek(0)
    filename = f"evaluaciones_{datetime.now().strftime('%Y%m%d')}.pdf"
    return StreamingResponse(output, media_type="application/pdf",
                             headers={"Content-Disposition": f"attachment; filename={filename}"})


# ============================================================================
# GENERACIÓN DE PDF - CON VISITAS Y TRATAMIENTOS
# ============================================================================

@router.get("/evaluaciones/{evaluacion_id}/diagnose")
async def diagnose_evaluacion(
    evaluacion_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Devuelve un JSON con el diagnóstico de por qué el PDF de una
    evaluación puede no encontrar sus visitas/tratamientos. No modifica
    nada. Útil para depurar producción sin acceso directo a MongoDB."""
    from database import visitas_collection, tratamientos_collection
    if not ObjectId.is_valid(evaluacion_id):
        raise HTTPException(status_code=400, detail="ID inválido")
    ev = await evaluaciones_collection.find_one({"_id": ObjectId(evaluacion_id)})
    if not ev:
        raise HTTPException(status_code=404, detail="Evaluación no encontrada")
    parcela_id = ev.get("parcela_id", "")
    codigo_plant = (ev.get("codigo_plantacion") or "").strip()
    contrato_id = ev.get("contrato_id")
    parcela_data = None
    if parcela_id and ObjectId.is_valid(parcela_id):
        parcela_data = await parcelas_collection.find_one({"_id": ObjectId(parcela_id)})
    parcela_por_codigo = None
    if codigo_plant:
        parcela_por_codigo = await parcelas_collection.find_one({"codigo_plantacion": codigo_plant}, {"_id": 1, "codigo_plantacion": 1, "proveedor": 1})
    # counts under each strategy
    n_vis_by_pid = await visitas_collection.count_documents({"parcela_id": parcela_id}) if parcela_id else 0
    n_vis_by_codigo = await visitas_collection.count_documents({"codigo_plantacion": codigo_plant}) if codigo_plant else 0
    n_vis_by_contrato = await visitas_collection.count_documents({"contrato_id": contrato_id}) if contrato_id else 0
    n_trat_by_pid = await tratamientos_collection.count_documents({"parcelas_ids": parcela_id}) if parcela_id else 0
    parcela_ids_by_codigo = []
    if codigo_plant:
        async for p in parcelas_collection.find({"codigo_plantacion": codigo_plant}, {"_id": 1}):
            parcela_ids_by_codigo.append(str(p["_id"]))
    n_trat_by_codigo = await tratamientos_collection.count_documents({"parcelas_ids": {"$in": parcela_ids_by_codigo}}) if parcela_ids_by_codigo else 0
    n_trat_by_contrato = await tratamientos_collection.count_documents({"contrato_id": contrato_id}) if contrato_id else 0
    return {
        "evaluacion_id": evaluacion_id,
        "evaluacion": {
            "parcela_id": parcela_id,
            "codigo_plantacion": codigo_plant,
            "contrato_id": contrato_id,
            "proveedor": ev.get("proveedor"),
            "cultivo": ev.get("cultivo"),
            "campana": ev.get("campana"),
        },
        "parcela_ref_by_id_existe": parcela_data is not None,
        "parcela_encontrada_por_codigo": {
            "_id": str(parcela_por_codigo["_id"]) if parcela_por_codigo else None,
            "codigo_plantacion": parcela_por_codigo.get("codigo_plantacion") if parcela_por_codigo else None,
            "proveedor": parcela_por_codigo.get("proveedor") if parcela_por_codigo else None,
        } if parcela_por_codigo else None,
        "counts": {
            "visitas_por_parcela_id": n_vis_by_pid,
            "visitas_por_codigo_plantacion": n_vis_by_codigo,
            "visitas_por_contrato_id": n_vis_by_contrato,
            "tratamientos_por_parcelas_ids": n_trat_by_pid,
            "tratamientos_por_codigo_plantacion(fallback)": n_trat_by_codigo,
            "tratamientos_por_contrato_id": n_trat_by_contrato,
        }
    }


@router.get("/evaluaciones/{evaluacion_id}/pdf")
async def generate_evaluacion_pdf(
    evaluacion_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Generar PDF de la hoja de evaluación con visitas y tratamientos"""
    from fastapi.responses import Response
    from weasyprint import HTML
    from database import visitas_collection, tratamientos_collection, maquinaria_collection
    import io
    
    # Collection de tecnicos aplicadores
    tecnicos_aplicadores_collection = db['tecnicos_aplicadores']
    
    if not ObjectId.is_valid(evaluacion_id):
        raise HTTPException(status_code=400, detail="ID inválido")
    
    evaluacion = await evaluaciones_collection.find_one({"_id": ObjectId(evaluacion_id)})
    if not evaluacion:
        raise HTTPException(status_code=404, detail="Evaluación no encontrada")
    
    parcela_id = evaluacion.get("parcela_id", "")
    
    # Ficheros temporales que se generan durante el render del PDF
    # (mapas satelitales de `staticmap`) y hay que borrar tras `write_pdf`.
    _pdf_temp_files: list[str] = []
    
    # Obtener datos completos de la parcela incluyendo geometría
    parcela_data = None
    if parcela_id and ObjectId.is_valid(parcela_id):
        parcela_data = await parcelas_collection.find_one({"_id": ObjectId(parcela_id)})
    
    # Fallback robusto: si la parcela referenciada por _id ya NO existe (fue
    # borrada y recreada, restaurada tras backup con distinto ObjectId, etc.)
    # intentamos localizarla por `codigo_plantacion` (que la evaluación guardó
    # como snapshot). Si la encontramos, reasociamos el nuevo _id — así el
    # PDF recupera automáticamente sus visitas y tratamientos.
    if not parcela_data:
        codigo_plant = (evaluacion.get("codigo_plantacion") or "").strip()
        if codigo_plant:
            parcela_data = await parcelas_collection.find_one({"codigo_plantacion": codigo_plant})
            if parcela_data:
                parcela_id = str(parcela_data["_id"])
    
    # Resolver Variedad en vivo desde el catálogo de Cultivos cuando la parcela
    # tiene `variedad` vacía pero el cultivo asociado tiene exactamente una
    # variedad en su catálogo (replica la lógica del frontend ParcelasForm /
    # EvaluacionesImpresos). Garantiza que el PDF muestre la variedad real
    # aunque nunca se haya guardado explícitamente en la parcela.
    variedad_resuelta = ""
    if parcela_data:
        variedad_resuelta = (parcela_data.get("variedad") or "").strip()
    if not variedad_resuelta:
        cultivo_nombre = (parcela_data or {}).get("cultivo") or evaluacion.get("cultivo") or ""
        if cultivo_nombre:
            cultivo_doc = await db['cultivos'].find_one({"nombre": cultivo_nombre})
            variedades = (cultivo_doc or {}).get("variedades") or []
            if len(variedades) == 1:
                variedad_resuelta = str(variedades[0]).strip()
    # Inyectar la variedad resuelta tanto en `evaluacion` como en `impresos`
    # para que todas las plantillas (cabecera Plantación, Impresos cabecera y
    # Datos Generales) la muestren de forma consistente.
    if variedad_resuelta:
        if not (evaluacion.get("variedad") or "").strip():
            evaluacion["variedad"] = variedad_resuelta
        imp_inline = evaluacion.get("impresos")
        if isinstance(imp_inline, dict) and not (imp_inline.get("variedad") or "").strip():
            imp_inline["variedad"] = variedad_resuelta
    
    # Obtener visitas de la parcela (ordenadas por número de visita ASC,
    # con fallback a fecha_visita cuando una visita antigua no tenga número).
    #
    # Fallback robusto: si NO encontramos visitas por `parcela_id` (por
    # ejemplo, porque la parcela fue borrada y recreada con otro ObjectId,
    # o los datos vienen de un import previo con IDs distintos), buscamos
    # por `codigo_plantacion` — que sí es un identificador de negocio
    # estable que compartimos entre evaluación, parcela y visita.
    visitas = []
    codigo_plant_lookup = (
        (parcela_data or {}).get("codigo_plantacion")
        or evaluacion.get("codigo_plantacion")
        or ""
    ).strip()
    if parcela_id:
        visitas = await visitas_collection.find({"parcela_id": parcela_id}).sort([
            ("numero_visita", 1),
            ("fecha_visita", 1),
        ]).to_list(100)
    if not visitas and codigo_plant_lookup:
        visitas = await visitas_collection.find({"codigo_plantacion": codigo_plant_lookup}).sort([
            ("numero_visita", 1),
            ("fecha_visita", 1),
        ]).to_list(100)
    # Último fallback: por contrato_id de la evaluación (algunas visitas
    # antiguas pueden no tener codigo_plantacion pero sí contrato_id).
    if not visitas:
        contrato_id_ev = evaluacion.get("contrato_id")
        if contrato_id_ev:
            visitas = await visitas_collection.find({"contrato_id": contrato_id_ev}).sort([
                ("numero_visita", 1),
                ("fecha_visita", 1),
            ]).to_list(100)
    
    # Obtener tratamientos de la parcela (ordenados de más antiguo a más nuevo).
    # Fallbacks robustos: (a) por `parcelas_ids` con el nuevo _id resuelto,
    # (b) buscando todas las parcelas con mismo `codigo_plantacion` y usando
    # sus _id en $in, (c) como último recurso por `contrato_id` de la
    # evaluación — que los tratamientos también almacenan.
    tratamientos = []
    if parcela_id:
        tratamientos = await tratamientos_collection.find({"parcelas_ids": parcela_id}).sort("fecha_tratamiento", 1).to_list(100)
    if not tratamientos and codigo_plant_lookup:
        # Recolectar todos los _id de parcelas que compartan codigo_plantacion
        parcela_ids_por_codigo = []
        async for pdoc in parcelas_collection.find({"codigo_plantacion": codigo_plant_lookup}, {"_id": 1}):
            parcela_ids_por_codigo.append(str(pdoc["_id"]))
        if parcela_ids_por_codigo:
            tratamientos = await tratamientos_collection.find({
                "parcelas_ids": {"$in": parcela_ids_por_codigo}
            }).sort("fecha_tratamiento", 1).to_list(100)
    if not tratamientos:
        contrato_id_ev = evaluacion.get("contrato_id")
        if contrato_id_ev:
            tratamientos = await tratamientos_collection.find({"contrato_id": contrato_id_ev}).sort("fecha_tratamiento", 1).to_list(100)
    
    # Para cada tratamiento, obtener los datos completos del aplicador y la máquina
    tratamientos_enriquecidos = []
    for trat in tratamientos:
        trat_data = dict(trat)
        
        # Obtener datos del aplicador (el campo es tecnico_aplicador_id)
        aplicador_id = trat.get("tecnico_aplicador_id") or trat.get("aplicador_id")
        if aplicador_id and ObjectId.is_valid(aplicador_id):
            aplicador = await tecnicos_aplicadores_collection.find_one({"_id": ObjectId(aplicador_id)})
            if aplicador:
                trat_data["aplicador_completo"] = serialize_doc(aplicador)
        
        # Obtener datos de la máquina
        maquina_id = trat.get("maquina_id")
        if maquina_id and ObjectId.is_valid(maquina_id):
            maquina = await maquinaria_collection.find_one({"_id": ObjectId(maquina_id)})
            if maquina:
                trat_data["maquina_completa"] = serialize_doc(maquina)
        
        tratamientos_enriquecidos.append(trat_data)
    
    tratamientos = tratamientos_enriquecidos
    
    # ------------------------------------------------------------------
    # Data-freshness: sobrescribir los campos "heredados" de la parcela
    # (proveedor, cultivo, variedad, finca, campaña, codigo_plantacion)
    # dentro de cada visita y tratamiento con los valores ACTUALES de la
    # parcela y del catálogo de cultivos. Así, cualquier edición posterior
    # de la parcela / cultivo / técnico / maquinaria queda reflejada en el
    # PDF de la Hoja de Evaluación de inmediato, sin depender del snapshot
    # denormalizado que se guardó al crear la visita/tratamiento.
    # ------------------------------------------------------------------
    inherited_fields = {}
    if parcela_data:
        inherited_fields = {
            "proveedor": (parcela_data.get("proveedor") or "").strip(),
            "cultivo": (parcela_data.get("cultivo") or "").strip(),
            "variedad": (variedad_resuelta or (parcela_data.get("variedad") or "")).strip(),
            "finca": (parcela_data.get("finca") or "").strip(),
            "campana": (parcela_data.get("campana") or "").strip(),
            "codigo_plantacion": (parcela_data.get("codigo_plantacion") or "").strip(),
        }
    def _overlay_inherited(doc):
        """Devuelve una copia del doc con los inherited_fields sobrescritos
        cuando la parcela sí trae ese valor. Preserva los originales si la
        parcela no tiene el campo (para no perder información legada)."""
        merged = dict(doc)
        for k, v in inherited_fields.items():
            if v:  # sólo si la parcela tiene un valor no vacío
                merged[k] = v
        return merged
    visitas = [_overlay_inherited(v) for v in visitas]
    tratamientos = [_overlay_inherited(t) for t in tratamientos]
    
    # Irrigaciones y Cosechas: eliminadas del cuaderno de campo (no se incluyen
    # en el PDF según requerimientos de usuario).
    # La paginación se calcula dinámicamente con counter(page)/counter(pages) en CSS.
    
    # Función helper para formatear respuestas
    def format_respuesta(resp):
        if resp is True:
            return "Sí"
        elif resp is False:
            return "No"
        elif resp is None or resp == "":
            return "—"
        return str(resp)
    
    def format_fecha(fecha):
        if not fecha:
            return "—"
        return str(fecha)
    
    # Logo FRUVECO (fragmento HTML pre-computado; posicionado como fixed
    # aparece en la esquina superior-izquierda de TODAS las páginas).
    _logo_data_uri = _get_fruveco_logo_data_uri()
    logo_header_html = (
        f'<div class="fruveco-logo-header"><img src="{_logo_data_uri}" alt="FRUVECO" /></div>'
        if _logo_data_uri else ''
    )

    # CSS común para todas las páginas
    css_styles = """
        @page {
            size: A4;
            margin: 3cm 1.5cm 1.5cm 1.5cm;
            @bottom-center {
                content: "Página " counter(page) " de " counter(pages);
                font-size: 8pt;
                color: #888;
            }
            @bottom-right {
                content: "FRUVECO · Cuaderno de Campo";
                font-size: 8pt;
                color: #aaa;
            }
        }
        body {
            font-family: 'Helvetica', 'Arial', sans-serif;
            font-size: 10pt;
            line-height: 1.4;
            color: #333;
        }
        /* Logo FRUVECO fijado en la esquina superior izquierda de cada página */
        .fruveco-logo-header {
            position: fixed;
            top: -2.2cm;
            left: -0.5cm;
            width: 4.5cm;
            height: auto;
        }
        .fruveco-logo-header img {
            width: 100%;
            height: auto;
            display: block;
        }
        .page-break {
            page-break-before: always;
        }
        .header {
            text-align: center;
            margin-bottom: 20px;
            padding-bottom: 15px;
            border-bottom: 2px solid #2d5a27;
        }
        .header h1 {
            color: #2d5a27;
            font-size: 18pt;
            margin: 0 0 5px 0;
        }
        .header h2 {
            color: #666;
            font-size: 12pt;
            margin: 0;
            font-weight: normal;
        }
        .header h3 {
            color: #2d5a27;
            font-size: 10pt;
            margin: 5px 0 0 0;
            font-weight: normal;
        }
        .section {
            margin-bottom: 15px;
            page-break-inside: avoid;
        }
        .section-title {
            background-color: #2d5a27;
            color: white;
            padding: 8px 12px;
            font-weight: bold;
            font-size: 11pt;
            margin-bottom: 0;
        }
        .section-title-blue {
            background-color: #1a5276;
        }
        .section-title-orange {
            background-color: #b9770e;
        }
        .section-content {
            border: 1px solid #ddd;
            border-top: none;
            padding: 12px;
        }
        .question-row {
            padding: 8px 0;
            border-bottom: 1px solid #eee;
        }
        .question-row:last-child {
            border-bottom: none;
        }
        .question-num {
            font-weight: bold;
            color: #2d5a27;
        }
        .question-text {
            margin-left: 5px;
        }
        .answer {
            margin-left: 20px;
            color: #555;
            font-style: italic;
        }
        .answer-yes {
            color: #28a745;
            font-weight: bold;
        }
        .answer-no {
            color: #dc3545;
            font-weight: bold;
        }
        .footer {
            margin-top: 30px;
            padding-top: 15px;
            border-top: 1px solid #ddd;
            font-size: 9pt;
            color: #666;
            text-align: center;
        }
        .datos-grid {
            display: grid;
            grid-template-columns: repeat(3, 1fr);
            gap: 10px;
        }
        .datos-grid-2 {
            display: grid;
            grid-template-columns: repeat(2, 1fr);
            gap: 10px;
        }
        .dato-item {
            padding: 5px;
        }
        .dato-label {
            font-weight: bold;
            font-size: 9pt;
            color: #666;
        }
        .dato-value {
            font-size: 10pt;
        }
        .visita-header {
            background-color: #1a5276;
            color: white;
            padding: 10px;
            margin-bottom: 15px;
        }
        .visita-header h3 {
            margin: 0;
            color: white;
        }
        .tratamiento-header {
            background-color: #b9770e;
            color: white;
            padding: 10px;
            margin-bottom: 15px;
        }
        .tratamiento-header h3 {
            margin: 0;
            color: white;
        }
        .irrigacion-header {
            background-color: #2874a6;
            color: white;
            padding: 10px;
            margin-bottom: 15px;
        }
        .irrigacion-header h3 {
            margin: 0;
            color: white;
        }
        .cosecha-header {
            background-color: #1e8449;
            color: white;
            padding: 10px;
            margin-bottom: 15px;
        }
        .cosecha-header h3 {
            margin: 0;
            color: white;
        }
        .section-title-water {
            background-color: #2874a6;
        }
        .section-title-harvest {
            background-color: #1e8449;
        }
        .summary-box {
            background-color: #f8f9fa;
            border: 1px solid #dee2e6;
            border-radius: 5px;
            padding: 15px;
            margin-bottom: 20px;
        }
        .summary-title {
            font-weight: bold;
            font-size: 12pt;
            margin-bottom: 10px;
            color: #2d5a27;
        }
        table.data-table {
            width: 100%;
            border-collapse: collapse;
            margin-top: 10px;
        }
        table.data-table th, table.data-table td {
            border: 1px solid #ddd;
            padding: 8px;
            text-align: left;
        }
        table.data-table th {
            background-color: #f2f2f2;
            font-weight: bold;
        }
        table.data-table tr:nth-child(even) {
            background-color: #f9f9f9;
        }
        .index-box {
            background-color: #fff;
            border: 2px solid #2d5a27;
            border-radius: 5px;
            padding: 15px;
            margin-bottom: 20px;
        }
        .index-title {
            font-weight: bold;
            font-size: 14pt;
            margin-bottom: 15px;
            color: #2d5a27;
            text-align: center;
            border-bottom: 2px solid #2d5a27;
            padding-bottom: 10px;
        }
        .index-section {
            margin-bottom: 15px;
        }
        .index-section-title {
            font-weight: bold;
            font-size: 11pt;
            margin-bottom: 8px;
            padding: 5px 10px;
            border-radius: 3px;
        }
        .index-section-title.visitas {
            background-color: #1a5276;
            color: white;
        }
        .index-section-title.tratamientos {
            background-color: #b9770e;
            color: white;
        }
        .index-section-title.irrigaciones {
            background-color: #2874a6;
            color: white;
        }
        .index-section-title.cosechas {
            background-color: #1e8449;
            color: white;
        }
        .index-item {
            display: flex;
            justify-content: space-between;
            padding: 5px 10px;
            border-bottom: 1px dotted #ddd;
            font-size: 10pt;
        }
        .index-item:last-child {
            border-bottom: none;
        }
        .index-item-name {
            flex: 1;
        }
        .index-item-date {
            color: #666;
            margin-left: 10px;
        }
        .index-item-page {
            font-weight: bold;
            margin-left: 10px;
            min-width: 50px;
            text-align: right;
        }
        .index-item-link {
            color: inherit;
            text-decoration: none;
            display: block;
        }
        .index-item-link:hover .index-item-name {
            color: #1e8449;
        }
        .index-empty {
            color: #999;
            font-style: italic;
            padding: 10px;
            text-align: center;
        }
        .aplicador-header {
            background-color: #7b2cbf;
            color: white;
            padding: 10px;
            margin-bottom: 15px;
        }
        .aplicador-header h3 {
            margin: 0;
            color: white;
        }
        .section-title-purple {
            background-color: #7b2cbf;
        }
        .maquina-header {
            background-color: #495057;
            color: white;
            padding: 10px;
            margin-bottom: 15px;
        }
        .maquina-header h3 {
            margin: 0;
            color: white;
        }
        .section-title-gray {
            background-color: #495057;
        }
        .certificate-image {
            max-width: 100%;
            max-height: 400px;
            margin: 15px auto;
            display: block;
            border: 2px solid #ddd;
            border-radius: 5px;
        }
        .placa-ce-image {
            max-width: 100%;
            max-height: 350px;
            margin: 15px auto;
            display: block;
            border: 2px solid #ddd;
            border-radius: 5px;
        }
        .ficha-datos {
            background-color: #f8f9fa;
            border: 1px solid #dee2e6;
            border-radius: 5px;
            padding: 15px;
            margin-bottom: 15px;
        }
        .ficha-titulo {
            font-size: 14pt;
            font-weight: bold;
            margin-bottom: 15px;
            padding-bottom: 10px;
            border-bottom: 2px solid;
        }
        .ficha-titulo-aplicador {
            color: #7b2cbf;
            border-color: #7b2cbf;
        }
        .ficha-titulo-maquina {
            color: #495057;
            border-color: #495057;
        }
        .no-image-box {
            background-color: #f0f0f0;
            border: 2px dashed #ccc;
            border-radius: 5px;
            padding: 40px;
            text-align: center;
            color: #999;
            font-style: italic;
        }
    """
    
    # Generar HTML para el PDF
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <style>{css_styles}</style>
    </head>
    <body>
        <!-- Logo FRUVECO fijado en la esquina superior izquierda de todas las páginas -->
        {logo_header_html}
        <!-- PÁGINA 1: HOJA DE EVALUACIÓN PRINCIPAL -->
        <div class="header">
            <h1>FRUVECO</h1>
            <h2>HOJA DE EVALUACIÓN · CUADERNO DE CAMPO</h2>
            <h3 style="font-size: 11pt; color: #2d5a27;">{evaluacion.get('proveedor', '')} · {evaluacion.get('cultivo', '')} · Campaña {evaluacion.get('campana', '')}</h3>
        </div>
        
        <!-- Resumen -->
        <div class="summary-box">
            <div class="summary-title">RESUMEN DEL CUADERNO DE CAMPO</div>
            <div class="datos-grid">
                <div class="dato-item">
                    <div class="dato-label">Visitas</div>
                    <div class="dato-value" style="font-size: 14pt; font-weight: bold; color: #1a5276;">{len(visitas)}</div>
                </div>
                <div class="dato-item">
                    <div class="dato-label">Tratamientos</div>
                    <div class="dato-value" style="font-size: 14pt; font-weight: bold; color: #b9770e;">{len(tratamientos)}</div>
                </div>
                <div class="dato-item">
                    <div class="dato-label">Total Páginas</div>
                    <div class="dato-value" style="font-size: 14pt; font-weight: bold;">Ver pie de página</div>
                </div>
            </div>
        </div>
        
        <!-- ÍNDICE DE CONTENIDOS -->
        <div class="index-box">
            <div class="index-title">ÍNDICE DE CONTENIDOS</div>
            
            <!-- Índice de Visitas -->
            <div class="index-section">
                <div class="index-section-title visitas">VISITAS ({len(visitas)})</div>
    """
    
    if visitas:
        for idx, visita in enumerate(visitas, 1):
            page_num = 1 + idx
            fecha = format_fecha(visita.get('fecha_visita'))
            objetivo = visita.get('objetivo', 'Sin objetivo')[:40]
            n_visita = visita.get('numero_visita') or idx
            html_content += f"""
                <a href="#visita-{idx}" class="index-item-link">
                    <div class="index-item">
                        <span class="index-item-name">Visita #{n_visita} · {objetivo}</span>
                        <span class="index-item-date">{fecha}</span>
                        <span class="index-item-page">Pág. {page_num}</span>
                    </div>
                </a>
            """
    else:
        html_content += """
                <div class="index-empty">No hay visitas registradas</div>
        """
    
    html_content += """
            </div>
            
            <!-- Índice de Tratamientos -->
            <div class="index-section">
                <div class="index-section-title tratamientos">TRATAMIENTOS (""" + str(len(tratamientos)) + """)</div>
    """
    
    if tratamientos:
        for idx, tratamiento in enumerate(tratamientos, 1):
            page_num = 1 + len(visitas) + idx
            fecha = format_fecha(tratamiento.get('fecha_tratamiento'))
            # El backend almacena el tipo en `tipo_tratamiento` (e.g. "FITOSANITARIOS")
            # y el subtipo (e.g. "Herbicida", "Fungicida"). El campo legacy
            # `tipo` raramente está poblado, por lo que mostrábamos "Sin tipo".
            tipo_principal = (tratamiento.get('tipo_tratamiento') or tratamiento.get('tipo') or 'Tratamiento').strip()
            subtipo = (tratamiento.get('subtipo') or '').strip()
            tipo_label = f"{tipo_principal}" + (f" — {subtipo}" if subtipo else '')
            descripcion = (tratamiento.get('producto_fitosanitario_nombre') or tratamiento.get('descripcion') or '')[:30]
            html_content += f"""
                <a href="#tratamiento-{idx}" class="index-item-link">
                    <div class="index-item">
                        <span class="index-item-name">{idx}. {tipo_label[:50]} {('· ' + descripcion) if descripcion else ''}</span>
                        <span class="index-item-date">{fecha}</span>
                        <span class="index-item-page">Pág. {page_num}</span>
                    </div>
                </a>
            """
    else:
        html_content += """
                <div class="index-empty">No hay tratamientos registrados</div>
        """
    
    # Índices de Irrigaciones y Cosechas: eliminados del cuaderno de campo
    # según requerimiento del usuario.
    
    html_content += f"""
            </div>
        </div>
        
        <!-- Datos Generales -->
        <div class="section">
            <div class="section-title">DATOS GENERALES</div>
            <div class="section-content">
                <div class="datos-grid">
                    <div class="dato-item">
                        <div class="dato-label">Fecha Inicio</div>
                        <div class="dato-value">{evaluacion.get('fecha_inicio', '—')}</div>
                    </div>
                    <div class="dato-item">
                        <div class="dato-label">Fecha Fin</div>
                        <div class="dato-value">{evaluacion.get('fecha_fin', '—')}</div>
                    </div>
                    <div class="dato-item">
                        <div class="dato-label">Técnico</div>
                        <div class="dato-value">{evaluacion.get('tecnico', '—')}</div>
                    </div>
                </div>
            </div>
        </div>
        
        <!-- Plantación -->
        <div class="section">
            <div class="section-title">PLANTACIÓN</div>
            <div class="section-content">
                <div class="datos-grid">
                    <div class="dato-item">
                        <div class="dato-label">Proveedor / Agricultor</div>
                        <div class="dato-value">{evaluacion.get('proveedor', '—')}</div>
                    </div>
                    <div class="dato-item">
                        <div class="dato-label">Código Plantación</div>
                        <div class="dato-value">{evaluacion.get('codigo_plantacion', '—')}</div>
                    </div>
                    <div class="dato-item">
                        <div class="dato-label">Finca</div>
                        <div class="dato-value">{evaluacion.get('finca', '—')}</div>
                    </div>
                    <div class="dato-item">
                        <div class="dato-label">Cultivo</div>
                        <div class="dato-value">{evaluacion.get('cultivo', '—')}</div>
                    </div>
                    <div class="dato-item">
                        <div class="dato-label">Variedad</div>
                        <div class="dato-value">{evaluacion.get('variedad', '—')}</div>
                    </div>
                    <div class="dato-item">
                        <div class="dato-label">Superficie</div>
                        <div class="dato-value">{evaluacion.get('superficie', 0)} ha</div>
                    </div>
                    <div class="dato-item">
                        <div class="dato-label">Campaña</div>
                        <div class="dato-value">{evaluacion.get('campana', '—')}</div>
                    </div>
                </div>
            </div>
        </div>
    """
    
    # Añadir sección de mapa de la parcela si hay geometría o imagen
    if parcela_data:
        geometria = None
        if parcela_data.get('recintos') and len(parcela_data['recintos']) > 0:
            geometria = parcela_data['recintos'][0].get('geometria', [])
        
        lat_parcela = parcela_data.get('latitud', 0)
        lng_parcela = parcela_data.get('longitud', 0)
        imagen_mapa_url = parcela_data.get('imagen_mapa_url', '')
        imagen_mapa_path = parcela_data.get('imagen_mapa_path', '')
        
        html_content += """
        <div class="section">
            <div class="section-title" style="background-color: #0d6efd;">UBICACIÓN DE LA PARCELA</div>
            <div class="section-content" style="text-align: center;">
        """
        
        # Primero intentar mostrar la imagen capturada del mapa
        if imagen_mapa_url or imagen_mapa_path:
            # Construir la ruta absoluta de la imagen
            if imagen_mapa_path and os.path.exists(imagen_mapa_path):
                img_src = f"file://{imagen_mapa_path}"
            elif imagen_mapa_url:
                # Convertir URL relativa a path absoluto
                if imagen_mapa_url.startswith('/api/uploads/'):
                    local_path = imagen_mapa_url.replace('/api/uploads/', '/app/uploads/')
                    if os.path.exists(local_path):
                        img_src = f"file://{local_path}"
                    else:
                        img_src = None
                else:
                    img_src = None
            else:
                img_src = None
            
            if img_src:
                html_content += f"""
                <div style="margin-bottom: 15px;">
                    <img src="{img_src}" style="max-width: 100%; max-height: 400px; border: 2px solid #2e7d32; border-radius: 8px; box-shadow: 0 4px 8px rgba(0,0,0,0.2);" alt="Mapa de la parcela" />
                </div>
                """
                
                # Añadir información de coordenadas si hay geometría
                if geometria and len(geometria) > 1:
                    lats = [p.get('lat', 0) for p in geometria]
                    lngs = [p.get('lng', 0) for p in geometria]
                    center_lat = sum(lats) / len(lats)
                    center_lng = sum(lngs) / len(lngs)
                    
                    html_content += f"""
                    <table style="width: 100%; margin-top: 10px; border-collapse: collapse;">
                        <tr>
                            <td style="padding: 8px; border: 1px solid #ddd; background: #f5f5f5; width: 50%;">
                                <strong>Centro del polígono:</strong><br/>
                                <span style="font-family: monospace;">{center_lat:.6f}, {center_lng:.6f}</span>
                            </td>
                            <td style="padding: 8px; border: 1px solid #ddd; background: #f5f5f5; width: 50%;">
                                <strong>Polígono:</strong><br/>
                                {len(geometria)} vértices
                            </td>
                        </tr>
                    </table>
                    """
            else:
                # No se pudo cargar la imagen, usar SVG como fallback
                pass
        
        elif geometria and len(geometria) > 1:
            # Calcular centro y bounds del polígono
            lats = [p.get('lat', 0) for p in geometria]
            lngs = [p.get('lng', 0) for p in geometria]
            center_lat = sum(lats) / len(lats)
            center_lng = sum(lngs) / len(lngs)
            
            min_lat, max_lat = min(lats), max(lats)
            min_lng, max_lng = min(lngs), max(lngs)
            
            # === Mapa satélite REAL con polígono superpuesto ===
            # Usamos `staticmap` (fetch de tiles Esri World Imagery) + Pillow
            # para renderizar la parcela sobre imagen aérea, igual que el
            # editor Leaflet Avanzado. Fallback a SVG si falla la red.
            satelite_ok = False
            try:
                from staticmap import StaticMap, Polygon as SmPolygon, CircleMarker
                import uuid as _uuid
                # Esri World Imagery (satélite gratuito)
                esri_url = "https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{z}/{y}/{x}"
                sm = StaticMap(900, 540, url_template=esri_url)
                pts = [(p.get('lng', 0), p.get('lat', 0)) for p in geometria]
                # Cerrar el anillo si no viene cerrado
                if pts and pts[0] != pts[-1]:
                    pts.append(pts[0])
                # Polígono verde translúcido con borde oscuro
                # NOTA: staticmap/PIL requieren color hex (#rrggbbaa) — no acepta rgba(..., 0.85)
                # Polygon(coords, fill_color, outline_color)
                sm.add_polygon(SmPolygon(pts, '#4CAF5066', '#2E7D32', simplify=False))
                # Marcador de centro (azul con núcleo blanco tipo Leaflet)
                sm.add_marker(CircleMarker((center_lng, center_lat), '#1565C0', 14))
                sm.add_marker(CircleMarker((center_lng, center_lat), '#FFFFFF', 8))
                img = sm.render()
                # === Overlay estilo SIGPAC: brújula N↑ y barra de escala ===
                try:
                    from PIL import ImageDraw, ImageFont
                    import math as _math
                    draw = ImageDraw.Draw(img, 'RGBA')
                    try:
                        font_bold = ImageFont.truetype("/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf", 16)
                        font_scale = ImageFont.truetype("/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf", 12)
                    except Exception:
                        font_bold = ImageFont.load_default()
                        font_scale = ImageFont.load_default()
                    W, H = img.size

                    # --- Brújula (esquina superior derecha) ---
                    cx, cy, r = W - 45, 45, 26
                    # Halo blanco semitransparente + círculo blanco con borde
                    draw.ellipse((cx-r-2, cy-r-2, cx+r+2, cy+r+2), fill=(255, 255, 255, 220), outline=(60, 60, 60, 255), width=2)
                    # Flecha norte (roja arriba, gris abajo) — polígono romboidal
                    draw.polygon([(cx, cy-r+6), (cx+7, cy+2), (cx, cy-2), (cx-7, cy+2)], fill=(198, 40, 40, 255))
                    draw.polygon([(cx, cy+r-6), (cx+7, cy-2), (cx, cy+2), (cx-7, cy-2)], fill=(90, 90, 90, 255))
                    # Etiqueta N en negrita encima
                    draw.text((cx-5, cy-r-2), "N", fill=(20, 20, 20, 255), font=font_bold)

                    # --- Barra de escala (esquina inferior izquierda) ---
                    # metros/pixel (Web Mercator): 156543.03392 * cos(lat) / 2^zoom
                    lat_rad = _math.radians(center_lat)
                    mpp = 156543.03392 * _math.cos(lat_rad) / (2 ** sm.zoom)
                    # Elegir una distancia "bonita" ≈ 150 px
                    target_px = 150
                    target_m = target_px * mpp
                    nice_values_m = [50, 100, 200, 500, 1000, 2000, 5000, 10000, 20000, 50000]
                    scale_m = min(nice_values_m, key=lambda v: abs(v - target_m))
                    scale_px = int(round(scale_m / mpp))
                    label = f"{scale_m/1000:.1f} km" if scale_m >= 1000 else f"{scale_m} m"

                    bx, by = 20, H - 40
                    bw, bh = scale_px, 10
                    # Fondo blanco semitransparente
                    draw.rectangle((bx-6, by-22, bx+bw+6, by+bh+6), fill=(255, 255, 255, 220), outline=(60, 60, 60, 255), width=1)
                    # Barra bicolor: mitad blanca / mitad negra
                    half = bw // 2
                    draw.rectangle((bx, by, bx+half, by+bh), fill=(255, 255, 255, 255), outline=(20, 20, 20, 255), width=1)
                    draw.rectangle((bx+half, by, bx+bw, by+bh), fill=(20, 20, 20, 255), outline=(20, 20, 20, 255), width=1)
                    # Etiquetas: 0 · mitad · total
                    draw.text((bx-3, by-16), "0", fill=(20, 20, 20, 255), font=font_scale)
                    draw.text((bx+bw-len(label)*6, by-16), label, fill=(20, 20, 20, 255), font=font_scale)
                except Exception as _ov_err:
                    print(f"[PDF] map overlay (compass/scale) failed: {_ov_err}")

                out_dir = "/app/uploads/evaluaciones/pdf_maps"
                os.makedirs(out_dir, exist_ok=True)
                out_path = os.path.join(out_dir, f"map_{_uuid.uuid4().hex}.png")
                img.save(out_path)
                _pdf_temp_files.append(out_path)
                map_html = f'''
                    <div style="border:2px solid #ccc; border-radius:8px; overflow:hidden;">
                        <img src="file://{out_path}" style="width:100%; display:block;" alt="Mapa satélite de la parcela" />
                    </div>
                    <div style="text-align:right; font-size:8pt; color:#888; margin-top:4px;">
                        Imagen satélite: Esri World Imagery · Polígono con {len(geometria)} vértices
                    </div>
                '''
                satelite_ok = True
            except Exception as _e:
                print(f"[PDF] staticmap satelite failed, falling back to SVG: {_e}")
                satelite_ok = False
            
            if not satelite_ok:
                # ---- Fallback SVG diagrama (código legacy) ----
                svg_width = 500
                svg_height = 300
                padding = 30
                lat_range = max_lat - min_lat if max_lat != min_lat else 0.001
                lng_range = max_lng - min_lng if max_lng != min_lng else 0.001
                
                def geo_to_svg(lat, lng):
                    x = padding + ((lng - min_lng) / lng_range) * (svg_width - 2 * padding)
                    y = padding + ((max_lat - lat) / lat_range) * (svg_height - 2 * padding)
                    return x, y
                
                polygon_points = ' '.join([f"{geo_to_svg(p.get('lat', 0), p.get('lng', 0))[0]:.1f},{geo_to_svg(p.get('lat', 0), p.get('lng', 0))[1]:.1f}" for p in geometria])
                map_html = f'''
                <svg width="{svg_width}" height="{svg_height}" viewBox="0 0 {svg_width} {svg_height}" style="border: 2px solid #ddd; border-radius: 8px; background: linear-gradient(135deg, #e8f5e9 0%, #c8e6c9 100%);">
                    <defs>
                        <pattern id="grid" width="20" height="20" patternUnits="userSpaceOnUse">
                            <path d="M 20 0 L 0 0 0 20" fill="none" stroke="#a5d6a7" stroke-width="0.5"/>
                        </pattern>
                    </defs>
                    <rect width="100%" height="100%" fill="url(#grid)"/>
                    <polygon points="{polygon_points}" fill="rgba(76, 175, 80, 0.4)" stroke="#2e7d32" stroke-width="3"/>
                    {"".join([f'<circle cx="{geo_to_svg(p.get("lat", 0), p.get("lng", 0))[0]:.1f}" cy="{geo_to_svg(p.get("lat", 0), p.get("lng", 0))[1]:.1f}" r="5" fill="#c62828" stroke="white" stroke-width="2"/>' for p in geometria])}
                    <circle cx="{geo_to_svg(center_lat, center_lng)[0]:.1f}" cy="{geo_to_svg(center_lat, center_lng)[1]:.1f}" r="8" fill="#1565c0" stroke="white" stroke-width="2"/>
                    <text x="{geo_to_svg(center_lat, center_lng)[0]:.1f}" y="{geo_to_svg(center_lat, center_lng)[1] + 25:.1f}" text-anchor="middle" font-size="11" fill="#1565c0" font-weight="bold">Centro</text>
                </svg>
                '''
            
            html_content += f"""
                <div style="margin-bottom: 15px;">
                    {map_html}
                </div>
                <table style="width: 100%; margin-top: 10px; border-collapse: collapse;">
                    <tr>
                        <td style="padding: 8px; border: 1px solid #ddd; background: #f5f5f5; width: 50%;">
                            <strong>Centro del polígono:</strong><br/>
                            <span style="font-family: monospace;">{center_lat:.6f}, {center_lng:.6f}</span>
                        </td>
                        <td style="padding: 8px; border: 1px solid #ddd; background: #f5f5f5; width: 50%;">
                            <strong>Polígono:</strong><br/>
                            {len(geometria)} vértices
                        </td>
                    </tr>
                    <tr>
                        <td colspan="2" style="padding: 8px; border: 1px solid #ddd; font-size: 9pt;">
                            <strong>Coordenadas de los vértices:</strong><br/>
                            {' → '.join([f"({p.get('lat', 0):.5f}, {p.get('lng', 0):.5f})" for p in geometria[:6]])}
                            {'...' if len(geometria) > 6 else ''}
                        </td>
                    </tr>
                </table>
            """
        elif lat_parcela and lng_parcela:
            # Si solo hay coordenadas puntuales, mostrar un marcador
            svg_point = '''
            <svg width="300" height="200" viewBox="0 0 300 200" style="border: 2px solid #ddd; border-radius: 8px; background: linear-gradient(135deg, #e8f5e9 0%, #c8e6c9 100%);">
                <defs>
                    <pattern id="grid2" width="20" height="20" patternUnits="userSpaceOnUse">
                        <path d="M 20 0 L 0 0 0 20" fill="none" stroke="#a5d6a7" stroke-width="0.5"/>
                    </pattern>
                </defs>
                <rect width="100%" height="100%" fill="url(#grid2)"/>
                
                <!-- Marcador de ubicación -->
                <circle cx="150" cy="100" r="12" fill="#c62828" stroke="white" stroke-width="3"/>
                <circle cx="150" cy="100" r="25" fill="none" stroke="#c62828" stroke-width="2" stroke-dasharray="5,5"/>
                
                <!-- Etiqueta -->
                <text x="150" y="145" text-anchor="middle" font-size="12" fill="#333" font-weight="bold">Ubicación de la parcela</text>
            </svg>
            '''
            
            html_content += f"""
                <div style="margin-bottom: 15px;">
                    {svg_point}
                </div>
                <div style="margin-top: 10px; padding: 10px; background: #f5f5f5; border-radius: 5px;">
                    <p style="font-size: 11pt; margin: 0;"><strong>Coordenadas:</strong> 
                    <span style="font-family: monospace;">{lat_parcela:.6f}, {lng_parcela:.6f}</span></p>
                    <p style="font-size: 9pt; color: #666; margin: 5px 0 0 0;">
                        (Punto de referencia - no hay polígono dibujado)
                    </p>
                </div>
            """
        else:
            html_content += """
                <div class="no-image-box">
                    <p>No hay coordenadas de ubicación disponibles para esta parcela</p>
                    <p style="font-size: 9pt; margin-top: 10px;">Añada las coordenadas o dibuje el polígono en el módulo de Mapas</p>
                </div>
            """
        
        html_content += """
            </div>
        </div>
        """
    
    # Cuestionario unificado: todas las secciones internas se muestran como
    # una única "TOMA DE DATOS" con numeración continua 1..N, respetando el
    # `orden_global` guardado en configuración (mismo orden que la vista plana
    # del frontend). Internamente los datos siguen almacenándose por sección
    # (toma_datos, analisis_suelo, calidad_cepellones, …) por compatibilidad.
    secciones_internas = [
        'toma_datos',
        'analisis_suelo',
        'pasos_precampana',
        'calidad_cepellones',
        'inspeccion_maquinaria',
        'observaciones',
        'calibracion_mantenimiento',
    ]
    todas_respuestas = []
    for _sk in secciones_internas:
        _resp = evaluacion.get(_sk, []) or []
        if isinstance(_resp, list):
            todas_respuestas.extend(_resp)
    
    # Aplicar orden_global (mismo que la UI): las preguntas con posición en
    # orden_global van primero según ese orden; las que no aparezcan mantienen
    # el orden de sección + índice original al final.
    try:
        _cfg = await evaluaciones_config_collection.find_one({"tipo": "preguntas"})
        _orden_global = (_cfg or {}).get("orden_global", []) if _cfg else []
    except Exception:
        _orden_global = []
    if _orden_global:
        _pos = {pid: i for i, pid in enumerate(_orden_global)}
        # Precompute stable fallback index (id(r) del objeto) para preguntas que
        # no aparecen en orden_global — evita `list.index()` durante sort que
        # falla porque timsort muta la lista y produce ValueError.
        _fallback = {id(r): i for i, r in enumerate(todas_respuestas)}
        todas_respuestas.sort(
            key=lambda r: _pos.get(r.get('pregunta_id'), 10_000 + _fallback[id(r)])
        )
    
    if todas_respuestas:
        html_content += """
        <div class="section">
            <div class="section-title">TOMA DE DATOS</div>
            <div class="section-content">
        """
        for idx, resp in enumerate(todas_respuestas, 1):
            pregunta = resp.get('pregunta', '')
            respuesta = resp.get('respuesta')
            respuesta_fmt = format_respuesta(respuesta)
            
            respuesta_class = ''
            if respuesta is True:
                respuesta_class = 'answer-yes'
            elif respuesta is False:
                respuesta_class = 'answer-no'
            
            html_content += f"""
                <div class="question-row">
                    <span class="question-num">{idx}.</span>
                    <span class="question-text">{pregunta}</span>
                    <div class="answer {respuesta_class}">R: {respuesta_fmt}</div>
                </div>
            """
        html_content += """
            </div>
        </div>
        """
    
    # Impresos — Cabecera + 6 secciones (esquema nuevo)
    impresos = evaluacion.get('impresos', {}) or {}
    # Mostrar Impresos siempre que haya una parcela vinculada al evaluación
    # (la cabecera se computa en vivo desde parcela/contrato; las 6 secciones
    # solo se muestran si el usuario ha rellenado algo en impresos).
    if impresos or evaluacion.get('parcela_id'):
        def _fmt_sn(v):
            """Format Sí/No tri-state (True/False/None)."""
            if v is True:
                return '<span class="answer-yes">Sí</span>'
            if v is False:
                return '<span class="answer-no">No</span>'
            return '<span style="color:#888;">—</span>'

        def _fmt_sintomas(s):
            if not s or not isinstance(s, dict):
                return '<span style="color:#888;">—</span>'
            marcados = [k.capitalize() for k in ('enfermedades', 'plagas', 'virus') if s.get(k)]
            return ', '.join(marcados) if marcados else '<span style="color:#888;">Ninguno</span>'

        def _txt(v, dash='—'):
            v = (v if v is not None else '')
            v = str(v).strip()
            return v if v else f'<span style="color:#888;">{dash}</span>'

        analisis = impresos.get('analisis_suelo', {}) or {}
        pasos = impresos.get('pasos_precampana', {}) or {}
        calibracion = impresos.get('calibracion', {}) or {}
        cepellones = impresos.get('calidad_cepellones', {}) or {}
        maquinaria = impresos.get('inspeccion_maquinaria', {}) or {}

        html_content += f"""
        <div class="page-break"></div>
        <div class="section">
            <div class="section-title">IMPRESOS — CABECERA / PLANTACIÓN</div>
            <div class="section-content">
                <div class="datos-grid">
                    <div class="dato-item"><div class="dato-label">La plantación (Proveedor)</div><div class="dato-value">{_txt(impresos.get('proveedor') or evaluacion.get('proveedor'))}</div></div>
                    <div class="dato-item"><div class="dato-label">Código Plantación</div><div class="dato-value">{_txt(impresos.get('codigo_plantacion') or evaluacion.get('codigo_plantacion'))}</div></div>
                    <div class="dato-item"><div class="dato-label">Finca</div><div class="dato-value">{_txt(impresos.get('finca') or evaluacion.get('finca'))}</div></div>
                    <div class="dato-item"><div class="dato-label">Cultivo</div><div class="dato-value">{_txt(impresos.get('cultivo') or evaluacion.get('cultivo'))}</div></div>
                    <div class="dato-item"><div class="dato-label">Variedad</div><div class="dato-value">{_txt(impresos.get('variedad') or evaluacion.get('variedad'))}</div></div>
                    <div class="dato-item"><div class="dato-label">Superficie</div><div class="dato-value">{_txt(impresos.get('superficie') if impresos.get('superficie') not in (None, '') else evaluacion.get('superficie'))}</div></div>
                </div>
                <div style="margin-top:8px;">
                    <div class="dato-label">Comentarios</div>
                    <div class="dato-value" style="white-space:pre-wrap;">{_txt(impresos.get('comentarios'))}</div>
                </div>
            </div>
        </div>

        <div class="section">
            <div class="section-title">1 · ANÁLISIS DE SUELO</div>
            <div class="section-content">
                <div class="question-row"><span class="question-text">¿Se ha archivado la hoja de los resultados de análisis con este impreso?</span><div class="answer">{_fmt_sn(analisis.get('hoja_archivada'))}</div></div>
                <div class="question-row"><span class="question-text">¿Los paquetes/envases de semillas están archivados?</span><div class="answer">{_fmt_sn(analisis.get('envases_archivados'))}</div></div>
                <div class="question-row"><span class="question-text">Medidas tomadas como consecuencia de los resultados de los análisis</span><div class="answer" style="white-space:pre-wrap;">{_txt(analisis.get('medidas_tomadas'))}</div></div>
                <div class="question-row"><span class="question-text">Este lote en el momento de entrega estaba libre de síntomas de:</span><div class="answer">{_fmt_sintomas(analisis.get('libre_sintomas'))}</div></div>
            </div>
        </div>

        <div class="section">
            <div class="section-title">2 · PASOS PRECAMPAÑA DESINFECCIÓN</div>
            <div class="section-content">
                <div class="question-row"><span class="question-text">Observaciones de interés</span><div class="answer" style="white-space:pre-wrap;">{_txt(pasos.get('observaciones'))}</div></div>
            </div>
        </div>

        <div class="section">
            <div class="section-title">3 · CALIBRACIÓN Y MANTENIMIENTO APARATOS MEDICIÓN FITO</div>
            <div class="section-content">
                <div class="datos-grid-2">
                    <div class="dato-item"><div class="dato-label">Vaso</div><div class="dato-value">{_txt(calibracion.get('vaso'))}</div></div>
                    <div class="dato-item"><div class="dato-label">Peso</div><div class="dato-value">{_txt(calibracion.get('peso'))}</div></div>
                </div>
            </div>
        </div>

        <div class="section">
            <div class="section-title">4 · CALIDAD DE CEPELLONES</div>
            <div class="section-content">
                <div class="question-row"><span class="question-text">Nº de referencia de lote de cepellones</span><div class="answer">{_txt(cepellones.get('numero_lote'))}</div></div>
                <div class="question-row"><span class="question-text">Anexo adjunto</span><div class="answer">{(lambda a: f'<span style="color:#1976d2;">📎 {a.get("filename","anexo")}</span> <span style="color:#888; font-size:0.85em;">({a.get("content_type","")})</span>' if isinstance(a, dict) and a.get('filename') else '<span style="color:#888;">—</span>')(cepellones.get('anexo'))}</div></div>
                {(lambda a: (
                    (lambda local: (
                        f'<div style="margin:10px 0; padding:8px; border:1px solid #ccc; border-radius:6px; background:#fafafa; text-align:center;">'
                        f'<img src="file://{local}" style="max-width:100%; max-height:380px; border:1px solid #d4d4d4; border-radius:4px; box-shadow:0 2px 6px rgba(0,0,0,0.08);" alt="Anexo adjunto" />'
                        f'<div style="font-size:8.5pt; color:#666; margin-top:6px; font-style:italic;">{a.get("filename", "Anexo")}</div>'
                        f'</div>'
                    ) if local and os.path.exists(local) else '')(
                        a.get('url', '').replace('/api/uploads/', '/app/uploads/') if a.get('url', '').startswith('/api/uploads/') else ''
                    )
                ) if isinstance(a, dict) and a.get('content_type', '').startswith('image/') else '')(cepellones.get('anexo') or {})}
                <div class="question-row"><span class="question-text">¿Los paquetes/envases de semillas están archivados con este impreso?</span><div class="answer">{_fmt_sn(cepellones.get('envases_archivados'))}</div></div>
                <div class="question-row"><span class="question-text">¿El semillero ha suministrado un certificado de sanidad vegetal?</span><div class="answer">{_fmt_sn(cepellones.get('certificado_sanidad'))}</div></div>
                <div class="question-row"><span class="question-text">Si existe el certificado de sanidad, ¿está archivado con este impreso?</span><div class="answer">{_fmt_sn(cepellones.get('certificado_archivado'))}</div></div>
                <div class="question-row"><span class="question-text">Este lote en el momento de entrega estaba libre de síntomas de:</span><div class="answer">{_fmt_sintomas(cepellones.get('libre_sintomas'))}</div></div>
            </div>
        </div>

        <div class="section">
            <div class="section-title">5 · INSPECCIÓN MAQUINARIA</div>
            <div class="section-content">
                <div class="datos-grid">
                    <div class="dato-item"><div class="dato-label">Tipo</div><div class="dato-value">{_txt(maquinaria.get('tipo'))}</div></div>
                    <div class="dato-item"><div class="dato-label">Modelo</div><div class="dato-value">{_txt(maquinaria.get('modelo'))}</div></div>
                    <div class="dato-item"><div class="dato-label">Nº de serie</div><div class="dato-value">{_txt(maquinaria.get('numero_serie'))}</div></div>
                </div>
                <div class="question-row"><span class="question-text">¿Se ha realizado la limpieza de los filtros?</span><div class="answer">{_fmt_sn(maquinaria.get('limpieza_filtros'))}</div></div>
                <div class="question-row"><span class="question-text">¿Se ha comprobado el estado de la manguera?</span><div class="answer">{_fmt_sn(maquinaria.get('estado_manguera'))}</div></div>
                <div class="question-row"><span class="question-text">¿Se han cambiado los diafragmas?</span><div class="answer">{_fmt_sn(maquinaria.get('diafragmas_cambiados'))}</div></div>
                <div class="question-row"><span class="question-text">¿Se han revisado todas las conexiones?</span><div class="answer">{_fmt_sn(maquinaria.get('conexiones_revisadas'))}</div></div>
            </div>
        </div>

        <div class="section">
            <div class="section-title">6 · OBSERVACIONES GENERALES</div>
            <div class="section-content">
                <div class="dato-value" style="white-space:pre-wrap;">{_txt(impresos.get('observaciones_generales'))}</div>
            </div>
        </div>
        """
    
    # ========================================================================
    # PÁGINAS DE VISITAS - Una página por cada visita
    # ========================================================================
    for idx, visita in enumerate(visitas, 1):
        page_num = 1 + idx
        html_content += f"""
        <div class="page-break" id="visita-{idx}"></div>
        <div class="header">
            <h1>FRUVECO</h1>
            <h2>REGISTRO DE VISITA</h2>
            <h3>Visita {idx} de {len(visitas)}</h3>
        </div>
        
        <div class="visita-header">
            <h3>VISITA #{visita.get('numero_visita') or idx} · {format_fecha(visita.get('fecha_visita'))}</h3>
        </div>
        
        <div class="section">
            <div class="section-title section-title-blue">DATOS DE LA VISITA</div>
            <div class="section-content">
                <div class="datos-grid">
                    <div class="dato-item">
                        <div class="dato-label">Fecha de Visita</div>
                        <div class="dato-value">{format_fecha(visita.get('fecha_visita'))}</div>
                    </div>
                    <div class="dato-item">
                        <div class="dato-label">Objetivo</div>
                        <div class="dato-value">{visita.get('objetivo', '—')}</div>
                    </div>
                    <div class="dato-item">
                        <div class="dato-label">Código Plantación</div>
                        <div class="dato-value">{visita.get('codigo_plantacion', '—')}</div>
                    </div>
                    <div class="dato-item">
                        <div class="dato-label">Proveedor</div>
                        <div class="dato-value">{visita.get('proveedor', '—')}</div>
                    </div>
                    <div class="dato-item">
                        <div class="dato-label">Cultivo</div>
                        <div class="dato-value">{visita.get('cultivo', '—')}</div>
                    </div>
                    <div class="dato-item">
                        <div class="dato-label">Variedad</div>
                        <div class="dato-value">{visita.get('variedad', '—')}</div>
                    </div>
                    <div class="dato-item">
                        <div class="dato-label">Finca</div>
                        <div class="dato-value">{visita.get('finca', '—')}</div>
                    </div>
                    <div class="dato-item">
                        <div class="dato-label">Campaña</div>
                        <div class="dato-value">{visita.get('campana', '—')}</div>
                    </div>
                    <div class="dato-item">
                        <div class="dato-label">Realizado</div>
                        <div class="dato-value">Sí</div>
                    </div>
                </div>
            </div>
        </div>
        
        <div class="section">
            <div class="section-title section-title-blue">OBSERVACIONES</div>
            <div class="section-content">
                <p>{visita.get('observaciones') or 'Sin observaciones registradas.'}</p>
            </div>
        </div>
        """
        
        # Cuestionario de Plagas y Enfermedades si existe.
        # NOTA: los valores son tri-estado 0/1/2 (Sin presencia / Baja / Alta).
        # El `if value:` original saltaba 0 → la sección salía vacía. Aquí
        # comprobamos `is not None` para incluir el 0.
        cuestionario_plagas = visita.get('cuestionario_plagas') or {}
        if isinstance(cuestionario_plagas, dict) and cuestionario_plagas:
            # Etiqueta legible para cada valor.
            def _label_presencia(v):
                try:
                    n = int(v)
                except (TypeError, ValueError):
                    return str(v) if v is not None else '—'
                if n == 0:
                    return '<span style="color:#2e7d32; font-weight:600;">0 · Sin presencia</span>'
                if n == 1:
                    return '<span style="color:#ed6c02; font-weight:600;">1 · Presencia baja</span>'
                if n == 2:
                    return '<span style="color:#c62828; font-weight:600;">2 · Presencia alta</span>'
                return str(v)
            
            html_content += """
        <div class="section">
            <div class="section-title section-title-blue">CUESTIONARIO PLAGAS Y ENFERMEDADES</div>
            <div class="section-content">
                <div style="font-size: 8.5pt; color: #666; margin-bottom: 6px;">
                    Escala: <strong>0</strong> Sin presencia · <strong>1</strong> Presencia baja · <strong>2</strong> Presencia alta
                </div>
                <table style="width:100%; border-collapse:collapse; font-size: 9.5pt;">
                    <tbody>
            """
            # Renderizamos en filas de 2 columnas (igual que la UI).
            items = list(cuestionario_plagas.items())
            for i in range(0, len(items), 2):
                pair = items[i:i + 2]
                html_content += "<tr>"
                for key, value in pair:
                    label = key.replace('_', ' ').title()
                    html_content += f"""
                        <td style="padding:6px 10px; border:1px solid #e0e0e0; width:50%;">
                            <div style="display:flex; justify-content:space-between; align-items:center; gap:8px;">
                                <span style="font-weight:600; color:#333;">{label}</span>
                                <span>{_label_presencia(value)}</span>
                            </div>
                        </td>
                    """
                if len(pair) == 1:
                    html_content += '<td style="border:1px solid #e0e0e0;"></td>'
                html_content += "</tr>"
            html_content += """
                    </tbody>
                </table>
            </div>
        </div>
            """
    
    # ========================================================================
    # PÁGINAS DE TRATAMIENTOS - Una página por cada tratamiento
    # ========================================================================
    for idx, tratamiento in enumerate(tratamientos, 1):
        page_num = 1 + len(visitas) + idx
        html_content += f"""
        <div class="page-break" id="tratamiento-{idx}"></div>
        <div class="header">
            <h1>FRUVECO</h1>
            <h2>REGISTRO DE TRATAMIENTO</h2>
            <h3>Tratamiento {idx} de {len(tratamientos)}</h3>
        </div>
        
        <div class="tratamiento-header">
            <h3>TRATAMIENTO #{idx} - {format_fecha(tratamiento.get('fecha_tratamiento'))}</h3>
        </div>
        
        <div class="section">
            <div class="section-title section-title-orange">DATOS DEL TRATAMIENTO</div>
            <div class="section-content">
                <div class="datos-grid">
                    <div class="dato-item">
                        <div class="dato-label">Fecha Tratamiento</div>
                        <div class="dato-value">{format_fecha(tratamiento.get('fecha_tratamiento')) or '—'}</div>
                    </div>
                    <div class="dato-item">
                        <div class="dato-label">Fecha Aplicación</div>
                        <div class="dato-value">{format_fecha(tratamiento.get('fecha_aplicacion')) or '—'}</div>
                    </div>
                    <div class="dato-item">
                        <div class="dato-label">Tipo</div>
                        <div class="dato-value">{((tratamiento.get('tipo_tratamiento') or tratamiento.get('tipo') or '—') + ((' — ' + tratamiento.get('subtipo')) if tratamiento.get('subtipo') else ''))}</div>
                    </div>
                    <div class="dato-item">
                        <div class="dato-label">Aplicador</div>
                        <div class="dato-value">{
                            (lambda ap: (f"{ap.get('nombre','').strip()} {ap.get('apellidos','').strip()}".strip() or None) if ap else None)(tratamiento.get('aplicador_completo'))
                            or tratamiento.get('aplicador_nombre') or '—'
                        }</div>
                    </div>
                    <div class="dato-item">
                        <div class="dato-label">Máquina</div>
                        <div class="dato-value">{
                            (tratamiento.get('maquina_completa') or {}).get('nombre')
                            or tratamiento.get('maquina_nombre') or '—'
                        }</div>
                    </div>
                    <div class="dato-item">
                        <div class="dato-label">Campaña</div>
                        <div class="dato-value">{tratamiento.get('campana') or '—'}</div>
                    </div>
                    <div class="dato-item">
                        <div class="dato-label">Realizado</div>
                        <div class="dato-value">Sí</div>
                    </div>
                    <div class="dato-item">
                        <div class="dato-label">Producto</div>
                        <div class="dato-value">{tratamiento.get('producto_fitosanitario_nombre') or '—'}</div>
                    </div>
                    <div class="dato-item">
                        <div class="dato-label">Dosis</div>
                        <div class="dato-value">{(f"{tratamiento.get('producto_fitosanitario_dosis')} {tratamiento.get('producto_fitosanitario_unidad') or ''}".strip()) if tratamiento.get('producto_fitosanitario_dosis') else '—'}</div>
                    </div>
                    <div class="dato-item">
                        <div class="dato-label">Superficie a Tratar</div>
                        <div class="dato-value">{(f"{tratamiento.get('superficie_aplicacion')} ha") if tratamiento.get('superficie_aplicacion') else '—'}</div>
                    </div>
                    <div class="dato-item">
                        <div class="dato-label">Caldo Recomendado</div>
                        <div class="dato-value">{(f"{tratamiento.get('caldo_superficie')} L/ha") if tratamiento.get('caldo_superficie') else '—'}</div>
                    </div>
                    <div class="dato-item">
                        <div class="dato-label">Nº Registro Producto</div>
                        <div class="dato-value">{tratamiento.get('producto_fitosanitario_num_registro') or '—'}</div>
                    </div>
                    <div class="dato-item">
                        <div class="dato-label">Plaga a Controlar</div>
                        <div class="dato-value">{tratamiento.get('plaga_a_controlar') or '—'}</div>
                    </div>
                </div>
            </div>
        </div>
        
        <div class="section">
            <div class="section-title section-title-orange">DESCRIPCIÓN / MOTIVO</div>
            <div class="section-content">
                <p>{tratamiento.get('descripcion') or tratamiento.get('motivo') or 'Sin descripción registrada.'}</p>
            </div>
        </div>
        
        <div class="section">
            <div class="section-title section-title-orange">OBSERVACIONES</div>
            <div class="section-content">
                <p>{tratamiento.get('observaciones') or 'Sin observaciones registradas.'}</p>
            </div>
        </div>
        """
        
        # Productos utilizados si existen
        productos = tratamiento.get('productos', [])
        if productos:
            html_content += """
        <div class="section">
            <div class="section-title section-title-orange">PRODUCTOS UTILIZADOS</div>
            <div class="section-content">
                <table class="data-table">
                    <thead>
                        <tr>
                            <th>Producto</th>
                            <th>Dosis</th>
                            <th>Cantidad</th>
                        </tr>
                    </thead>
                    <tbody>
            """
            for prod in productos:
                html_content += f"""
                        <tr>
                            <td>{prod.get('nombre', '—')}</td>
                            <td>{prod.get('dosis', '—')}</td>
                            <td>{prod.get('cantidad', '—')}</td>
                        </tr>
                """
            html_content += """
                    </tbody>
                </table>
            </div>
        </div>
            """
    
    # ========================================================================
    # PÁGINA FINAL — FICHA DEL APLICADOR Y MAQUINARIA (consolidada)
    # Mostramos los aplicadores y máquinas únicos usados en todos los
    # tratamientos del cuaderno de campo, una sola vez al final del documento.
    # ========================================================================
    aplicadores_unicos = {}
    maquinas_unicas = {}
    
    # PRIMERA PASADA: recoger fichas completas (con _id) para tener el set
    # canónico de nombres que ya están representados.
    for trat in tratamientos:
        ap = trat.get("aplicador_completo")
        if ap and ap.get("_id"):
            aplicadores_unicos.setdefault(str(ap["_id"]), ap)
        mq = trat.get("maquina_completa")
        if mq and mq.get("_id"):
            maquinas_unicas.setdefault(str(mq["_id"]), mq)
    
    # Nombres ya cubiertos por fichas completas (normalizado a lowercase para
    # evitar duplicados por mayúsculas/tildes).
    def _norm(s):
        return (str(s or '').strip().lower())
    nombres_ap_cubiertos = {
        _norm(f"{ap.get('nombre','')} {ap.get('apellidos','')}")
        for ap in aplicadores_unicos.values()
    }
    nombres_mq_cubiertos = {_norm(mq.get('nombre')) for mq in maquinas_unicas.values()}
    
    # SEGUNDA PASADA: añadir sólo minimal-by-name si el nombre NO coincide con
    # ninguna ficha completa ya presente. Así evitamos que un tratamiento con
    # solo `aplicador_nombre` (texto libre) duplique al mismo técnico que ya
    # está registrado con su ficha vía `tecnico_aplicador_id`.
    for trat in tratamientos:
        nombre_ap = trat.get("aplicador_nombre")
        if nombre_ap and _norm(nombre_ap) not in nombres_ap_cubiertos:
            key = f"name:{_norm(nombre_ap)}"
            aplicadores_unicos.setdefault(key, {"nombre": nombre_ap, "_minimal": True})
        nombre_mq = trat.get("maquina_nombre")
        if nombre_mq and _norm(nombre_mq) not in nombres_mq_cubiertos:
            key = f"name:{_norm(nombre_mq)}"
            maquinas_unicas.setdefault(key, {"nombre": nombre_mq, "_minimal": True})
    
    if aplicadores_unicos or maquinas_unicas:
        html_content += """
        <div class="page-break"></div>
        <div class="header">
            <h1>FRUVECO</h1>
            <h2>FICHA DEL APLICADOR Y MAQUINARIA</h2>
            <h3>Resumen de aplicadores y maquinaria utilizada</h3>
        </div>
        """
        
        # Aplicadores
        if aplicadores_unicos:
            html_content += """
        <div class="section">
            <div class="section-title section-title-purple">TÉCNICOS APLICADORES</div>
            <div class="section-content">
            """
            for ap in aplicadores_unicos.values():
                nombre_completo = f"{ap.get('nombre', '')} {ap.get('apellidos', '')}".strip() or "—"
                img_path = ap.get('imagen_certificado_path', '')
                img_url = ap.get('imagen_certificado_url', '')
                if img_path:
                    img_certificado = f"file://{img_path}"
                elif img_url and img_url.startswith('/api/uploads/'):
                    local = img_url.replace('/api/uploads/', '/app/uploads/')
                    import os as _os
                    img_certificado = f"file://{local}" if _os.path.exists(local) else ''
                else:
                    img_certificado = ''
                html_content += f"""
                <div style="border:1px solid #d4cfeb; border-radius:8px; padding:12px; margin-bottom:14px; background:#faf9ff; page-break-inside:avoid;">
                    <div style="display:grid; grid-template-columns:1fr 1fr; gap:8px 16px;">
                        <div class="dato-item"><div class="dato-label">Nombre Completo</div><div class="dato-value" style="font-weight:bold; font-size:11pt;">{nombre_completo}</div></div>
                        <div class="dato-item"><div class="dato-label">DNI/NIF</div><div class="dato-value">{ap.get('dni', '—') or '—'}</div></div>
                        <div class="dato-item"><div class="dato-label">Nivel Capacitación</div><div class="dato-value" style="font-weight:bold; color:#7b2cbf;">{ap.get('nivel_capacitacion', '—') or '—'}</div></div>
                        <div class="dato-item"><div class="dato-label">Nº Carnet</div><div class="dato-value">{ap.get('num_carnet', '—') or '—'}</div></div>
                        <div class="dato-item"><div class="dato-label">Fecha Certificación</div><div class="dato-value">{format_fecha(ap.get('fecha_certificacion'))}</div></div>
                        <div class="dato-item"><div class="dato-label">Fecha Validez</div><div class="dato-value">{format_fecha(ap.get('fecha_validez'))}</div></div>
                    </div>
                """
                if img_certificado:
                    html_content += f"""
                    <div style="margin-top:10px; text-align:center;">
                        <img src="{img_certificado}" style="max-width:100%; max-height:260px; border:1px solid #ccc; border-radius:4px;" alt="Certificado de {nombre_completo}" />
                        <div style="font-size:8.5pt; color:#777; margin-top:4px;">Certificado de capacitación</div>
                    </div>
                    """
                html_content += "</div>"
            html_content += """
            </div>
        </div>
            """
        
        # Maquinaria
        if maquinas_unicas:
            html_content += """
        <div class="section">
            <div class="section-title section-title-gray">MAQUINARIA UTILIZADA</div>
            <div class="section-content">
            """
            for mq in maquinas_unicas.values():
                img_path = mq.get('imagen_placa_ce_path', '')
                img_url = mq.get('imagen_placa_ce_url', '')
                if img_path:
                    img_placa = f"file://{img_path}"
                elif img_url and img_url.startswith('/api/uploads/'):
                    local = img_url.replace('/api/uploads/', '/app/uploads/')
                    import os as _os
                    img_placa = f"file://{local}" if _os.path.exists(local) else ''
                else:
                    img_placa = ''
                html_content += f"""
                <div style="border:1px solid #d9d9d9; border-radius:8px; padding:12px; margin-bottom:14px; background:#fafafa; page-break-inside:avoid;">
                    <div style="display:grid; grid-template-columns:1fr 1fr; gap:8px 16px;">
                        <div class="dato-item"><div class="dato-label">Nombre</div><div class="dato-value" style="font-weight:bold; font-size:11pt;">{mq.get('nombre', '—') or '—'}</div></div>
                        <div class="dato-item"><div class="dato-label">Tipo</div><div class="dato-value">{mq.get('tipo', '—') or '—'}</div></div>
                        <div class="dato-item"><div class="dato-label">Marca</div><div class="dato-value">{mq.get('marca', '—') or '—'}</div></div>
                        <div class="dato-item"><div class="dato-label">Modelo</div><div class="dato-value">{mq.get('modelo', '—') or '—'}</div></div>
                        <div class="dato-item"><div class="dato-label">Matrícula</div><div class="dato-value">{mq.get('matricula', '—') or '—'}</div></div>
                        <div class="dato-item"><div class="dato-label">Nº Serie</div><div class="dato-value">{mq.get('num_serie', '—') or '—'}</div></div>
                        <div class="dato-item"><div class="dato-label">Registro ROMA</div><div class="dato-value">{mq.get('registro_roma', '—') or '—'}</div></div>
                        <div class="dato-item"><div class="dato-label">Nº Bastidor</div><div class="dato-value">{mq.get('numero_bastidor', '—') or '—'}</div></div>
                    </div>
                """
                if img_placa:
                    html_content += f"""
                    <div style="margin-top:10px; text-align:center;">
                        <img src="{img_placa}" style="max-width:100%; max-height:200px; border:1px solid #ccc; border-radius:4px;" alt="Placa CE de {mq.get('nombre', 'Máquina')}" />
                        <div style="font-size:8.5pt; color:#777; margin-top:4px;">Placa CE (marcado CE)</div>
                    </div>
                    """
                html_content += "</div>"
            html_content += """
            </div>
        </div>
            """
    
    # Footer final
    html_content += f"""
        <div class="footer">
            <p>Documento generado automáticamente por FRUVECO - Cuaderno de Campo</p>
            <p>Fecha de generación: {datetime.now().strftime('%d/%m/%Y %H:%M')}</p>
            <p>{len(visitas)} visitas | {len(tratamientos)} tratamientos</p>
        </div>
    </body>
    </html>
    """
    
    # Generar PDF
    try:
        pdf_buffer = io.BytesIO()
        HTML(string=html_content).write_pdf(pdf_buffer)
        pdf_buffer.seek(0)
        
        filename = f"cuaderno_campo_{evaluacion.get('codigo_plantacion', 'sin_codigo')}_{evaluacion.get('campana', 'sin_campana')}.pdf"
        
        return Response(
            content=pdf_buffer.getvalue(),
            media_type="application/pdf",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"'
            }
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generando PDF: {str(e)}")
    finally:
        # Limpieza inmediata de mapas temporales — el PNG solo hace falta
        # mientras WeasyPrint lo lee vía file://. Después se puede borrar
        # para evitar que /app/uploads/evaluaciones/pdf_maps/ crezca sin
        # control con un UUID por cada exportación.
        for _tf in _pdf_temp_files:
            try:
                if _tf and os.path.exists(_tf):
                    os.remove(_tf)
            except Exception as _cleanup_err:
                print(f"[PDF] cleanup failed for {_tf}: {_cleanup_err}")


def _extract_emails_from_proveedor(proveedor: Optional[dict]) -> List[str]:
    """Extrae emails validos de un proveedor.
    Soporta ambos formatos:
      - `emails`: List[dict] con clave `valor` (nuevo formato)
      - `email`: str (legacy)
    """
    if not proveedor:
        return []
    result: List[str] = []
    raw_emails = proveedor.get("emails")
    if isinstance(raw_emails, list):
        for item in raw_emails:
            val = None
            if isinstance(item, dict):
                val = item.get("valor") or item.get("email")
            elif isinstance(item, str):
                val = item
            if val and "@" in val:
                result.append(val.strip())
    legacy = proveedor.get("email")
    if isinstance(legacy, str) and "@" in legacy and legacy not in result:
        result.append(legacy.strip())
    # Deduplicate manteniendo orden
    seen = set()
    unique = []
    for e in result:
        if e.lower() not in seen:
            seen.add(e.lower())
            unique.append(e)
    return unique


@router.get("/evaluaciones/{evaluacion_id}/email-suggestion")
async def suggest_evaluacion_email_recipients(
    evaluacion_id: str,
    current_user: dict = Depends(get_current_user),
):
    """Sugiere destinatarios de email a partir del proveedor asociado a la
    evaluacion (via su parcela). Devuelve emails encontrados y meta para que
    el frontend decida si enviar directamente o preguntar al usuario.
    """
    if not ObjectId.is_valid(evaluacion_id):
        raise HTTPException(status_code=400, detail="ID de evaluacion invalido")
    ev = await evaluaciones_collection.find_one({"_id": ObjectId(evaluacion_id)})
    if not ev:
        raise HTTPException(status_code=404, detail="Evaluacion no encontrada")

    proveedor_doc = None
    proveedor_nombre = None
    proveedor_id_str = None

    parcela_id = ev.get("parcela_id")
    if parcela_id and ObjectId.is_valid(parcela_id):
        parcela = await parcelas_collection.find_one({"_id": ObjectId(parcela_id)})
        if parcela:
            proveedor_id_str = parcela.get("proveedor_id")
            proveedor_nombre = parcela.get("proveedor")

    proveedores_col = db["proveedores"]
    # Buscar por id primero, si no existe intentar por nombre
    if proveedor_id_str and ObjectId.is_valid(proveedor_id_str):
        proveedor_doc = await proveedores_col.find_one({"_id": ObjectId(proveedor_id_str)})
    if not proveedor_doc and proveedor_nombre:
        proveedor_doc = await proveedores_col.find_one({"nombre": proveedor_nombre})

    emails = _extract_emails_from_proveedor(proveedor_doc)

    return {
        "evaluacion_id": evaluacion_id,
        "proveedor_id": str(proveedor_doc["_id"]) if proveedor_doc else None,
        "proveedor_nombre": (proveedor_doc or {}).get("nombre") if proveedor_doc else proveedor_nombre,
        "emails": emails,
        "has_email": bool(emails),
    }



@router.post("/evaluaciones/{evaluacion_id}/email")
async def send_evaluacion_by_email(
    evaluacion_id: str,
    payload: dict,
    current_user: dict = Depends(get_current_user),
):
    """Envia el PDF de la Hoja de Evaluacion por email a uno o varios destinatarios.

    Body: { recipients: ["email@..."], cc?: [...], subject?: str, message?: str }
    """
    from email_service import send_email, _is_smtp_configured, _smtp_config_from_user, get_email_template

    # Debe haber config del usuario o config global
    if not _smtp_config_from_user(current_user) and not _is_smtp_configured():
        raise HTTPException(
            status_code=503,
            detail="SMTP no configurado. Añade tus credenciales SMTP en tu perfil de usuario o pide al Admin que configure el envío global.",
        )

    recipients = (payload or {}).get("recipients") or []
    if isinstance(recipients, str):
        recipients = [recipients]
    recipients = [r for r in recipients if r and isinstance(r, str)]
    if not recipients:
        raise HTTPException(status_code=422, detail="Se requiere al menos un destinatario")

    cc = (payload or {}).get("cc") or []
    subject = (payload or {}).get("subject") or "Hoja de Evaluacion / Cuaderno de Campo"
    body_msg = (payload or {}).get("message") or ""

    # Generar el PDF reutilizando la funcion existente (misma logica que el GET)
    resp = await generate_evaluacion_pdf(evaluacion_id, current_user)
    pdf_bytes = resp.body if hasattr(resp, "body") else b""
    if not pdf_bytes:
        raise HTTPException(status_code=500, detail="No se pudo generar el PDF de la evaluacion")

    # Meta de la evaluacion para nombre de archivo + cuerpo del email
    evaluacion = await evaluaciones_collection.find_one({"_id": ObjectId(evaluacion_id)})
    codigo = (evaluacion or {}).get("codigo_plantacion", "sin_codigo")
    campana = (evaluacion or {}).get("campana", "sin_campana")
    filename = f"cuaderno_campo_{codigo}_{campana}.pdf"

    html_content = get_email_template(
        title="Hoja de Evaluacion / Cuaderno de Campo",
        content=(
            f"<p>Se adjunta la hoja de evaluacion correspondiente a:</p>"
            f"<ul>"
            f"<li><b>Codigo Plantacion:</b> {codigo}</li>"
            f"<li><b>Campana:</b> {campana}</li>"
            f"</ul>"
            + (f"<p style='margin-top:16px'>{body_msg}</p>" if body_msg else "")
            + "<p style='margin-top:16px'>Un saludo,<br/>Equipo FRUVECO</p>"
        ),
    )

    result = await send_email(
        recipient_email=recipients,
        subject=subject,
        html_content=html_content,
        cc=cc or None,
        attachments=[{"filename": filename, "content": pdf_bytes, "content_type": "application/pdf"}],
        user=current_user,
    )

    if result.get("status") == "error":
        raise HTTPException(status_code=500, detail=result.get("message", "Fallo al enviar email"))

    return {
        "success": True,
        "recipients": recipients,
        "cc": cc,
        "filename": filename,
        "status": result.get("status"),
        "message": result.get("message"),
    }




@router.get("/evaluaciones/export/excel")
async def export_evaluaciones_excel(
    current_user: dict = Depends(get_current_user)
):
    """Export evaluaciones to Excel"""
    from fastapi.responses import StreamingResponse
    import io as _io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    evaluaciones_list = await evaluaciones_collection.find({}).sort("created_at", -1).to_list(5000)
    evaluaciones_raw = serialize_docs(evaluaciones_list)

    wb = Workbook()
    ws = wb.active
    ws.title = "Evaluaciones"

    headers = ["Titulo", "Parcela", "Cultivo", "Proveedor", "Campana", "Estado", "Puntuacion", "Fecha", "Evaluador"]
    header_fill = PatternFill(start_color="4A148C", end_color="4A148C", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for idx, e in enumerate(evaluaciones_raw, 2):
        row = [
            e.get("titulo", ""), e.get("parcela_codigo", ""), e.get("cultivo", ""),
            e.get("proveedor", ""), e.get("campana", ""), e.get("estado", ""),
            e.get("puntuacion_total", 0),
            e.get("created_at", "")[:10] if e.get("created_at") else "",
            e.get("evaluador", "")
        ]
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=idx, column=col, value=val)
            cell.border = thin_border

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 18

    output = _io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"evaluaciones_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename={filename}"})

