"""
RRHH - Gestión de Prenóminas
"""
from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from typing import Optional
from datetime import datetime
from bson import ObjectId
import io

router = APIRouter(prefix="/api/rrhh", tags=["RRHH - Prenóminas"])

db = None

def set_database(database):
    global db
    db = database

def get_db():
    if db is None:
        raise HTTPException(status_code=500, detail="Database not initialized")
    return db


@router.get("/prenominas")
async def get_prenominas(
    empleado_id: Optional[str] = None,
    mes: Optional[int] = None,
    ano: Optional[int] = None,
    estado: Optional[str] = None
):
    """Obtener prenóminas"""
    database = get_db()
    
    query = {}
    if empleado_id:
        query["empleado_id"] = empleado_id
    if mes:
        query["periodo_mes"] = mes
    if ano:
        query["periodo_ano"] = ano
    if estado:
        query["estado"] = estado
    
    prenominas = []
    cursor = database.prenominas.find(query).sort([("periodo_ano", -1), ("periodo_mes", -1)])
    
    async for p in cursor:
        p["_id"] = str(p["_id"])
        emp = await database.empleados.find_one({"_id": ObjectId(p["empleado_id"])})
        if emp:
            p["empleado_nombre"] = f"{emp.get('nombre', '')} {emp.get('apellidos', '')}"
            p["empleado_dni"] = emp.get("dni_nie")
        prenominas.append(p)
    
    return {"success": True, "prenominas": prenominas, "total": len(prenominas)}


@router.post("/prenominas/calcular")
async def calcular_prenomina(data: dict):
    """Calcular prenómina de un empleado para un periodo"""
    database = get_db()
    
    empleado_id = data.get("empleado_id")
    mes = data.get("mes")
    ano = data.get("ano")
    
    if not all([empleado_id, mes, ano]):
        raise HTTPException(status_code=400, detail="Se requiere empleado_id, mes y año")
    
    empleado = await database.empleados.find_one({"_id": ObjectId(empleado_id)})
    if not empleado:
        raise HTTPException(status_code=404, detail="Empleado no encontrado")
    
    # Calcular rango de fechas
    fecha_desde = f"{ano}-{str(mes).zfill(2)}-01"
    if mes == 12:
        fecha_hasta = f"{ano + 1}-01-01"
    else:
        fecha_hasta = f"{ano}-{str(mes + 1).zfill(2)}-01"
    
    # Obtener fichajes del periodo
    fichajes = []
    cursor = database.fichajes.find({
        "empleado_id": empleado_id,
        "fecha": {"$gte": fecha_desde, "$lt": fecha_hasta}
    }).sort([("fecha", 1), ("hora", 1)])
    
    async for f in cursor:
        fichajes.append(f)
    
    # Calcular horas trabajadas
    horas_normales = 0.0
    horas_extra = 0.0
    horas_nocturnas = 0.0
    horas_festivos = 0.0
    dias_trabajados = set()
    
    fichajes_por_dia = {}
    for f in fichajes:
        fecha = f["fecha"]
        if fecha not in fichajes_por_dia:
            fichajes_por_dia[fecha] = []
        fichajes_por_dia[fecha].append(f)
    
    for fecha, fichajes_dia in fichajes_por_dia.items():
        entradas = [f for f in fichajes_dia if f["tipo"] == "entrada"]
        salidas = [f for f in fichajes_dia if f["tipo"] == "salida"]
        
        if entradas and salidas:
            dias_trabajados.add(fecha)
            entrada = datetime.strptime(f"{fecha} {entradas[0]['hora']}", "%Y-%m-%d %H:%M:%S")
            salida = datetime.strptime(f"{fecha} {salidas[-1]['hora']}", "%Y-%m-%d %H:%M:%S")
            horas_dia = (salida - entrada).seconds / 3600
            
            if horas_dia > 8:
                horas_normales += 8
                horas_extra += horas_dia - 8
            else:
                horas_normales += horas_dia
    
    # Calcular importes
    salario_hora = empleado.get("salario_hora", 0)
    salario_hora_extra = empleado.get("salario_hora_extra") or salario_hora * 1.25
    salario_hora_nocturna = empleado.get("salario_hora_nocturna") or salario_hora * 1.25
    salario_hora_festivo = empleado.get("salario_hora_festivo") or salario_hora * 1.5
    
    conceptos = []
    
    if horas_normales > 0:
        conceptos.append({
            "concepto": "horas_normales",
            "cantidad": round(horas_normales, 2),
            "precio_unitario": salario_hora,
            "importe": round(horas_normales * salario_hora, 2)
        })
    
    if horas_extra > 0:
        conceptos.append({
            "concepto": "horas_extra",
            "cantidad": round(horas_extra, 2),
            "precio_unitario": salario_hora_extra,
            "importe": round(horas_extra * salario_hora_extra, 2)
        })
    
    if horas_nocturnas > 0:
        conceptos.append({
            "concepto": "horas_nocturnas",
            "cantidad": round(horas_nocturnas, 2),
            "precio_unitario": salario_hora_nocturna,
            "importe": round(horas_nocturnas * salario_hora_nocturna, 2)
        })
    
    if horas_festivos > 0:
        conceptos.append({
            "concepto": "horas_festivos",
            "cantidad": round(horas_festivos, 2),
            "precio_unitario": salario_hora_festivo,
            "importe": round(horas_festivos * salario_hora_festivo, 2)
        })
    
    # Plus productividad
    prod_cursor = database.productividad.find({
        "empleado_id": empleado_id,
        "fecha": {"$gte": fecha_desde, "$lt": fecha_hasta}
    })
    
    kilos_totales = 0.0
    async for p in prod_cursor:
        kilos_totales += p.get("kilos_recogidos", 0)
    
    importe_bruto = sum(c["importe"] for c in conceptos)
    
    prenomina = {
        "empleado_id": empleado_id,
        "periodo_mes": mes,
        "periodo_ano": ano,
        "horas_normales": round(horas_normales, 2),
        "horas_extra": round(horas_extra, 2),
        "horas_nocturnas": round(horas_nocturnas, 2),
        "horas_festivos": round(horas_festivos, 2),
        "total_horas": round(horas_normales + horas_extra + horas_nocturnas + horas_festivos, 2),
        "dias_trabajados": len(dias_trabajados),
        "conceptos": conceptos,
        "importe_bruto": round(importe_bruto, 2),
        "deducciones": 0,
        "importe_neto": round(importe_bruto, 2),
        "kilos_totales": round(kilos_totales, 2),
        "estado": "borrador",
        "created_at": datetime.now(),
        "updated_at": datetime.now()
    }
    
    existing = await database.prenominas.find_one({
        "empleado_id": empleado_id,
        "periodo_mes": mes,
        "periodo_ano": ano
    })
    
    if existing:
        await database.prenominas.update_one(
            {"_id": existing["_id"]},
            {"$set": prenomina}
        )
        prenomina["_id"] = str(existing["_id"])
    else:
        result = await database.prenominas.insert_one(prenomina)
        prenomina["_id"] = str(result.inserted_id)
    
    prenomina["empleado_nombre"] = f"{empleado.get('nombre', '')} {empleado.get('apellidos', '')}"
    
    return {"success": True, "data": prenomina}


@router.post("/prenominas/calcular-todos")
async def calcular_prenominas_todos(data: dict):
    """Calcular prenóminas de todos los empleados activos para un periodo"""
    database = get_db()
    
    mes = data.get("mes")
    ano = data.get("ano")
    
    if not all([mes, ano]):
        raise HTTPException(status_code=400, detail="Se requiere mes y año")
    
    empleados = []
    cursor = database.empleados.find({"activo": True})
    async for emp in cursor:
        empleados.append(emp)
    
    resultados = []
    for emp in empleados:
        try:
            result = await calcular_prenomina({
                "empleado_id": str(emp["_id"]),
                "mes": mes,
                "ano": ano
            })
            resultados.append(result["data"])
        except Exception as e:
            resultados.append({
                "empleado_id": str(emp["_id"]),
                "empleado_nombre": f"{emp.get('nombre', '')} {emp.get('apellidos', '')}",
                "error": str(e)
            })
    
    return {"success": True, "prenominas": resultados, "total": len(resultados)}


@router.put("/prenominas/{prenomina_id}/validar")
async def validar_prenomina(prenomina_id: str, validador: dict):
    """Validar una prenómina"""
    database = get_db()
    
    result = await database.prenominas.update_one(
        {"_id": ObjectId(prenomina_id)},
        {"$set": {
            "estado": "validada",
            "validada_por": validador.get("validada_por"),
            "fecha_validacion": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "updated_at": datetime.now()
        }}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Prenómina no encontrada")
    
    return {"success": True}


@router.get("/prenominas/{prenomina_id}/excel")
async def export_prenomina_excel(prenomina_id: str):
    """Exportar prenómina individual a Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    database = get_db()
    
    prenomina = await database.prenominas.find_one({"_id": ObjectId(prenomina_id)})
    if not prenomina:
        raise HTTPException(status_code=404, detail="Prenómina no encontrada")
    
    empleado = await database.empleados.find_one({"_id": ObjectId(prenomina["empleado_id"])})
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Prenómina"
    
    # Estilos
    title_font = Font(bold=True, size=14, color="FFFFFF")
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    money_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Título
    ws.merge_cells('A1:E1')
    ws['A1'] = "PRENÓMINA"
    ws['A1'].font = title_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = Alignment(horizontal="center")
    
    # Info empleado
    ws['A3'] = "Empleado:"
    ws['B3'] = f"{empleado.get('nombre', '')} {empleado.get('apellidos', '')}" if empleado else "N/A"
    ws['A4'] = "DNI/NIE:"
    ws['B4'] = empleado.get('dni_nie', '') if empleado else "N/A"
    ws['D3'] = "Período:"
    ws['E3'] = f"{prenomina.get('periodo_mes', '')}/{prenomina.get('periodo_ano', '')}"
    
    # Resumen de horas
    ws.merge_cells('A7:E7')
    ws['A7'] = "RESUMEN DE HORAS"
    ws['A7'].font = header_font
    ws['A7'].fill = header_fill
    ws['A7'].alignment = Alignment(horizontal="center")
    
    horas_data = [
        ["Concepto", "Horas", "€/Hora", "Importe"],
        ["Horas Normales", prenomina.get('horas_normales', 0), empleado.get('salario_hora', 0) if empleado else 0, ""],
        ["Horas Extra", prenomina.get('horas_extra', 0), (empleado.get('salario_hora_extra') or (empleado.get('salario_hora', 0) * 1.25)) if empleado else 0, ""],
        ["Horas Nocturnas", prenomina.get('horas_nocturnas', 0), (empleado.get('salario_hora_nocturna') or (empleado.get('salario_hora', 0) * 1.25)) if empleado else 0, ""],
        ["Horas Festivos", prenomina.get('horas_festivos', 0), (empleado.get('salario_hora_festivo') or (empleado.get('salario_hora', 0) * 1.5)) if empleado else 0, ""]
    ]
    
    for row_idx, row_data in enumerate(horas_data, 8):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if row_idx == 8:
                cell.font = header_font
                cell.fill = header_fill
            if col_idx == 4 and row_idx > 8:
                horas = ws.cell(row=row_idx, column=2).value or 0
                precio = ws.cell(row=row_idx, column=3).value or 0
                cell.value = round(horas * precio, 2)
                cell.number_format = '#,##0.00 €'
    
    # Totales
    row = 13
    ws.merge_cells(f'A{row}:C{row}')
    ws[f'A{row}'] = "TOTAL HORAS:"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'D{row}'] = prenomina.get('total_horas', 0)
    
    row += 2
    ws.merge_cells(f'A{row}:C{row}')
    ws[f'A{row}'] = "IMPORTE BRUTO:"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'D{row}'] = prenomina.get('importe_bruto', 0)
    ws[f'D{row}'].fill = money_fill
    ws[f'D{row}'].number_format = '#,##0.00 €'
    
    row += 2
    ws.merge_cells(f'A{row}:C{row}')
    ws[f'A{row}'] = "IMPORTE NETO:"
    ws[f'A{row}'].font = Font(bold=True, size=14, color="2E7D32")
    ws[f'D{row}'] = prenomina.get('importe_neto', 0)
    ws[f'D{row}'].font = Font(bold=True, size=14, color="2E7D32")
    ws[f'D{row}'].fill = money_fill
    ws[f'D{row}'].number_format = '#,##0.00 €'
    
    # Ajustar anchos
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    apellidos = empleado.get('apellidos', 'empleado').replace(' ', '_') if empleado else 'empleado'
    filename = f"prenomina_{apellidos}_{prenomina.get('periodo_mes', '')}_{prenomina.get('periodo_ano', '')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/prenominas/{prenomina_id}/pdf")
async def export_prenomina_pdf(prenomina_id: str):
    """Exportar prenómina individual a PDF"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    
    database = get_db()
    
    prenomina = await database.prenominas.find_one({"_id": ObjectId(prenomina_id)})
    if not prenomina:
        raise HTTPException(status_code=404, detail="Prenómina no encontrada")
    
    empleado = await database.empleados.find_one({"_id": ObjectId(prenomina["empleado_id"])})
    
    output = io.BytesIO()
    pdf = SimpleDocTemplate(output, pagesize=A4, topMargin=20*mm, bottomMargin=20*mm, leftMargin=20*mm, rightMargin=20*mm)
    
    elements = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18, textColor=colors.HexColor('#2E7D32'), alignment=1, spaceAfter=5*mm)
    elements.append(Paragraph("PRENÓMINA", title_style))
    
    info_style = ParagraphStyle('Info', parent=styles['Normal'], fontSize=10, spaceAfter=2*mm)
    emp_nombre = f"{empleado.get('nombre', '')} {empleado.get('apellidos', '')}" if empleado else "N/A"
    elements.append(Paragraph(f"<b>Empleado:</b> {emp_nombre}", info_style))
    elements.append(Paragraph(f"<b>Período:</b> {prenomina.get('periodo_mes', '')}/{prenomina.get('periodo_ano', '')}", info_style))
    elements.append(Spacer(1, 8*mm))
    
    salario_hora = empleado.get('salario_hora', 0) if empleado else 0
    salario_extra = (empleado.get('salario_hora_extra') or salario_hora * 1.25) if empleado else 0
    
    table_data = [
        ["Concepto", "Horas", "€/Hora", "Importe"],
        ["Horas Normales", f"{prenomina.get('horas_normales', 0):.2f}", f"{salario_hora:.2f} €", f"{prenomina.get('horas_normales', 0) * salario_hora:.2f} €"],
        ["Horas Extra", f"{prenomina.get('horas_extra', 0):.2f}", f"{salario_extra:.2f} €", f"{prenomina.get('horas_extra', 0) * salario_extra:.2f} €"],
    ]
    
    col_widths = [50*mm, 30*mm, 35*mm, 40*mm]
    horas_table = Table(table_data, colWidths=col_widths)
    horas_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E7D32')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.gray),
    ]))
    elements.append(horas_table)
    elements.append(Spacer(1, 8*mm))
    
    resumen_data = [
        ["IMPORTE NETO:", f"{prenomina.get('importe_neto', 0):.2f} €"],
    ]
    
    resumen_table = Table(resumen_data, colWidths=[80*mm, 50*mm])
    resumen_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 14),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#2E7D32')),
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#E8F5E9')),
    ]))
    elements.append(resumen_table)
    
    pdf.build(elements)
    output.seek(0)
    
    apellidos = empleado.get('apellidos', 'empleado').replace(' ', '_') if empleado else 'empleado'
    filename = f"prenomina_{apellidos}_{prenomina.get('periodo_mes', '')}_{prenomina.get('periodo_ano', '')}.pdf"
    
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/prenominas/export")
async def export_prenominas(mes: int, ano: int):
    """Exportar prenóminas para software de nóminas (formato JSON)"""
    database = get_db()
    
    prenominas = []
    cursor = database.prenominas.find({
        "periodo_mes": mes,
        "periodo_ano": ano,
        "estado": {"$in": ["validada", "exportada"]}
    })
    
    async for p in cursor:
        emp = await database.empleados.find_one({"_id": ObjectId(p["empleado_id"])})
        if emp:
            prenominas.append({
                "codigo_empleado": emp.get("codigo"),
                "dni": emp.get("dni_nie"),
                "nombre": f"{emp.get('nombre', '')} {emp.get('apellidos', '')}",
                "horas_normales": p.get("horas_normales", 0),
                "horas_extra": p.get("horas_extra", 0),
                "total_horas": p.get("total_horas", 0),
                "importe_bruto": p.get("importe_bruto", 0),
                "importe_neto": p.get("importe_neto", 0)
            })
    
    await database.prenominas.update_many(
        {"periodo_mes": mes, "periodo_ano": ano, "estado": "validada"},
        {"$set": {"estado": "exportada", "updated_at": datetime.now()}}
    )
    
    return {"success": True, "prenominas": prenominas, "total": len(prenominas)}


@router.get("/prenominas/export/csv")
async def export_prenominas_csv(mes: int, ano: int, formato: Optional[str] = "estandar"):
    """
    Exportar prenóminas masivas a CSV para importar en software de nóminas.
    
    Formatos disponibles:
    - estandar: Formato genérico compatible con la mayoría de software
    - a3nom: Formato compatible con A3NOM
    - sage: Formato compatible con Sage
    - nominaplus: Formato compatible con NominaPlus
    """
    import csv
    
    database = get_db()
    
    # Obtener todas las prenóminas del periodo (validadas o borradores)
    prenominas = []
    cursor = database.prenominas.find({
        "periodo_mes": mes,
        "periodo_ano": ano
    }).sort("empleado_id", 1)
    
    async for p in cursor:
        emp = await database.empleados.find_one({"_id": ObjectId(p["empleado_id"])})
        if emp:
            prenominas.append({
                "prenomina": p,
                "empleado": emp
            })
    
    if not prenominas:
        raise HTTPException(status_code=404, detail=f"No hay prenóminas para {mes}/{ano}")
    
    output = io.StringIO()
    
    if formato == "a3nom":
        # Formato A3NOM
        headers = [
            "CODIGO", "DNI", "APELLIDOS", "NOMBRE", "HORAS_NORM", "HORAS_EXTRA",
            "HORAS_NOCT", "HORAS_FEST", "TOTAL_HORAS", "DIAS_TRAB", "BRUTO", "NETO"
        ]
        writer = csv.DictWriter(output, fieldnames=headers, delimiter=';', quoting=csv.QUOTE_MINIMAL)
        writer.writeheader()
        
        for item in prenominas:
            emp = item["empleado"]
            p = item["prenomina"]
            writer.writerow({
                "CODIGO": emp.get("codigo", ""),
                "DNI": emp.get("dni_nie", ""),
                "APELLIDOS": emp.get("apellidos", ""),
                "NOMBRE": emp.get("nombre", ""),
                "HORAS_NORM": f"{p.get('horas_normales', 0):.2f}".replace('.', ','),
                "HORAS_EXTRA": f"{p.get('horas_extra', 0):.2f}".replace('.', ','),
                "HORAS_NOCT": f"{p.get('horas_nocturnas', 0):.2f}".replace('.', ','),
                "HORAS_FEST": f"{p.get('horas_festivos', 0):.2f}".replace('.', ','),
                "TOTAL_HORAS": f"{p.get('total_horas', 0):.2f}".replace('.', ','),
                "DIAS_TRAB": p.get("dias_trabajados", 0),
                "BRUTO": f"{p.get('importe_bruto', 0):.2f}".replace('.', ','),
                "NETO": f"{p.get('importe_neto', 0):.2f}".replace('.', ',')
            })
    
    elif formato == "sage":
        # Formato SAGE
        headers = [
            "CodEmpleado", "NIF", "Empleado", "HorasOrdinarias", "HorasExtras",
            "HorasNocturnas", "TotalHoras", "DiasTrabajos", "ImporteBruto", "ImporteNeto"
        ]
        writer = csv.DictWriter(output, fieldnames=headers, delimiter=';', quoting=csv.QUOTE_ALL)
        writer.writeheader()
        
        for item in prenominas:
            emp = item["empleado"]
            p = item["prenomina"]
            nombre_completo = f"{emp.get('apellidos', '')}, {emp.get('nombre', '')}"
            writer.writerow({
                "CodEmpleado": emp.get("codigo", ""),
                "NIF": emp.get("dni_nie", ""),
                "Empleado": nombre_completo,
                "HorasOrdinarias": p.get("horas_normales", 0),
                "HorasExtras": p.get("horas_extra", 0),
                "HorasNocturnas": p.get("horas_nocturnas", 0),
                "TotalHoras": p.get("total_horas", 0),
                "DiasTrabajos": p.get("dias_trabajados", 0),
                "ImporteBruto": p.get("importe_bruto", 0),
                "ImporteNeto": p.get("importe_neto", 0)
            })
    
    elif formato == "nominaplus":
        # Formato NominaPlus
        headers = [
            "Empresa", "CodTrabajador", "DNI", "Apellido1", "Apellido2", "Nombre",
            "HorasNormales", "HorasExtra", "DiasLaborables", "Bruto", "Liquido"
        ]
        writer = csv.DictWriter(output, fieldnames=headers, delimiter=',', quoting=csv.QUOTE_MINIMAL)
        writer.writeheader()
        
        for item in prenominas:
            emp = item["empleado"]
            p = item["prenomina"]
            apellidos = emp.get("apellidos", "").split(" ", 1)
            apellido1 = apellidos[0] if len(apellidos) > 0 else ""
            apellido2 = apellidos[1] if len(apellidos) > 1 else ""
            writer.writerow({
                "Empresa": "001",
                "CodTrabajador": emp.get("codigo", ""),
                "DNI": emp.get("dni_nie", ""),
                "Apellido1": apellido1,
                "Apellido2": apellido2,
                "Nombre": emp.get("nombre", ""),
                "HorasNormales": p.get("horas_normales", 0),
                "HorasExtra": p.get("horas_extra", 0),
                "DiasLaborables": p.get("dias_trabajados", 0),
                "Bruto": p.get("importe_bruto", 0),
                "Liquido": p.get("importe_neto", 0)
            })
    
    else:  # formato == "estandar"
        # Formato estándar genérico
        headers = [
            "Codigo_Empleado", "DNI_NIE", "Apellidos", "Nombre", "Puesto", "Departamento",
            "Tipo_Contrato", "Fecha_Alta", "Salario_Hora", "Horas_Normales", "Horas_Extra",
            "Horas_Nocturnas", "Horas_Festivos", "Total_Horas", "Dias_Trabajados",
            "Kilos_Producidos", "Importe_Bruto", "Deducciones", "Importe_Neto",
            "Periodo_Mes", "Periodo_Ano", "Estado", "IBAN", "Numero_SS"
        ]
        writer = csv.DictWriter(output, fieldnames=headers, delimiter=',', quoting=csv.QUOTE_MINIMAL)
        writer.writeheader()
        
        for item in prenominas:
            emp = item["empleado"]
            p = item["prenomina"]
            writer.writerow({
                "Codigo_Empleado": emp.get("codigo", ""),
                "DNI_NIE": emp.get("dni_nie", ""),
                "Apellidos": emp.get("apellidos", ""),
                "Nombre": emp.get("nombre", ""),
                "Puesto": emp.get("puesto", ""),
                "Departamento": emp.get("departamento", ""),
                "Tipo_Contrato": emp.get("tipo_contrato", ""),
                "Fecha_Alta": emp.get("fecha_alta", ""),
                "Salario_Hora": emp.get("salario_hora", 0),
                "Horas_Normales": p.get("horas_normales", 0),
                "Horas_Extra": p.get("horas_extra", 0),
                "Horas_Nocturnas": p.get("horas_nocturnas", 0),
                "Horas_Festivos": p.get("horas_festivos", 0),
                "Total_Horas": p.get("total_horas", 0),
                "Dias_Trabajados": p.get("dias_trabajados", 0),
                "Kilos_Producidos": p.get("kilos_totales", 0),
                "Importe_Bruto": p.get("importe_bruto", 0),
                "Deducciones": p.get("deducciones", 0),
                "Importe_Neto": p.get("importe_neto", 0),
                "Periodo_Mes": mes,
                "Periodo_Ano": ano,
                "Estado": p.get("estado", "borrador"),
                "IBAN": emp.get("iban", ""),
                "Numero_SS": emp.get("numero_ss", "")
            })
    
    # Marcar como exportadas
    await database.prenominas.update_many(
        {"periodo_mes": mes, "periodo_ano": ano, "estado": {"$in": ["validada", "borrador"]}},
        {"$set": {"estado": "exportada", "fecha_exportacion": datetime.now(), "updated_at": datetime.now()}}
    )
    
    csv_content = output.getvalue()
    output.close()
    
    # Crear respuesta con el archivo CSV
    response_output = io.BytesIO(csv_content.encode('utf-8-sig'))  # BOM para Excel
    response_output.seek(0)
    
    meses = ["", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", 
             "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    mes_nombre = meses[mes] if 1 <= mes <= 12 else str(mes)
    filename = f"prenominas_{mes_nombre}_{ano}_{formato}.csv"
    
    return StreamingResponse(
        response_output,
        media_type="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Content-Type": "text/csv; charset=utf-8"
        }
    )


@router.get("/prenominas/export/excel-masivo")
async def export_prenominas_excel_masivo(mes: int, ano: int):
    """Exportar todas las prenóminas del periodo a un único archivo Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    database = get_db()
    
    prenominas = []
    cursor = database.prenominas.find({
        "periodo_mes": mes,
        "periodo_ano": ano
    }).sort("empleado_id", 1)
    
    async for p in cursor:
        emp = await database.empleados.find_one({"_id": ObjectId(p["empleado_id"])})
        if emp:
            prenominas.append({
                "prenomina": p,
                "empleado": emp
            })
    
    if not prenominas:
        raise HTTPException(status_code=404, detail=f"No hay prenóminas para {mes}/{ano}")
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"Prenóminas {mes}-{ano}"
    
    # Estilos
    title_font = Font(bold=True, size=16, color="FFFFFF")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    money_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    alt_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Título
    meses = ["", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", 
             "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    mes_nombre = meses[mes] if 1 <= mes <= 12 else str(mes)
    
    ws.merge_cells('A1:L1')
    ws['A1'] = f"PRENÓMINAS - {mes_nombre} {ano}"
    ws['A1'].font = title_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 30
    
    # Info resumen
    ws['A3'] = f"Total empleados: {len(prenominas)}"
    ws['A3'].font = Font(bold=True)
    total_bruto = sum(item["prenomina"].get("importe_bruto", 0) for item in prenominas)
    total_neto = sum(item["prenomina"].get("importe_neto", 0) for item in prenominas)
    ws['C3'] = f"Total Bruto: {total_bruto:,.2f} €"
    ws['C3'].font = Font(bold=True, color="2E7D32")
    ws['F3'] = f"Total Neto: {total_neto:,.2f} €"
    ws['F3'].font = Font(bold=True, color="2E7D32")
    ws['I3'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    
    # Encabezados
    headers = [
        "Código", "DNI/NIE", "Apellidos", "Nombre", "Puesto",
        "H. Normales", "H. Extra", "H. Nocturnas", "H. Festivos", "Total Horas",
        "Días Trab.", "Importe Bruto", "Deducciones", "Importe Neto", "Estado"
    ]
    
    row = 5
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws.row_dimensions[row].height = 25
    
    # Datos
    for idx, item in enumerate(prenominas):
        row = 6 + idx
        emp = item["empleado"]
        p = item["prenomina"]
        
        data = [
            emp.get("codigo", ""),
            emp.get("dni_nie", ""),
            emp.get("apellidos", ""),
            emp.get("nombre", ""),
            emp.get("puesto", ""),
            p.get("horas_normales", 0),
            p.get("horas_extra", 0),
            p.get("horas_nocturnas", 0),
            p.get("horas_festivos", 0),
            p.get("total_horas", 0),
            p.get("dias_trabajados", 0),
            p.get("importe_bruto", 0),
            p.get("deducciones", 0),
            p.get("importe_neto", 0),
            p.get("estado", "borrador").upper()
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border
            
            # Alternar colores de fila
            if idx % 2 == 1:
                cell.fill = alt_fill
            
            # Formato para columnas de dinero
            if col in [12, 13, 14]:
                cell.number_format = '#,##0.00 €'
                if col == 14:  # Importe Neto
                    cell.fill = money_fill
                    cell.font = Font(bold=True)
            
            # Formato para horas
            if col in [6, 7, 8, 9, 10]:
                cell.number_format = '#,##0.00'
            
            # Centrar números
            if col >= 6:
                cell.alignment = Alignment(horizontal="center")
    
    # Fila de totales
    total_row = 6 + len(prenominas)
    ws.merge_cells(f'A{total_row}:E{total_row}')
    ws[f'A{total_row}'] = "TOTALES"
    ws[f'A{total_row}'].font = Font(bold=True, size=12)
    ws[f'A{total_row}'].alignment = Alignment(horizontal="right")
    
    # Sumar columnas numéricas
    for col in range(6, 15):
        cell = ws.cell(row=total_row, column=col)
        start_row = 6
        end_row = total_row - 1
        col_letter = get_column_letter(col)
        cell.value = f"=SUM({col_letter}{start_row}:{col_letter}{end_row})"
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
        
        if col in [12, 13, 14]:
            cell.number_format = '#,##0.00 €'
        elif col in [6, 7, 8, 9, 10]:
            cell.number_format = '#,##0.00'
    
    # Ajustar anchos de columna
    column_widths = [12, 12, 18, 15, 15, 12, 10, 12, 12, 12, 10, 14, 12, 14, 12]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Guardar
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"prenominas_{mes_nombre}_{ano}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

