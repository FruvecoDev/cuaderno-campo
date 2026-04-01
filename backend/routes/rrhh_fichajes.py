"""
RRHH - Fichajes / Control Horario
Includes: fichajes CRUD, QR/NFC/facial fichaje, informes Excel/PDF
"""
from fastapi import APIRouter, HTTPException
from typing import List, Optional
from datetime import datetime, timedelta
from bson import ObjectId
import io

router = APIRouter(prefix="/api/rrhh", tags=["RRHH - Fichajes"])

db = None

def set_database(database):
    global db
    db = database

def get_db():
    if db is None:
        raise HTTPException(status_code=500, detail="Database not initialized")
    return db


@router.get("/fichajes")
async def get_fichajes(
    empleado_id: Optional[str] = None,
    fecha_desde: Optional[str] = None,
    fecha_hasta: Optional[str] = None,
    parcela_id: Optional[str] = None
):
    database = get_db()
    query = {}
    if empleado_id:
        query["empleado_id"] = empleado_id
    if fecha_desde:
        query["fecha"] = {"$gte": fecha_desde}
    if fecha_hasta:
        if "fecha" in query:
            query["fecha"]["$lte"] = fecha_hasta
        else:
            query["fecha"] = {"$lte": fecha_hasta}
    if parcela_id:
        query["parcela_id"] = parcela_id
    
    fichajes = []
    cursor = database.fichajes.find(query).sort([("fecha", -1), ("hora", -1)])
    async for f in cursor:
        f["_id"] = str(f["_id"])
        emp = await database.empleados.find_one({"_id": ObjectId(f["empleado_id"])})
        if emp:
            f["empleado_nombre"] = f"{emp.get('nombre', '')} {emp.get('apellidos', '')}"
        fichajes.append(f)
    return {"success": True, "fichajes": fichajes, "total": len(fichajes)}


@router.get("/fichajes/hoy")
async def get_fichajes_hoy():
    database = get_db()
    hoy = datetime.now().strftime("%Y-%m-%d")
    fichajes = []
    cursor = database.fichajes.find({"fecha": hoy}).sort("hora", -1)
    async for f in cursor:
        f["_id"] = str(f["_id"])
        emp = await database.empleados.find_one({"_id": ObjectId(f["empleado_id"])})
        if emp:
            f["empleado_nombre"] = f"{emp.get('nombre', '')} {emp.get('apellidos', '')}"
            f["empleado_foto"] = emp.get("foto_url")
        fichajes.append(f)
    
    empleados_activos = await database.empleados.count_documents({"activo": True})
    empleados_fichados = len(set([f["empleado_id"] for f in fichajes if f["tipo"] == "entrada"]))
    return {
        "success": True, "fichajes": fichajes,
        "estadisticas": {
            "empleados_activos": empleados_activos,
            "empleados_fichados": empleados_fichados,
            "pendientes_fichar": empleados_activos - empleados_fichados
        }
    }


@router.post("/fichajes")
async def create_fichaje(fichaje: dict):
    database = get_db()
    empleado = await database.empleados.find_one({"_id": ObjectId(fichaje["empleado_id"])})
    if not empleado:
        raise HTTPException(status_code=404, detail="Empleado no encontrado")
    fichaje["created_at"] = datetime.now()
    fichaje["sincronizado"] = True
    result = await database.fichajes.insert_one(fichaje)
    fichaje["_id"] = str(result.inserted_id)
    fichaje["empleado_nombre"] = f"{empleado.get('nombre', '')} {empleado.get('apellidos', '')}"
    return {"success": True, "data": fichaje}


@router.post("/fichajes/qr")
async def fichaje_por_qr(data: dict):
    database = get_db()
    qr_code = data.get("qr_code")
    tipo = data.get("tipo", "entrada")
    empleado = await database.empleados.find_one({"qr_code": qr_code, "activo": True})
    if not empleado:
        raise HTTPException(status_code=404, detail="Codigo QR no valido o empleado inactivo")
    now = datetime.now()
    fichaje = {
        "empleado_id": str(empleado["_id"]), "tipo": tipo,
        "fecha": now.strftime("%Y-%m-%d"), "hora": now.strftime("%H:%M:%S"),
        "metodo_identificacion": "qr",
        "latitud": data.get("latitud"), "longitud": data.get("longitud"),
        "ubicacion_nombre": data.get("ubicacion_nombre"),
        "parcela_id": data.get("parcela_id"),
        "offline": data.get("offline", False), "sincronizado": True, "created_at": now
    }
    result = await database.fichajes.insert_one(fichaje)
    fichaje["_id"] = str(result.inserted_id)
    fichaje["empleado_nombre"] = f"{empleado.get('nombre', '')} {empleado.get('apellidos', '')}"
    fichaje["empleado_foto"] = empleado.get("foto_url")
    return {"success": True, "data": fichaje}


@router.post("/fichajes/nfc")
async def fichaje_por_nfc(data: dict):
    database = get_db()
    nfc_id = data.get("nfc_id")
    tipo = data.get("tipo", "entrada")
    empleado = await database.empleados.find_one({"nfc_id": nfc_id, "activo": True})
    if not empleado:
        raise HTTPException(status_code=404, detail="Tarjeta NFC no registrada o empleado inactivo")
    now = datetime.now()
    fichaje = {
        "empleado_id": str(empleado["_id"]), "tipo": tipo,
        "fecha": now.strftime("%Y-%m-%d"), "hora": now.strftime("%H:%M:%S"),
        "metodo_identificacion": "nfc",
        "latitud": data.get("latitud"), "longitud": data.get("longitud"),
        "ubicacion_nombre": data.get("ubicacion_nombre"),
        "parcela_id": data.get("parcela_id"),
        "offline": data.get("offline", False), "sincronizado": True, "created_at": now
    }
    result = await database.fichajes.insert_one(fichaje)
    fichaje["_id"] = str(result.inserted_id)
    fichaje["empleado_nombre"] = f"{empleado.get('nombre', '')} {empleado.get('apellidos', '')}"
    return {"success": True, "data": fichaje}


@router.post("/fichajes/facial")
async def fichaje_por_facial(data: dict):
    database = get_db()
    empleado_id = data.get("empleado_id")
    tipo = data.get("tipo", "entrada")
    if not empleado_id:
        raise HTTPException(status_code=400, detail="Se requiere ID de empleado")
    empleado = await database.empleados.find_one({"_id": ObjectId(empleado_id), "activo": True})
    if not empleado:
        raise HTTPException(status_code=404, detail="Empleado no encontrado")
    if not empleado.get("foto_url"):
        raise HTTPException(status_code=400, detail="El empleado no tiene foto registrada")
    now = datetime.now()
    fichaje = {
        "empleado_id": empleado_id, "tipo": tipo,
        "fecha": now.strftime("%Y-%m-%d"), "hora": now.strftime("%H:%M:%S"),
        "metodo_identificacion": "facial",
        "foto_verificacion_url": data.get("foto_capturada"),
        "latitud": data.get("latitud"), "longitud": data.get("longitud"),
        "ubicacion_nombre": data.get("ubicacion_nombre"),
        "parcela_id": data.get("parcela_id"),
        "offline": data.get("offline", False), "sincronizado": True, "created_at": now
    }
    result = await database.fichajes.insert_one(fichaje)
    fichaje["_id"] = str(result.inserted_id)
    fichaje["empleado_nombre"] = f"{empleado.get('nombre', '')} {empleado.get('apellidos', '')}"
    return {"success": True, "data": fichaje}


@router.post("/fichajes/sync")
async def sync_fichajes_offline(fichajes: List[dict]):
    database = get_db()
    synced = 0
    errors = []
    for fichaje in fichajes:
        try:
            fichaje["sincronizado"] = True
            fichaje["created_at"] = datetime.now()
            await database.fichajes.insert_one(fichaje)
            synced += 1
        except Exception as e:
            errors.append({"fichaje": fichaje, "error": str(e)})
    return {"success": True, "synced": synced, "errors": errors}


@router.put("/fichajes/{fichaje_id}/validar")
async def validar_fichaje(fichaje_id: str, validador: dict):
    database = get_db()
    result = await database.fichajes.update_one(
        {"_id": ObjectId(fichaje_id)},
        {"$set": {
            "validado": True,
            "validado_por": validador.get("validado_por"),
            "validado_fecha": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Fichaje no encontrado")
    return {"success": True}


# Helper: agrupar fichajes por dia y calcular horas
def _agrupar_fichajes_por_dia(fichajes):
    dias = {}
    for f in fichajes:
        fecha = f["fecha"]
        if fecha not in dias:
            dias[fecha] = {"entradas": [], "salidas": []}
        if f["tipo"] == "entrada":
            dias[fecha]["entradas"].append(f["hora"])
        else:
            dias[fecha]["salidas"].append(f["hora"])
    return dias


def _calcular_horas(entrada_str, salida_str):
    h_e = datetime.strptime(entrada_str, "%H:%M:%S") if len(entrada_str) > 5 else datetime.strptime(entrada_str, "%H:%M")
    h_s = datetime.strptime(salida_str, "%H:%M:%S") if len(salida_str) > 5 else datetime.strptime(salida_str, "%H:%M")
    diff = h_s - h_e
    secs = diff.total_seconds()
    return int(secs // 3600), int((secs % 3600) // 60)


@router.get("/fichajes/informe")
async def get_informe_control_horario(empleado_id: str, fecha_desde: str, fecha_hasta: str):
    database = get_db()
    empleado = await database.empleados.find_one({"_id": ObjectId(empleado_id)})
    if not empleado:
        raise HTTPException(status_code=404, detail="Empleado no encontrado")
    
    fichajes_raw = []
    cursor = database.fichajes.find({"empleado_id": empleado_id, "fecha": {"$gte": fecha_desde, "$lte": fecha_hasta}}).sort([("fecha", 1), ("hora", 1)])
    async for f in cursor:
        fichajes_raw.append(f)
    
    dias_trabajados = _agrupar_fichajes_por_dia(fichajes_raw)
    
    registros = []
    total_horas = 0
    total_minutos = 0
    dias_completos = 0
    dias_con_ausencia = 0
    
    fecha_inicio = datetime.strptime(fecha_desde, "%Y-%m-%d").date()
    fecha_fin = datetime.strptime(fecha_hasta, "%Y-%m-%d").date()
    current_date = fecha_inicio
    
    while current_date <= fecha_fin:
        if current_date.weekday() < 5:
            fecha_str = current_date.strftime("%Y-%m-%d")
            if fecha_str in dias_trabajados:
                dia_data = dias_trabajados[fecha_str]
                entrada = dia_data["entradas"][0] if dia_data["entradas"] else None
                salida = dia_data["salidas"][-1] if dia_data["salidas"] else None
                horas_dia = minutos_dia = 0
                if entrada and salida:
                    horas_dia, minutos_dia = _calcular_horas(entrada, salida)
                    total_horas += horas_dia
                    total_minutos += minutos_dia
                    if horas_dia >= 8:
                        dias_completos += 1
                registros.append({
                    "fecha": fecha_str, "dia_semana": current_date.strftime("%A"),
                    "entrada": entrada, "salida": salida,
                    "horas": horas_dia, "minutos": minutos_dia,
                    "horas_decimal": round(horas_dia + minutos_dia / 60, 2),
                    "completo": horas_dia >= 8, "ausencia": False
                })
            else:
                dias_con_ausencia += 1
                registros.append({
                    "fecha": fecha_str, "dia_semana": current_date.strftime("%A"),
                    "entrada": None, "salida": None,
                    "horas": 0, "minutos": 0, "horas_decimal": 0,
                    "completo": False, "ausencia": True
                })
        current_date += timedelta(days=1)
    
    total_horas += total_minutos // 60
    total_minutos = total_minutos % 60
    
    return {
        "success": True,
        "empleado": {
            "id": str(empleado["_id"]),
            "nombre": empleado.get("nombre", ""), "apellidos": empleado.get("apellidos", ""),
            "dni": empleado.get("dni_nie", ""), "puesto": empleado.get("puesto", "")
        },
        "periodo": {"desde": fecha_desde, "hasta": fecha_hasta},
        "registros": registros,
        "resumen": {
            "total_horas": total_horas, "total_minutos": total_minutos,
            "total_horas_decimal": round(total_horas + total_minutos / 60, 2),
            "dias_trabajados": len([r for r in registros if not r["ausencia"]]),
            "dias_completos": dias_completos, "dias_con_ausencia": dias_con_ausencia,
            "horas_esperadas": len([r for r in registros if not r["ausencia"]]) * 8,
            "diferencia_horas": round((total_horas + total_minutos / 60) - (len([r for r in registros if not r["ausencia"]]) * 8), 2)
        }
    }


@router.get("/fichajes/informe/excel")
async def export_informe_control_horario_excel(empleado_id: str, fecha_desde: str, fecha_hasta: str):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from fastapi.responses import StreamingResponse
    
    database = get_db()
    empleado = await database.empleados.find_one({"_id": ObjectId(empleado_id)})
    if not empleado:
        raise HTTPException(status_code=404, detail="Empleado no encontrado")
    
    fichajes_raw = []
    cursor = database.fichajes.find({"empleado_id": empleado_id, "fecha": {"$gte": fecha_desde, "$lte": fecha_hasta}}).sort([("fecha", 1), ("hora", 1)])
    async for f in cursor:
        fichajes_raw.append(f)
    
    dias_trabajados = _agrupar_fichajes_por_dia(fichajes_raw)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Control Horario"
    
    title_font = Font(bold=True, size=14, color="FFFFFF")
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    warning_fill = PatternFill(start_color="FFEB3B", end_color="FFEB3B", fill_type="solid")
    error_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
    ok_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    ws.merge_cells('A1:G1')
    ws['A1'] = "INFORME DE CONTROL HORARIO"
    ws['A1'].font = title_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = Alignment(horizontal="center")
    
    ws['A3'] = "Empleado:"
    ws['B3'] = f"{empleado.get('nombre', '')} {empleado.get('apellidos', '')}"
    ws['A4'] = "DNI/NIE:"
    ws['B4'] = empleado.get('dni_nie', '')
    ws['A5'] = "Puesto:"
    ws['B5'] = empleado.get('puesto', '')
    ws['D3'] = "Periodo:"
    ws['E3'] = f"{fecha_desde} a {fecha_hasta}"
    
    headers = ["Fecha", "Dia", "Entrada", "Salida", "Horas", "Min", "Total Horas", "Estado"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
    
    dias_semana_es = {'Monday': 'Lunes', 'Tuesday': 'Martes', 'Wednesday': 'Miercoles', 'Thursday': 'Jueves', 'Friday': 'Viernes', 'Saturday': 'Sabado', 'Sunday': 'Domingo'}
    
    row = 8
    total_horas = total_minutos = dias_ausencia = dias_incompletos = 0
    fecha_inicio = datetime.strptime(fecha_desde, "%Y-%m-%d").date()
    fecha_fin = datetime.strptime(fecha_hasta, "%Y-%m-%d").date()
    current_date = fecha_inicio
    
    while current_date <= fecha_fin:
        if current_date.weekday() < 5:
            fecha_str = current_date.strftime("%Y-%m-%d")
            dia_semana = dias_semana_es.get(current_date.strftime("%A"), current_date.strftime("%A"))
            
            if fecha_str in dias_trabajados:
                dia_data = dias_trabajados[fecha_str]
                entrada = dia_data["entradas"][0] if dia_data["entradas"] else "-"
                salida = dia_data["salidas"][-1] if dia_data["salidas"] else "-"
                horas = minutos = 0
                if entrada != "-" and salida != "-":
                    horas, minutos = _calcular_horas(entrada, salida)
                    total_horas += horas
                    total_minutos += minutos
                estado = "OK" if horas >= 8 else "Incompleto"
                if horas < 8:
                    dias_incompletos += 1
                
                ws.cell(row=row, column=1, value=fecha_str).border = thin_border
                ws.cell(row=row, column=2, value=dia_semana).border = thin_border
                ws.cell(row=row, column=3, value=entrada).border = thin_border
                ws.cell(row=row, column=4, value=salida).border = thin_border
                ws.cell(row=row, column=5, value=horas).border = thin_border
                ws.cell(row=row, column=6, value=minutos).border = thin_border
                ws.cell(row=row, column=7, value=f"{horas}:{minutos:02d}").border = thin_border
                estado_cell = ws.cell(row=row, column=8, value=estado)
                estado_cell.border = thin_border
                estado_cell.fill = ok_fill if horas >= 8 else warning_fill
            else:
                dias_ausencia += 1
                ws.cell(row=row, column=1, value=fecha_str).border = thin_border
                ws.cell(row=row, column=2, value=dia_semana).border = thin_border
                for c in range(3, 8):
                    ws.cell(row=row, column=c, value="-" if c < 7 else "0:00" if c == 7 else "AUSENCIA").border = thin_border
                ws.cell(row=row, column=5, value=0).border = thin_border
                ws.cell(row=row, column=6, value=0).border = thin_border
                ws.cell(row=row, column=7, value="0:00").border = thin_border
                ausencia_cell = ws.cell(row=row, column=8, value="AUSENCIA")
                ausencia_cell.border = thin_border
                ausencia_cell.fill = error_fill
            row += 1
        current_date += timedelta(days=1)
    
    total_horas += total_minutos // 60
    total_minutos = total_minutos % 60
    
    row += 1
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = "RESUMEN"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    for c in range(5, 9):
        ws.cell(row=row, column=c).fill = header_fill
    
    row += 1
    ws[f'A{row}'] = "Total Horas Trabajadas:"
    ws[f'G{row}'] = f"{total_horas}:{total_minutos:02d}"
    row += 1
    ws[f'A{row}'] = "Dias con Ausencia:"
    ws[f'G{row}'] = dias_ausencia
    row += 1
    ws[f'A{row}'] = "Dias Incompletos (<8h):"
    ws[f'G{row}'] = dias_incompletos
    
    for col_letter, width in [('A', 12), ('B', 12), ('C', 10), ('D', 10), ('E', 8), ('F', 6), ('G', 12), ('H', 12)]:
        ws.column_dimensions[col_letter].width = width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    nombre_archivo = f"control_horario_{empleado.get('apellidos', '')}_{fecha_desde}_{fecha_hasta}.xlsx".replace(" ", "_")
    
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename={nombre_archivo}"})


@router.get("/fichajes/informe/pdf")
async def export_informe_control_horario_pdf(empleado_id: str, fecha_desde: str, fecha_hasta: str):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from fastapi.responses import StreamingResponse
    
    database = get_db()
    empleado = await database.empleados.find_one({"_id": ObjectId(empleado_id)})
    if not empleado:
        raise HTTPException(status_code=404, detail="Empleado no encontrado")
    
    fichajes_raw = []
    cursor = database.fichajes.find({"empleado_id": empleado_id, "fecha": {"$gte": fecha_desde, "$lte": fecha_hasta}}).sort([("fecha", 1), ("hora", 1)])
    async for f in cursor:
        fichajes_raw.append(f)
    
    dias_trabajados = _agrupar_fichajes_por_dia(fichajes_raw)
    
    output = io.BytesIO()
    pdf = SimpleDocTemplate(output, pagesize=A4, topMargin=15*mm, bottomMargin=15*mm, leftMargin=15*mm, rightMargin=15*mm)
    elements = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=16, textColor=colors.HexColor('#2E7D32'), alignment=1, spaceAfter=5*mm)
    elements.append(Paragraph("INFORME DE CONTROL HORARIO", title_style))
    
    info_style = ParagraphStyle('Info', parent=styles['Normal'], fontSize=10, spaceAfter=2*mm)
    elements.append(Paragraph(f"<b>Empleado:</b> {empleado.get('nombre', '')} {empleado.get('apellidos', '')}", info_style))
    elements.append(Paragraph(f"<b>DNI/NIE:</b> {empleado.get('dni_nie', '')} | <b>Puesto:</b> {empleado.get('puesto', '')}", info_style))
    elements.append(Paragraph(f"<b>Periodo:</b> {fecha_desde} a {fecha_hasta}", info_style))
    elements.append(Spacer(1, 5*mm))
    
    dias_semana_es = {'Monday': 'Lun', 'Tuesday': 'Mar', 'Wednesday': 'Mie', 'Thursday': 'Jue', 'Friday': 'Vie'}
    table_data = [["Fecha", "Dia", "Entrada", "Salida", "Total", "Estado"]]
    total_horas = total_minutos = dias_ausencia = 0
    
    fecha_inicio = datetime.strptime(fecha_desde, "%Y-%m-%d").date()
    fecha_fin = datetime.strptime(fecha_hasta, "%Y-%m-%d").date()
    current_date = fecha_inicio
    
    while current_date <= fecha_fin:
        if current_date.weekday() < 5:
            fecha_str = current_date.strftime("%Y-%m-%d")
            dia_semana = dias_semana_es.get(current_date.strftime("%A"), "")
            if fecha_str in dias_trabajados:
                dia_data = dias_trabajados[fecha_str]
                entrada = dia_data["entradas"][0][:5] if dia_data["entradas"] else "-"
                salida = dia_data["salidas"][-1][:5] if dia_data["salidas"] else "-"
                horas = minutos = 0
                if entrada != "-" and salida != "-":
                    try:
                        horas, minutos = _calcular_horas(dia_data["entradas"][0], dia_data["salidas"][-1])
                        total_horas += horas
                        total_minutos += minutos
                    except Exception:
                        pass
                estado = "OK" if horas >= 8 else "<8h"
                table_data.append([fecha_str, dia_semana, entrada, salida, f"{horas}:{minutos:02d}", estado])
            else:
                dias_ausencia += 1
                table_data.append([fecha_str, dia_semana, "-", "-", "0:00", "AUSENCIA"])
        current_date += timedelta(days=1)
    
    total_horas += total_minutos // 60
    total_minutos = total_minutos % 60
    
    col_widths = [25*mm, 12*mm, 18*mm, 18*mm, 18*mm, 22*mm]
    doc_table = Table(table_data, colWidths=col_widths)
    table_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E7D32')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('TOPPADDING', (0, 0), (-1, 0), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 3),
        ('TOPPADDING', (0, 1), (-1, -1), 3),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.gray),
    ])
    for i, row_data in enumerate(table_data[1:], 1):
        if row_data[5] == "AUSENCIA":
            table_style.add('BACKGROUND', (0, i), (-1, i), colors.HexColor('#FFCDD2'))
        elif row_data[5] == "<8h":
            table_style.add('BACKGROUND', (0, i), (-1, i), colors.HexColor('#FFF9C4'))
        else:
            table_style.add('BACKGROUND', (0, i), (-1, i), colors.HexColor('#E8F5E9'))
    doc_table.setStyle(table_style)
    elements.append(doc_table)
    
    elements.append(Spacer(1, 8*mm))
    resumen_style = ParagraphStyle('Resumen', parent=styles['Normal'], fontSize=11, spaceAfter=3*mm)
    elements.append(Paragraph(f"<b>TOTAL HORAS TRABAJADAS: {total_horas}:{total_minutos:02d}</b>", resumen_style))
    elements.append(Paragraph(f"Dias con ausencia: {dias_ausencia}", info_style))
    elements.append(Paragraph(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", info_style))
    
    pdf.build(elements)
    output.seek(0)
    nombre_archivo = f"control_horario_{empleado.get('apellidos', '')}_{fecha_desde}_{fecha_hasta}.pdf".replace(" ", "_")
    
    return StreamingResponse(output, media_type="application/pdf",
                             headers={"Content-Disposition": f"attachment; filename={nombre_archivo}"})
