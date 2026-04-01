"""
RRHH - Documentos de Empleados
"""
from fastapi import APIRouter, HTTPException, UploadFile, File, Form
from typing import Optional
from datetime import datetime, timedelta
from bson import ObjectId
import os
import uuid
import io

router = APIRouter(prefix="/api/rrhh", tags=["RRHH - Documentos"])

db = None
send_documento_notification = None

def set_database(database):
    global db
    db = database

def set_email_service(email_func):
    global send_documento_notification
    send_documento_notification = email_func

def get_db():
    if db is None:
        raise HTTPException(status_code=500, detail="Database not initialized")
    return db


@router.get("/documentos")
async def get_documentos(
    empleado_id: Optional[str] = None,
    tipo: Optional[str] = None,
    estado: Optional[str] = None,
    fecha_desde: Optional[str] = None,
    fecha_hasta: Optional[str] = None
):
    database = get_db()
    query = {}
    if empleado_id:
        query["empleado_id"] = empleado_id
    if tipo:
        query["tipo"] = tipo
    if estado:
        if estado == "firmado":
            query["firmado"] = True
        elif estado == "pendiente":
            query["requiere_firma"] = True
            query["firmado"] = False
        elif estado == "no_requiere":
            query["requiere_firma"] = False
    if fecha_desde or fecha_hasta:
        query["created_at"] = {}
        if fecha_desde:
            query["created_at"]["$gte"] = datetime.strptime(fecha_desde, "%Y-%m-%d")
        if fecha_hasta:
            query["created_at"]["$lt"] = datetime.strptime(fecha_hasta, "%Y-%m-%d") + timedelta(days=1)
    
    documentos = []
    cursor = database.documentos_empleados.find(query).sort("created_at", -1)
    async for doc in cursor:
        doc["_id"] = str(doc["_id"])
        if doc.get("created_at"):
            doc["created_at"] = doc["created_at"].isoformat()
        if doc.get("updated_at"):
            doc["updated_at"] = doc["updated_at"].isoformat()
        documentos.append(doc)
    return {"success": True, "documentos": documentos, "total": len(documentos)}


@router.get("/documentos/export/excel")
async def export_documentos_excel(
    empleado_id: Optional[str] = None,
    fecha_desde: Optional[str] = None,
    fecha_hasta: Optional[str] = None,
    tipo: Optional[str] = None,
    estado: Optional[str] = None
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from fastapi.responses import StreamingResponse
    
    database = get_db()
    query = {}
    if empleado_id:
        query["empleado_id"] = empleado_id
    if tipo:
        query["tipo"] = tipo
    if estado:
        if estado == "firmado":
            query["firmado"] = True
        elif estado == "pendiente":
            query["requiere_firma"] = True
            query["firmado"] = False
        elif estado == "no_requiere":
            query["requiere_firma"] = False
    if fecha_desde or fecha_hasta:
        query["created_at"] = {}
        if fecha_desde:
            query["created_at"]["$gte"] = datetime.strptime(fecha_desde, "%Y-%m-%d")
        if fecha_hasta:
            query["created_at"]["$lt"] = datetime.strptime(fecha_hasta, "%Y-%m-%d") + timedelta(days=1)
    
    documentos = []
    cursor = database.documentos_empleados.find(query).sort("created_at", -1)
    async for doc in cursor:
        documentos.append(doc)
    
    empleados_dict = {}
    async for emp in database.empleados.find({}):
        empleados_dict[str(emp["_id"])] = f"{emp.get('nombre', '')} {emp.get('apellidos', '')}"
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Documentos"
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    headers = ["Documento", "Empleado", "Tipo", "Fecha Documento", "Fecha Registro", "Estado", "Archivo Adjunto"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    
    tipos_doc = {'contrato': 'Contrato de Trabajo', 'anexo': 'Anexo Contrato', 'nomina': 'Nomina', 'certificado': 'Certificado', 'formacion': 'Formacion PRL', 'epi': 'Entrega EPI', 'otro': 'Otro'}
    
    for row, doc in enumerate(documentos, 2):
        empleado_nombre = empleados_dict.get(doc.get("empleado_id", ""), "Desconocido")
        tipo_label = tipos_doc.get(doc.get("tipo", "otro"), doc.get("tipo", ""))
        fecha_doc = doc.get("fecha_creacion", "")
        fecha_registro = doc.get("created_at").strftime("%d/%m/%Y %H:%M") if doc.get("created_at") else ""
        estado_val = "Firmado" if doc.get("firmado") else ("Pendiente" if doc.get("requiere_firma") else "No requiere firma")
        archivo = "Si" if doc.get("archivo_url") else "No"
        
        for col, val in enumerate([doc.get("nombre", ""), empleado_nombre, tipo_label, fecha_doc, fecha_registro, estado_val, archivo], 1):
            ws.cell(row=row, column=col, value=val).border = thin_border
    
    for col_letter, width in [('A', 35), ('B', 25), ('C', 20), ('D', 15), ('E', 18), ('F', 18), ('G', 15)]:
        ws.column_dimensions[col_letter].width = width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"documentos_rrhh_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename={filename}"})


@router.get("/documentos/export/pdf")
async def export_documentos_pdf(
    empleado_id: Optional[str] = None,
    fecha_desde: Optional[str] = None,
    fecha_hasta: Optional[str] = None,
    tipo: Optional[str] = None,
    estado: Optional[str] = None
):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from fastapi.responses import StreamingResponse
    
    database = get_db()
    query = {}
    if empleado_id:
        query["empleado_id"] = empleado_id
    if tipo:
        query["tipo"] = tipo
    if estado:
        if estado == "firmado":
            query["firmado"] = True
        elif estado == "pendiente":
            query["requiere_firma"] = True
            query["firmado"] = False
        elif estado == "no_requiere":
            query["requiere_firma"] = False
    if fecha_desde or fecha_hasta:
        query["created_at"] = {}
        if fecha_desde:
            query["created_at"]["$gte"] = datetime.strptime(fecha_desde, "%Y-%m-%d")
        if fecha_hasta:
            query["created_at"]["$lt"] = datetime.strptime(fecha_hasta, "%Y-%m-%d") + timedelta(days=1)
    
    documentos = []
    cursor = database.documentos_empleados.find(query).sort("created_at", -1)
    async for doc in cursor:
        documentos.append(doc)
    
    empleados_dict = {}
    async for emp in database.empleados.find({}):
        empleados_dict[str(emp["_id"])] = f"{emp.get('nombre', '')} {emp.get('apellidos', '')}"
    
    output = io.BytesIO()
    pdf_doc = SimpleDocTemplate(output, pagesize=landscape(A4), topMargin=20*mm, bottomMargin=20*mm)
    elements = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=18, textColor=colors.HexColor('#2E7D32'), spaceAfter=10*mm, alignment=1)
    elements.append(Paragraph("Informe de Documentos - RRHH", title_style))
    
    subtitle = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    if fecha_desde:
        subtitle += f" | Desde: {fecha_desde}"
    if fecha_hasta:
        subtitle += f" | Hasta: {fecha_hasta}"
    subtitle_style = ParagraphStyle('Subtitle', parent=styles['Normal'], fontSize=10, textColor=colors.gray, alignment=1)
    elements.append(Paragraph(subtitle, subtitle_style))
    elements.append(Spacer(1, 10*mm))
    
    total_docs = len(documentos)
    firmados = sum(1 for d in documentos if d.get("firmado"))
    pendientes = sum(1 for d in documentos if d.get("requiere_firma") and not d.get("firmado"))
    con_archivo = sum(1 for d in documentos if d.get("archivo_url"))
    
    resumen_data = [["Total Documentos", "Firmados", "Pendientes Firma", "Con Archivo"], [str(total_docs), str(firmados), str(pendientes), str(con_archivo)]]
    resumen_table = Table(resumen_data, colWidths=[60*mm, 50*mm, 50*mm, 50*mm])
    resumen_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E7D32')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#E8F5E9')),
        ('FONTSIZE', (0, 1), (-1, -1), 12),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#2E7D32'))
    ]))
    elements.append(resumen_table)
    elements.append(Spacer(1, 10*mm))
    
    tipos_doc = {'contrato': 'Contrato', 'anexo': 'Anexo', 'nomina': 'Nomina', 'certificado': 'Certificado', 'formacion': 'Formacion', 'epi': 'EPI', 'otro': 'Otro'}
    
    if documentos:
        table_data = [["Documento", "Empleado", "Tipo", "Fecha Doc.", "Fecha Registro", "Estado"]]
        for doc in documentos[:50]:
            empleado_nombre = empleados_dict.get(doc.get("empleado_id", ""), "Desconocido")
            table_data.append([
                doc.get("nombre", "")[:40], empleado_nombre[:25],
                tipos_doc.get(doc.get("tipo", "otro"), doc.get("tipo", "")),
                doc.get("fecha_creacion", "-"),
                doc.get("created_at").strftime("%d/%m/%Y") if doc.get("created_at") else "-",
                "Firmado" if doc.get("firmado") else ("Pendiente" if doc.get("requiere_firma") else "OK")
            ])
        doc_table = Table(table_data, colWidths=[70*mm, 55*mm, 35*mm, 30*mm, 35*mm, 25*mm])
        doc_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E7D32')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (3, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.gray),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')])
        ]))
        elements.append(doc_table)
        if len(documentos) > 50:
            elements.append(Spacer(1, 5*mm))
            elements.append(Paragraph(f"Mostrando 50 de {len(documentos)} documentos.", styles['Normal']))
    else:
        elements.append(Paragraph("No hay documentos con los filtros seleccionados.", styles['Normal']))
    
    pdf_doc.build(elements)
    output.seek(0)
    filename = f"informe_documentos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    
    return StreamingResponse(output, media_type="application/pdf",
                             headers={"Content-Disposition": f"attachment; filename={filename}"})


@router.post("/documentos")
async def create_documento(documento: dict):
    database = get_db()
    documento["created_at"] = datetime.now()
    documento["updated_at"] = datetime.now()
    documento["firmado"] = False
    documento["activo"] = True
    result = await database.documentos_empleados.insert_one(documento)
    documento["_id"] = str(result.inserted_id)
    return {"success": True, "data": documento}


@router.post("/documentos/upload")
async def upload_documento(
    file: UploadFile = File(...),
    empleado_id: str = Form(...),
    nombre: str = Form(...),
    tipo: str = Form("otro"),
    descripcion: str = Form(""),
    requiere_firma: str = Form("true"),
    fecha_creacion: str = Form(None)
):
    database = get_db()
    requiere_firma_bool = requiere_firma.lower() in ('true', '1', 'yes')
    
    allowed_extensions = ['.pdf', '.doc', '.docx', '.jpg', '.jpeg', '.png', '.gif']
    file_ext = os.path.splitext(file.filename)[1].lower()
    if file_ext not in allowed_extensions:
        raise HTTPException(status_code=400, detail=f"Tipo de archivo no permitido. Use: {', '.join(allowed_extensions)}")
    
    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="El archivo es demasiado grande. Maximo 10MB.")
    
    upload_dir = "/app/uploads/documentos_empleados"
    os.makedirs(upload_dir, exist_ok=True)
    unique_filename = f"{uuid.uuid4().hex}{file_ext}"
    file_path = os.path.join(upload_dir, unique_filename)
    with open(file_path, "wb") as f:
        f.write(content)
    
    documento = {
        "empleado_id": empleado_id, "nombre": nombre, "tipo": tipo, "descripcion": descripcion,
        "requiere_firma": requiere_firma_bool,
        "fecha_creacion": fecha_creacion or datetime.now().strftime("%Y-%m-%d"),
        "archivo_url": f"/api/uploads/documentos_empleados/{unique_filename}",
        "archivo_nombre_original": file.filename,
        "archivo_tipo": file.content_type, "archivo_tamano": len(content),
        "firmado": False, "activo": True,
        "created_at": datetime.now(), "updated_at": datetime.now()
    }
    result = await database.documentos_empleados.insert_one(documento)
    documento["_id"] = str(result.inserted_id)
    
    empleado = await database.empleados.find_one({"_id": ObjectId(empleado_id)})
    if empleado and empleado.get("email"):
        empleado_nombre = f"{empleado.get('nombre', '')} {empleado.get('apellidos', '')}"
        notificacion = {
            "titulo": "Nuevo Documento Disponible",
            "mensaje": f"Se ha subido el documento '{nombre}' {'que requiere tu firma' if requiere_firma_bool else ''}.",
            "tipo": "warning" if requiere_firma_bool else "info",
            "enlace": "/portal-empleado",
            "destinatarios": [empleado.get("email")],
            "prioridad": "alta" if requiere_firma_bool else "normal",
            "datos_extra": {"documento_id": str(result.inserted_id), "tipo": "documento"},
            "created_at": datetime.now(), "leida_por": []
        }
        await database.notificaciones.insert_one(notificacion)
        
        if send_documento_notification:
            try:
                await send_documento_notification(
                    recipient_email=empleado.get("email"),
                    empleado_nombre=empleado_nombre,
                    documento_nombre=nombre,
                    tipo_documento=tipo,
                    requiere_firma=requiere_firma_bool
                )
            except Exception as e:
                print(f"Error sending documento email: {e}")
    
    return {"success": True, "data": documento}


@router.put("/documentos/{documento_id}/firmar")
async def firmar_documento(documento_id: str, firma_data: dict):
    database = get_db()
    result = await database.documentos_empleados.update_one(
        {"_id": ObjectId(documento_id)},
        {"$set": {
            "firmado": True, "firma_empleado_url": firma_data.get("firma_url"),
            "fecha_firma": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "updated_at": datetime.now()
        }}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Documento no encontrado")
    return {"success": True}


@router.delete("/documentos/{documento_id}")
async def delete_documento(documento_id: str):
    database = get_db()
    result = await database.documentos_empleados.delete_one({"_id": ObjectId(documento_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Documento no encontrado")
    return {"success": True}
