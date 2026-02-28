"""
Rutas para el módulo de Recursos Humanos (RRHH)
- Gestión de empleados
- Control horario (fichajes)
- Productividad
- Documentos

Módulos extraídos a archivos separados:
- Prenóminas: rrhh_prenominas.py
- Ausencias: rrhh_ausencias.py
"""

from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Query, Form
from typing import List, Optional
from datetime import datetime, timedelta
from bson import ObjectId
import os
import uuid
import qrcode
import io
import base64

# Import email service for notifications
from email_service import send_ausencia_notification, send_documento_notification

# Import sub-routers
from routes.rrhh_ausencias import router as ausencias_router, set_database as set_ausencias_db, set_email_service as set_ausencias_email
from routes.rrhh_prenominas import router as prenominas_router, set_database as set_prenominas_db

router = APIRouter(prefix="/api/rrhh", tags=["RRHH"])

# Database will be injected
db = None

def set_database(database):
    global db
    db = database
    # Also set database for sub-routers
    set_ausencias_db(database)
    set_prenominas_db(database)
    # Set email service for ausencias
    set_ausencias_email(send_ausencia_notification)

def get_db():
    if db is None:
        raise HTTPException(status_code=500, detail="Database not initialized")
    return db

# ============================================================================
# EMPLEADOS
# ============================================================================

@router.get("/empleados")
async def get_empleados(
    activo: Optional[bool] = None,
    puesto: Optional[str] = None,
    departamento: Optional[str] = None
):
    """Obtener lista de empleados con filtros opcionales"""
    database = get_db()
    
    query = {}
    if activo is not None:
        query["activo"] = activo
    if puesto:
        query["puesto"] = puesto
    if departamento:
        query["departamento"] = departamento
    
    empleados = []
    cursor = database.empleados.find(query).sort("apellidos", 1)
    
    async for emp in cursor:
        emp["_id"] = str(emp["_id"])
        empleados.append(emp)
    
    return {"success": True, "empleados": empleados, "total": len(empleados)}

@router.get("/empleados/stats")
async def get_empleados_stats():
    """Estadísticas de empleados"""
    database = get_db()
    
    total = await database.empleados.count_documents({})
    activos = await database.empleados.count_documents({"activo": True})
    
    # Por puesto
    pipeline = [
        {"$match": {"activo": True}},
        {"$group": {"_id": "$puesto", "count": {"$sum": 1}}}
    ]
    por_puesto = {}
    async for doc in database.empleados.aggregate(pipeline):
        por_puesto[doc["_id"] or "Sin asignar"] = doc["count"]
    
    # Por tipo de contrato
    pipeline = [
        {"$match": {"activo": True}},
        {"$group": {"_id": "$tipo_contrato", "count": {"$sum": 1}}}
    ]
    por_contrato = {}
    async for doc in database.empleados.aggregate(pipeline):
        por_contrato[doc["_id"] or "Sin asignar"] = doc["count"]
    
    return {
        "success": True,
        "total": total,
        "activos": activos,
        "inactivos": total - activos,
        "por_puesto": por_puesto,
        "por_contrato": por_contrato
    }

@router.get("/empleados/{empleado_id}")
async def get_empleado(empleado_id: str):
    """Obtener un empleado por ID"""
    database = get_db()
    
    empleado = await database.empleados.find_one({"_id": ObjectId(empleado_id)})
    if not empleado:
        raise HTTPException(status_code=404, detail="Empleado no encontrado")
    
    empleado["_id"] = str(empleado["_id"])
    return {"success": True, "empleado": empleado}

@router.post("/empleados")
async def create_empleado(empleado: dict):
    """Crear un nuevo empleado"""
    database = get_db()
    
    # Generar código único
    count = await database.empleados.count_documents({})
    empleado["codigo"] = f"EMP-{str(count + 1).zfill(4)}"
    
    # Generar QR único
    qr_data = f"EMP:{empleado['codigo']}:{uuid.uuid4().hex[:8]}"
    empleado["qr_code"] = qr_data
    
    # Timestamps
    empleado["created_at"] = datetime.now()
    empleado["updated_at"] = datetime.now()
    empleado["activo"] = True
    
    result = await database.empleados.insert_one(empleado)
    empleado["_id"] = str(result.inserted_id)
    
    return {"success": True, "data": empleado}

@router.put("/empleados/{empleado_id}")
async def update_empleado(empleado_id: str, empleado: dict):
    """Actualizar un empleado"""
    database = get_db()
    
    empleado["updated_at"] = datetime.now()
    
    # No permitir cambiar código ni qr_code
    empleado.pop("codigo", None)
    empleado.pop("qr_code", None)
    empleado.pop("_id", None)
    
    result = await database.empleados.update_one(
        {"_id": ObjectId(empleado_id)},
        {"$set": empleado}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Empleado no encontrado")
    
    return {"success": True}

@router.delete("/empleados/{empleado_id}")
async def delete_empleado(empleado_id: str):
    """Dar de baja un empleado (soft delete)"""
    database = get_db()
    
    result = await database.empleados.update_one(
        {"_id": ObjectId(empleado_id)},
        {"$set": {
            "activo": False,
            "fecha_baja": datetime.now().strftime("%Y-%m-%d"),
            "updated_at": datetime.now()
        }}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Empleado no encontrado")
    
    return {"success": True}

@router.delete("/empleados/{empleado_id}/permanente")
async def delete_empleado_permanente(empleado_id: str):
    """Eliminar permanentemente un empleado (solo si está en baja)"""
    database = get_db()
    
    # Verificar que el empleado existe y está en baja
    empleado = await database.empleados.find_one({"_id": ObjectId(empleado_id)})
    if not empleado:
        raise HTTPException(status_code=404, detail="Empleado no encontrado")
    
    if empleado.get("activo", True):
        raise HTTPException(status_code=400, detail="Solo se pueden eliminar empleados en baja")
    
    # Eliminar permanentemente
    result = await database.empleados.delete_one({"_id": ObjectId(empleado_id)})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=500, detail="Error al eliminar empleado")
    
    # También eliminar datos relacionados
    await database.fichajes.delete_many({"empleado_id": empleado_id})
    await database.productividad.delete_many({"empleado_id": empleado_id})
    await database.documentos_empleado.delete_many({"empleado_id": empleado_id})
    await database.prenominas.delete_many({"empleado_id": empleado_id})
    
    return {"success": True, "message": "Empleado eliminado permanentemente"}

@router.put("/empleados/{empleado_id}/reactivar")
async def reactivar_empleado(empleado_id: str):
    """Reactivar un empleado que estaba en baja"""
    database = get_db()
    
    # Verificar que el empleado existe y está en baja
    empleado = await database.empleados.find_one({"_id": ObjectId(empleado_id)})
    if not empleado:
        raise HTTPException(status_code=404, detail="Empleado no encontrado")
    
    if empleado.get("activo", True):
        raise HTTPException(status_code=400, detail="El empleado ya está activo")
    
    result = await database.empleados.update_one(
        {"_id": ObjectId(empleado_id)},
        {"$set": {
            "activo": True,
            "fecha_baja": None,
            "updated_at": datetime.now()
        }}
    )
    
    return {"success": True, "message": "Empleado reactivado"}

@router.get("/empleados/{empleado_id}/qr")
async def get_empleado_qr(empleado_id: str):
    """Generar imagen QR para un empleado"""
    database = get_db()
    
    empleado = await database.empleados.find_one({"_id": ObjectId(empleado_id)})
    if not empleado:
        raise HTTPException(status_code=404, detail="Empleado no encontrado")
    
    # Generar QR
    qr = qrcode.QRCode(version=1, box_size=10, border=5)
    qr.add_data(empleado.get("qr_code", f"EMP:{empleado.get('codigo', 'UNKNOWN')}"))
    qr.make(fit=True)
    
    img = qr.make_image(fill_color="black", back_color="white")
    
    # Convertir a base64
    buffer = io.BytesIO()
    img.save(buffer, format="PNG")
    img_str = base64.b64encode(buffer.getvalue()).decode()
    
    return {
        "success": True,
        "qr_image": f"data:image/png;base64,{img_str}",
        "qr_code": empleado.get("qr_code"),
        "empleado_nombre": f"{empleado.get('nombre', '')} {empleado.get('apellidos', '')}"
    }

# ============================================================================
# FICHAJES / CONTROL HORARIO
# ============================================================================

@router.get("/fichajes")
async def get_fichajes(
    empleado_id: Optional[str] = None,
    fecha_desde: Optional[str] = None,
    fecha_hasta: Optional[str] = None,
    parcela_id: Optional[str] = None
):
    """Obtener fichajes con filtros"""
    database = get_db()
    
    query = {}
    if empleado_id:
        query["empleado_id"] = empleado_id
    if fecha_desde:
        query["fecha"] = {"$gte": fecha_desde}
    if fecha_hasta:
        if "fecha" in query:
            query["fecha"]["$lte"] = fecha_hasta
        else:
            query["fecha"] = {"$lte": fecha_hasta}
    if parcela_id:
        query["parcela_id"] = parcela_id
    
    fichajes = []
    cursor = database.fichajes.find(query).sort([("fecha", -1), ("hora", -1)])
    
    async for f in cursor:
        f["_id"] = str(f["_id"])
        # Obtener nombre del empleado
        emp = await database.empleados.find_one({"_id": ObjectId(f["empleado_id"])})
        if emp:
            f["empleado_nombre"] = f"{emp.get('nombre', '')} {emp.get('apellidos', '')}"
        fichajes.append(f)
    
    return {"success": True, "fichajes": fichajes, "total": len(fichajes)}

@router.get("/fichajes/hoy")
async def get_fichajes_hoy():
    """Obtener fichajes del día actual"""
    database = get_db()
    
    hoy = datetime.now().strftime("%Y-%m-%d")
    
    fichajes = []
    cursor = database.fichajes.find({"fecha": hoy}).sort("hora", -1)
    
    async for f in cursor:
        f["_id"] = str(f["_id"])
        emp = await database.empleados.find_one({"_id": ObjectId(f["empleado_id"])})
        if emp:
            f["empleado_nombre"] = f"{emp.get('nombre', '')} {emp.get('apellidos', '')}"
            f["empleado_foto"] = emp.get("foto_url")
        fichajes.append(f)
    
    # Estadísticas del día
    empleados_activos = await database.empleados.count_documents({"activo": True})
    empleados_fichados = len(set([f["empleado_id"] for f in fichajes if f["tipo"] == "entrada"]))
    
    return {
        "success": True,
        "fichajes": fichajes,
        "estadisticas": {
            "empleados_activos": empleados_activos,
            "empleados_fichados": empleados_fichados,
            "pendientes_fichar": empleados_activos - empleados_fichados
        }
    }

@router.post("/fichajes")
async def create_fichaje(fichaje: dict):
    """Registrar un fichaje"""
    database = get_db()
    
    # Validar que existe el empleado
    empleado = await database.empleados.find_one({"_id": ObjectId(fichaje["empleado_id"])})
    if not empleado:
        raise HTTPException(status_code=404, detail="Empleado no encontrado")
    
    fichaje["created_at"] = datetime.now()
    fichaje["sincronizado"] = True
    
    result = await database.fichajes.insert_one(fichaje)
    fichaje["_id"] = str(result.inserted_id)
    fichaje["empleado_nombre"] = f"{empleado.get('nombre', '')} {empleado.get('apellidos', '')}"
    
    return {"success": True, "data": fichaje}

@router.post("/fichajes/qr")
async def fichaje_por_qr(data: dict):
    """Fichaje mediante código QR"""
    database = get_db()
    
    qr_code = data.get("qr_code")
    tipo = data.get("tipo", "entrada")
    
    # Buscar empleado por QR
    empleado = await database.empleados.find_one({"qr_code": qr_code, "activo": True})
    if not empleado:
        raise HTTPException(status_code=404, detail="Código QR no válido o empleado inactivo")
    
    # Crear fichaje
    now = datetime.now()
    fichaje = {
        "empleado_id": str(empleado["_id"]),
        "tipo": tipo,
        "fecha": now.strftime("%Y-%m-%d"),
        "hora": now.strftime("%H:%M:%S"),
        "metodo_identificacion": "qr",
        "latitud": data.get("latitud"),
        "longitud": data.get("longitud"),
        "ubicacion_nombre": data.get("ubicacion_nombre"),
        "parcela_id": data.get("parcela_id"),
        "offline": data.get("offline", False),
        "sincronizado": True,
        "created_at": now
    }
    
    result = await database.fichajes.insert_one(fichaje)
    fichaje["_id"] = str(result.inserted_id)
    fichaje["empleado_nombre"] = f"{empleado.get('nombre', '')} {empleado.get('apellidos', '')}"
    fichaje["empleado_foto"] = empleado.get("foto_url")
    
    return {"success": True, "data": fichaje}

@router.post("/fichajes/nfc")
async def fichaje_por_nfc(data: dict):
    """Fichaje mediante tarjeta NFC"""
    database = get_db()
    
    nfc_id = data.get("nfc_id")
    tipo = data.get("tipo", "entrada")
    
    # Buscar empleado por NFC
    empleado = await database.empleados.find_one({"nfc_id": nfc_id, "activo": True})
    if not empleado:
        raise HTTPException(status_code=404, detail="Tarjeta NFC no registrada o empleado inactivo")
    
    # Crear fichaje
    now = datetime.now()
    fichaje = {
        "empleado_id": str(empleado["_id"]),
        "tipo": tipo,
        "fecha": now.strftime("%Y-%m-%d"),
        "hora": now.strftime("%H:%M:%S"),
        "metodo_identificacion": "nfc",
        "latitud": data.get("latitud"),
        "longitud": data.get("longitud"),
        "ubicacion_nombre": data.get("ubicacion_nombre"),
        "parcela_id": data.get("parcela_id"),
        "offline": data.get("offline", False),
        "sincronizado": True,
        "created_at": now
    }
    
    result = await database.fichajes.insert_one(fichaje)
    fichaje["_id"] = str(result.inserted_id)
    fichaje["empleado_nombre"] = f"{empleado.get('nombre', '')} {empleado.get('apellidos', '')}"
    
    return {"success": True, "data": fichaje}

@router.post("/fichajes/facial")
async def fichaje_por_facial(data: dict):
    """Fichaje mediante reconocimiento facial (comparación simple de fotos)"""
    database = get_db()
    
    empleado_id = data.get("empleado_id")
    foto_capturada = data.get("foto_capturada")  # Base64 de la foto capturada
    tipo = data.get("tipo", "entrada")
    
    if not empleado_id:
        raise HTTPException(status_code=400, detail="Se requiere ID de empleado")
    
    # Verificar que el empleado existe y tiene foto
    empleado = await database.empleados.find_one({"_id": ObjectId(empleado_id), "activo": True})
    if not empleado:
        raise HTTPException(status_code=404, detail="Empleado no encontrado")
    
    if not empleado.get("foto_url"):
        raise HTTPException(status_code=400, detail="El empleado no tiene foto registrada")
    
    # Aquí iría la lógica de comparación facial
    # Por ahora, simplemente registramos el fichaje con la foto de verificación
    
    now = datetime.now()
    fichaje = {
        "empleado_id": empleado_id,
        "tipo": tipo,
        "fecha": now.strftime("%Y-%m-%d"),
        "hora": now.strftime("%H:%M:%S"),
        "metodo_identificacion": "facial",
        "foto_verificacion_url": foto_capturada,  # Guardar foto de verificación
        "latitud": data.get("latitud"),
        "longitud": data.get("longitud"),
        "ubicacion_nombre": data.get("ubicacion_nombre"),
        "parcela_id": data.get("parcela_id"),
        "offline": data.get("offline", False),
        "sincronizado": True,
        "created_at": now
    }
    
    result = await database.fichajes.insert_one(fichaje)
    fichaje["_id"] = str(result.inserted_id)
    fichaje["empleado_nombre"] = f"{empleado.get('nombre', '')} {empleado.get('apellidos', '')}"
    
    return {"success": True, "data": fichaje}

@router.post("/fichajes/sync")
async def sync_fichajes_offline(fichajes: List[dict]):
    """Sincronizar fichajes creados offline"""
    database = get_db()
    
    synced = 0
    errors = []
    
    for fichaje in fichajes:
        try:
            fichaje["sincronizado"] = True
            fichaje["created_at"] = datetime.now()
            await database.fichajes.insert_one(fichaje)
            synced += 1
        except Exception as e:
            errors.append({"fichaje": fichaje, "error": str(e)})
    
    return {
        "success": True,
        "synced": synced,
        "errors": errors
    }

@router.put("/fichajes/{fichaje_id}/validar")
async def validar_fichaje(fichaje_id: str, validador: dict):
    """Validar un fichaje"""
    database = get_db()
    
    result = await database.fichajes.update_one(
        {"_id": ObjectId(fichaje_id)},
        {"$set": {
            "validado": True,
            "validado_por": validador.get("validado_por"),
            "validado_fecha": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Fichaje no encontrado")
    
    return {"success": True}


# ============================================================================
# INFORME DE CONTROL HORARIO
# ============================================================================

@router.get("/fichajes/informe")
async def get_informe_control_horario(
    empleado_id: str,
    fecha_desde: str,
    fecha_hasta: str
):
    """Obtener informe de control horario por empleado y rango de fechas"""
    database = get_db()
    
    # Obtener datos del empleado
    empleado = await database.empleados.find_one({"_id": ObjectId(empleado_id)})
    if not empleado:
        raise HTTPException(status_code=404, detail="Empleado no encontrado")
    
    # Obtener fichajes del empleado en el rango de fechas
    query = {
        "empleado_id": empleado_id,
        "fecha": {"$gte": fecha_desde, "$lte": fecha_hasta}
    }
    
    fichajes = []
    cursor = database.fichajes.find(query).sort([("fecha", 1), ("hora", 1)])
    async for f in cursor:
        fichajes.append(f)
    
    # Agrupar fichajes por día y calcular horas
    dias_trabajados = {}
    for fichaje in fichajes:
        fecha = fichaje["fecha"]
        if fecha not in dias_trabajados:
            dias_trabajados[fecha] = {"entradas": [], "salidas": []}
        
        if fichaje["tipo"] == "entrada":
            dias_trabajados[fecha]["entradas"].append(fichaje["hora"])
        else:
            dias_trabajados[fecha]["salidas"].append(fichaje["hora"])
    
    # Calcular horas por día
    registros = []
    total_horas = 0
    total_minutos = 0
    dias_completos = 0
    dias_con_ausencia = 0
    
    # Generar lista de todos los días laborables en el rango
    from datetime import date as date_class
    fecha_inicio = datetime.strptime(fecha_desde, "%Y-%m-%d").date()
    fecha_fin = datetime.strptime(fecha_hasta, "%Y-%m-%d").date()
    
    current_date = fecha_inicio
    while current_date <= fecha_fin:
        # Excluir fines de semana (sábado=5, domingo=6)
        if current_date.weekday() < 5:  # Día laborable
            fecha_str = current_date.strftime("%Y-%m-%d")
            
            if fecha_str in dias_trabajados:
                dia_data = dias_trabajados[fecha_str]
                entrada = dia_data["entradas"][0] if dia_data["entradas"] else None
                salida = dia_data["salidas"][-1] if dia_data["salidas"] else None
                
                horas_dia = 0
                minutos_dia = 0
                
                if entrada and salida:
                    # Calcular diferencia de tiempo
                    h_entrada = datetime.strptime(entrada, "%H:%M:%S") if len(entrada) > 5 else datetime.strptime(entrada, "%H:%M")
                    h_salida = datetime.strptime(salida, "%H:%M:%S") if len(salida) > 5 else datetime.strptime(salida, "%H:%M")
                    
                    diff = h_salida - h_entrada
                    total_segundos = diff.total_seconds()
                    horas_dia = int(total_segundos // 3600)
                    minutos_dia = int((total_segundos % 3600) // 60)
                    
                    total_horas += horas_dia
                    total_minutos += minutos_dia
                    
                    if horas_dia >= 8:
                        dias_completos += 1
                
                registros.append({
                    "fecha": fecha_str,
                    "dia_semana": current_date.strftime("%A"),
                    "entrada": entrada,
                    "salida": salida,
                    "horas": horas_dia,
                    "minutos": minutos_dia,
                    "horas_decimal": round(horas_dia + minutos_dia / 60, 2),
                    "completo": horas_dia >= 8,
                    "ausencia": False
                })
            else:
                # Día sin fichaje - ausencia
                dias_con_ausencia += 1
                registros.append({
                    "fecha": fecha_str,
                    "dia_semana": current_date.strftime("%A"),
                    "entrada": None,
                    "salida": None,
                    "horas": 0,
                    "minutos": 0,
                    "horas_decimal": 0,
                    "completo": False,
                    "ausencia": True
                })
        
        current_date += timedelta(days=1)
    
    # Convertir minutos totales a horas
    total_horas += total_minutos // 60
    total_minutos = total_minutos % 60
    
    return {
        "success": True,
        "empleado": {
            "id": str(empleado["_id"]),
            "nombre": empleado.get("nombre", ""),
            "apellidos": empleado.get("apellidos", ""),
            "dni": empleado.get("dni_nie", ""),
            "puesto": empleado.get("puesto", "")
        },
        "periodo": {
            "desde": fecha_desde,
            "hasta": fecha_hasta
        },
        "registros": registros,
        "resumen": {
            "total_horas": total_horas,
            "total_minutos": total_minutos,
            "total_horas_decimal": round(total_horas + total_minutos / 60, 2),
            "dias_trabajados": len([r for r in registros if not r["ausencia"]]),
            "dias_completos": dias_completos,
            "dias_con_ausencia": dias_con_ausencia,
            "horas_esperadas": len([r for r in registros if not r["ausencia"]]) * 8,
            "diferencia_horas": round((total_horas + total_minutos / 60) - (len([r for r in registros if not r["ausencia"]]) * 8), 2)
        }
    }


@router.get("/fichajes/informe/excel")
async def export_informe_control_horario_excel(
    empleado_id: str,
    fecha_desde: str,
    fecha_hasta: str
):
    """Exportar informe de control horario a Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from fastapi.responses import StreamingResponse
    
    # Obtener datos del informe
    database = get_db()
    
    empleado = await database.empleados.find_one({"_id": ObjectId(empleado_id)})
    if not empleado:
        raise HTTPException(status_code=404, detail="Empleado no encontrado")
    
    # Reutilizar lógica del endpoint principal
    query = {
        "empleado_id": empleado_id,
        "fecha": {"$gte": fecha_desde, "$lte": fecha_hasta}
    }
    
    fichajes = []
    cursor = database.fichajes.find(query).sort([("fecha", 1), ("hora", 1)])
    async for f in cursor:
        fichajes.append(f)
    
    # Agrupar por día
    dias_trabajados = {}
    for fichaje in fichajes:
        fecha = fichaje["fecha"]
        if fecha not in dias_trabajados:
            dias_trabajados[fecha] = {"entradas": [], "salidas": []}
        if fichaje["tipo"] == "entrada":
            dias_trabajados[fecha]["entradas"].append(fichaje["hora"])
        else:
            dias_trabajados[fecha]["salidas"].append(fichaje["hora"])
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Control Horario"
    
    # Estilos
    title_font = Font(bold=True, size=14, color="FFFFFF")
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    warning_fill = PatternFill(start_color="FFEB3B", end_color="FFEB3B", fill_type="solid")
    error_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
    ok_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    ws.merge_cells('A1:G1')
    ws['A1'] = "INFORME DE CONTROL HORARIO"
    ws['A1'].font = title_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = Alignment(horizontal="center")
    
    # Info empleado
    ws['A3'] = "Empleado:"
    ws['B3'] = f"{empleado.get('nombre', '')} {empleado.get('apellidos', '')}"
    ws['A4'] = "DNI/NIE:"
    ws['B4'] = empleado.get('dni_nie', '')
    ws['A5'] = "Puesto:"
    ws['B5'] = empleado.get('puesto', '')
    ws['D3'] = "Período:"
    ws['E3'] = f"{fecha_desde} a {fecha_hasta}"
    
    # Headers de tabla
    headers = ["Fecha", "Día", "Entrada", "Salida", "Horas", "Min", "Total Horas", "Estado"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
    
    # Datos
    dias_semana_es = {
        'Monday': 'Lunes', 'Tuesday': 'Martes', 'Wednesday': 'Miércoles',
        'Thursday': 'Jueves', 'Friday': 'Viernes', 'Saturday': 'Sábado', 'Sunday': 'Domingo'
    }
    
    row = 8
    total_horas = 0
    total_minutos = 0
    dias_ausencia = 0
    dias_incompletos = 0
    
    fecha_inicio = datetime.strptime(fecha_desde, "%Y-%m-%d").date()
    fecha_fin = datetime.strptime(fecha_hasta, "%Y-%m-%d").date()
    current_date = fecha_inicio
    
    while current_date <= fecha_fin:
        if current_date.weekday() < 5:  # Día laborable
            fecha_str = current_date.strftime("%Y-%m-%d")
            dia_semana = dias_semana_es.get(current_date.strftime("%A"), current_date.strftime("%A"))
            
            if fecha_str in dias_trabajados:
                dia_data = dias_trabajados[fecha_str]
                entrada = dia_data["entradas"][0] if dia_data["entradas"] else "-"
                salida = dia_data["salidas"][-1] if dia_data["salidas"] else "-"
                
                horas = 0
                minutos = 0
                
                if entrada != "-" and salida != "-":
                    h_entrada = datetime.strptime(entrada, "%H:%M:%S") if len(entrada) > 5 else datetime.strptime(entrada, "%H:%M")
                    h_salida = datetime.strptime(salida, "%H:%M:%S") if len(salida) > 5 else datetime.strptime(salida, "%H:%M")
                    diff = h_salida - h_entrada
                    total_segundos = diff.total_seconds()
                    horas = int(total_segundos // 3600)
                    minutos = int((total_segundos % 3600) // 60)
                    total_horas += horas
                    total_minutos += minutos
                
                estado = "OK" if horas >= 8 else "Incompleto"
                if horas < 8:
                    dias_incompletos += 1
                
                ws.cell(row=row, column=1, value=fecha_str).border = thin_border
                ws.cell(row=row, column=2, value=dia_semana).border = thin_border
                ws.cell(row=row, column=3, value=entrada).border = thin_border
                ws.cell(row=row, column=4, value=salida).border = thin_border
                ws.cell(row=row, column=5, value=horas).border = thin_border
                ws.cell(row=row, column=6, value=minutos).border = thin_border
                ws.cell(row=row, column=7, value=f"{horas}:{minutos:02d}").border = thin_border
                estado_cell = ws.cell(row=row, column=8, value=estado)
                estado_cell.border = thin_border
                if horas >= 8:
                    estado_cell.fill = ok_fill
                else:
                    estado_cell.fill = warning_fill
            else:
                # Ausencia
                dias_ausencia += 1
                ws.cell(row=row, column=1, value=fecha_str).border = thin_border
                ws.cell(row=row, column=2, value=dia_semana).border = thin_border
                ws.cell(row=row, column=3, value="-").border = thin_border
                ws.cell(row=row, column=4, value="-").border = thin_border
                ws.cell(row=row, column=5, value=0).border = thin_border
                ws.cell(row=row, column=6, value=0).border = thin_border
                ws.cell(row=row, column=7, value="0:00").border = thin_border
                ausencia_cell = ws.cell(row=row, column=8, value="AUSENCIA")
                ausencia_cell.border = thin_border
                ausencia_cell.fill = error_fill
            
            row += 1
        current_date += timedelta(days=1)
    
    # Convertir minutos totales
    total_horas += total_minutos // 60
    total_minutos = total_minutos % 60
    
    # Resumen
    row += 1
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = "RESUMEN"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws.cell(row=row, column=5).fill = header_fill
    ws.cell(row=row, column=6).fill = header_fill
    ws.cell(row=row, column=7).fill = header_fill
    ws.cell(row=row, column=8).fill = header_fill
    
    row += 1
    ws[f'A{row}'] = "Total Horas Trabajadas:"
    ws[f'E{row}'] = total_horas
    ws[f'F{row}'] = total_minutos
    ws[f'G{row}'] = f"{total_horas}:{total_minutos:02d}"
    
    row += 1
    ws[f'A{row}'] = "Días con Ausencia:"
    ws[f'G{row}'] = dias_ausencia
    
    row += 1
    ws[f'A{row}'] = "Días Incompletos (<8h):"
    ws[f'G{row}'] = dias_incompletos
    
    # Ajustar anchos
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 6
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    
    # Guardar
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    nombre_archivo = f"control_horario_{empleado.get('apellidos', '')}_{fecha_desde}_{fecha_hasta}.xlsx"
    nombre_archivo = nombre_archivo.replace(" ", "_")
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nombre_archivo}"}
    )


@router.get("/fichajes/informe/pdf")
async def export_informe_control_horario_pdf(
    empleado_id: str,
    fecha_desde: str,
    fecha_hasta: str
):
    """Exportar informe de control horario a PDF"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from fastapi.responses import StreamingResponse
    
    database = get_db()
    
    empleado = await database.empleados.find_one({"_id": ObjectId(empleado_id)})
    if not empleado:
        raise HTTPException(status_code=404, detail="Empleado no encontrado")
    
    # Obtener fichajes
    query = {
        "empleado_id": empleado_id,
        "fecha": {"$gte": fecha_desde, "$lte": fecha_hasta}
    }
    
    fichajes = []
    cursor = database.fichajes.find(query).sort([("fecha", 1), ("hora", 1)])
    async for f in cursor:
        fichajes.append(f)
    
    # Agrupar por día
    dias_trabajados = {}
    for fichaje in fichajes:
        fecha = fichaje["fecha"]
        if fecha not in dias_trabajados:
            dias_trabajados[fecha] = {"entradas": [], "salidas": []}
        if fichaje["tipo"] == "entrada":
            dias_trabajados[fecha]["entradas"].append(fichaje["hora"])
        else:
            dias_trabajados[fecha]["salidas"].append(fichaje["hora"])
    
    # Crear PDF
    output = io.BytesIO()
    pdf = SimpleDocTemplate(output, pagesize=A4, topMargin=15*mm, bottomMargin=15*mm, leftMargin=15*mm, rightMargin=15*mm)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Título
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=16, textColor=colors.HexColor('#2E7D32'), alignment=1, spaceAfter=5*mm)
    elements.append(Paragraph("INFORME DE CONTROL HORARIO", title_style))
    
    # Info empleado
    info_style = ParagraphStyle('Info', parent=styles['Normal'], fontSize=10, spaceAfter=2*mm)
    elements.append(Paragraph(f"<b>Empleado:</b> {empleado.get('nombre', '')} {empleado.get('apellidos', '')}", info_style))
    elements.append(Paragraph(f"<b>DNI/NIE:</b> {empleado.get('dni_nie', '')} | <b>Puesto:</b> {empleado.get('puesto', '')}", info_style))
    elements.append(Paragraph(f"<b>Período:</b> {fecha_desde} a {fecha_hasta}", info_style))
    elements.append(Spacer(1, 5*mm))
    
    # Tabla de registros
    dias_semana_es = {
        'Monday': 'Lun', 'Tuesday': 'Mar', 'Wednesday': 'Mié',
        'Thursday': 'Jue', 'Friday': 'Vie', 'Saturday': 'Sáb', 'Sunday': 'Dom'
    }
    
    table_data = [["Fecha", "Día", "Entrada", "Salida", "Total", "Estado"]]
    
    total_horas = 0
    total_minutos = 0
    dias_ausencia = 0
    
    fecha_inicio = datetime.strptime(fecha_desde, "%Y-%m-%d").date()
    fecha_fin = datetime.strptime(fecha_hasta, "%Y-%m-%d").date()
    current_date = fecha_inicio
    
    while current_date <= fecha_fin:
        if current_date.weekday() < 5:
            fecha_str = current_date.strftime("%Y-%m-%d")
            dia_semana = dias_semana_es.get(current_date.strftime("%A"), "")
            
            if fecha_str in dias_trabajados:
                dia_data = dias_trabajados[fecha_str]
                entrada = dia_data["entradas"][0][:5] if dia_data["entradas"] else "-"
                salida = dia_data["salidas"][-1][:5] if dia_data["salidas"] else "-"
                
                horas = 0
                minutos = 0
                
                if entrada != "-" and salida != "-":
                    try:
                        h_entrada = datetime.strptime(dia_data["entradas"][0], "%H:%M:%S") if len(dia_data["entradas"][0]) > 5 else datetime.strptime(dia_data["entradas"][0], "%H:%M")
                        h_salida = datetime.strptime(dia_data["salidas"][-1], "%H:%M:%S") if len(dia_data["salidas"][-1]) > 5 else datetime.strptime(dia_data["salidas"][-1], "%H:%M")
                        diff = h_salida - h_entrada
                        total_segundos = diff.total_seconds()
                        horas = int(total_segundos // 3600)
                        minutos = int((total_segundos % 3600) // 60)
                        total_horas += horas
                        total_minutos += minutos
                    except Exception:
                        pass
                
                estado = "OK" if horas >= 8 else "<8h"
                table_data.append([fecha_str, dia_semana, entrada, salida, f"{horas}:{minutos:02d}", estado])
            else:
                dias_ausencia += 1
                table_data.append([fecha_str, dia_semana, "-", "-", "0:00", "AUSENCIA"])
        
        current_date += timedelta(days=1)
    
    # Convertir minutos
    total_horas += total_minutos // 60
    total_minutos = total_minutos % 60
    
    # Crear tabla
    col_widths = [25*mm, 12*mm, 18*mm, 18*mm, 18*mm, 22*mm]
    doc_table = Table(table_data, colWidths=col_widths)
    
    table_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E7D32')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('TOPPADDING', (0, 0), (-1, 0), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 3),
        ('TOPPADDING', (0, 1), (-1, -1), 3),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.gray),
    ])
    
    # Colorear filas según estado
    for i, row_data in enumerate(table_data[1:], 1):
        if row_data[5] == "AUSENCIA":
            table_style.add('BACKGROUND', (0, i), (-1, i), colors.HexColor('#FFCDD2'))
        elif row_data[5] == "<8h":
            table_style.add('BACKGROUND', (0, i), (-1, i), colors.HexColor('#FFF9C4'))
        else:
            table_style.add('BACKGROUND', (0, i), (-1, i), colors.HexColor('#E8F5E9'))
    
    doc_table.setStyle(table_style)
    elements.append(doc_table)
    
    # Resumen
    elements.append(Spacer(1, 8*mm))
    resumen_style = ParagraphStyle('Resumen', parent=styles['Normal'], fontSize=11, spaceAfter=3*mm)
    elements.append(Paragraph(f"<b>TOTAL HORAS TRABAJADAS: {total_horas}:{total_minutos:02d}</b>", resumen_style))
    elements.append(Paragraph(f"Días con ausencia: {dias_ausencia}", info_style))
    elements.append(Paragraph(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", info_style))
    
    pdf.build(elements)
    output.seek(0)
    
    nombre_archivo = f"control_horario_{empleado.get('apellidos', '')}_{fecha_desde}_{fecha_hasta}.pdf"
    nombre_archivo = nombre_archivo.replace(" ", "_")
    
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={nombre_archivo}"}
    )


# ============================================================================
# PRODUCTIVIDAD
# ============================================================================

@router.get("/productividad")
async def get_registros_productividad(
    empleado_id: Optional[str] = None,
    fecha_desde: Optional[str] = None,
    fecha_hasta: Optional[str] = None,
    parcela_id: Optional[str] = None,
    tipo_trabajo: Optional[str] = None
):
    """Obtener registros de productividad"""
    database = get_db()
    
    query = {}
    if empleado_id:
        query["empleado_id"] = empleado_id
    if fecha_desde:
        query["fecha"] = {"$gte": fecha_desde}
    if fecha_hasta:
        if "fecha" in query:
            query["fecha"]["$lte"] = fecha_hasta
        else:
            query["fecha"] = {"$lte": fecha_hasta}
    if parcela_id:
        query["parcela_id"] = parcela_id
    if tipo_trabajo:
        query["tipo_trabajo"] = tipo_trabajo
    
    registros = []
    cursor = database.productividad.find(query).sort("fecha", -1)
    
    async for r in cursor:
        r["_id"] = str(r["_id"])
        # Obtener nombre del empleado
        emp = await database.empleados.find_one({"_id": ObjectId(r["empleado_id"])})
        if emp:
            r["empleado_nombre"] = f"{emp.get('nombre', '')} {emp.get('apellidos', '')}"
        registros.append(r)
    
    return {"success": True, "registros": registros, "total": len(registros)}

@router.get("/productividad/stats")
async def get_productividad_stats(
    fecha_desde: Optional[str] = None,
    fecha_hasta: Optional[str] = None
):
    """Estadísticas de productividad"""
    database = get_db()
    
    # Fechas por defecto: último mes
    if not fecha_desde:
        fecha_desde = (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")
    if not fecha_hasta:
        fecha_hasta = datetime.now().strftime("%Y-%m-%d")
    
    query = {"fecha": {"$gte": fecha_desde, "$lte": fecha_hasta}}
    
    # Totales
    pipeline = [
        {"$match": query},
        {"$group": {
            "_id": None,
            "total_kilos": {"$sum": {"$ifNull": ["$kilos_recogidos", 0]}},
            "total_hectareas": {"$sum": {"$ifNull": ["$hectareas_trabajadas", 0]}},
            "total_horas": {"$sum": {"$ifNull": ["$horas_trabajadas", 0]}},
            "total_registros": {"$sum": 1}
        }}
    ]
    
    totales = {"total_kilos": 0, "total_hectareas": 0, "total_horas": 0, "total_registros": 0}
    async for doc in database.productividad.aggregate(pipeline):
        totales = {
            "total_kilos": doc.get("total_kilos", 0),
            "total_hectareas": doc.get("total_hectareas", 0),
            "total_horas": doc.get("total_horas", 0),
            "total_registros": doc.get("total_registros", 0)
        }
    
    # Top empleados por kilos
    pipeline = [
        {"$match": query},
        {"$group": {
            "_id": "$empleado_id",
            "total_kilos": {"$sum": {"$ifNull": ["$kilos_recogidos", 0]}},
            "total_horas": {"$sum": {"$ifNull": ["$horas_trabajadas", 0]}}
        }},
        {"$sort": {"total_kilos": -1}},
        {"$limit": 10}
    ]
    
    top_empleados = []
    async for doc in database.productividad.aggregate(pipeline):
        emp = await database.empleados.find_one({"_id": ObjectId(doc["_id"])})
        if emp:
            top_empleados.append({
                "empleado_id": doc["_id"],
                "empleado_nombre": f"{emp.get('nombre', '')} {emp.get('apellidos', '')}",
                "total_kilos": doc["total_kilos"],
                "total_horas": doc["total_horas"],
                "kilos_hora": round(doc["total_kilos"] / doc["total_horas"], 2) if doc["total_horas"] > 0 else 0
            })
    
    # Por tipo de trabajo
    pipeline = [
        {"$match": query},
        {"$group": {
            "_id": "$tipo_trabajo",
            "count": {"$sum": 1},
            "horas": {"$sum": {"$ifNull": ["$horas_trabajadas", 0]}}
        }}
    ]
    
    por_tipo = {}
    async for doc in database.productividad.aggregate(pipeline):
        por_tipo[doc["_id"] or "otros"] = {
            "registros": doc["count"],
            "horas": doc["horas"]
        }
    
    return {
        "success": True,
        "periodo": {"desde": fecha_desde, "hasta": fecha_hasta},
        "totales": totales,
        "top_empleados": top_empleados,
        "por_tipo_trabajo": por_tipo
    }

@router.get("/productividad/tiempo-real")
async def get_productividad_tiempo_real():
    """Productividad en tiempo real (día actual)"""
    database = get_db()
    
    hoy = datetime.now().strftime("%Y-%m-%d")
    
    # Empleados trabajando hoy (con fichaje de entrada sin salida)
    pipeline = [
        {"$match": {"fecha": hoy}},
        {"$sort": {"hora": -1}},
        {"$group": {
            "_id": "$empleado_id",
            "ultimo_fichaje": {"$first": "$tipo"},
            "hora_entrada": {"$last": "$hora"}
        }},
        {"$match": {"ultimo_fichaje": "entrada"}}
    ]
    
    empleados_trabajando = []
    async for doc in database.fichajes.aggregate(pipeline):
        emp = await database.empleados.find_one({"_id": ObjectId(doc["_id"])})
        if emp:
            # Buscar productividad del día
            prod = await database.productividad.find_one({
                "empleado_id": doc["_id"],
                "fecha": hoy
            })
            
            empleados_trabajando.append({
                "empleado_id": doc["_id"],
                "empleado_nombre": f"{emp.get('nombre', '')} {emp.get('apellidos', '')}",
                "empleado_foto": emp.get("foto_url"),
                "puesto": emp.get("puesto"),
                "hora_entrada": doc["hora_entrada"],
                "kilos_hoy": prod.get("kilos_recogidos", 0) if prod else 0,
                "horas_hoy": prod.get("horas_trabajadas", 0) if prod else 0
            })
    
    # Totales del día
    pipeline = [
        {"$match": {"fecha": hoy}},
        {"$group": {
            "_id": None,
            "total_kilos": {"$sum": {"$ifNull": ["$kilos_recogidos", 0]}},
            "total_horas": {"$sum": {"$ifNull": ["$horas_trabajadas", 0]}}
        }}
    ]
    
    totales_hoy = {"total_kilos": 0, "total_horas": 0}
    async for doc in database.productividad.aggregate(pipeline):
        totales_hoy = {
            "total_kilos": doc.get("total_kilos", 0),
            "total_horas": doc.get("total_horas", 0)
        }
    
    return {
        "success": True,
        "fecha": hoy,
        "empleados_trabajando": empleados_trabajando,
        "total_empleados_trabajando": len(empleados_trabajando),
        "totales_hoy": totales_hoy
    }

@router.post("/productividad")
async def create_registro_productividad(registro: dict):
    """Crear registro de productividad"""
    database = get_db()
    
    # Calcular horas trabajadas
    if registro.get("hora_inicio") and registro.get("hora_fin"):
        h_inicio = datetime.strptime(registro["hora_inicio"], "%H:%M")
        h_fin = datetime.strptime(registro["hora_fin"], "%H:%M")
        diff = (h_fin - h_inicio).seconds / 3600
        descanso = registro.get("minutos_descanso", 0) / 60
        registro["horas_trabajadas"] = round(diff - descanso, 2)
    
    registro["created_at"] = datetime.now()
    registro["validado"] = False
    
    result = await database.productividad.insert_one(registro)
    registro["_id"] = str(result.inserted_id)
    
    # Crear notificación para el empleado
    empleado_id = registro.get("empleado_id")
    if empleado_id:
        empleado = await database.empleados.find_one({"_id": ObjectId(empleado_id)})
        if empleado and empleado.get("email"):
            kilos = registro.get("kilos", 0)
            tipo = registro.get("tipo_trabajo", "trabajo")
            
            # Crear notificación de productividad
            notificacion = {
                "tipo": "productividad",
                "titulo": "Nuevo registro de productividad",
                "mensaje": f"Se ha registrado tu actividad: {kilos} kg en {tipo}",
                "destinatarios": [empleado["email"]],
                "leida_por": [],
                "datos": {
                    "kilos": kilos,
                    "tipo_trabajo": tipo,
                    "fecha": registro.get("fecha", "")
                },
                "created_at": datetime.now()
            }
            await database.notificaciones.insert_one(notificacion)
    
    return {"success": True, "data": registro}

@router.put("/productividad/{registro_id}")
async def update_registro_productividad(registro_id: str, registro: dict):
    """Actualizar registro de productividad"""
    database = get_db()
    
    registro.pop("_id", None)
    
    # Recalcular horas si se cambiaron
    if registro.get("hora_inicio") and registro.get("hora_fin"):
        h_inicio = datetime.strptime(registro["hora_inicio"], "%H:%M")
        h_fin = datetime.strptime(registro["hora_fin"], "%H:%M")
        diff = (h_fin - h_inicio).seconds / 3600
        descanso = registro.get("minutos_descanso", 0) / 60
        registro["horas_trabajadas"] = round(diff - descanso, 2)
    
    result = await database.productividad.update_one(
        {"_id": ObjectId(registro_id)},
        {"$set": registro}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Registro no encontrado")
    
    return {"success": True}

@router.delete("/productividad/{registro_id}")
async def delete_registro_productividad(registro_id: str):
    """Eliminar registro de productividad"""
    database = get_db()
    
    result = await database.productividad.delete_one({"_id": ObjectId(registro_id)})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Registro no encontrado")
    
    return {"success": True}

# ============================================================================
# DOCUMENTOS
# ============================================================================

@router.get("/documentos")
async def get_documentos(
    empleado_id: Optional[str] = None, 
    tipo: Optional[str] = None,
    estado: Optional[str] = None,
    fecha_desde: Optional[str] = None,
    fecha_hasta: Optional[str] = None
):
    """Obtener documentos de empleados con filtros de fecha, tipo y estado"""
    database = get_db()
    
    query = {}
    if empleado_id:
        query["empleado_id"] = empleado_id
    if tipo:
        query["tipo"] = tipo
    
    # Filtro por estado de firma
    if estado:
        if estado == "firmado":
            query["firmado"] = True
        elif estado == "pendiente":
            query["requiere_firma"] = True
            query["firmado"] = False
        elif estado == "no_requiere":
            query["requiere_firma"] = False
    
    # Filtros de fecha de registro (created_at)
    if fecha_desde or fecha_hasta:
        query["created_at"] = {}
        if fecha_desde:
            query["created_at"]["$gte"] = datetime.strptime(fecha_desde, "%Y-%m-%d")
        if fecha_hasta:
            # Añadir un día para incluir todo el día hasta
            fecha_hasta_dt = datetime.strptime(fecha_hasta, "%Y-%m-%d") + timedelta(days=1)
            query["created_at"]["$lt"] = fecha_hasta_dt
    
    documentos = []
    cursor = database.documentos_empleados.find(query).sort("created_at", -1)
    
    async for doc in cursor:
        doc["_id"] = str(doc["_id"])
        # Convertir datetime a string ISO para JSON
        if doc.get("created_at"):
            doc["created_at"] = doc["created_at"].isoformat()
        if doc.get("updated_at"):
            doc["updated_at"] = doc["updated_at"].isoformat()
        documentos.append(doc)
    
    return {"success": True, "documentos": documentos, "total": len(documentos)}


@router.get("/documentos/export/excel")
async def export_documentos_excel(
    empleado_id: Optional[str] = None,
    fecha_desde: Optional[str] = None,
    fecha_hasta: Optional[str] = None,
    tipo: Optional[str] = None,
    estado: Optional[str] = None
):
    """Exportar documentos a Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from fastapi.responses import StreamingResponse
    
    database = get_db()
    
    # Construir query
    query = {}
    if empleado_id:
        query["empleado_id"] = empleado_id
    if tipo:
        query["tipo"] = tipo
    if estado:
        if estado == "firmado":
            query["firmado"] = True
        elif estado == "pendiente":
            query["requiere_firma"] = True
            query["firmado"] = False
        elif estado == "no_requiere":
            query["requiere_firma"] = False
    if fecha_desde or fecha_hasta:
        query["created_at"] = {}
        if fecha_desde:
            query["created_at"]["$gte"] = datetime.strptime(fecha_desde, "%Y-%m-%d")
        if fecha_hasta:
            fecha_hasta_dt = datetime.strptime(fecha_hasta, "%Y-%m-%d") + timedelta(days=1)
            query["created_at"]["$lt"] = fecha_hasta_dt
    
    # Obtener documentos
    documentos = []
    cursor = database.documentos_empleados.find(query).sort("created_at", -1)
    async for doc in cursor:
        documentos.append(doc)
    
    # Obtener empleados para nombres
    empleados_dict = {}
    emp_cursor = database.empleados.find({})
    async for emp in emp_cursor:
        empleados_dict[str(emp["_id"])] = f"{emp.get('nombre', '')} {emp.get('apellidos', '')}"
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Documentos"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ["Documento", "Empleado", "Tipo", "Fecha Documento", "Fecha Registro", "Estado", "Archivo Adjunto"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Tipos de documento
    tipos_doc = {
        'contrato': 'Contrato de Trabajo',
        'anexo': 'Anexo Contrato',
        'nomina': 'Nómina',
        'certificado': 'Certificado',
        'formacion': 'Formación PRL',
        'epi': 'Entrega EPI',
        'otro': 'Otro'
    }
    
    # Datos
    for row, doc in enumerate(documentos, 2):
        empleado_nombre = empleados_dict.get(doc.get("empleado_id", ""), "Desconocido")
        tipo_label = tipos_doc.get(doc.get("tipo", "otro"), doc.get("tipo", ""))
        fecha_doc = doc.get("fecha_creacion", "")
        fecha_registro = doc.get("created_at").strftime("%d/%m/%Y %H:%M") if doc.get("created_at") else ""
        estado = "Firmado" if doc.get("firmado") else ("Pendiente" if doc.get("requiere_firma") else "No requiere firma")
        archivo = "Sí" if doc.get("archivo_url") else "No"
        
        ws.cell(row=row, column=1, value=doc.get("nombre", "")).border = thin_border
        ws.cell(row=row, column=2, value=empleado_nombre).border = thin_border
        ws.cell(row=row, column=3, value=tipo_label).border = thin_border
        ws.cell(row=row, column=4, value=fecha_doc).border = thin_border
        ws.cell(row=row, column=5, value=fecha_registro).border = thin_border
        ws.cell(row=row, column=6, value=estado).border = thin_border
        ws.cell(row=row, column=7, value=archivo).border = thin_border
    
    # Ajustar anchos
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 15
    
    # Guardar en memoria
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"documentos_rrhh_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/documentos/export/pdf")
async def export_documentos_pdf(
    empleado_id: Optional[str] = None,
    fecha_desde: Optional[str] = None,
    fecha_hasta: Optional[str] = None,
    tipo: Optional[str] = None,
    estado: Optional[str] = None
):
    """Generar informe PDF de documentos"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from fastapi.responses import StreamingResponse
    
    database = get_db()
    
    # Construir query
    query = {}
    if empleado_id:
        query["empleado_id"] = empleado_id
    if tipo:
        query["tipo"] = tipo
    if estado:
        if estado == "firmado":
            query["firmado"] = True
        elif estado == "pendiente":
            query["requiere_firma"] = True
            query["firmado"] = False
        elif estado == "no_requiere":
            query["requiere_firma"] = False
    if fecha_desde or fecha_hasta:
        query["created_at"] = {}
        if fecha_desde:
            query["created_at"]["$gte"] = datetime.strptime(fecha_desde, "%Y-%m-%d")
        if fecha_hasta:
            fecha_hasta_dt = datetime.strptime(fecha_hasta, "%Y-%m-%d") + timedelta(days=1)
            query["created_at"]["$lt"] = fecha_hasta_dt
    
    # Obtener documentos
    documentos = []
    cursor = database.documentos_empleados.find(query).sort("created_at", -1)
    async for doc in cursor:
        documentos.append(doc)
    
    # Obtener empleados
    empleados_dict = {}
    emp_cursor = database.empleados.find({})
    async for emp in emp_cursor:
        empleados_dict[str(emp["_id"])] = f"{emp.get('nombre', '')} {emp.get('apellidos', '')}"
    
    # Crear PDF
    output = io.BytesIO()
    pdf_doc = SimpleDocTemplate(output, pagesize=landscape(A4), topMargin=20*mm, bottomMargin=20*mm)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#2E7D32'),
        spaceAfter=10*mm,
        alignment=1  # Center
    )
    elements.append(Paragraph("Informe de Documentos - RRHH", title_style))
    
    # Subtítulo con filtros
    subtitle = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    if fecha_desde:
        subtitle += f" | Desde: {fecha_desde}"
    if fecha_hasta:
        subtitle += f" | Hasta: {fecha_hasta}"
    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.gray,
        alignment=1
    )
    elements.append(Paragraph(subtitle, subtitle_style))
    elements.append(Spacer(1, 10*mm))
    
    # Resumen
    total_docs = len(documentos)
    firmados = sum(1 for d in documentos if d.get("firmado"))
    pendientes = sum(1 for d in documentos if d.get("requiere_firma") and not d.get("firmado"))
    con_archivo = sum(1 for d in documentos if d.get("archivo_url"))
    
    resumen_data = [
        ["Total Documentos", "Firmados", "Pendientes Firma", "Con Archivo"],
        [str(total_docs), str(firmados), str(pendientes), str(con_archivo)]
    ]
    resumen_table = Table(resumen_data, colWidths=[60*mm, 50*mm, 50*mm, 50*mm])
    resumen_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E7D32')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#E8F5E9')),
        ('FONTSIZE', (0, 1), (-1, -1), 12),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 10),
        ('TOPPADDING', (0, 1), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#2E7D32'))
    ]))
    elements.append(resumen_table)
    elements.append(Spacer(1, 10*mm))
    
    # Tipos de documento
    tipos_doc = {
        'contrato': 'Contrato',
        'anexo': 'Anexo',
        'nomina': 'Nómina',
        'certificado': 'Certificado',
        'formacion': 'Formación',
        'epi': 'EPI',
        'otro': 'Otro'
    }
    
    # Tabla de documentos
    if documentos:
        table_data = [["Documento", "Empleado", "Tipo", "Fecha Doc.", "Fecha Registro", "Estado"]]
        
        for doc in documentos[:50]:  # Limitar a 50 documentos para el PDF
            empleado_nombre = empleados_dict.get(doc.get("empleado_id", ""), "Desconocido")
            tipo_label = tipos_doc.get(doc.get("tipo", "otro"), doc.get("tipo", ""))
            fecha_doc = doc.get("fecha_creacion", "-")
            fecha_registro = doc.get("created_at").strftime("%d/%m/%Y") if doc.get("created_at") else "-"
            estado = "Firmado" if doc.get("firmado") else ("Pendiente" if doc.get("requiere_firma") else "OK")
            
            # Truncar nombre si es muy largo
            nombre_doc = doc.get("nombre", "")[:40]
            empleado_nombre = empleado_nombre[:25]
            
            table_data.append([nombre_doc, empleado_nombre, tipo_label, fecha_doc, fecha_registro, estado])
        
        doc_table = Table(table_data, colWidths=[70*mm, 55*mm, 35*mm, 30*mm, 35*mm, 25*mm])
        doc_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E7D32')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (3, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('TOPPADDING', (0, 0), (-1, 0), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
            ('TOPPADDING', (0, 1), (-1, -1), 4),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.gray),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')])
        ]))
        elements.append(doc_table)
        
        if len(documentos) > 50:
            elements.append(Spacer(1, 5*mm))
            elements.append(Paragraph(f"Mostrando 50 de {len(documentos)} documentos. Exportar a Excel para ver todos.", styles['Normal']))
    else:
        elements.append(Paragraph("No hay documentos con los filtros seleccionados.", styles['Normal']))
    
    pdf_doc.build(elements)
    output.seek(0)
    
    filename = f"informe_documentos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.post("/documentos")
async def create_documento(documento: dict):
    """Crear documento de empleado"""
    database = get_db()
    
    documento["created_at"] = datetime.now()
    documento["updated_at"] = datetime.now()
    documento["firmado"] = False
    documento["activo"] = True
    
    result = await database.documentos_empleados.insert_one(documento)
    documento["_id"] = str(result.inserted_id)
    
    return {"success": True, "data": documento}

@router.post("/documentos/upload")
async def upload_documento(
    file: UploadFile = File(...),
    empleado_id: str = Form(...),
    nombre: str = Form(...),
    tipo: str = Form("otro"),
    descripcion: str = Form(""),
    requiere_firma: str = Form("true"),
    fecha_creacion: str = Form(None)
):
    """Subir documento con archivo adjunto"""
    database = get_db()
    
    # Convertir requiere_firma a boolean
    requiere_firma_bool = requiere_firma.lower() in ('true', '1', 'yes')
    
    # Validar tipo de archivo
    allowed_extensions = ['.pdf', '.doc', '.docx', '.jpg', '.jpeg', '.png', '.gif']
    file_ext = os.path.splitext(file.filename)[1].lower()
    
    if file_ext not in allowed_extensions:
        raise HTTPException(
            status_code=400, 
            detail=f"Tipo de archivo no permitido. Use: {', '.join(allowed_extensions)}"
        )
    
    # Validar tamaño (10MB)
    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="El archivo es demasiado grande. Máximo 10MB.")
    
    # Crear directorio si no existe
    upload_dir = "/app/uploads/documentos_empleados"
    os.makedirs(upload_dir, exist_ok=True)
    
    # Generar nombre único
    unique_filename = f"{uuid.uuid4().hex}{file_ext}"
    file_path = os.path.join(upload_dir, unique_filename)
    
    # Guardar archivo
    with open(file_path, "wb") as f:
        f.write(content)
    
    # Crear documento en base de datos
    documento = {
        "empleado_id": empleado_id,
        "nombre": nombre,
        "tipo": tipo,
        "descripcion": descripcion,
        "requiere_firma": requiere_firma_bool,
        "fecha_creacion": fecha_creacion or datetime.now().strftime("%Y-%m-%d"),
        "archivo_url": f"/api/uploads/documentos_empleados/{unique_filename}",
        "archivo_nombre_original": file.filename,
        "archivo_tipo": file.content_type,
        "archivo_tamano": len(content),
        "firmado": False,
        "activo": True,
        "created_at": datetime.now(),
        "updated_at": datetime.now()
    }
    
    result = await database.documentos_empleados.insert_one(documento)
    documento["_id"] = str(result.inserted_id)
    
    # Crear notificación para el empleado
    empleado = await database.empleados.find_one({"_id": ObjectId(empleado_id)})
    if empleado and empleado.get("email"):
        titulo = "Nuevo Documento Disponible"
        empleado_nombre = f"{empleado.get('nombre', '')} {empleado.get('apellidos', '')}"
        
        if requiere_firma_bool:
            mensaje = f"Se ha subido el documento '{nombre}' que requiere tu firma."
            tipo_notif = "warning"
        else:
            mensaje = f"Se ha subido un nuevo documento: '{nombre}'."
            tipo_notif = "info"
        
        notificacion = {
            "titulo": titulo,
            "mensaje": mensaje,
            "tipo": tipo_notif,
            "enlace": "/portal-empleado",
            "destinatarios": [empleado.get("email")],
            "prioridad": "normal" if not requiere_firma_bool else "alta",
            "datos_extra": {"documento_id": str(result.inserted_id), "tipo": "documento"},
            "created_at": datetime.now(),
            "leida_por": []
        }
        await database.notificaciones.insert_one(notificacion)
        
        # Enviar email de notificación
        try:
            await send_documento_notification(
                recipient_email=empleado.get("email"),
                empleado_nombre=empleado_nombre,
                documento_nombre=nombre,
                tipo_documento=tipo,
                requiere_firma=requiere_firma_bool
            )
        except Exception as e:
            # Log error but don't fail the request
            print(f"Error sending documento email: {e}")
    
    return {"success": True, "data": documento}

@router.put("/documentos/{documento_id}/firmar")
async def firmar_documento(documento_id: str, firma_data: dict):
    """Firmar un documento digitalmente"""
    database = get_db()
    
    result = await database.documentos_empleados.update_one(
        {"_id": ObjectId(documento_id)},
        {"$set": {
            "firmado": True,
            "firma_empleado_url": firma_data.get("firma_url"),
            "fecha_firma": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "updated_at": datetime.now()
        }}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Documento no encontrado")
    
    return {"success": True}

@router.delete("/documentos/{documento_id}")
async def delete_documento(documento_id: str):
    """Eliminar documento"""
    database = get_db()
    
    result = await database.documentos_empleados.delete_one({"_id": ObjectId(documento_id)})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Documento no encontrado")
    
    return {"success": True}

# ============================================================================
# PRENÓMINA
# ============================================================================

@router.get("/prenominas")
async def get_prenominas(
    empleado_id: Optional[str] = None,
    mes: Optional[int] = None,
    ano: Optional[int] = None,
    estado: Optional[str] = None
):
    """Obtener prenóminas"""
    database = get_db()
    
    query = {}
    if empleado_id:
        query["empleado_id"] = empleado_id
    if mes:
        query["periodo_mes"] = mes
    if ano:
        query["periodo_ano"] = ano
    if estado:
        query["estado"] = estado
    
    prenominas = []
    cursor = database.prenominas.find(query).sort([("periodo_ano", -1), ("periodo_mes", -1)])
    
    async for p in cursor:
        p["_id"] = str(p["_id"])
        # Obtener nombre del empleado
        emp = await database.empleados.find_one({"_id": ObjectId(p["empleado_id"])})
        if emp:
            p["empleado_nombre"] = f"{emp.get('nombre', '')} {emp.get('apellidos', '')}"
            p["empleado_dni"] = emp.get("dni_nie")
        prenominas.append(p)
    
    return {"success": True, "prenominas": prenominas, "total": len(prenominas)}

@router.post("/prenominas/calcular")
async def calcular_prenomina(data: dict):
    """Calcular prenómina de un empleado para un periodo"""
    database = get_db()
    
    empleado_id = data.get("empleado_id")
    mes = data.get("mes")
    ano = data.get("ano")
    
    if not all([empleado_id, mes, ano]):
        raise HTTPException(status_code=400, detail="Se requiere empleado_id, mes y año")
    
    # Obtener empleado
    empleado = await database.empleados.find_one({"_id": ObjectId(empleado_id)})
    if not empleado:
        raise HTTPException(status_code=404, detail="Empleado no encontrado")
    
    # Calcular rango de fechas del mes
    fecha_desde = f"{ano}-{str(mes).zfill(2)}-01"
    if mes == 12:
        fecha_hasta = f"{ano + 1}-01-01"
    else:
        fecha_hasta = f"{ano}-{str(mes + 1).zfill(2)}-01"
    
    # Obtener fichajes del periodo
    fichajes = []
    cursor = database.fichajes.find({
        "empleado_id": empleado_id,
        "fecha": {"$gte": fecha_desde, "$lt": fecha_hasta}
    }).sort([("fecha", 1), ("hora", 1)])
    
    async for f in cursor:
        fichajes.append(f)
    
    # Calcular horas trabajadas por día
    horas_normales = 0.0
    horas_extra = 0.0
    horas_nocturnas = 0.0
    horas_festivos = 0.0
    dias_trabajados = set()
    
    # Agrupar fichajes por día
    fichajes_por_dia = {}
    for f in fichajes:
        fecha = f["fecha"]
        if fecha not in fichajes_por_dia:
            fichajes_por_dia[fecha] = []
        fichajes_por_dia[fecha].append(f)
    
    for fecha, fichajes_dia in fichajes_por_dia.items():
        entradas = [f for f in fichajes_dia if f["tipo"] == "entrada"]
        salidas = [f for f in fichajes_dia if f["tipo"] == "salida"]
        
        if entradas and salidas:
            dias_trabajados.add(fecha)
            
            # Calcular horas del día (simplificado)
            entrada = datetime.strptime(f"{fecha} {entradas[0]['hora']}", "%Y-%m-%d %H:%M:%S")
            salida = datetime.strptime(f"{fecha} {salidas[-1]['hora']}", "%Y-%m-%d %H:%M:%S")
            horas_dia = (salida - entrada).seconds / 3600
            
            # Por simplicidad, todo como horas normales (se puede mejorar)
            if horas_dia > 8:
                horas_normales += 8
                horas_extra += horas_dia - 8
            else:
                horas_normales += horas_dia
    
    # Calcular importes
    salario_hora = empleado.get("salario_hora", 0)
    salario_hora_extra = empleado.get("salario_hora_extra") or salario_hora * 1.25
    salario_hora_nocturna = empleado.get("salario_hora_nocturna") or salario_hora * 1.25
    salario_hora_festivo = empleado.get("salario_hora_festivo") or salario_hora * 1.5
    
    conceptos = []
    
    if horas_normales > 0:
        conceptos.append({
            "concepto": "horas_normales",
            "cantidad": round(horas_normales, 2),
            "precio_unitario": salario_hora,
            "importe": round(horas_normales * salario_hora, 2)
        })
    
    if horas_extra > 0:
        conceptos.append({
            "concepto": "horas_extra",
            "cantidad": round(horas_extra, 2),
            "precio_unitario": salario_hora_extra,
            "importe": round(horas_extra * salario_hora_extra, 2)
        })
    
    if horas_nocturnas > 0:
        conceptos.append({
            "concepto": "horas_nocturnas",
            "cantidad": round(horas_nocturnas, 2),
            "precio_unitario": salario_hora_nocturna,
            "importe": round(horas_nocturnas * salario_hora_nocturna, 2)
        })
    
    if horas_festivos > 0:
        conceptos.append({
            "concepto": "horas_festivos",
            "cantidad": round(horas_festivos, 2),
            "precio_unitario": salario_hora_festivo,
            "importe": round(horas_festivos * salario_hora_festivo, 2)
        })
    
    # Plus productividad (opcional)
    prod_cursor = database.productividad.find({
        "empleado_id": empleado_id,
        "fecha": {"$gte": fecha_desde, "$lt": fecha_hasta}
    })
    
    kilos_totales = 0.0
    async for p in prod_cursor:
        kilos_totales += p.get("kilos_recogidos", 0)
    
    importe_bruto = sum(c["importe"] for c in conceptos)
    
    # Crear o actualizar prenómina
    prenomina = {
        "empleado_id": empleado_id,
        "periodo_mes": mes,
        "periodo_ano": ano,
        "horas_normales": round(horas_normales, 2),
        "horas_extra": round(horas_extra, 2),
        "horas_nocturnas": round(horas_nocturnas, 2),
        "horas_festivos": round(horas_festivos, 2),
        "total_horas": round(horas_normales + horas_extra + horas_nocturnas + horas_festivos, 2),
        "dias_trabajados": len(dias_trabajados),
        "conceptos": conceptos,
        "importe_bruto": round(importe_bruto, 2),
        "deducciones": 0,
        "importe_neto": round(importe_bruto, 2),
        "kilos_totales": round(kilos_totales, 2),
        "estado": "borrador",
        "created_at": datetime.now(),
        "updated_at": datetime.now()
    }
    
    # Verificar si ya existe una prenómina para este periodo
    existing = await database.prenominas.find_one({
        "empleado_id": empleado_id,
        "periodo_mes": mes,
        "periodo_ano": ano
    })
    
    if existing:
        await database.prenominas.update_one(
            {"_id": existing["_id"]},
            {"$set": prenomina}
        )
        prenomina["_id"] = str(existing["_id"])
    else:
        result = await database.prenominas.insert_one(prenomina)
        prenomina["_id"] = str(result.inserted_id)
    
    prenomina["empleado_nombre"] = f"{empleado.get('nombre', '')} {empleado.get('apellidos', '')}"
    
    return {"success": True, "data": prenomina}

@router.post("/prenominas/calcular-todos")
async def calcular_prenominas_todos(data: dict):
    """Calcular prenóminas de todos los empleados activos para un periodo"""
    database = get_db()
    
    mes = data.get("mes")
    ano = data.get("ano")
    
    if not all([mes, ano]):
        raise HTTPException(status_code=400, detail="Se requiere mes y año")
    
    # Obtener todos los empleados activos
    empleados = []
    cursor = database.empleados.find({"activo": True})
    async for emp in cursor:
        empleados.append(emp)
    
    resultados = []
    for emp in empleados:
        try:
            # Calcular prenómina de cada empleado
            result = await calcular_prenomina({
                "empleado_id": str(emp["_id"]),
                "mes": mes,
                "ano": ano
            })
            resultados.append(result["data"])
        except Exception as e:
            resultados.append({
                "empleado_id": str(emp["_id"]),
                "empleado_nombre": f"{emp.get('nombre', '')} {emp.get('apellidos', '')}",
                "error": str(e)
            })
    
    return {"success": True, "prenominas": resultados, "total": len(resultados)}

@router.put("/prenominas/{prenomina_id}/validar")
async def validar_prenomina(prenomina_id: str, validador: dict):
    """Validar una prenómina"""
    database = get_db()
    
    result = await database.prenominas.update_one(
        {"_id": ObjectId(prenomina_id)},
        {"$set": {
            "estado": "validada",
            "validada_por": validador.get("validada_por"),
            "fecha_validacion": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "updated_at": datetime.now()
        }}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Prenómina no encontrada")
    
    return {"success": True}


@router.get("/prenominas/{prenomina_id}/excel")
async def export_prenomina_excel(prenomina_id: str):
    """Exportar prenómina individual a Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from fastapi.responses import StreamingResponse
    
    database = get_db()
    
    prenomina = await database.prenominas.find_one({"_id": ObjectId(prenomina_id)})
    if not prenomina:
        raise HTTPException(status_code=404, detail="Prenómina no encontrada")
    
    empleado = await database.empleados.find_one({"_id": ObjectId(prenomina["empleado_id"])})
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Prenómina"
    
    # Estilos
    title_font = Font(bold=True, size=14, color="FFFFFF")
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    money_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    ws.merge_cells('A1:E1')
    ws['A1'] = "PRENÓMINA"
    ws['A1'].font = title_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = Alignment(horizontal="center")
    
    # Info empleado
    ws['A3'] = "Empleado:"
    ws['B3'] = f"{empleado.get('nombre', '')} {empleado.get('apellidos', '')}" if empleado else "N/A"
    ws['A4'] = "DNI/NIE:"
    ws['B4'] = empleado.get('dni_nie', '') if empleado else "N/A"
    ws['A5'] = "Puesto:"
    ws['B5'] = empleado.get('puesto', '') if empleado else "N/A"
    ws['D3'] = "Período:"
    ws['E3'] = f"{prenomina.get('periodo_mes', '')}/{prenomina.get('periodo_ano', '')}"
    ws['D4'] = "Estado:"
    ws['E4'] = prenomina.get('estado', 'borrador').upper()
    
    # Resumen de horas
    ws.merge_cells('A7:E7')
    ws['A7'] = "RESUMEN DE HORAS"
    ws['A7'].font = header_font
    ws['A7'].fill = header_fill
    ws['A7'].alignment = Alignment(horizontal="center")
    
    horas_data = [
        ["Concepto", "Horas", "€/Hora", "Importe"],
        ["Horas Normales", prenomina.get('horas_normales', 0), empleado.get('salario_hora', 0) if empleado else 0, ""],
        ["Horas Extra", prenomina.get('horas_extra', 0), (empleado.get('salario_hora_extra') or (empleado.get('salario_hora', 0) * 1.25)) if empleado else 0, ""],
        ["Horas Nocturnas", prenomina.get('horas_nocturnas', 0), (empleado.get('salario_hora_nocturna') or (empleado.get('salario_hora', 0) * 1.25)) if empleado else 0, ""],
        ["Horas Festivos", prenomina.get('horas_festivos', 0), (empleado.get('salario_hora_festivo') or (empleado.get('salario_hora', 0) * 1.5)) if empleado else 0, ""]
    ]
    
    for row_idx, row_data in enumerate(horas_data, 8):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if row_idx == 8:
                cell.font = header_font
                cell.fill = header_fill
            if col_idx == 4 and row_idx > 8:
                # Calcular importe
                horas = ws.cell(row=row_idx, column=2).value or 0
                precio = ws.cell(row=row_idx, column=3).value or 0
                cell.value = round(horas * precio, 2)
                cell.number_format = '#,##0.00 €'
    
    # Totales
    row = 13
    ws.merge_cells(f'A{row}:C{row}')
    ws[f'A{row}'] = "TOTAL HORAS:"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'D{row}'] = prenomina.get('total_horas', 0)
    ws[f'D{row}'].font = Font(bold=True)
    
    row += 1
    ws.merge_cells(f'A{row}:C{row}')
    ws[f'A{row}'] = "DÍAS TRABAJADOS:"
    ws[f'D{row}'] = prenomina.get('dias_trabajados', 0)
    
    row += 2
    ws.merge_cells(f'A{row}:C{row}')
    ws[f'A{row}'] = "IMPORTE BRUTO:"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'D{row}'] = prenomina.get('importe_bruto', 0)
    ws[f'D{row}'].font = Font(bold=True, size=12)
    ws[f'D{row}'].fill = money_fill
    ws[f'D{row}'].number_format = '#,##0.00 €'
    
    row += 1
    ws.merge_cells(f'A{row}:C{row}')
    ws[f'A{row}'] = "DEDUCCIONES:"
    ws[f'D{row}'] = prenomina.get('deducciones', 0)
    ws[f'D{row}'].number_format = '#,##0.00 €'
    
    row += 1
    ws.merge_cells(f'A{row}:C{row}')
    ws[f'A{row}'] = "IMPORTE NETO:"
    ws[f'A{row}'].font = Font(bold=True, size=14, color="2E7D32")
    ws[f'D{row}'] = prenomina.get('importe_neto', 0)
    ws[f'D{row}'].font = Font(bold=True, size=14, color="2E7D32")
    ws[f'D{row}'].fill = money_fill
    ws[f'D{row}'].number_format = '#,##0.00 €'
    
    # Ajustar anchos
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    
    # Guardar
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    apellidos = empleado.get('apellidos', 'empleado').replace(' ', '_') if empleado else 'empleado'
    filename = f"prenomina_{apellidos}_{prenomina.get('periodo_mes', '')}_{prenomina.get('periodo_ano', '')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/prenominas/{prenomina_id}/pdf")
async def export_prenomina_pdf(prenomina_id: str):
    """Exportar prenómina individual a PDF"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from fastapi.responses import StreamingResponse
    
    database = get_db()
    
    prenomina = await database.prenominas.find_one({"_id": ObjectId(prenomina_id)})
    if not prenomina:
        raise HTTPException(status_code=404, detail="Prenómina no encontrada")
    
    empleado = await database.empleados.find_one({"_id": ObjectId(prenomina["empleado_id"])})
    
    # Crear PDF
    output = io.BytesIO()
    pdf = SimpleDocTemplate(output, pagesize=A4, topMargin=20*mm, bottomMargin=20*mm, leftMargin=20*mm, rightMargin=20*mm)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Título
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18, textColor=colors.HexColor('#2E7D32'), alignment=1, spaceAfter=5*mm)
    elements.append(Paragraph("PRENÓMINA", title_style))
    
    # Info empleado
    info_style = ParagraphStyle('Info', parent=styles['Normal'], fontSize=10, spaceAfter=2*mm)
    emp_nombre = f"{empleado.get('nombre', '')} {empleado.get('apellidos', '')}" if empleado else "N/A"
    elements.append(Paragraph(f"<b>Empleado:</b> {emp_nombre}", info_style))
    elements.append(Paragraph(f"<b>DNI/NIE:</b> {empleado.get('dni_nie', '') if empleado else 'N/A'} | <b>Puesto:</b> {empleado.get('puesto', '') if empleado else 'N/A'}", info_style))
    elements.append(Paragraph(f"<b>Período:</b> {prenomina.get('periodo_mes', '')}/{prenomina.get('periodo_ano', '')} | <b>Estado:</b> {prenomina.get('estado', 'borrador').upper()}", info_style))
    elements.append(Spacer(1, 8*mm))
    
    # Tabla de horas
    salario_hora = empleado.get('salario_hora', 0) if empleado else 0
    salario_extra = (empleado.get('salario_hora_extra') or salario_hora * 1.25) if empleado else 0
    salario_nocturna = (empleado.get('salario_hora_nocturna') or salario_hora * 1.25) if empleado else 0
    salario_festivo = (empleado.get('salario_hora_festivo') or salario_hora * 1.5) if empleado else 0
    
    table_data = [
        ["Concepto", "Horas", "€/Hora", "Importe"],
        ["Horas Normales", f"{prenomina.get('horas_normales', 0):.2f}", f"{salario_hora:.2f} €", f"{prenomina.get('horas_normales', 0) * salario_hora:.2f} €"],
        ["Horas Extra", f"{prenomina.get('horas_extra', 0):.2f}", f"{salario_extra:.2f} €", f"{prenomina.get('horas_extra', 0) * salario_extra:.2f} €"],
        ["Horas Nocturnas", f"{prenomina.get('horas_nocturnas', 0):.2f}", f"{salario_nocturna:.2f} €", f"{prenomina.get('horas_nocturnas', 0) * salario_nocturna:.2f} €"],
        ["Horas Festivos", f"{prenomina.get('horas_festivos', 0):.2f}", f"{salario_festivo:.2f} €", f"{prenomina.get('horas_festivos', 0) * salario_festivo:.2f} €"],
    ]
    
    col_widths = [50*mm, 30*mm, 35*mm, 40*mm]
    horas_table = Table(table_data, colWidths=col_widths)
    horas_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E7D32')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (0, 1), (0, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.gray),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')])
    ]))
    elements.append(horas_table)
    elements.append(Spacer(1, 8*mm))
    
    # Resumen
    resumen_data = [
        ["Total Horas:", f"{prenomina.get('total_horas', 0):.2f}"],
        ["Días Trabajados:", f"{prenomina.get('dias_trabajados', 0)}"],
        ["", ""],
        ["IMPORTE BRUTO:", f"{prenomina.get('importe_bruto', 0):.2f} €"],
        ["Deducciones:", f"{prenomina.get('deducciones', 0):.2f} €"],
        ["IMPORTE NETO:", f"{prenomina.get('importe_neto', 0):.2f} €"],
    ]
    
    resumen_table = Table(resumen_data, colWidths=[80*mm, 50*mm])
    resumen_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('FONTNAME', (0, 3), (-1, 3), 'Helvetica-Bold'),
        ('FONTNAME', (0, 5), (-1, 5), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 5), (-1, 5), 14),
        ('TEXTCOLOR', (0, 5), (-1, 5), colors.HexColor('#2E7D32')),
        ('BACKGROUND', (0, 5), (-1, 5), colors.HexColor('#E8F5E9')),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(resumen_table)
    
    # Fecha generación
    elements.append(Spacer(1, 15*mm))
    elements.append(Paragraph(f"<i>Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}</i>", ParagraphStyle('Small', fontSize=8, textColor=colors.gray)))
    
    pdf.build(elements)
    output.seek(0)
    
    apellidos = empleado.get('apellidos', 'empleado').replace(' ', '_') if empleado else 'empleado'
    filename = f"prenomina_{apellidos}_{prenomina.get('periodo_mes', '')}_{prenomina.get('periodo_ano', '')}.pdf"
    
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/prenominas/export")
async def export_prenominas(mes: int, ano: int):
    """Exportar prenóminas para software de nóminas (formato CSV)"""
    database = get_db()
    
    prenominas = []
    cursor = database.prenominas.find({
        "periodo_mes": mes,
        "periodo_ano": ano,
        "estado": {"$in": ["validada", "exportada"]}
    })
    
    async for p in cursor:
        emp = await database.empleados.find_one({"_id": ObjectId(p["empleado_id"])})
        if emp:
            prenominas.append({
                "codigo_empleado": emp.get("codigo"),
                "dni": emp.get("dni_nie"),
                "nombre": f"{emp.get('nombre', '')} {emp.get('apellidos', '')}",
                "horas_normales": p.get("horas_normales", 0),
                "horas_extra": p.get("horas_extra", 0),
                "horas_nocturnas": p.get("horas_nocturnas", 0),
                "horas_festivos": p.get("horas_festivos", 0),
                "total_horas": p.get("total_horas", 0),
                "dias_trabajados": p.get("dias_trabajados", 0),
                "importe_bruto": p.get("importe_bruto", 0),
                "deducciones": p.get("deducciones", 0),
                "importe_neto": p.get("importe_neto", 0)
            })
    
    # Marcar como exportadas
    await database.prenominas.update_many(
        {"periodo_mes": mes, "periodo_ano": ano, "estado": "validada"},
        {"$set": {"estado": "exportada", "updated_at": datetime.now()}}
    )
    
    return {"success": True, "prenominas": prenominas, "total": len(prenominas)}

# ============================================================================
# AUSENCIAS / VACACIONES
# ============================================================================

@router.get("/ausencias")
async def get_ausencias(
    empleado_id: Optional[str] = None,
    estado: Optional[str] = None,
    tipo: Optional[str] = None
):
    """Obtener ausencias"""
    database = get_db()
    
    query = {}
    if empleado_id:
        query["empleado_id"] = empleado_id
    if estado:
        query["estado"] = estado
    if tipo:
        query["tipo"] = tipo
    
    ausencias = []
    cursor = database.ausencias.find(query).sort("fecha_inicio", -1)
    
    async for a in cursor:
        a["_id"] = str(a["_id"])
        emp = await database.empleados.find_one({"_id": ObjectId(a["empleado_id"])})
        if emp:
            a["empleado_nombre"] = f"{emp.get('nombre', '')} {emp.get('apellidos', '')}"
        ausencias.append(a)
    
    return {"success": True, "ausencias": ausencias, "total": len(ausencias)}

@router.post("/ausencias")
async def create_ausencia(ausencia: dict):
    """Crear solicitud de ausencia"""
    database = get_db()
    
    # Calcular días totales
    fecha_inicio = datetime.strptime(ausencia["fecha_inicio"], "%Y-%m-%d")
    fecha_fin = datetime.strptime(ausencia["fecha_fin"], "%Y-%m-%d")
    ausencia["dias_totales"] = (fecha_fin - fecha_inicio).days + 1
    
    ausencia["estado"] = "pendiente"
    ausencia["created_at"] = datetime.now()
    
    result = await database.ausencias.insert_one(ausencia)
    ausencia["_id"] = str(result.inserted_id)
    
    return {"success": True, "data": ausencia}

@router.put("/ausencias/{ausencia_id}/aprobar")
async def aprobar_ausencia(ausencia_id: str, aprobador: dict):
    """Aprobar o rechazar una ausencia"""
    database = get_db()
    
    estado = aprobador.get("estado", "aprobada")  # aprobada o rechazada
    
    # Obtener la ausencia antes de actualizar para notificar
    ausencia = await database.ausencias.find_one({"_id": ObjectId(ausencia_id)})
    if not ausencia:
        raise HTTPException(status_code=404, detail="Ausencia no encontrada")
    
    result = await database.ausencias.update_one(
        {"_id": ObjectId(ausencia_id)},
        {"$set": {
            "estado": estado,
            "aprobada_por": aprobador.get("aprobada_por"),
            "fecha_aprobacion": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "comentario_aprobador": aprobador.get("comentario", "")
        }}
    )
    
    # Crear notificación para el empleado
    empleado = await database.empleados.find_one({"_id": ObjectId(ausencia["empleado_id"])})
    if empleado and empleado.get("email"):
        tipo_ausencia = ausencia.get("tipo", "ausencia").replace("_", " ").capitalize()
        fecha_inicio = ausencia.get("fecha_inicio", "")
        fecha_fin = ausencia.get("fecha_fin", "")
        empleado_nombre = f"{empleado.get('nombre', '')} {empleado.get('apellidos', '')}"
        comentario = aprobador.get("comentario", "")
        
        if estado == "aprobada":
            titulo = f"Solicitud de {tipo_ausencia} Aprobada"
            mensaje = f"Tu solicitud de {tipo_ausencia} del {fecha_inicio} al {fecha_fin} ha sido aprobada."
            tipo_notif = "success"
        else:
            titulo = f"Solicitud de {tipo_ausencia} Rechazada"
            mensaje = f"Tu solicitud de {tipo_ausencia} del {fecha_inicio} al {fecha_fin} ha sido rechazada."
            if comentario:
                mensaje += f" Motivo: {comentario}"
            tipo_notif = "warning"
        
        # Insertar notificación en BD
        notificacion = {
            "titulo": titulo,
            "mensaje": mensaje,
            "tipo": tipo_notif,
            "enlace": "/portal-empleado",
            "destinatarios": [empleado.get("email")],
            "prioridad": "alta",
            "datos_extra": {"ausencia_id": ausencia_id, "tipo": "ausencia"},
            "created_at": datetime.now(),
            "leida_por": []
        }
        await database.notificaciones.insert_one(notificacion)
        
        # Enviar email de notificación
        try:
            await send_ausencia_notification(
                recipient_email=empleado.get("email"),
                empleado_nombre=empleado_nombre,
                tipo_ausencia=ausencia.get("tipo", "ausencia"),
                fecha_inicio=fecha_inicio,
                fecha_fin=fecha_fin,
                estado=estado,
                comentario=comentario
            )
        except Exception as e:
            # Log error but don't fail the request
            print(f"Error sending ausencia email: {e}")
    
    return {"success": True}

@router.delete("/ausencias/{ausencia_id}")
async def delete_ausencia(ausencia_id: str):
    """Eliminar solicitud de ausencia"""
    database = get_db()
    
    result = await database.ausencias.delete_one({"_id": ObjectId(ausencia_id)})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Ausencia no encontrada")
    
    return {"success": True}

# ============================================================================
# PORTAL DEL EMPLEADO
# ============================================================================

@router.get("/portal/mi-perfil")
async def get_mi_perfil(empleado_id: str):
    """Obtener perfil del empleado para el portal"""
    database = get_db()
    
    empleado = await database.empleados.find_one({"_id": ObjectId(empleado_id)})
    if not empleado:
        raise HTTPException(status_code=404, detail="Empleado no encontrado")
    
    empleado["_id"] = str(empleado["_id"])
    
    # Ocultar datos sensibles
    empleado.pop("salario_hora", None)
    empleado.pop("salario_hora_extra", None)
    empleado.pop("iban", None)
    
    return {"success": True, "empleado": empleado}

@router.get("/portal/mis-fichajes")
async def get_mis_fichajes(empleado_id: str, mes: Optional[int] = None, ano: Optional[int] = None):
    """Obtener fichajes del empleado"""
    database = get_db()
    
    if not mes:
        mes = datetime.now().month
    if not ano:
        ano = datetime.now().year
    
    fecha_desde = f"{ano}-{str(mes).zfill(2)}-01"
    if mes == 12:
        fecha_hasta = f"{ano + 1}-01-01"
    else:
        fecha_hasta = f"{ano}-{str(mes + 1).zfill(2)}-01"
    
    fichajes = []
    cursor = database.fichajes.find({
        "empleado_id": empleado_id,
        "fecha": {"$gte": fecha_desde, "$lt": fecha_hasta}
    }).sort([("fecha", -1), ("hora", -1)])
    
    async for f in cursor:
        f["_id"] = str(f["_id"])
        fichajes.append(f)
    
    return {"success": True, "fichajes": fichajes, "periodo": {"mes": mes, "ano": ano}}

@router.get("/portal/mis-documentos")
async def get_mis_documentos(empleado_id: str):
    """Obtener documentos del empleado"""
    database = get_db()
    
    documentos = []
    cursor = database.documentos_empleados.find({
        "empleado_id": empleado_id,
        "activo": True
    }).sort("created_at", -1)
    
    async for doc in cursor:
        doc["_id"] = str(doc["_id"])
        documentos.append(doc)
    
    return {"success": True, "documentos": documentos}

@router.get("/portal/mis-ausencias")
async def get_mis_ausencias(empleado_id: str):
    """Obtener ausencias del empleado"""
    database = get_db()
    
    ausencias = []
    cursor = database.ausencias.find({"empleado_id": empleado_id}).sort("fecha_inicio", -1)
    
    async for a in cursor:
        a["_id"] = str(a["_id"])
        ausencias.append(a)
    
    return {"success": True, "ausencias": ausencias}
