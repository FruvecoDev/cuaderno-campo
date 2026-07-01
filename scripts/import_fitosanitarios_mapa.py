"""Importa productos fitosanitarios MAPA + sus usos autorizados desde el
Excel oficial (~60k filas).

Estructura Excel:
- Cada fila = combinación (producto + cultivo + plaga) con dosis específica.
- Múltiples filas comparten el mismo numero_registro (producto).

Estrategia:
1. Recorrer el Excel una vez, agrupando por numero_registro:
   - Crear/upsert 1 doc en `fitosanitarios` por numero_registro único (metadata general).
   - Bulk-insert de N docs en `fitosanitarios_usos` por producto.
2. Bulk-write en chunks de 5000 para performance.
"""
import sys
import time
sys.path.insert(0, '/app/backend')

from openpyxl import load_workbook
from pymongo import ReplaceOne, InsertOne
from bson import ObjectId

import asyncio
from database import db

fitosanitarios_collection = db['fitosanitarios']
fitosanitarios_usos_collection = db['fitosanitarios_usos']


HEADERS_IDX = {
    'numero_registro': 0,
    'nombre_comercial': 1,
    'denominacion_comun': 2,
    'empresa': 3,
    'tipo': 4,
    'materia_activa': 5,
    'dosis_min': 6,
    'dosis_max': 7,
    'unidad_dosis': 8,
    'volumen_agua_min': 9,
    'volumen_agua_max': 10,
    'plagas_objetivo': 11,
    'plazo_seguridad': 12,
    'observaciones': 13,
    'cultivo': 14,
    'codigo_cultivo': 15,
    'codigo_agente': 16,
    'aplicaciones': 17,
    'intervalo_aplicaciones': 18,
    'bbch': 19,
    'volumen_caldo': 20,
    'condicionamiento_especifico': 21,
    'estado': 22,
    'fecha_caducidad': 23,
}


def _v(row, key):
    idx = HEADERS_IDX.get(key)
    if idx is None or idx >= len(row):
        return None
    val = row[idx]
    if val is None or val == '':
        return None
    return val


def _f(v):
    """Cast a float o None"""
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


async def main(path='/tmp/fitos.xlsx'):
    t0 = time.time()
    print(f'Abriendo {path}...')
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb['Plantilla']
    print(f'Sheet Plantilla · rows≈{ws.max_row}')

    # Limpiar antes (opcional, para import limpio). Comentar si quieres delta.
    print('Limpiando colecciones fitosanitarios y fitosanitarios_usos...')
    d1 = await fitosanitarios_collection.delete_many({})
    d2 = await fitosanitarios_usos_collection.delete_many({})
    print(f'  Borrados: {d1.deleted_count} productos + {d2.deleted_count} usos')

    productos_creados = {}  # numero_registro → {_id, usos_count}
    productos_ops = []
    usos_ops = []
    processed = 0
    filas_omitidas = 0

    print('Procesando filas...')
    for row in ws.iter_rows(min_row=2, values_only=True):
        numero = _v(row, 'numero_registro')
        if not numero:
            filas_omitidas += 1
            continue
        numero = str(numero).strip()
        nombre = str(_v(row, 'nombre_comercial') or '').strip()

        # 1) Producto (metadata única por numero_registro)
        if numero not in productos_creados:
            producto_id = ObjectId()
            productos_creados[numero] = {'_id': producto_id, 'usos_count': 0}
            productos_ops.append(InsertOne({
                '_id': producto_id,
                'numero_registro': numero,
                'nombre_comercial': nombre,
                'denominacion_comun': _v(row, 'denominacion_comun'),
                'empresa': _v(row, 'empresa'),
                'tipo': _v(row, 'tipo') or 'Otro',
                'materia_activa': _v(row, 'materia_activa'),
                'estado': _v(row, 'estado') or 'Vigente',
                'fecha_caducidad': str(_v(row, 'fecha_caducidad') or '') or None,
                'observaciones': _v(row, 'observaciones'),
                'activo': (str(_v(row, 'estado') or 'Vigente').lower() == 'vigente'),
                'usos_count': 0,  # se actualiza al final
            }))
        producto_id = productos_creados[numero]['_id']
        productos_creados[numero]['usos_count'] += 1

        # 2) Uso (combinación producto + cultivo + plaga con dosis específica)
        usos_ops.append(InsertOne({
            'fitosanitario_id': str(producto_id),
            'numero_registro': numero,
            'nombre_comercial': nombre,
            'cultivo': str(_v(row, 'cultivo') or '').strip() or None,
            'codigo_cultivo': str(_v(row, 'codigo_cultivo') or '').strip() or None,
            'plaga': str(_v(row, 'plagas_objetivo') or '').strip() or None,
            'codigo_agente': str(_v(row, 'codigo_agente') or '').strip() or None,
            'dosis_min': _f(_v(row, 'dosis_min')),
            'dosis_max': _f(_v(row, 'dosis_max')),
            'unidad_dosis': _v(row, 'unidad_dosis'),
            'volumen_agua_min': _f(_v(row, 'volumen_agua_min')),
            'volumen_agua_max': _f(_v(row, 'volumen_agua_max')),
            'volumen_caldo': _v(row, 'volumen_caldo'),
            'plazo_seguridad': str(_v(row, 'plazo_seguridad') or '').strip() or None,
            'bbch': str(_v(row, 'bbch') or '').strip() or None,
            'aplicaciones': str(_v(row, 'aplicaciones') or '').strip() or None,
            'intervalo_aplicaciones': str(_v(row, 'intervalo_aplicaciones') or '').strip() or None,
            'condicionamiento_especifico': _v(row, 'condicionamiento_especifico'),
        }))

        processed += 1

        # Flush por lotes cada 5000 usos
        if len(usos_ops) >= 5000:
            await fitosanitarios_usos_collection.bulk_write(usos_ops, ordered=False)
            usos_ops = []
            if productos_ops:
                await fitosanitarios_collection.bulk_write(productos_ops, ordered=False)
                productos_ops = []
            print(f'  {processed} filas procesadas... productos={len(productos_creados)}, elapsed={time.time()-t0:.1f}s')

    # Flush final
    if usos_ops:
        await fitosanitarios_usos_collection.bulk_write(usos_ops, ordered=False)
    if productos_ops:
        await fitosanitarios_collection.bulk_write(productos_ops, ordered=False)

    # 3) Actualizar usos_count final por producto (usando aggregation)
    print('Actualizando usos_count por producto...')
    for numero, meta in productos_creados.items():
        await fitosanitarios_collection.update_one(
            {'_id': meta['_id']},
            {'$set': {'usos_count': meta['usos_count']}},
        )

    # 4) Crear índices para búsquedas rápidas
    print('Creando índices...')
    await fitosanitarios_usos_collection.create_index([('fitosanitario_id', 1)])
    await fitosanitarios_usos_collection.create_index([('numero_registro', 1)])
    await fitosanitarios_usos_collection.create_index([('cultivo', 1), ('plaga', 1)])
    await fitosanitarios_collection.create_index([('numero_registro', 1)], unique=False)
    await fitosanitarios_collection.create_index([('nombre_comercial', 1)])
    await fitosanitarios_collection.create_index([('tipo', 1)])

    print(f'\n=== TERMINADO ===')
    print(f'Productos únicos:    {len(productos_creados)}')
    print(f'Usos totales:        {processed}')
    print(f'Filas sin registro:  {filas_omitidas}')
    print(f'Tiempo total:        {time.time()-t0:.1f}s')


if __name__ == '__main__':
    asyncio.run(main())
